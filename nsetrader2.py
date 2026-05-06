#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
╔══════════════════════════════════════════════════════════════════════════════════════╗
║  NSE SWING TRADER  v10.0  ·  UNIFIED SINGLE FILE                                    ║
║  Terminal Output  +  Streamlit Dashboard  —  ONE file, TWO modes                    ║
╠══════════════════════════════════════════════════════════════════════════════════════╣
║  INSTALL  :  pip install yfinance pandas numpy rich streamlit plotly                 ║
║                                                                                      ║
║  ── TERMINAL MODE ──────────────────────────────────────────────────────────────────║
║     python nse_trader.py                     # full live scan, Rich output          ║
║     python nse_trader.py --sample            # offline demo                         ║
║     python nse_trader.py --group "FO STOCKS"                                        ║
║     python nse_trader.py --threshold 0.18    # bear market mode                     ║
║     python nse_trader.py --capital 2000000   # ₹20L portfolio                       ║
║                                                                                      ║
║  ── STREAMLIT DASHBOARD ─────────────────────────────────────────────────────────── ║
║     streamlit run nse_trader.py                                                      ║
║                                                                                      ║
║  DISCLAIMER : Research / educational use only.  Not financial advice.               ║
╚══════════════════════════════════════════════════════════════════════════════════════╝
"""

from __future__ import annotations

# ─────────────────────────────────────────────────────────────────────────────
# DETECT MODE  (must happen before any Streamlit calls)
# ─────────────────────────────────────────────────────────────────────────────
def _is_streamlit() -> bool:
    """True when this script is executed by the Streamlit runner."""
    try:
        from streamlit.runtime.scriptrunner import get_script_run_ctx
        return get_script_run_ctx() is not None
    except Exception:
        pass
    try:                                    # older Streamlit versions
        import streamlit.runtime.scriptrunner as _sr
        return _sr.get_script_run_ctx() is not None
    except Exception:
        return False

_STREAMLIT = _is_streamlit()

# ─────────────────────────────────────────────────────────────────────────────
# STREAMLIT PAGE CONFIG  — MUST be the very first st.* call
# ─────────────────────────────────────────────────────────────────────────────
if _STREAMLIT:
    import streamlit as st
    st.set_page_config(
        page_title="NSE Swing Trader Pro",
        page_icon="📈",
        layout="wide",
        initial_sidebar_state="expanded",
    )

def _safe_int(v, default=0):
    """Convert float to int safely, returning default for NaN/None/inf."""
    try:
        if v is None: return default
        f = float(v)
        if f != f or f == float('inf') or f == float('-inf'):  # NaN or inf check
            return default
        return int(f)
    except (TypeError, ValueError):
        return default

def _safe_float(v, default=0.0):
    """Convert to float safely, returning default for NaN/None/inf."""
    try:
        if v is None: return default
        f = float(v)
        if f != f or f == float('inf') or f == float('-inf'):
            return default
        return f
    except (TypeError, ValueError):
        return default


# ─────────────────────────────────────────────────────────────────────────────
# SHARED STDLIB
# ─────────────────────────────────────────────────────────────────────────────
import argparse
import contextlib
import io
import json
import logging
import os
import sys
import time
import warnings
from collections import defaultdict
from dataclasses import dataclass, field
from datetime import datetime
from math import sqrt
from pathlib import Path
from typing import Optional

import numpy as np
import pandas as pd

warnings.filterwarnings("ignore")

# ─────────────────────────────────────────────────────────────────────────────
# OPTIONAL HEAVY DEPS
# ─────────────────────────────────────────────────────────────────────────────
try:
    import yfinance as yf
    _HAS_YF = True
except ImportError:
    yf = None; _HAS_YF = False

_HAS_TA = False
try:
    with contextlib.redirect_stdout(io.StringIO()), contextlib.redirect_stderr(io.StringIO()):
        import pandas_ta as _pta
    _HAS_TA = True
except Exception:
    _pta = None

# Rich (terminal mode only)
_HAS_RICH = False
_con = None
if not _STREAMLIT:
    try:
        from rich.align   import Align
        from rich.columns import Columns
        from rich.console import Console
        from rich         import box as rbox
        from rich.padding import Padding
        from rich.panel   import Panel
        from rich.rule    import Rule
        from rich.table   import Table
        from rich.text    import Text
        from rich.progress import (Progress, SpinnerColumn, TextColumn,
                                   BarColumn, TimeElapsedColumn)
        _HAS_RICH = True
        _con = Console(highlight=False)
    except ImportError:
        pass

# Streamlit + Plotly (dashboard mode only)
if _STREAMLIT:
    import plotly.graph_objects as go
    import plotly.express as px
    from plotly.subplots import make_subplots

LOG = logging.getLogger("NSEv10")
for _nm in ("yfinance","peewee","urllib3","requests","charset_normalizer"):
    logging.getLogger(_nm).setLevel(logging.CRITICAL)


# ══════════════════════════════════════════════════════════════════════════════
# §1  UNIVERSE
# ══════════════════════════════════════════════════════════════════════════════

_YF_OVERRIDE: dict[str,str] = {
    "M&M":       "M%26M",
    "BAJAJ-AUTO":"BAJAJ-AUTO",
    "LTM":       "LTIM",
}

_SKIP_SYMBOLS: set[str] = {
    "APL","ETERNAL","JIOFIN","PATANJALI","ASHOKLEY",
}

_UNIVERSE: dict[str,list[str]] = {
    "NIFTY 50 LEADERS": [
        "RELIANCE","HDFCBANK","ICICIBANK","TCS","INFY","SBIN","BHARTIARTL","LT",
        "AXISBANK","KOTAKBANK","ITC","HINDUNILVR","BAJFINANCE","SUNPHARMA","TITAN",
        "MARUTI","M&M","NTPC","POWERGRID","ADANIPORTS",
    ],
    "NIFTY 50": [
        "RELIANCE","HDFCBANK","BHARTIARTL","SBIN","ICICIBANK","TCS","BAJFINANCE",
        "INFY","HINDUNILVR","LT","SUNPHARMA","MARUTI","M&M","AXISBANK","ITC",
        "KOTAKBANK","NTPC","TITAN","HCLTECH","ONGC","ULTRACEMCO","BEL","ADANIPORTS",
        "COALINDIA","JSWSTEEL","POWERGRID","BAJAJFINSV","ADANIENT","BAJAJ-AUTO",
        "TATASTEEL","NESTLEIND","ASIANPAINT","HINDALCO","WIPRO","SBILIFE","EICHERMOT",
        "SHRIRAMFIN","GRASIM","INDIGO","HDFCLIFE","TECHM","TRENT","TATAMOTORS",
        "DRREDDY","APOLLOHOSP","TATACONSUM","CIPLA","MAXHEALTH",
    ],
    "NIFTY NEXT 50": [
        "LICI","ADANIGREEN","ADANIPOWER","VEDL","HAL","SIEMENS","GODREJCP","DABUR",
        "PIDILITIND","DMART","MARICO","BRITANNIA","HAVELLS","AMBUJACEM","GAIL","BHEL",
        "SAIL","BPCL","HINDPETRO","IOC","PETRONET","CONCOR","NMDC","RECLTD","PFC",
        "IRFC","IREDA","RVNL","NHPC","SUZLON","TATAPOWER","JSWENERGY","POLYCAB",
        "CUMMINSIND","VOLTAS","DLF","LODHA","GODREJPROP","OBEROIRLTY","PRESTIGE",
        "PHOENIXLTD","INDHOTEL","JUBLFOOD","NAUKRI","MPHASIS","COFORGE","PERSISTENT",
        "KPITTECH","LTM",
    ],
    "NIFTY MIDCAP 100": [
        "TVSMOTOR","CHOLAFIN","MUTHOOTFIN","LUPIN","AUROPHARMA","DIVISLAB","ALKEM",
        "TORNTPHARM","BIOCON","GLENMARK","MANKIND","ZYDUSLIFE","LAURUSLABS","FORTIS",
        "SYNGENE","AUBANK","FEDERALBNK","BANDHANBNK","RBLBANK","IDFCFIRSTB","PNB",
        "BANKBARODA","CANBK","INDIANB","UNIONBANK","INDUSINDBK","SRF","ASTRAL",
        "CROMPTON","BLUESTARCO","KEI","ABB","BHARATFORG","BDL","TIINDIA","SONACOMS",
        "UNOMINDA","EXIDEIND","KALYANKJIL","PAGEIND","VBL","MCX","BSE","CDSL","CAMS",
        "HDFCAMC","KFINTECH","ANGELONE","NUVAMA","POLICYBZR","DIXON","AMBER","KAYNES",
        "DALBHARAT","SHREECEM","JKCEMENT","HINDZINC","NATIONALUM","JINDALSTEL",
        "APLAPOLLO","GMRAIRPORT","DELHIVERY",
    ],
    "NIFTY SMALLCAP 250": [
        "IREDA","RVNL","IRFC","NHPC","HUDCO","SJVN","NBCC","PNBHOUSING","LICHSGFIN",
        "MANAPPURAM","ABCAPITAL","LTF","TATAELXSI","OFSS","INOXWIND","WAAREEENER",
        "TATAPOWER","TORNTPOWER","OIL","ZOMATO","NYKAA","PAYTM","COLPAL","EMAMILTD",
        "PIIND","UPL","DEEPAKNTR","SUPREMEIND","SOLARINDS","MAZDOCK","BOSCHLTD","MOTHERSON",
    ],
    "NIFTY BANK": [
        "HDFCBANK","ICICIBANK","SBIN","AXISBANK","KOTAKBANK","INDUSINDBK","BANDHANBNK",
        "FEDERALBNK","AUBANK","IDFCFIRSTB","PNB","BANKBARODA","CANBK","INDIANB",
        "UNIONBANK","RBLBANK",
    ],
    "NIFTY IT": [
        "TCS","INFY","HCLTECH","WIPRO","TECHM","LTM","MPHASIS","COFORGE","PERSISTENT",
        "KPITTECH","TATAELXSI","OFSS","NAUKRI",
    ],
    "NIFTY ENERGY": [
        "RELIANCE","ONGC","NTPC","POWERGRID","TATAPOWER","ADANIGREEN","ADANIPOWER",
        "JSWENERGY","NHPC","IREDA","SUZLON","INOXWIND","WAAREEENER","TORNTPOWER",
        "BPCL","IOC","HINDPETRO","OIL","GAIL","PETRONET","COALINDIA",
    ],
    "NIFTY AUTO": [
        "MARUTI","M&M","TATAMOTORS","BAJAJ-AUTO","EICHERMOT","HEROMOTOCO","TVSMOTOR",
        "MOTHERSON","BOSCHLTD","TIINDIA","SONACOMS","UNOMINDA","EXIDEIND","BHARATFORG",
    ],
    "NIFTY INFRA": [
        "LT","ADANIPORTS","POWERGRID","NTPC","COALINDIA","BHEL","SIEMENS","ABB",
        "HAVELLS","POLYCAB","KEI","CUMMINSIND","RVNL","NBCC","HUDCO","IRFC",
        "GMRAIRPORT","CONCOR","DELHIVERY","DLF","LODHA","GODREJPROP","OBEROIRLTY",
        "PRESTIGE","PHOENIXLTD",
    ],
    "FO STOCKS": [
        "RELIANCE","HDFCBANK","ICICIBANK","SBIN","TCS","INFY","AXISBANK","KOTAKBANK",
        "LT","BAJFINANCE","WIPRO","HCLTECH","SUNPHARMA","MARUTI","M&M","ITC","TITAN",
        "BHARTIARTL","ADANIPORTS","ADANIENT","BAJAJ-AUTO","BAJAJFINSV","NTPC","POWERGRID",
        "COALINDIA","ONGC","JSWSTEEL","TATASTEEL","HINDALCO","GRASIM","NESTLEIND",
        "ASIANPAINT","HINDUNILVR","TRENT","TATAMOTORS","TATACONSUM","DRREDDY","CIPLA",
        "EICHERMOT","SHRIRAMFIN","TECHM","INDUSINDBK","ULTRACEMCO","DIVISLAB","BEL",
        "HDFCLIFE","SBILIFE","MAXHEALTH","APOLLOHOSP","INDIGO","TATAPOWER","RECLTD",
        "PFC","IRFC","IREDA","RVNL","NHPC","SUZLON","JSWENERGY","ADANIGREEN",
        "WAAREEENER","BPCL","IOC","GAIL","PETRONET","TVSMOTOR","HEROMOTOCO","BOSCHLTD",
        "CHOLAFIN","MUTHOOTFIN","AUBANK","FEDERALBNK","BANDHANBNK","RBLBANK",
        "IDFCFIRSTB","PNB","BANKBARODA","CANBK","UNIONBANK","LTM","MPHASIS","COFORGE",
        "PERSISTENT","KPITTECH","TATAELXSI","OFSS","NAUKRI","DLF","LODHA","GODREJPROP",
        "OBEROIRLTY","PRESTIGE","PHOENIXLTD","POLYCAB","HAVELLS","SIEMENS","ABB",
        "CUMMINSIND","BHEL","AMBUJACEM","DMART","PIDILITIND","MARICO","DABUR",
        "BRITANNIA","COLPAL","GODREJCP","VBL","JUBLFOOD","HDFCAMC","CDSL","BSE","CAMS",
        "ANGELONE","MCX","NUVAMA","DIXON","KAYNES","AMBER","HAL","BDL","MAZDOCK",
        "SOLARINDS","JINDALSTEL","SAIL","NMDC","HINDZINC","NATIONALUM","VEDL","CONCOR",
        "DELHIVERY","GMRAIRPORT","INDHOTEL","BHARATFORG","TIINDIA","EXIDEIND","UNOMINDA",
    ],
}

_FO_SET   = set(_UNIVERSE["FO STOCKS"])
_ALL_SYMS = sorted({s for v in _UNIVERSE.values() for s in v} - _SKIP_SYMBOLS)

_SYM_GROUPS: dict[str,list[str]] = defaultdict(list)
for _g, _sl in _UNIVERSE.items():
    for _s in _sl:
        if _s not in _SKIP_SYMBOLS:
            _SYM_GROUPS[_s].append(_g)

_GRP_SHORT = {
    "NIFTY 50 LEADERS":"N50L","NIFTY 50":"N50","NIFTY NEXT 50":"NN50",
    "NIFTY MIDCAP 100":"MC100","NIFTY SMALLCAP 250":"SC250",
    "NIFTY BANK":"BNK","NIFTY IT":"IT","NIFTY ENERGY":"NRG",
    "NIFTY AUTO":"AUTO","NIFTY INFRA":"INFRA","FO STOCKS":"F&O",
}

def symbol_tags(sym: str) -> str:
    tags = list(dict.fromkeys(_GRP_SHORT.get(g, g[:4]) for g in _SYM_GROUPS.get(sym,[])))
    return " · ".join(tags[:4]) or "—"

def yf_ticker(sym: str) -> str:
    return f"{_YF_OVERRIDE.get(sym, sym)}.NS"

# ══════════════════════════════════════════════════════════════════════════════
# §2  CONFIG
# ══════════════════════════════════════════════════════════════════════════════

@dataclass
class Cfg:
    live_period:        str   = "8mo"
    live_interval:      str   = "1d"
    output_dir:         Path  = Path("nse_v10_output")
    use_sample:         bool  = False
    prices_csv:         Path  = Path("data/prices.csv")
    fetch_fundamentals: bool  = True
    symbols:            list  = field(default_factory=list)
    min_avg_vol:        int   = 750_000
    min_price:          float = 30.0
    min_traded_val_cr:  float = 2.0
    top_n:              int   = 10
    ema_spans:          tuple = (9, 21, 50, 200)
    rsi_period:         int   = 14
    atr_period:         int   = 14
    bb_period:          int   = 20
    adx_period:         int   = 14
    breakout_window:    int   = 20
    min_bars:           int   = 60
    base_threshold:     float = 0.16
    bear_threshold:     float = 0.30
    min_categories:     int   = 2
    weights: dict       = field(default_factory=lambda: {
        "trend":0.24,"momentum":0.16,"breakout":0.17,"pullback":0.11,
        "volume":0.10,"pattern":0.10,"fundamental":0.08,"sentiment":0.04,
    })
    min_atr_pct:        float = 0.012
    max_atr_pct:        float = 0.09
    st_sl_mult:         float = 1.0
    st_tp_mult:         float = 1.8
    lt_sl_mult:         float = 1.5
    lt_tp_mult:         float = 3.5
    min_rr:             float = 1.2
    bt_capital:         float = 1_000_000.0
    bt_max_pos:         int   = 5
    bt_pos_pct:         float = 0.20
    bt_sl_pct:          float = 0.04
    bt_tp_pct:          float = 0.09
    bt_max_hold:        int   = 12
    bt_min_hold:        int   = 2
    bt_cost_bps:        float = 12.0
    bt_slip_bps:        float = 5.0
    # display
    capital:            float = 1_000_000.0

# ══════════════════════════════════════════════════════════════════════════════
# §3  DATA LAYER
# ══════════════════════════════════════════════════════════════════════════════

_NIFTY_CACHE: dict = {}

def _norm_dates(s: pd.Series) -> pd.Series:
    p = pd.to_datetime(s)
    return p.dt.tz_convert(None) if getattr(p.dt,"tz",None) is not None else p

def _safe_dl(ticker: str, period: str, interval: str) -> pd.DataFrame:
    if not _HAS_YF: return pd.DataFrame()
    try:
        buf = io.StringIO()
        with contextlib.redirect_stdout(buf), contextlib.redirect_stderr(buf):
            df = yf.download(ticker, period=period, interval=interval,
                             progress=False, auto_adjust=True, timeout=20)
        return df if df is not None else pd.DataFrame()
    except Exception:
        return pd.DataFrame()

def nifty50_state() -> dict:
    global _NIFTY_CACHE
    c = _NIFTY_CACHE
    if c.get("ts") and (datetime.now()-c["ts"]).seconds < 3600:
        return c
    df = _safe_dl("^NSEI","4mo","1d")
    if df is None or df.empty: return {}
    try:
        if isinstance(df.columns, pd.MultiIndex): df.columns = df.columns.get_level_values(0)
        df.columns = [str(x).lower().strip() for x in df.columns]
        cl = df["close"].dropna()
        if len(cl)<10: return {}
        e9=cl.ewm(span=9,adjust=False).mean(); e21=cl.ewm(span=21,adjust=False).mean()
        e50=cl.ewm(span=50,adjust=False).mean()
        rsi_val=float(_rsi_s(cl,14).iloc[-1]); last=float(cl.iloc[-1])
        l9=float(e9.iloc[-1]); l21=float(e21.iloc[-1]); l50=float(e50.iloc[-1])
        if   last>l9>l21>l50:  trend,lbl=1.0,"🐂 Strong Bull"
        elif last>l9>l21:       trend,lbl=0.7,"📈 Mild Bull"
        elif last<l9<l21<l50:   trend,lbl=-1.0,"🐻 Strong Bear"
        elif last<l9<l21:       trend,lbl=-0.7,"📉 Mild Bear"
        else:                    trend,lbl=0.0,"↔️  Sideways"
        chg1m=float((cl.iloc[-1]/cl.iloc[-21]-1)*100) if len(cl)>=21 else 0.0
        chg3m=float((cl.iloc[-1]/cl.iloc[0]-1)*100)
        _NIFTY_CACHE=dict(ts=datetime.now(),trend=trend,label=lbl,
                          rsi=round(rsi_val,1),chg_1m=round(chg1m,2),chg_3m=round(chg3m,2),
                          last=round(last,2),ema9=round(l9,2),ema21=round(l21,2),ema50=round(l50,2))
        return _NIFTY_CACHE
    except Exception as e:
        LOG.debug("Nifty: %s",e); return {}

def fetch_ohlcv(sym: str, period: str, interval: str) -> pd.DataFrame:
    df = _safe_dl(yf_ticker(sym), period, interval)
    if df is None or df.empty: return pd.DataFrame()
    try:
        if isinstance(df.columns,pd.MultiIndex): df.columns=df.columns.get_level_values(0)
        df.columns=[str(x).lower().strip() for x in df.columns]
        df=df.reset_index()
        dc=next((c for c in df.columns if c.lower() in {"date","datetime"}),df.columns[0])
        df=df.rename(columns={dc:"date"})
        df["date"]=_norm_dates(df["date"]).dt.normalize()
        df["symbol"]=sym.upper()
        df["volume"]=pd.to_numeric(df.get("volume",0),errors="coerce").fillna(0)
        need=["date","symbol","open","high","low","close","volume"]
        if any(c not in df.columns for c in need): return pd.DataFrame()
        return df[need].dropna(subset=["open","high","low","close"]).reset_index(drop=True)
    except Exception as e:
        LOG.debug("OHLCV %s: %s",sym,e); return pd.DataFrame()

def fetch_fundamentals(sym: str) -> dict:
    if not _HAS_YF: return {}
    try:
        buf=io.StringIO()
        with contextlib.redirect_stdout(buf),contextlib.redirect_stderr(buf):
            info=yf.Ticker(yf_ticker(sym)).info
        mc=info.get("marketCap",0) or 0
        return dict(pe=info.get("trailingPE"),pb=info.get("priceToBook"),
                    roe=info.get("returnOnEquity"),eps_g=info.get("earningsGrowth"),
                    rev_g=info.get("revenueGrowth"),de=info.get("debtToEquity"),
                    sector=info.get("sector","N/A"),indust=info.get("industry","N/A"),
                    mcap=round(mc/1e7,1) if mc else None,
                    w52h=info.get("fiftyTwoWeekHigh"),w52l=info.get("fiftyTwoWeekLow"),
                    beta=info.get("beta"),peg=info.get("pegRatio"),
                    div_y=info.get("dividendYield"))
    except Exception: return {}

def sample_ohlcv(syms: list, start: pd.Timestamp, end: pd.Timestamp, seed:int=42) -> pd.DataFrame:
    rng=np.random.default_rng(seed); dates=pd.bdate_range(start,end)
    profiles={"RELIANCE":{"p":2850,"d":0.0010,"v":0.018,"vol":8_000_000},
               "TCS":{"p":4180,"d":0.0006,"v":0.013,"vol":2_500_000},
               "INFY":{"p":1625,"d":0.0007,"v":0.016,"vol":6_000_000},
               "HDFCBANK":{"p":1540,"d":0.0005,"v":0.012,"vol":9_000_000},
               "ICICIBANK":{"p":1125,"d":0.0008,"v":0.017,"vol":10_000_000},
               "SBIN":{"p":800,"d":0.0009,"v":0.020,"vol":15_000_000},
               "BAJFINANCE":{"p":7200,"d":0.0007,"v":0.019,"vol":2_000_000},
               "LT":{"p":3800,"d":0.0008,"v":0.016,"vol":3_000_000},
               "BHARTIARTL":{"p":1700,"d":0.0009,"v":0.015,"vol":5_000_000},
               "KOTAKBANK":{"p":2000,"d":0.0005,"v":0.013,"vol":3_500_000}}
    rows=[]
    for s in syms:
        pr=profiles.get(s,{"p":1000,"d":0.0006,"v":0.018,"vol":2_000_000})
        c=float(pr["p"])
        for i,dt in enumerate(dates):
            rb=0.0009 if i%22<12 else -0.0003
            dr=pr["d"]+rb+rng.normal(0,pr["v"]); op=c*(1+rng.normal(0,pr["v"]/3))
            nc=max(50.0,c*(1+dr)); sp=abs(rng.normal(0.013,pr["v"]/2))
            hi=max(op,nc)*(1+sp); lo=min(op,nc)*max(0.93,1-sp)
            vol=int(max(200_000,pr["vol"]*(1+rng.normal(0,0.25))))
            rows.append({"date":dt,"symbol":s,"open":round(op,2),"high":round(hi,2),
                          "low":round(lo,2),"close":round(nc,2),"volume":vol})
            c=nc
    return pd.DataFrame(rows).sort_values(["date","symbol"]).reset_index(drop=True)

# ══════════════════════════════════════════════════════════════════════════════
# §4  INDICATOR ENGINE
# ══════════════════════════════════════════════════════════════════════════════

def _ema(s:pd.Series,n:int)->pd.Series: return s.ewm(span=n,adjust=False).mean()

def _rsi_s(close:pd.Series,n:int=14)->pd.Series:
    d=close.diff(); g=d.clip(lower=0).ewm(alpha=1/n,adjust=False,min_periods=n).mean()
    l=(-d.clip(upper=0)).ewm(alpha=1/n,adjust=False,min_periods=n).mean()
    return (100-100/(1+g/l.replace(0,np.nan))).fillna(50)

def _atr_s(df:pd.DataFrame,n:int=14)->pd.Series:
    pc=df["close"].shift(1)
    tr=pd.concat([df["high"]-df["low"],(df["high"]-pc).abs(),(df["low"]-pc).abs()],axis=1).max(axis=1)
    return tr.ewm(alpha=1/n,adjust=False,min_periods=n).mean()

def _adx_di(df:pd.DataFrame,n:int=14):
    hd=df["high"].diff(); ld=-df["low"].diff()
    pdm=pd.Series(np.where((hd>ld)&(hd>0),hd,0.0),index=df.index)
    mdm=pd.Series(np.where((ld>hd)&(ld>0),ld,0.0),index=df.index)
    pc=df["close"].shift(1)
    tr=pd.concat([df["high"]-df["low"],(df["high"]-pc).abs(),(df["low"]-pc).abs()],axis=1).max(axis=1)
    atr_s=tr.ewm(alpha=1/n,adjust=False,min_periods=n).mean().replace(0,np.nan)
    pdi=pdm.ewm(alpha=1/n,adjust=False,min_periods=n).mean()/atr_s*100
    mdi=mdm.ewm(alpha=1/n,adjust=False,min_periods=n).mean()/atr_s*100
    dx=((pdi-mdi).abs()/(pdi+mdi).replace(0,np.nan)*100)
    return dx.ewm(alpha=1/n,adjust=False,min_periods=n).mean().fillna(20),pdi.fillna(0),mdi.fillna(0)

def _supertrend(df:pd.DataFrame,mult:float=3.0,n:int=10):
    atr_s=_atr_s(df,n); hl2=(df["high"]+df["low"])/2
    up=hl2+mult*atr_s; dn=hl2-mult*atr_s
    fi_up=up.copy(); fi_dn=dn.copy()
    for i in range(1,len(df)):
        pc=df["close"].iat[i-1]
        fi_up.iat[i]=min(up.iat[i],fi_up.iat[i-1]) if pc<=fi_up.iat[i-1] else up.iat[i]
        fi_dn.iat[i]=max(dn.iat[i],fi_dn.iat[i-1]) if pc>=fi_dn.iat[i-1] else dn.iat[i]
    direction=pd.Series(1.0,index=df.index)
    for i in range(1,len(df)):
        pd_=direction.iat[i-1]
        if pd_==-1 and df["close"].iat[i]>fi_up.iat[i]: direction.iat[i]=1
        elif pd_==1 and df["close"].iat[i]<fi_dn.iat[i]: direction.iat[i]=-1
        else: direction.iat[i]=pd_
    flip=((direction==1)&(direction.shift(1).fillna(-1)==-1)).astype(int)
    return direction,flip

def compute_indicators(raw:pd.DataFrame,cfg:Cfg)->pd.DataFrame:
    df=raw.copy().sort_values("date").reset_index(drop=True)
    if len(df)<cfg.min_bars: return pd.DataFrame()
    c=df["close"]; h=df["high"]; l=df["low"]; o=df["open"]; v=df["volume"]
    for sp in cfg.ema_spans: df[f"ema{sp}"]=_ema(c,sp)
    df["ema_gap"]=(df["ema9"]/df["ema21"].replace(0,np.nan)-1)*100
    df["macd"]=_ema(c,12)-_ema(c,26); df["macd_sig"]=_ema(df["macd"],9)
    df["macd_h"]=df["macd"]-df["macd_sig"]; df["macd_h_p"]=df["macd_h"].shift(1)
    df["rsi14"]=_rsi_s(c,14); df["rsi9"]=_rsi_s(c,9)
    df["atr14"]=_atr_s(df,14); df["atr_pct"]=df["atr14"]/c.replace(0,np.nan)*100
    df["adx"],df["plus_di"],df["minus_di"]=_adx_di(df,14)
    df["st_dir"],df["st_flip"]=_supertrend(df,mult=3.0,n=10)
    bb_mid=c.rolling(cfg.bb_period,min_periods=10).mean()
    bb_std=c.rolling(cfg.bb_period,min_periods=10).std()
    df["bb_up"]=bb_mid+2*bb_std; df["bb_dn"]=bb_mid-2*bb_std
    bb_rng=(df["bb_up"]-df["bb_dn"]).replace(0,np.nan)
    df["bb_pct"]=(c-df["bb_dn"])/bb_rng; df["bb_bw"]=bb_rng/bb_mid.replace(0,np.nan)
    lo14=l.rolling(14,min_periods=7).min(); hi14=h.rolling(14,min_periods=7).max()
    df["stoch_k"]=((c-lo14)/(hi14-lo14).replace(0,np.nan)*100).fillna(50)
    df["stoch_d"]=df["stoch_k"].rolling(3,min_periods=1).mean()
    tp=(h+l+c)/3; tp_ma=tp.rolling(20,min_periods=10).mean()
    tp_md=tp.rolling(20,min_periods=10).apply(lambda x:np.mean(np.abs(x-x.mean())),raw=True)
    df["cci"]=((tp-tp_ma)/(0.015*tp_md.replace(0,np.nan))).fillna(0)
    hi14r=h.rolling(14,min_periods=7).max(); lo14r=l.rolling(14,min_periods=7).min()
    df["willr"]=((hi14r-c)/(hi14r-lo14r).replace(0,np.nan)*-100).fillna(-50)
    tp2=(h+l+c)/3; rmf=tp2*v
    pos_mf=rmf.where(tp2>tp2.shift(1),0.0); neg_mf=rmf.where(tp2<tp2.shift(1),0.0)
    mfr=(pos_mf.rolling(14,min_periods=7).sum()/neg_mf.rolling(14,min_periods=7).sum().replace(0,np.nan))
    df["mfi"]=(100-100/(1+mfr)).fillna(50)
    df["obv"]=(np.sign(c.diff())*v).cumsum(); df["obv_ema"]=_ema(df["obv"],20)
    df["avg_vol20"]=v.rolling(20,min_periods=10).mean()
    df["vol_ratio"]=v/df["avg_vol20"].replace(0,np.nan)
    df["vol_z"]=(v-df["avg_vol20"])/v.rolling(20,min_periods=10).std().replace(0,np.nan)
    df["vol_z"]=df["vol_z"].fillna(0)
    df["med_tv20"]=(c*v).rolling(20,min_periods=10).median()
    df["ret1"]=c.pct_change(); df["ret5"]=c.pct_change(5); df["ret20"]=c.pct_change(20)
    df["hi20"]=h.shift(1).rolling(cfg.breakout_window,min_periods=8).max()
    df["lo20"]=l.shift(1).rolling(cfg.breakout_window,min_periods=8).min()
    df["hi50"]=h.shift(1).rolling(50,min_periods=15).max()
    df["hi52"]=h.rolling(252,min_periods=50).max(); df["lo52"]=l.rolling(252,min_periods=50).min()
    df["bo20"]=(c>df["hi20"]).astype(int); df["bo50"]=(c>df["hi50"]).astype(int)
    df["n52h"]=(c>=df["hi52"]*0.97).astype(int); df["n52l"]=(c<=df["lo52"]*1.03).astype(int)
    df["bo_d"]=(c/df["hi20"].replace(0,np.nan)-1)*100; df["bd_d"]=(c/df["lo20"].replace(0,np.nan)-1)*100
    df["pull_slow"]=(c/df["ema21"].replace(0,np.nan)-1)*100
    df["pull_mid"]=(c/df["ema50"].replace(0,np.nan)-1)*100
    po=o.shift(1); pc_=c.shift(1); ph=h.shift(1); pl=l.shift(1)
    body=(c-o).abs(); rng_=(h-l).replace(0,np.nan)
    ls=df[["open","close"]].min(axis=1)-l; us=h-df[["open","close"]].max(axis=1)
    df["cdl_bull_eng"]=((c>o)&(pc_<po)&(o<=pc_)&(c>=po)).astype(int)
    df["cdl_bear_eng"]=((c<o)&(pc_>po)&(o>=pc_)&(c<=po)).astype(int)
    df["cdl_hammer"]=((ls>=2.0*body)&(us<=0.3*body)&(c>o)).astype(int)
    df["cdl_inv_hammer"]=((us>=2.0*body)&(ls<=0.3*body)&(c>o)).astype(int)
    df["cdl_doji"]=((body/rng_).fillna(1.0)<0.10).astype(int)
    df["cdl_marubozu"]=((body/rng_).fillna(0.0)>=0.80).astype(int)
    df["cdl_inside"]=((h<ph)&(l>pl)).astype(int)
    df["cdl_hh"]=((h>ph)&(l>pl)).astype(int)
    df["cdl_morn_star"]=((pc_.shift(1)<po.shift(1))&
                         ((c.shift(1)-o.shift(1)).abs()<(rng_.shift(2).fillna(1)*0.35))&
                         (c>(po.shift(1)+pc_.shift(1))/2)).astype(int)
    df["cdl_3ws"]=((c>o)&(pc_>po)&(c.shift(2)>o.shift(2))&(c>pc_)&(pc_>c.shift(2))).astype(int)
    df["cdl_pier"]=((pc_<po)&(c>o)&(c>(po+pc_)/2)&(c<po)).astype(int)
    df["cdl_harami"]=((h<ph)&(l>pl)&(c>o)&(pc_<po)).astype(int)
    df["cdl_bo_candle"]=((c>df["hi20"])&(df["vol_ratio"].fillna(0)>1.3)).astype(int)
    df["cdl_sup_bounce"]=((df["ret1"].fillna(0)>0.005)&(c>c.shift(1))&(df["bd_d"].fillna(100)<3.0)).astype(int)
    return df.replace([np.inf,-np.inf],np.nan)

def engineer_all(prices:pd.DataFrame,cfg:Cfg)->pd.DataFrame:
    parts=[]
    for sym,grp in prices.groupby("symbol",sort=False):
        r=compute_indicators(grp,cfg)
        if not r.empty: parts.append(r)
    return pd.concat(parts,ignore_index=True) if parts else pd.DataFrame()

# ══════════════════════════════════════════════════════════════════════════════
# §5  PATTERN ENGINE
# ══════════════════════════════════════════════════════════════════════════════

CAT_COL = {
    "Trend":"cyan","Momentum":"yellow","Candlestick":"magenta",
    "Breakout":"bright_green","Volume":"blue","Volatility":"white",
    "Price Action":"bright_white","Structure":"bright_cyan",
}

def _g(row,k:str,d:float=0.0)->float:
    try: v=row[k] if isinstance(row,dict) else getattr(row,k,d)
    except Exception: return d
    if v is None or (isinstance(v,float) and np.isnan(v)): return d
    return float(v)

def detect_patterns(row,tail:pd.DataFrame)->list:
    hits:dict[str,tuple]={}
    def add(sc,lb,cat):
        if lb not in hits or sc>hits[lb][0]: hits[lb]=(sc,lb,cat)
    prev=tail.iloc[-2] if len(tail)>=2 else row
    prev2=tail.iloc[-3] if len(tail)>=3 else prev
    c=_g(row,"close"); e9=_g(row,"ema9"); e21=_g(row,"ema21")
    e50=_g(row,"ema50"); e200=_g(row,"ema200",e50)
    rsi=_g(row,"rsi14",50); mh=_g(row,"macd_h"); macd=_g(row,"macd")
    macds=_g(row,"macd_sig"); mhp=_g(row,"macd_h_p")
    atr=_g(row,"atr14",c*0.02) or c*0.02
    vol=_g(row,"vol_ratio",1.0); adx=_g(row,"adx",20)
    stk=_g(row,"stoch_k",50); std_=_g(row,"stoch_d",50)
    cci=_g(row,"cci",0); wr=_g(row,"willr",-50); mfi=_g(row,"mfi",50)
    bbp=_g(row,"bb_pct",0.5); bbw=_g(row,"bb_bw",0.04)
    st=_g(row,"st_dir",0); stf=_g(row,"st_flip",0)
    obv=_g(row,"obv",0); obve=_g(row,"obv_ema",0); vz=_g(row,"vol_z",0)
    pdi=_g(row,"plus_di",0); mdi=_g(row,"minus_di",0)
    pull=_g(row,"pull_slow",0)
    pc=_g(prev,"close"); pe9=_g(prev,"ema9"); pe21=_g(prev,"ema21")
    pmh=_g(prev,"macd_h"); prsi=_g(prev,"rsi14",50)
    pstk=_g(prev,"stoch_k",50); pbbp=_g(prev,"bb_pct",0.5)
    pobv=_g(prev,"obv",0); p2obv=_g(prev2,"obv",0); pst=_g(prev,"st_dir",0)

    if c>e9>e21>e50>e200: add(0.96,"🌟 Full EMA Bull Stack (All 4 EMAs)","Trend")
    elif c>e9>e21>e50:    add(0.85,"📈 EMA Bull Stack (9>21>50)","Trend")
    elif c>e9>e21:        add(0.70,"📊 Short EMA Bullish (9>21)","Trend")
    if c>e200 and e50>e200: add(0.72,"📊 Price & EMA50 Above 200","Trend")
    if pe9<=pe21 and e9>e21: add(0.92,"⚡ Golden Cross: EMA9/EMA21","Trend")
    if pe21<=e50*1.005 and e9>e21 and e21>e50: add(0.94,"🌟 Full EMA Bullish Alignment","Trend")
    if adx>28 and e9>e21 and pdi>mdi: add(0.88,"💪 Strong Trend: ADX>28, +DI>-DI","Trend")
    if adx>40: add(0.92,"🔥 Very Strong Trend: ADX>40","Trend")
    if stf==1: add(0.97,"🟩 SuperTrend BUY Flip (Bear→Bull!)","Trend")
    elif st==1: add(0.74,"🟩 SuperTrend Bullish Mode","Trend")
    if mhp<=0 and mh>0: add(0.90,"🔀 MACD Histogram Bull Cross","Momentum")
    if macd<0 and mh>0 and macd>macds: add(0.93,"🚀 MACD Bull Cross Below Zero","Momentum")
    if macd>0 and mh>0 and mh>pmh: add(0.76,"✅ MACD Positive + Accelerating","Momentum")
    if prsi<30 and rsi>30: add(0.95,"🚀 RSI Oversold Bounce (prsi<30→rsi>30)","Momentum")
    elif rsi<30: add(0.87,"🟢 RSI Oversold <30","Momentum")
    elif rsi<38: add(0.73,"🟡 RSI Near-Oversold <38","Momentum")
    if pstk<20 and stk>20 and stk>std_: add(0.90,"🔀 Stochastic Bull Cross Oversold","Momentum")
    elif stk<20: add(0.78,"📉 Stochastic Oversold Zone","Momentum")
    if cci<-100: add(0.74,"📉 CCI Oversold <-100","Momentum")
    if _g(prev,"willr",-50)<-80 and wr>-80: add(0.86,"🔀 Williams %R Bounce","Momentum")
    if mfi<25 and vol>1.2: add(0.80,"💰 MFI Oversold <25 + Volume","Momentum")
    if pbbp<0.05 and bbp>0.10: add(0.90,"↩️  BB Lower Band Bounce","Volatility")
    elif bbp<0.05: add(0.80,"📌 BB Lower Band Touch","Volatility")
    if bbw<0.035 and vol>1.2: add(0.84,"💥 BB Squeeze Breakout","Volatility")
    if vol>=2.5 and c>pc: add(0.95,"🔊 Volume Surge 2.5× (Institutional)","Volume")
    elif vol>=2.0 and c>pc: add(0.90,"🔊 Volume 2× on Up Day","Volume")
    elif vol>=1.5 and c>pc: add(0.78,"📢 Volume 1.5× Bullish","Volume")
    if vz>2.5 and c>pc: add(0.88,"🌊 Volume Z-Score >2.5σ","Volume")
    if obv>obve and c>e21: add(0.68,"📊 OBV Above 20d EMA","Volume")
    if obv>pobv>p2obv: add(0.65,"📈 OBV 3-Bar Rising (Accumulation)","Volume")
    if _g(row,"bo50")==1:
        sc=0.95 if vol>1.5 else 0.80
        add(sc,f"🏆 50-Day Breakout{' + Volume' if vol>1.5 else ''}","Breakout")
    if _g(row,"bo20")==1:
        sc=0.90 if vol>1.5 else 0.74
        add(sc,f"🚀 20-Day Breakout{' + Volume' if vol>1.5 else ''}","Breakout")
    if _g(row,"n52h")==1 and vol>1.2: add(0.87,"🏔️  Near 52W High + Volume","Breakout")
    if _g(row,"n52l")==1 and vol>1.3 and rsi<42: add(0.82,"🪃 52W Low Base Accumulation","Breakout")
    if _g(row,"cdl_bo_candle")==1: add(0.89,"💡 Breakout Candle (Vol-Confirmed)","Breakout")
    if _g(row,"cdl_inside")==1 and c>pc and e9>e21: add(0.84,"📦 Inside Bar Breakout","Price Action")
    if _g(row,"cdl_sup_bounce")==1: add(0.82,"🪨 Support Bounce (20D Low Zone)","Price Action")
    if abs(pull)<2.0 and rsi>45 and e9>e50: add(0.76,"📐 Pullback to EMA21 (Retest)","Price Action")
    if _g(row,"cdl_bull_eng")==1: add(0.92,"🕯️  Bullish Engulfing","Candlestick")
    if _g(row,"cdl_hammer")==1: add(0.84,"🔨 Hammer Candle","Candlestick")
    if _g(row,"cdl_morn_star")==1: add(0.95,"🌅 Morning Star (3-Bar Reversal)","Candlestick")
    if _g(row,"cdl_3ws")==1: add(0.93,"⚔️  Three White Soldiers","Candlestick")
    if _g(row,"cdl_pier")==1: add(0.86,"🗡️  Piercing Line","Candlestick")
    if _g(row,"cdl_harami")==1 and e9>e21: add(0.72,"🤱 Bullish Harami","Candlestick")
    if _g(row,"cdl_marubozu")==1 and c>pc: add(0.82,"📊 Bullish Marubozu","Candlestick")
    if _g(row,"cdl_hh")==1 and e9>e21: add(0.66,"📶 Higher High + Higher Low","Structure")
    return sorted(hits.values(),key=lambda x:-x[0])

def pat_confidence(hits:list)->float:
    if not hits: return 0.0
    s=hits[0][0]
    for i,h in enumerate(hits[1:6],1): s+=h[0]*(0.60**i)
    return round(min(s/1.9,0.98),4)

def n_cats(hits:list)->int: return len({h[2] for h in hits})

# ══════════════════════════════════════════════════════════════════════════════
# §6  COMPOSITE SCORE
# ══════════════════════════════════════════════════════════════════════════════

def row_score(row,weights:dict)->float:
    c=_g(row,"close"); e9=_g(row,"ema9"); e21=_g(row,"ema21")
    e50=_g(row,"ema50"); e200=_g(row,"ema200",e50)
    bull=((c>e9)+(e9>e21)+(e21>e50)+(e50>e200))/4.0
    bear=((c<e9)+(e9<e21)+(e21<e50)+(e50<e200))/4.0
    trend=float(np.clip(bull-bear,-1,1))
    rsi=_g(row,"rsi14",50); mh=_g(row,"macd_h")
    atr=_g(row,"atr14",c*0.02) or c*0.02
    rsi_s=float(np.clip((rsi-50)/18,-1,1)); macd_s=float(np.clip((mh/atr)*3,-1,1))
    mom=float(np.clip(0.6*rsi_s+0.4*macd_s,-1,1))
    bd=_g(row,"bo20"); vol=_g(row,"vol_ratio",1.0)
    bod=float(np.clip(_g(row,"bo_d",0)/8,-1,1))
    brk=float(np.clip(max(float(bd),bod)*min(vol/1.5,1.2),-1,1))
    pull=_g(row,"pull_slow",0)
    pull_s=(0.75 if (abs(pull)<2.5 and c>e50 and 40<rsi<62) else
            0.55 if (abs(pull)<2.5 and c>e21 and 40<rsi<62) else 0.0)
    vol_s=float(np.clip((vol-1)/1.5,-1,1))
    eng=_g(row,"cdl_bull_eng"); ham=_g(row,"cdl_hammer")
    morn=_g(row,"cdl_morn_star"); sol=_g(row,"cdl_3ws")
    pat_raw=(1.0 if (morn or sol) else 0.85 if eng else 0.65 if ham else 0.0)
    pat_s=float(np.clip(pat_raw,-1,1))
    pe=_g(row,"_pe",0); roe=_g(row,"_roe",0); eg=_g(row,"_epsg",0)
    fund_s=0.0
    if pe>0: fund_s+=0.35 if pe<22 else(-0.20 if pe>60 else 0.10)
    if roe>0: fund_s+=0.35 if roe>0.18 else(-0.10 if roe<0.05 else 0.10)
    if eg!=0: fund_s+=0.20 if eg>0.12 else(-0.10 if eg<0 else 0.05)
    fund_s=float(np.clip(fund_s,-1,1))
    r5=_g(row,"ret5",0); sent_s=float(np.clip(r5/0.05,-1,1))
    w=weights; ws=sum(abs(v) for v in w.values())
    sc=(w.get("trend",0)*trend+w.get("momentum",0)*mom+w.get("breakout",0)*brk+
        w.get("pullback",0)*pull_s+w.get("volume",0)*vol_s+w.get("pattern",0)*pat_s+
        w.get("fundamental",0)*fund_s+w.get("sentiment",0)*sent_s)
    return float(np.clip(sc/ws,-1,1))

def add_scores(feat:pd.DataFrame,cfg:Cfg,nifty_trend:float)->pd.DataFrame:
    threshold=cfg.bear_threshold if nifty_trend<=-0.5 else cfg.base_threshold
    df=feat.copy()
    df["score"]=df.apply(lambda r:row_score(r,cfg.weights),axis=1)
    df["signal"]=np.where(df["score"]>=threshold,"LONG","NEUTRAL")
    return df

# ══════════════════════════════════════════════════════════════════════════════
# §7  CONFIDENCE MODULES
# ══════════════════════════════════════════════════════════════════════════════

def ai_confidence(row,fund:dict,hits:list)->dict:
    W={"trend":0.24,"momentum":0.16,"breakout":0.17,"pullback":0.11,
       "volume":0.10,"pattern":0.10,"fundamental":0.08,"sentiment":0.04}
    c=_g(row,"close"); e9=_g(row,"ema9"); e21=_g(row,"ema21")
    e50=_g(row,"ema50"); e200=_g(row,"ema200",e50)
    bull=((c>e9)+(e9>e21)+(e21>e50)+(e50>e200))/4.0
    bear=((c<e9)+(e9<e21)+(e21<e50)+(e50<e200))/4.0
    trend_s=float(np.clip(bull-bear,-1,1))
    rsi=_g(row,"rsi14",50); mh=_g(row,"macd_h")
    atr=_g(row,"atr14",c*0.02) or c*0.02
    rsi_s=float(np.clip((rsi-50)/18,-1,1)); macd_s=float(np.clip((mh/atr)*3,-1,1))
    mom_s=float(np.clip(0.6*rsi_s+0.4*macd_s,-1,1))
    bd=_g(row,"bo20"); vol=_g(row,"vol_ratio",1.0)
    brk_s=float(np.clip(bd*(vol/1.5),-1,1))
    pull=_g(row,"pull_slow",0)
    pull_s=0.75 if (abs(pull)<2.5 and c>e50 and 40<rsi<62) else 0.0
    vol_s=float(np.clip((vol-1)/1.5,-1,1))
    pat_s=min(sum(h[0]*0.15 for h in hits[:6]),1.0)
    pe=fund.get("pe"); roe=fund.get("roe"); eg=fund.get("eps_g")
    rg=fund.get("rev_g"); de=fund.get("de"); peg=fund.get("peg")
    fund_s=0.0
    if pe and pe>0: fund_s+=0.30 if pe<18 else(0.15 if pe<28 else(-0.20 if pe>55 else 0.05))
    if roe: fund_s+=0.30 if roe>0.20 else(0.10 if roe>0.12 else(-0.10 if roe<0.05 else 0))
    if eg: fund_s+=0.20 if eg>0.15 else(-0.10 if eg<0 else 0.05)
    if rg: fund_s+=0.10 if rg>0.10 else 0
    if de: fund_s-=0.15 if de>3 else(0.05 if de>1.5 else 0)
    if peg and peg>0: fund_s+=0.10 if peg<1 else(-0.05 if peg>2 else 0)
    fund_s=float(np.clip(fund_s,-1,1))
    r5=_g(row,"ret5",0); r20=_g(row,"ret20",0)
    sent_s=float(np.clip((r5*0.7+r20*0.3)/0.05,-1,1))
    total=float(np.clip(
        W["trend"]*trend_s+W["momentum"]*mom_s+W["breakout"]*brk_s+
        W["pullback"]*pull_s+W["volume"]*vol_s+W["pattern"]*pat_s+
        W["fundamental"]*fund_s+W["sentiment"]*sent_s,-1,1))
    ai_pct=round(((total+1)/2)*100,1)
    bonus=0.0
    if trend_s>0.75: bonus+=0.03
    if vol_s>0.40: bonus+=0.02
    if pat_s>0.50: bonus+=0.02
    ai_pct=min(round(ai_pct+bonus*50,1),99.0)
    return dict(ai_pct=ai_pct,total=round(total,4),trend_s=round(trend_s,3),
                mom_s=round(mom_s,3),brk_s=round(brk_s,3),vol_s=round(vol_s,3),
                fund_s=round(fund_s,3),sent_s=round(sent_s,3),pat_s=round(pat_s,3))

def mkt_confidence(nifty:dict)->dict:
    if not nifty: return dict(pct=50.0,label="Unknown",align="⚪ N/A",nifty_last=0,chg_1m=0,rsi=50)
    nt=nifty.get("trend",0.0); lbl=nifty.get("label","N/A")
    nl=nifty.get("last",0.0); c1m=nifty.get("chg_1m",0.0); nr=nifty.get("rsi",50.0)
    pct=float(np.clip((nt+1)/2*100,0,100))
    align=("✅ Favorable" if nt>0.5 else "🟡 Supportive" if nt>0.2 else
           "⚠️  Headwind" if nt<-0.3 else "🔄 Neutral")
    return dict(pct=round(pct,1),label=lbl,align=align,
                nifty_last=round(nl,2),chg_1m=round(c1m,2),rsi=round(nr,1))

# ══════════════════════════════════════════════════════════════════════════════
# §8  TRADE LEVELS
# ══════════════════════════════════════════════════════════════════════════════

def trade_levels(close:float,atr:float,cfg:Cfg)->Optional[dict]:
    if atr<=0 or close<=0: return None
    def rr(sl,tp): return round(abs(tp-close)/abs(close-sl),2) if abs(close-sl)>0 else 0.0
    st_sl=round(close-cfg.st_sl_mult*atr,2); st_tp=round(close+cfg.st_tp_mult*atr,2)
    lt_sl=round(close-cfg.lt_sl_mult*atr,2); lt_tp=round(close+cfg.lt_tp_mult*atr,2)
    st_rr=rr(st_sl,st_tp); lt_rr=rr(lt_sl,lt_tp)
    if st_rr<cfg.min_rr and lt_rr<cfg.min_rr: return None
    def pkg(sl,tp,rr_v,win):
        return dict(entry=round(close,2),sl=sl,tp=tp,risk=round(abs(close-sl),2),
                    reward=round(abs(tp-close),2),rr=rr_v,rr_str=f"1:{rr_v}",window=win)
    return dict(short_term=pkg(st_sl,st_tp,st_rr,"2–5 trading days"),
                long_term =pkg(lt_sl,lt_tp,lt_rr,"10–20 trading days"))

# ══════════════════════════════════════════════════════════════════════════════
# §9  BACKTEST
# ══════════════════════════════════════════════════════════════════════════════

@dataclass
class _Pos:
    sym:str; qty:int; entry_date:pd.Timestamp; entry_p:float; fees_in:float; bars:int=0

def run_backtest(feat:pd.DataFrame,cfg:Cfg)->dict:
    empty=dict(ret=0.0,sharpe=0.0,maxdd=0.0,winrate=0.0,trades=0,
               final=cfg.bt_capital,avg_ret=0.0,avg_bars=0.0,trades_df=pd.DataFrame())
    need={"date","symbol","open","close","score","signal"}
    if not need.issubset(feat.columns): return empty
    data=feat.copy(); data["date"]=_norm_dates(data["date"])
    data=data.sort_values(["date","symbol"]).reset_index(drop=True)
    cost=cfg.bt_cost_bps/10_000; slip=cfg.bt_slip_bps/10_000
    by_d={d:g.set_index("symbol") for d,g in data.groupby("date")}
    dates=sorted(by_d.keys())
    buys=defaultdict(list); sells=defaultdict(list)
    poss:dict[str,_Pos]={}; trades=[]; eq_rows=[]; cash=cfg.bt_capital
    for idx,date in enumerate(dates):
        day=by_d[date]
        for sym,reason in list(sells.pop(date,[])):
            p=poss.get(sym)
            if not p or sym not in day.index: continue
            fp=float(day.loc[sym,"open"])*(1-slip); tv=p.qty*fp; ef=abs(tv)*cost; cash+=tv-ef
            pnl=p.qty*(fp-p.entry_p)-p.fees_in-ef; basis=p.qty*p.entry_p
            trades.append(dict(sym=sym,entry=p.entry_date,exit=date,ep=round(p.entry_p,2),
                               xp=round(fp,2),pnl=round(pnl,2),ret=round(pnl/basis if basis else 0,4),
                               bars=p.bars,reason=reason))
            del poss[sym]
        for sym,conf in sorted(buys.pop(date,[]),key=lambda x:-x[1]):
            if sym in poss or sym not in day.index: continue
            if len(poss)>=cfg.bt_max_pos: break
            fp=float(day.loc[sym,"open"])*(1+slip)
            budget=min(cfg.bt_capital*cfg.bt_pos_pct,max(cash,0)/max(1,cfg.bt_max_pos-len(poss)))
            qty=int(budget//fp)
            if qty<=0 or cash<qty*fp: continue
            tv=qty*fp; ef=abs(tv)*cost; cash-=tv+ef
            poss[sym]=_Pos(sym,qty,date,fp,ef)
        equity=cash+sum(p.qty*float(day.loc[s,"close"]) for s,p in poss.items() if s in day.index)
        eq_rows.append(dict(date=date,equity=round(equity,2)))
        nd=dates[idx+1] if idx+1<len(dates) else None
        if not nd: continue
        for sym,p in list(poss.items()):
            if sym not in day.index: continue
            p.bars+=1; pnl_pct=float(day.loc[sym,"close"])/p.entry_p-1
            sig=str(day.loc[sym,"signal"]) if "signal" in day.columns else "NEUTRAL"
            reason=None
            if pnl_pct<=-cfg.bt_sl_pct: reason="stop_loss"
            elif pnl_pct>=cfg.bt_tp_pct: reason="take_profit"
            elif p.bars>=cfg.bt_max_hold: reason="max_hold"
            elif p.bars>=cfg.bt_min_hold and sig!="LONG": reason="signal_exit"
            if reason and not any(s==sym for s,_ in sells[nd]): sells[nd].append((sym,reason))
        blocked=set(poss)|{s for s,_ in sells[nd]}|{s for s,_ in buys[nd]}
        slots=cfg.bt_max_pos-len(poss)+len({s for s,_ in sells[nd]})-len(buys[nd])
        if slots>0:
            cands=day.reset_index()
            cands=cands[(cands.get("signal","NEUTRAL")=="LONG")&
                        (cands.get("score",pd.Series(dtype=float)).fillna(0)>=cfg.base_threshold)&
                        (~cands["symbol"].isin(blocked))].sort_values("score",ascending=False)
            for r in cands.head(slots).itertuples(index=False):
                buys[nd].append((r.symbol,float(getattr(r,"score",0))))
    ld=by_d[dates[-1]]
    for sym,p in list(poss.items()):
        if sym not in ld.index: continue
        fp=float(ld.loc[sym,"close"])*(1-slip); tv=p.qty*fp; ef=abs(tv)*cost; cash+=tv-ef
        pnl=p.qty*(fp-p.entry_p)-p.fees_in-ef; basis=p.qty*p.entry_p
        trades.append(dict(sym=sym,entry=p.entry_date,exit=dates[-1],ep=round(p.entry_p,2),
                           xp=round(fp,2),pnl=round(pnl,2),ret=round(pnl/basis if basis else 0,4),
                           bars=p.bars,reason="eop"))
    eq=pd.DataFrame(eq_rows); trd=pd.DataFrame(trades)
    if not eq.empty: eq["dr"]=eq["equity"].pct_change().fillna(0); eq["dd"]=(eq["equity"]/eq["equity"].cummax())-1
    std=float(eq["dr"].std(ddof=0)) if not eq.empty and "dr" in eq.columns else 0
    sharpe=float((eq["dr"].mean()/std)*sqrt(252)) if std else 0.0
    final=float(eq["equity"].iloc[-1]) if not eq.empty else cfg.bt_capital
    return dict(ret=round(final/cfg.bt_capital-1,4),sharpe=round(sharpe,3),
                maxdd=round(float(eq["dd"].min()) if not eq.empty and "dd" in eq.columns else 0,4),
                winrate=round(float((trd["pnl"]>0).mean()) if not trd.empty else 0,3),
                trades=len(trd),final=round(final,2),
                avg_ret=round(float(trd["ret"].mean()) if not trd.empty else 0,4),
                avg_bars=round(float(trd["bars"].mean()) if not trd.empty else 0,1),
                trades_df=trd)

# ══════════════════════════════════════════════════════════════════════════════
# §10  ALERT BUILDER
# ══════════════════════════════════════════════════════════════════════════════

def _sel_reason(row,hits,fund,ai,mkt)->str:
    parts=[]; e9=_g(row,"ema9"); e21=_g(row,"ema21"); e50=_g(row,"ema50")
    rsi=_g(row,"rsi14",50); vol=_g(row,"vol_ratio",1.0)
    mh=_g(row,"macd_h"); c=_g(row,"close"); adx=_g(row,"adx",0); stf=int(_g(row,"st_flip",0))
    if hits: parts.append(f"Primary: {hits[0][1]} ({hits[0][0]*100:.0f}% conf)")
    if stf==1: parts.append("⚡ SuperTrend just flipped BULLISH")
    if e9>e21>e50: parts.append(f"Full EMA alignment ({e9:.0f}>{e21:.0f}>{e50:.0f})")
    elif e9>e21: parts.append(f"EMA bullish ({e9:.0f}>{e21:.0f})")
    if rsi<35: parts.append(f"RSI={rsi:.1f} oversold")
    elif 45<rsi<65: parts.append(f"RSI={rsi:.1f} healthy")
    if mh>0: parts.append("MACD histogram positive")
    if adx>25: parts.append(f"ADX={adx:.0f} strong trend")
    if vol>=1.5: parts.append(f"Volume {vol:.1f}× avg")
    pe=fund.get("pe"); roe=fund.get("roe")
    if pe and pe>0: parts.append(f"P/E={pe:.1f}")
    if roe and roe>0: parts.append(f"ROE={roe*100:.1f}%")
    parts.append(f"Market: {mkt.get('label','N/A')} | {mkt.get('align','N/A')}")
    return "  •  ".join(parts[:7])

def build_alerts(feat:pd.DataFrame,nifty:dict,fund_cache:dict,cfg:Cfg)->tuple:
    latest=(feat.sort_values("date").groupby("symbol",sort=False).tail(1).reset_index(drop=True))
    results=[]; rej=defaultdict(int)
    for _,row in latest.iterrows():
        sym=str(row["symbol"]); c=float(row["close"]); sig=str(row.get("signal","NEUTRAL"))
        score=float(row.get("score",0)); atr=float(row.get("atr14",c*0.02) or c*0.02)
        atr_p=atr/c*100 if c else 0; avg_v=float(row.get("avg_vol20",0) or 0)
        tv=float(row.get("med_tv20",0) or 0)/1e7
        if sig!="LONG": continue
        if c<cfg.min_price:              rej["price"]+=1;  continue
        if atr_p<cfg.min_atr_pct*100:   rej["atr_lo"]+=1; continue
        if atr_p>cfg.max_atr_pct*100:   rej["atr_hi"]+=1; continue
        if avg_v<cfg.min_avg_vol:        rej["vol"]+=1;    continue
        if tv<cfg.min_traded_val_cr:     rej["tv"]+=1;     continue
        hits=detect_patterns(row,latest[latest["symbol"]==sym].tail(3))
        pc=pat_confidence(hits); cats=n_cats(hits)
        if cats<cfg.min_categories:      rej["cats"]+=1;   continue
        lvl=trade_levels(c,atr,cfg)
        if lvl is None:                  rej["rr"]+=1;     continue
        fund=fund_cache.get(sym,{}); ai=ai_confidence(row,fund,hits); mk=mkt_confidence(nifty)
        sel=_sel_reason(row,hits,fund,ai,mk)
        results.append(dict(
            symbol=sym,last_close=round(c,2),score=round(score,4),
            atr=round(atr,2),atr_pct=round(atr_p,2),
            rsi=round(float(row.get("rsi14",50) or 50),1),
            macd_h=round(float(row.get("macd_h",0) or 0),4),
            adx=round(float(row.get("adx",0) or 0),1),
            vol_ratio=round(float(row.get("vol_ratio",1) or 1),2),
            vol_z=round(float(row.get("vol_z",0) or 0),2),
            avg_vol=int(avg_v),traded_val_cr=round(tv,2),
            ema9=round(float(row.get("ema9",0) or 0),2),
            ema21=round(float(row.get("ema21",0) or 0),2),
            ema50=round(float(row.get("ema50",0) or 0),2),
            ema200=round(float(row.get("ema200",0) or 0),2),
            st_flip=int(_g(row,"st_flip",0)),
            is_fo=sym in _FO_SET,indices=symbol_tags(sym),
            sector=fund.get("sector","N/A"),industry=fund.get("indust","N/A"),
            pe=fund.get("pe"),pb=fund.get("pb"),roe=fund.get("roe"),
            mcap=fund.get("mcap"),w52h=fund.get("w52h"),w52l=fund.get("w52l"),
            beta=fund.get("beta"),
            hits=hits,pat_conf=pc,n_cats=cats,levels=lvl,ai=ai,mkt=mk,
            reason=sel,scan_ts=datetime.now().strftime("%Y-%m-%d %H:%M")))
    results.sort(key=lambda r:(-r["ai"]["ai_pct"],-abs(r["score"])))
    LOG.info("Alerts: %d passed | rej: %s",len(results),dict(rej))
    return results,dict(rej)

# ══════════════════════════════════════════════════════════════════════════════
# §11  SAVE OUTPUTS
# ══════════════════════════════════════════════════════════════════════════════

def save_all(alerts:list,bt:dict,nifty:dict,cfg:Cfg)->dict:
    od=cfg.output_dir; od.mkdir(parents=True,exist_ok=True)
    ts=datetime.now().strftime("%Y%m%d_%H%M")
    rows=[]
    for r in alerts:
        st=r["levels"]["short_term"]; lt=r["levels"]["long_term"]; ai=r["ai"]; mk=r["mkt"]
        rows.append({"scan_ts":r["scan_ts"],"symbol":r["symbol"],"last_close":r["last_close"],
                     "score":r["score"],"rsi":r["rsi"],"adx":r["adx"],"atr_pct":r["atr_pct"],
                     "vol_ratio":r["vol_ratio"],"avg_vol":r["avg_vol"],"traded_val_cr":r["traded_val_cr"],
                     "ai_pct":ai["ai_pct"],"mkt_pct":mk["pct"],
                     "pat_conf_pct":round(r["pat_conf"]*100,1),"n_cats":r["n_cats"],
                     "top_signal":r["hits"][0][1] if r["hits"] else "",
                     "st_entry":st["entry"],"st_target":st["tp"],"st_sl":st["sl"],"st_rr":st["rr"],
                     "lt_entry":lt["entry"],"lt_target":lt["tp"],"lt_sl":lt["sl"],"lt_rr":lt["rr"],
                     "is_fo":r["is_fo"],"indices":r["indices"],"sector":r["sector"],
                     "pe":r["pe"],"roe":r["roe"],"mcap":r["mcap"],"reason":r["reason"][:250]})
    alerts_p=od/f"alerts_{ts}.csv"
    pd.DataFrame(rows).to_csv(alerts_p,index=False)
    bt.get("trades_df",pd.DataFrame()).to_csv(od/f"trades_{ts}.csv",index=False)
    with open(od/f"summary_{ts}.json","w") as f:
        json.dump({"run_ts":ts,"nifty":{k:v for k,v in nifty.items() if k!="ts"},
                   "backtest":{k:v for k,v in bt.items() if k!="trades_df"},
                   "top10":[{"symbol":r["symbol"],"ai_pct":r["ai"]["ai_pct"],
                              "st":r["levels"]["short_term"],"lt":r["levels"]["long_term"]}
                             for r in alerts[:10]]},f,indent=2,default=str)
    return dict(alerts=alerts_p,output_dir=od)

# ══════════════════════════════════════════════════════════════════════════════
# §12  DISPLAY HELPERS
# ══════════════════════════════════════════════════════════════════════════════

def _sparkline(prices: list, width: int = 28) -> str:
    B = "▁▂▃▄▅▆▇█"
    if len(prices) < 2:
        return "─" * width
    tail = prices[-width:]
    mn, mx = min(tail), max(tail)
    span   = mx - mn if mx != mn else 1
    chars  = [B[int((v - mn) / span * (len(B) - 1))] for v in tail]
    return " " * (width - len(chars)) + "".join(chars)

def _grade(pct: float) -> tuple:
    if pct >= 88: return "A+", "bold bright_green"
    if pct >= 78: return "A",  "bold green"
    if pct >= 68: return "B+", "bold yellow"
    if pct >= 58: return "B",  "bold yellow"
    if pct >= 48: return "C+", "bold red"
    return "C", "bold red"

def _gauge(pct: float, width: int = 18) -> str:
    f = max(0, min(int(pct / 100 * width), width))
    if pct >= 80:   col = "bold bright_green"
    elif pct >= 65: col = "green"
    elif pct >= 50: col = "yellow"
    else:           col = "red"
    return f"[{col}]{'█' * f}[/{col}][dim]{'░' * (width - f)}[/dim]  [{col}]{pct:.1f}%[/{col}]"

def _sc(s: float) -> str:
    if s >= 0.35: return "bold bright_green"
    if s >= 0.22: return "bold green"
    if s >= 0.10: return "yellow"
    return "dim"

def _rsi_r(r: float) -> str:
    if r < 30:  return f"[bold bright_green]{r:.1f}  ◀ OVERSOLD[/bold bright_green]"
    if r < 42:  return f"[green]{r:.1f}  ◀ Near Oversold[/green]"
    if r < 60:  return f"[yellow]{r:.1f}[/yellow]"
    if r < 75:  return f"[orange3]{r:.1f}  ▲ Elevated[/orange3]"
    return f"[bold red]{r:.1f}  ▲ OVERBOUGHT[/bold red]"

def _adx_r(a: float) -> str:
    if a >= 40: return f"[bold bright_green]{a:.0f}  ◆ Very Strong[/bold bright_green]"
    if a >= 28: return f"[green]{a:.0f}  ◆ Strong[/green]"
    if a >= 20: return f"[yellow]{a:.0f}  ◆ Moderate[/yellow]"
    return f"[red]{a:.0f}  ◆ Weak[/red]"

def _vol_r(v: float) -> str:
    if v >= 2.5: return f"[bold bright_green]{v:.2f}×  ▲ SURGE[/bold bright_green]"
    if v >= 1.5: return f"[green]{v:.2f}×  ▲ HIGH[/green]"
    if v >= 1.0: return f"[yellow]{v:.2f}×[/yellow]"
    return f"[red]{v:.2f}×  ▼ LOW[/red]"

def _pct_m(entry: float, target: float) -> str:
    if entry <= 0: return "—"
    pct = (target / entry - 1) * 100
    col = "bright_green" if pct >= 0 else "red"
    return f"[{col}]{pct:+.2f}%[/{col}]"

def _kelly(wr: float, aw: float, al: float) -> float:
    if al == 0: return 0.0
    b = aw / al
    k = wr - (1 - wr) / b if b > 0 else 0
    return round(min(max(k, 0), 0.25), 4)

def _sec_emoji(s: str) -> str:
    M = {"Technology": "💻", "Financial Services": "🏦", "Energy": "⚡",
         "Automobile": "🚗", "Infrastructure": "🏗️", "Consumer": "🛍️",
         "Healthcare": "🏥", "Materials": "⚙️", "Utilities": "💡",
         "Communication": "📡", "Real Estate": "🏢", "Industrial": "🏭"}
    for k, e in M.items():
        if k.lower() in str(s).lower():
            return e
    return "📊"

# ══════════════════════════════════════════════════════════════════════════════
# §13  SEPARATOR HELPERS
# ══════════════════════════════════════════════════════════════════════════════

def _rule_major(title: str = "") -> None:
    _con.print()
    _con.print(Rule(
        f"[bright_cyan]{'  ' + title + '  ' if title else ''}[/bright_cyan]",
        style="bright_cyan", characters="═"))
    _con.print()

def _rule_minor(title: str = "") -> None:
    _con.print(Rule(
        f"[dim cyan]{'  ' + title + '  ' if title else ''}[/dim cyan]",
        style="dim cyan", characters="─"))

def _rule_sub() -> None:
    _con.print(Rule(style="dim", characters="·"))

def _section_hdr(icon: str, title: str, sub: str = "", colour: str = "bright_cyan") -> None:
    _con.print()
    _con.print(Rule(characters="▀", style=colour))
    line = f"[bold {colour}]  {icon}  {title.upper()}  [/bold {colour}]"
    if sub:
        line += f"[dim]  ·  {sub}[/dim]"
    _con.print(Align.center(line))
    _con.print(Rule(characters="▄", style=colour))
    _con.print()

def _card_div(sym: str, rank: int, ai_p: float, score: float, ltr: str, lcol: str) -> None:
    _con.print()
    _con.print(Rule(characters="█", style="green"))
    _con.print(Rule(characters="▓", style="bright_green"))
    h = (f"[bold bright_green]  #{rank}  [/bold bright_green]"
         f"[bold bright_white on green]   {sym}   [/bold bright_white on green]"
         f"[bold bright_green]  ●  LONG SIGNAL  [/bold bright_green]"
         f"[dim green]│[/dim green]"
         f"  [bold cyan]AI Confidence: {ai_p:.1f}%[/bold cyan]"
         f"  [bold white]Composite Score: {score:+.4f}[/bold white]"
         f"  [{lcol}]Conviction Grade: {ltr}[/{lcol}]")
    _con.print(Align.center(h))
    _con.print(Rule(characters="▓", style="bright_green"))
    _con.print(Rule(characters="█", style="green"))
    _con.print()

def _sub_label(icon: str, label: str, colour: str = "bright_cyan") -> None:
    _con.print(f"\n  [{colour}]{icon}  {label}[/{colour}]")
    _con.print(f"  [dim]{'─' * 130}[/dim]")

# ══════════════════════════════════════════════════════════════════════════════
# §14  MARKET BANNER
# ══════════════════════════════════════════════════════════════════════════════

def render_market_banner(nifty: dict, bt: dict, alerts: list) -> None:
    _section_hdr("🌐", "Live Market Dashboard", colour="bright_cyan")

    trend = nifty.get("trend", 0); label = nifty.get("label", "N/A")
    last  = nifty.get("last", 0);  rsi_n  = nifty.get("rsi", 50)
    chg1m = nifty.get("chg_1m", 0); chg3m = nifty.get("chg_3m", 0)
    e9    = nifty.get("ema9", 0); e21 = nifty.get("ema21", 0); e50 = nifty.get("ema50", 0)

    if trend >= 0.7:    mc, mi = "bold bright_green", "🟢"
    elif trend >= 0.2:  mc, mi = "bold green",         "🟡"
    elif trend <= -0.7: mc, mi = "bold bright_red",    "🔴"
    elif trend <= -0.2: mc, mi = "bold red",            "🟠"
    else:               mc, mi = "bold yellow",         "⚪"

    rsent = (
        "[bold bright_green]OVERSOLD — High Reversal Probability[/bold bright_green]" if rsi_n < 30 else
        "[green]Recovering — Momentum Building[/green]"  if rsi_n < 45 else
        "[yellow]Neutral Zone[/yellow]"                   if rsi_n < 60 else
        "[orange3]Elevated — Caution[/orange3]"           if rsi_n < 75 else
        "[bold red]OVERBOUGHT — Pullback Risk[/bold red]"
    )
    if last > e9 > e21 > e50:    emas = "[bold bright_green]▲ Full Bullish Stack (9 > 21 > 50)[/bold bright_green]"
    elif last < e9 < e21 < e50:  emas = "[bold bright_red]▼ Full Bearish Stack (9 < 21 < 50)[/bold bright_red]"
    elif last > e21:              emas = "[yellow]◆ Mixed / Recovering Bias[/yellow]"
    else:                          emas = "[red]◆ Mixed / Weakening Bias[/red]"

    c1c = "bright_green" if chg1m >= 0 else "red"
    c3c = "bright_green" if chg3m >= 0 else "red"
    n_sig  = len(alerts)
    avg_ai = sum(r["ai"]["ai_pct"] for r in alerts) / max(n_sig, 1)
    n_fo   = sum(1 for r in alerts if r.get("is_fo"))
    top_s  = alerts[0]["symbol"] if alerts else "—"
    top_ai = alerts[0]["ai"]["ai_pct"] if alerts else 0

    bt_ret = bt.get("ret", 0); bt_sh = bt.get("sharpe", 0)
    bt_dd  = bt.get("maxdd", 0); bt_wr = bt.get("winrate", 0)
    bt_tr  = bt.get("trades", 0); bt_ar = bt.get("avg_ret", 0)
    bc  = "bright_green" if bt_ret >= 0 else "red"
    sc_ = "bright_green" if bt_sh > 1 else "yellow" if bt_sh > 0 else "red"
    wc  = "bright_green" if bt_wr > 0.5 else "red"
    now = datetime.now().strftime("%d %b %Y  %H:%M IST")

    t = Table(box=rbox.SIMPLE_HEAD, expand=True, padding=(0, 3),
              show_header=False, border_style="dim cyan")
    t.add_column("N",  ratio=3)
    t.add_column("S1", width=1, style="dim")
    t.add_column("B",  ratio=3)
    t.add_column("S2", width=1, style="dim")
    t.add_column("R",  ratio=3)

    mkt = (
        f"[{mc}]{mi}  NIFTY 50   ₹{last:,.2f}[/{mc}]\n\n"
        f"  [dim]Trend    [/dim]  [{mc}]{label}[/{mc}]\n"
        f"  [dim]RSI      [/dim]  {rsent}\n"
        f"  [dim]EMA Stack[/dim]  {emas}\n"
        f"  [dim]Levels   [/dim]  [dim]EMA9 {e9:.1f}  ·  EMA21 {e21:.1f}  ·  EMA50 {e50:.1f}[/dim]\n"
        f"  [dim]1M Change[/dim]  [{c1c}]{chg1m:+.2f}%[/{c1c}]"
        f"      [dim]3M Change[/dim]  [{c3c}]{chg3m:+.2f}%[/{c3c}]"
    )
    bkt = (
        f"[bold bright_magenta]📊  BACKTEST SUMMARY (8-Month)[/bold bright_magenta]\n\n"
        f"  [dim]Return      [/dim]  [{bc}]{bt_ret:+.2%}[/{bc}]\n"
        f"  [dim]Sharpe Ratio[/dim]  [{sc_}]{bt_sh:.3f}[/{sc_}]\n"
        f"  [dim]Max Drawdown[/dim]  [bold red]{bt_dd:.2%}[/bold red]\n"
        f"  [dim]Win Rate    [/dim]  [{wc}]{bt_wr:.1%}[/{wc}]\n"
        f"  [dim]Trade Count [/dim]  {bt_tr}      [dim]Avg Return[/dim]  [{bc}]{bt_ar:+.2%}[/{bc}]"
    )
    scn = (
        f"[bold bright_yellow]🔍  SCAN RESULTS[/bold bright_yellow]  [dim]{now}[/dim]\n\n"
        f"  [dim]Signals Found[/dim]  [bold bright_green]{n_sig}[/bold bright_green] LONG alerts\n"
        f"  [dim]F&O Eligible [/dim]  [bold cyan]{n_fo}[/bold cyan] of {n_sig} stocks\n"
        f"  [dim]Avg AI Score [/dim]  [bold yellow]{avg_ai:.1f}%[/bold yellow]\n"
        f"  [dim]Top Pick     [/dim]  [bold bright_white]{top_s}[/bold bright_white]"
        f"   [dim]AI Score[/dim]  [bold yellow]{top_ai:.1f}%[/bold yellow]"
    )

    t.add_row(mkt, "[dim]│[/dim]", bkt, "[dim]│[/dim]", scn)
    _con.print(Panel(t, border_style="bright_cyan", padding=(1, 2),
                     title="[bold bright_cyan]  ── LIVE MARKET STATUS ──  [/bold bright_cyan]"))

# ══════════════════════════════════════════════════════════════════════════════
# §15  SECTOR HEATMAP
# ══════════════════════════════════════════════════════════════════════════════

def render_sector_heatmap(alerts: list) -> None:
    _section_hdr("🗺️", "Sector Signal Heatmap", colour="bright_yellow")

    gs: dict[str, dict] = {}
    for r in alerts:
        for tag in r["indices"].split(" · "):
            tag = tag.strip()
            if not tag or tag == "—":
                continue
            d = gs.setdefault(tag, {"count": 0, "ai": [], "syms": []})
            d["count"] += 1
            d["ai"].append(r["ai"]["ai_pct"])
            d["syms"].append(r["symbol"])

    if not gs:
        _con.print("[dim]  No sector data.[/dim]\n")
        return

    GF = {
        "N50L": "Nifty 50 Leaders",   "N50":  "Nifty 50",
        "NN50": "Nifty Next 50",       "MC100":"Nifty Midcap 100",
        "SC250":"Nifty Smallcap 250",  "BNK":  "Nifty Bank",
        "IT":   "Nifty IT",            "NRG":  "Nifty Energy",
        "AUTO": "Nifty Auto",          "INFRA":"Nifty Infra",
        "F&O":  "F&O Stocks",
    }
    GI = {
        "N50L":"👑","N50":"📊","NN50":"🔵","MC100":"🟡","SC250":"🟠",
        "BNK":"🏦","IT":"💻","NRG":"⚡","AUTO":"🚗","INFRA":"🏗️","F&O":"🔰",
    }

    t = Table(box=rbox.ROUNDED, expand=True, header_style="bold bright_yellow",
              border_style="yellow", show_lines=True, padding=(0, 1))
    t.add_column("",          width=3,  justify="center")
    t.add_column("Index Group", width=22, style="bold white")
    t.add_column("Signals",   width=9,  justify="center")
    t.add_column("Avg AI Score", width=32)
    t.add_column("Signal Heat", width=28)
    t.add_column("Top Picks (Bold = Strongest)",  min_width=40)

    for tag in ["N50L","N50","NN50","MC100","SC250","BNK","IT","NRG","AUTO","INFRA","F&O"]:
        ico  = GI.get(tag, "📊")
        full = GF.get(tag, tag)
        if tag not in gs:
            t.add_row(ico, f"[dim]{full}[/dim]", "[dim]0[/dim]",
                      "[dim]No signals today[/dim]",
                      "[dim]░░░░░░░░░░░░░░░░░░░░[/dim]", "[dim]—[/dim]")
            continue

        d   = gs[tag]; cnt = d["count"]; avg = sum(d["ai"]) / cnt
        if   avg >= 80 and cnt >= 3: hc, hi = "bold bright_green", "🔥"
        elif avg >= 70 or cnt >= 3:  hc, hi = "bold green",         "✅"
        elif avg >= 55 or cnt >= 2:  hc, hi = "yellow",             "🟡"
        else:                         hc, hi = "dim",                "⬜"

        bw     = 20
        filled = max(1, int(min(cnt / max(len(alerts) * 0.3, 1), 1.0) * bw))
        heat   = f"[{hc}]{'█' * filled}[/{hc}][dim]{'░' * (bw - filled)}[/dim]   [{hc}]{avg:.1f}%[/{hc}]"
        syms   = "  ".join(f"[bold bright_white]{s}[/bold bright_white]"
                            for s in d["syms"][:6])
        if len(d["syms"]) > 6:
            syms += f"  [dim]+ {len(d['syms'])-6} more[/dim]"

        t.add_row(
            f"{hi}", f"[bold]{full}[/bold]",
            f"[{hc}]{cnt}[/{hc}]",
            _gauge(avg, 18), heat, syms,
        )
    _con.print(t)

# ══════════════════════════════════════════════════════════════════════════════
# §16  PER-STOCK PRO CARD SECTIONS
# ══════════════════════════════════════════════════════════════════════════════

def _narrative(r: dict) -> str:
    sym  = r["symbol"]; hits = r.get("hits", [])
    ai   = r["ai"]; mk = r["mkt"]
    st   = r["levels"]["short_term"]; lt = r["levels"]["long_term"]
    rsi  = r["rsi"]; adx = r["adx"]; vol = r["vol_ratio"]; atr = r["atr_pct"]
    score = r["score"]; stf = r.get("st_flip", 0)
    e9 = r["ema9"]; e21 = r["ema21"]; e50 = r["ema50"]
    sig = (hits[0][1].lstrip("🟢📈⚡🔀🚀🌟⭐🏆🕯️🔨↩️📊🔊🏔️💡💥🏄📌🌅⚔️").strip()
           if hits else "composite setup")

    ema_w  = ("strong uptrend"      if e9 > e21 > e50 else
              "recovering trend"    if e9 > e21       else "developing base")
    rsi_w  = ("oversold territory — historically a high-probability mean-reversion zone" if rsi < 35 else
              "neutral-to-positive momentum with room to extend"                          if rsi < 58 else
              "building momentum though nearing elevated readings")
    vol_w  = (f"significantly elevated at {vol:.2f}× the 20-day average — institutional accumulation signal"
              if vol >= 1.8 else
              f"above-average at {vol:.2f}× the 20-day average"
              if vol >= 1.2 else
              f"near-average volume ({vol:.2f}×) — no unusual activity")
    st_w   = ("Notably, the [bold bright_green]SuperTrend indicator just flipped to bullish[/bold bright_green] — "
              "a high-conviction buy confirmation signal. " if stf else "")

    return (
        f"[bold bright_white]{sym}[/bold bright_white] is exhibiting a "
        f"[italic bright_cyan]{sig}[/italic bright_cyan] setup within a [bold]{ema_w}[/bold] "
        f"(EMA9 = [bold cyan]{e9:.2f}[/bold cyan]  /  EMA21 = [bold cyan]{e21:.2f}[/bold cyan]  /  "
        f"EMA50 = [bold cyan]{e50:.2f}[/bold cyan]). "
        f"RSI stands at [bold]{rsi:.1f}[/bold], in {rsi_w}. "
        f"The latest session recorded {vol_w}. {st_w}"
        f"ADX = [bold]{adx:.1f}[/bold] — "
        f"{'[green]confirms trend strength sufficient to carry the move[/green]' if adx > 25 else '[yellow]trend is still establishing itself[/yellow]'}. "
        f"ATR = [bold]{atr:.2f}%[/bold] of price — "
        f"{'[green]ideal swing-trade volatility band[/green]' if 1.5 < atr < 6 else '[yellow]elevated — consider smaller position size[/yellow]'}. "
        f"Broader market backdrop: [bold]{mk.get('label','N/A')}[/bold] ({mk.get('align','N/A')}). "
        f"Composite model score [bold bright_yellow]{score:+.4f}[/bold bright_yellow]  "
        f"·  AI conviction [bold bright_cyan]{ai['ai_pct']:.1f}%[/bold bright_cyan].  "
        f"Short-term target [bold bright_green]₹{st['tp']}[/bold bright_green] "
        f"(Stop [bold red]₹{st['sl']}[/bold red], R:R [bold]{st['rr_str']}[/bold])  ·  "
        f"Swing target [bold bright_green]₹{lt['tp']}[/bold bright_green] "
        f"(Stop [bold red]₹{lt['sl']}[/bold red], R:R [bold]{lt['rr_str']}[/bold])."
    )


def _quick_stats(r: dict, spark: str) -> Table:
    c_   = r["last_close"]; sym = r["symbol"]; fo = r["is_fo"]; sec = str(r["sector"])
    mcap = r.get("mcap"); pe = r.get("pe"); roe = r.get("roe"); beta = r.get("beta")
    w52h = r.get("w52h"); w52l = r.get("w52l")
    e9 = r["ema9"]; e21 = r["ema21"]; e50 = r["ema50"]; e200 = r["ema200"]
    ec  = "bright_green" if c_ > e9 > e21 > e50 else "yellow" if c_ > e21 else "red"
    pec = "bright_green" if pe and pe < 20 else "green" if pe and pe < 30 else "yellow" if pe and pe < 45 else "red"
    rec = "bright_green" if roe and roe > 0.18 else "green" if roe and roe > 0.10 else "red"

    t = Table(box=rbox.SIMPLE_HEAD, expand=True, padding=(0, 2),
              show_header=False, border_style="dim bright_cyan")
    t.add_column("Label", style="dim",         ratio=1)
    t.add_column("Value", style="bold white",  ratio=2)
    t.add_column("Label", style="dim",         ratio=1)
    t.add_column("Value", style="bold white",  ratio=2)
    t.add_column("Label", style="dim",         ratio=1)
    t.add_column("Value", style="bold white",  ratio=2)

    t.add_row(
        "Symbol",
        f"[bold bright_white on dark_green]   {sym}   [/bold bright_white on dark_green]",
        "Last Traded Price",
        f"[bold bright_cyan]₹{c_:,.2f}[/bold bright_cyan]",
        "Sector",
        f"{_sec_emoji(sec)}  {sec}",
    )
    t.add_row(
        "20-Bar Sparkline",
        f"[{ec}]{spark}[/{ec}]",
        "Index Groups",
        f"[dim]{r['indices']}[/dim]",
        "F&O Eligible",
        "[bold bright_green]✅  YES[/bold bright_green]" if fo else "[dim]✗  No[/dim]",
    )
    t.add_row(
        "RSI (14)",         _rsi_r(r["rsi"]),
        "ADX Strength",     _adx_r(r["adx"]),
        "Volume Ratio",     _vol_r(r["vol_ratio"]),
    )
    t.add_row(
        "ATR %",
        f"[{'green' if 1.5<r['atr_pct']<4.5 else 'yellow'}]{r['atr_pct']:.2f}%  "
        f"{'✓ Ideal' if 1.5<r['atr_pct']<4.5 else '⚠ High'}[/{'green' if 1.5<r['atr_pct']<4.5 else 'yellow'}]",
        "MACD Histogram",
        f"[{'bright_green' if r['macd_h']>0 else 'red'}]{r['macd_h']:+.4f}  "
        f"({'Rising' if r['macd_h']>0 else 'Falling'})[/{'bright_green' if r['macd_h']>0 else 'red'}]",
        "Composite Score",
        f"[{_sc(r['score'])}]{r['score']:+.4f}[/{_sc(r['score'])}]",
    )
    t.add_row(
        "EMA 9 / 21",
        f"[{ec}]{e9:.2f}[/{ec}]  /  [{ec}]{e21:.2f}[/{ec}]",
        "EMA 50 / 200",
        (f"[{ec}]{e50:.2f}[/{ec}]  /  [dim]{e200:.2f}[/dim]"
         if e200 else f"[{ec}]{e50:.2f}[/{ec}]  /  [dim]N/A[/dim]"),
        "ATR (₹ absolute)",
        f"[dim]₹{r['atr']:.2f}[/dim]",
    )
    t.add_row(
        "P/E Ratio",
        f"[{pec}]{pe:.1f}[/{pec}]" if pe else "[dim]N/A[/dim]",
        "ROE",
        f"[{rec}]{roe*100:.1f}%[/{rec}]" if roe else "[dim]N/A[/dim]",
        "Market Cap / Beta",
        f"[dim]₹{mcap:,.0f} Cr  /  β {beta:.2f}[/dim]" if mcap and beta else "[dim]N/A[/dim]",
    )
    t.add_row(
        "52-Week High",
        f"[dim]₹{w52h:,.2f}[/dim]" if w52h else "[dim]—[/dim]",
        "52-Week Low",
        f"[dim]₹{w52l:,.2f}[/dim]" if w52l else "[dim]—[/dim]",
        "Median Traded Value",
        f"[green]₹{r['traded_val_cr']:.2f} Cr / day[/green]",
    )
    return t


def _price_levels(r: dict) -> Table:
    st  = r["levels"]["short_term"]; lt = r["levels"]["long_term"]
    c_  = r["last_close"]; atr = r["atr"]
    # Dip entry: −2% from current price
    dip_e  = round(c_ * 0.98, 2)
    dip_sl = round(dip_e - atr * 1.2, 2)
    dip_tp = round(dip_e + atr * 2.5, 2)
    dip_ri = round(dip_e - dip_sl, 2)
    dip_rw = round(dip_tp - dip_e, 2)
    dip_rr = round(dip_rw / dip_ri, 2) if dip_ri > 0 else 0

    t = Table(box=rbox.DOUBLE, expand=True,
              header_style="bold bright_cyan", border_style="bright_cyan",
              show_lines=True, padding=(0, 1))
    for col, w, jus in [
        ("Trade Scenario",      28, "left"),
        ("Entry Price ₹",       14, "right"),
        ("Target Price ₹",      14, "right"),
        ("Stop Loss ₹",         14, "right"),
        ("Risk per Share ₹",    16, "right"),
        ("Reward per Share ₹",  18, "right"),
        ("Risk : Reward",       12, "center"),
        ("% Move to Target",    16, "right"),
        ("Holding Window",      22, "left"),
    ]:
        t.add_column(col, width=w, justify=jus)

    t.add_row(
        "[bold bright_yellow]⚡  Aggressive Entry (Short-Term)[/bold bright_yellow]",
        f"[bold cyan]₹{st['entry']:,.2f}[/bold cyan]",
        f"[bold bright_green]₹{st['tp']:,.2f}[/bold bright_green]",
        f"[bold red]₹{st['sl']:,.2f}[/bold red]",
        f"[red]₹{st['risk']:.2f}[/red]",
        f"[green]₹{st['reward']:.2f}[/green]",
        f"[bold white]{st['rr_str']}[/bold white]",
        _pct_m(st["entry"], st["tp"]),
        "[yellow]2 – 5 trading days[/yellow]",
    )
    t.add_row(
        "[bold bright_green]📅  Swing Entry (Long-Term)[/bold bright_green]",
        f"[bold cyan]₹{lt['entry']:,.2f}[/bold cyan]",
        f"[bold bright_green]₹{lt['tp']:,.2f}[/bold bright_green]",
        f"[bold red]₹{lt['sl']:,.2f}[/bold red]",
        f"[red]₹{lt['risk']:.2f}[/red]",
        f"[bright_green]₹{lt['reward']:.2f}[/bright_green]",
        f"[bold white]{lt['rr_str']}[/bold white]",
        _pct_m(lt["entry"], lt["tp"]),
        "[green]10 – 20 trading days[/green]",
    )
    t.add_row(
        "[dim]📌  Limit / Dip Entry (−2% from current)[/dim]",
        f"[dim cyan]₹{dip_e:,.2f}[/dim cyan]",
        f"[dim green]₹{dip_tp:,.2f}[/dim green]",
        f"[dim red]₹{dip_sl:,.2f}[/dim red]",
        f"[dim]₹{dip_ri:.2f}[/dim]",
        f"[dim]₹{dip_rw:.2f}[/dim]",
        f"[dim]1 : {dip_rr}[/dim]",
        _pct_m(dip_e, dip_tp),
        "[dim]Limit order placed at −2%[/dim]",
    )
    return t


def _sizing(r: dict, capital: float) -> Table:
    st  = r["levels"]["short_term"]; c_ = r["last_close"]; sl = st["risk"]
    kf  = _kelly(0.55, 0.06, 0.04)

    rows_data = [
        ("1% Portfolio Risk Rule",      max(1, int(capital * 0.01 / sl)) if sl else 0, "green"),
        ("2% Portfolio Risk Rule",      max(1, int(capital * 0.02 / sl)) if sl else 0, "yellow"),
        (f"Half-Kelly Criterion ({kf:.1%})",
         max(1, int(capital * kf / c_)) if c_ else 0,  "cyan"),
        ("Fixed 20% Capital Allocation",int(capital * 0.20 / c_) if c_ else 0,  "dim"),
    ]

    t = Table(box=rbox.DOUBLE, expand=True,
              header_style="bold bright_magenta", border_style="bright_magenta",
              show_lines=True, padding=(0, 1))
    for col, w, jus in [
        ("Position Sizing Rule",    28, "left"),
        ("Shares (Qty)",            14, "right"),
        ("Capital Deployed",        18, "right"),
        ("Maximum Loss at SL",      20, "right"),
        ("Target Profit (ST)",      20, "right"),
        ("% of Total Portfolio",    20, "right"),
    ]:
        t.add_column(col, width=w, justify=jus)

    for lbl, qty, col in rows_data:
        inv  = qty * c_; ml = qty * sl
        tp_  = qty * st["reward"]; pct = inv / capital * 100
        pc   = "bright_green" if pct <= 20 else "yellow" if pct <= 30 else "red"
        t.add_row(
            f"[bold {col}]{lbl}[/bold {col}]",
            f"[bold cyan]{qty:,}[/bold cyan]",
            f"[white]₹{inv:>14,.2f}[/white]",
            f"[bold red]₹{ml:>12,.2f}[/bold red]",
            f"[bold bright_green]₹{tp_:>12,.2f}[/bold bright_green]",
            f"[{pc}]{pct:.1f}%[/{pc}]",
        )
    return t


def _factor_tbl(r: dict) -> Table:
    ai     = r["ai"]
    pull_s = 0.75 if (
        abs(r.get("ema21", 0) - r["last_close"]) / max(r.get("ema21", 1), 1) < 0.025
        and r["last_close"] > r.get("ema50", 0)
        and 40 < r["rsi"] < 62
    ) else 0.0

    W = {"Trend":0.24,"Momentum":0.16,"Breakout":0.17,"Pullback":0.11,
         "Volume":0.10,"Pattern":0.10,"Fundamental":0.08,"Sentiment":0.04}

    facts = [
        ("📈  Trend",        ai.get("trend_s",  0), 0.24, "Relative EMA positioning — how many of all 4 EMAs are bullishly stacked"),
        ("⚡  Momentum",     ai.get("mom_s",    0), 0.16, "RSI(14) normalised around 50  +  MACD histogram magnitude / ATR scaled"),
        ("🚀  Breakout",     ai.get("brk_s",    0), 0.17, "Distance above 20-day high × volume ratio — measures force of breakout"),
        ("📐  Pullback",     pull_s,                0.11, "Clean retest of slow EMA inside an uptrend — high-quality entry timing"),
        ("🔊  Volume",       ai.get("vol_s",    0), 0.10, "Volume ratio Z-score vs 20-day average — smart-money participation proxy"),
        ("🕯️   Pattern",     ai.get("pat_s",    0), 0.10, "Candlestick + price-action hit density scored across all 8 categories"),
        ("🏦  Fundamental",  ai.get("fund_s",   0), 0.08, "P/E · ROE · EPS growth · Revenue growth · Debt/Equity scoring overlay"),
        ("📡  Sentiment",    ai.get("sent_s",   0), 0.04, "5-day + 20-day price return used as a price-momentum sentiment proxy"),
    ]

    t = Table(box=rbox.DOUBLE, expand=True,
              header_style="bold bright_cyan", border_style="cyan",
              show_lines=True, padding=(0, 1))
    for col, w, jus in [
        ("Factor",          16, "left"),
        ("Score",           8,  "right"),
        ("Direction",       16, "center"),
        ("Weighted Contribution", 20, "center"),
        ("Weight",          8,  "right"),
        ("Score Bar",       22, "left"),
        ("What This Factor Measures", 46, "left"),
    ]:
        t.add_column(col, width=w, justify=jus,
                     style="" if col != "What This Factor Measures" else "dim")

    for name, sc, wt, desc in facts:
        pct    = (float(sc) + 1) / 2 * 100
        bw     = 18; filled = max(0, min(int(pct / 100 * bw), bw))
        if   sc >= 0.50: fc, ds = "bold bright_green", "[bold bright_green]▲▲  STRONG BULLISH[/bold bright_green]"
        elif sc >= 0.20: fc, ds = "green",              "[bold green]▲  BULLISH[/bold green]"
        elif sc >= 0.05: fc, ds = "yellow",             "[yellow]◆  Mild Bullish[/yellow]"
        elif sc >= -0.05:fc, ds = "dim",                "[dim]─  NEUTRAL[/dim]"
        elif sc >= -0.20:fc, ds = "orange3",            "[orange3]▽  Mild Bearish[/orange3]"
        else:             fc, ds = "bold red",           "[bold red]▼▼  STRONG BEARISH[/bold red]"
        bar = f"[{fc}]{'█' * filled}[/{fc}][dim]{'░' * (bw - filled)}[/dim]"
        contrib_val = sc * wt
        cc = "bright_green" if contrib_val > 0 else "red" if contrib_val < 0 else "dim"
        t.add_row(
            name,
            f"[{fc}]{sc:+.3f}[/{fc}]",
            ds,
            f"[{cc}]{contrib_val:+.4f}[/{cc}]",
            f"[dim]{wt:.0%}[/dim]",
            bar,
            desc,
        )
    return t


def _confidence_trio(r: dict) -> Panel:
    ai  = r["ai"]; mk = r["mkt"]
    ai_p = ai["ai_pct"]; mk_p = mk["pct"]; pt_p = r["pat_conf"] * 100
    ltr, col = _grade(ai_p)
    body = Text()
    body.append("  ┌─── Confidence Scores ──────────────────────────────────────────────────────────\n", "dim")
    body.append(f"  │  🤖  AI Confidence        {_gauge(ai_p, 20)}\n", "bold cyan")
    body.append(
        f"  │       Trend:[cyan]{ai['trend_s']:+.3f}[/cyan]   Mom:[cyan]{ai['mom_s']:+.3f}[/cyan]"
        f"   Brk:[cyan]{ai['brk_s']:+.3f}[/cyan]   Vol:[cyan]{ai['vol_s']:+.3f}[/cyan]"
        f"   Fund:[cyan]{ai['fund_s']:+.3f}[/cyan]   Pattern:[cyan]{ai['pat_s']:+.3f}[/cyan]\n",
        "dim",
    )
    body.append("  │\n", "dim")
    body.append(f"  │  📊  Market Confidence     {_gauge(mk_p, 20)}\n", "bold blue")
    body.append(
        f"  │       {mk['label']}   Nifty50 ₹{mk['nifty_last']:.2f}   "
        f"1M Change:[{'green' if mk['chg_1m']>=0 else 'red'}]{mk['chg_1m']:+.2f}%[/{'green' if mk['chg_1m']>=0 else 'red'}]"
        f"   RSI:{mk['rsi']:.1f}   Alignment:[bold]{mk['align']}[/bold]\n",
        "dim",
    )
    body.append("  │\n", "dim")
    body.append(f"  │  🎯  Pattern Confidence    {_gauge(pt_p, 20)}\n", "bold green")
    body.append(
        f"  │       [green]{len(r['hits'])}[/green] signals detected"
        f"   across [green]{r['n_cats']}[/green] distinct categories   "
        f"Top: [white]{r['hits'][0][1] if r['hits'] else '—'}[/white]\n",
        "dim",
    )
    if r.get("st_flip"):
        body.append("  │\n", "dim")
        body.append("  │  ⚡  [bold bright_green]SuperTrend just FLIPPED to BULLISH on this bar — high-conviction confirmation![/bold bright_green]\n")
    body.append("  └────────────────────────────────────────────────────────────────────────────────\n", "dim")
    return Panel(
        body,
        title=f"[bold]📈  Confidence Dashboard[/bold]   [{col}]AI Conviction Grade: {ltr}[/{col}]",
        border_style="cyan", padding=(0, 0),
    )


def _pattern_panel(r: dict) -> Panel:
    hits = r["hits"]
    body = Text()
    body.append("  ┌─── All Detected Signals ───────────────────────────────────────────────────────\n", "dim")
    cm: dict[str, list] = defaultdict(list)
    for sc, lb, cat in hits[:18]:
        cm[cat].append((sc, lb))
    for cat, items in cm.items():
        col = CAT_COL.get(cat, "white")
        body.append(f"  │  [ {cat.upper()} ]\n", f"bold {col}")
        for sc, lb in items:
            bar = f"[{col}]{'█' * int(sc * 10)}[/{col}][dim]{'░' * (10 - int(sc * 10))}[/dim]"
            body.append(f"  │      {lb}\n", "bright_white")
            body.append(f"  │      {bar}  [{col}]{sc * 100:.0f}% confidence[/{col}]\n")
    body.append(
        f"  │\n  │  [bold]Summary:[/bold]  [bright_white]{len(hits)}[/bright_white] total signals"
        f"  ·  [bright_white]{r['n_cats']}[/bright_white] distinct categories activated\n",
        "dim",
    )
    body.append("  └────────────────────────────────────────────────────────────────────────────────\n", "dim")
    return Panel(
        body,
        title="[bold bright_green]🎯  Pattern & Signal Hits[/bold bright_green]",
        border_style="bright_green", padding=(0, 0),
    )


def _conviction_scorecard(r: dict) -> Panel:
    ai  = r["ai"]; mk = r["mkt"]; cats = r["n_cats"]; stf = r.get("st_flip", 0)
    gd  = [
        ("🤖  AI Model",          r["ai"]["ai_pct"]),
        ("📊  Market Alignment",  r["mkt"]["pct"]),
        ("🎯  Pattern Strength",  r["pat_conf"] * 100),
        ("📈  Trend Quality",     (ai.get("trend_s", 0) + 1) / 2 * 100),
        ("🔊  Volume Quality",    (ai.get("vol_s",   0) + 1) / 2 * 100),
        ("🏦  Fundamental Score", (ai.get("fund_s",  0) + 1) / 2 * 100),
    ]
    body = Text()
    body.append("  ┌─── Dimension Grades ───────────────────────────────────────────────────────────\n", "dim cyan")
    for name, pct in gd:
        ltr, col = _grade(pct)
        body.append(f"  │  {name:<30}", "dim")
        body.append(f"[{col}]{ltr:<4}[/{col}]  ")
        body.append(f"{_gauge(pct, 18)}\n")
    body.append("  ├─── Positive Signals ───────────────────────────────────────────────────────────\n", "dim green")
    bonuses = []
    if stf:                    bonuses.append(("🟩", "SuperTrend just flipped BULLISH — rare buy signal"))
    if cats >= 4:              bonuses.append(("📊", f"{cats} distinct pattern categories — strong signal breadth"))
    if r["is_fo"]:             bonuses.append(("🔰", "F&O eligible — position can be hedged with options/futures"))
    if r["adx"] > 30:          bonuses.append(("💪", f"ADX = {r['adx']:.0f} — trend strength well above 30 threshold"))
    if r["vol_ratio"] >= 2:    bonuses.append(("🔊", f"Volume {r['vol_ratio']:.2f}× above average — institutional footprint"))
    if mk.get("trend", 0) >= 0.5: bonuses.append(("🐂", "Nifty50 is bullish — market tailwind supporting trade direction"))
    if r["rsi"] < 35:          bonuses.append(("🟢", f"RSI = {r['rsi']:.1f} deeply oversold — high-probability mean reversion"))
    for ico, txt in bonuses[:6]:
        body.append(f"  │  {ico}  ", "green")
        body.append(f"{txt}\n", "bright_white")
    body.append("  ├─── Risk Flags ─────────────────────────────────────────────────────────────────\n", "dim yellow")
    risks = []
    if mk.get("trend", 0) <= -0.5: risks.append(("⚠️", "Bear market headwind — trade is counter-trend to broader Nifty"))
    if r["atr_pct"] > 5:            risks.append(("⚠️", f"ATR = {r['atr_pct']:.1f}% is elevated — reduce position size accordingly"))
    if r.get("rsi", 50) > 72:       risks.append(("⚠️", f"RSI = {r.get('rsi',50):.1f} — overbought, consider waiting for pullback entry"))
    if not r["is_fo"]:              risks.append(("ℹ️",  "Not F&O listed — cash-only position, no derivative hedge available"))
    for ico, txt in (risks or [("✅", "No major risk flags detected — relatively clean setup")]):
        c2 = "yellow" if ico.startswith("⚠️") else "dim green" if ico.startswith("✅") else "dim"
        body.append(f"  │  {ico}  ", c2)
        body.append(f"{txt}\n", c2)
    overall = sum(v for _, v in gd) / len(gd)
    ltr_o, col_o = _grade(overall)
    body.append("  └────────────────────────────────────────────────────────────────────────────────\n", "dim cyan")
    return Panel(
        body,
        title=f"[bold]📋  Conviction Scorecard[/bold]   [{col_o}]Overall Grade: {ltr_o}  ({overall:.1f}%)[/{col_o}]",
        border_style="bright_cyan", padding=(0, 0),
    )


def render_pro_card(r: dict, rank: int, feat_df: Optional[pd.DataFrame], capital: float) -> None:
    sym  = r["symbol"]; ai_p = r["ai"]["ai_pct"]; score = r["score"]
    ltr, lcol = _grade(ai_p)

    _card_div(sym, rank, ai_p, score, ltr, lcol)

    # Sparkline
    spark = "─" * 28
    if feat_df is not None and not feat_df.empty:
        if all(c in feat_df.columns for c in ("close", "symbol", "date")):
            sub = feat_df[feat_df["symbol"] == sym].sort_values("date")["close"].tolist()
            if sub:
                spark = _sparkline(sub, 28)

    _sub_label("📊", "STOCK OVERVIEW  ·  All Key Metrics at a Glance", "bright_cyan")
    _con.print(Panel(
        _quick_stats(r, spark),
        title=(f"[bold bright_white]  ●  {sym}  [/bold bright_white]"
               f"[dim]  {r['indices']}  ·  {r.get('industry','N/A')}[/dim]"),
        border_style="bright_cyan", padding=(0, 0),
    ))

    _rule_sub()
    _sub_label("📝", "ANALYST NARRATIVE  ·  Quantitative Research Summary", "bright_white")
    _con.print(Panel(
        Padding(_narrative(r), (0, 2)),
        title="[bold dim]  Research Note  [/bold dim]",
        border_style="dim", padding=(0, 0),
    ))

    _rule_sub()
    _sub_label("📈", "CONFIDENCE DASHBOARD  ·  AI  ·  Market  ·  Pattern", "cyan")
    _con.print(_confidence_trio(r))

    _rule_sub()
    _sub_label("📐", "TRADE SCENARIOS  ·  Entry  ·  Target  ·  Stop Loss", "bright_cyan")
    _con.print(Panel(
        _price_levels(r),
        title="[bold bright_cyan]  Entry / Target / Stop Loss  ─  Three Trade Scenarios  [/bold bright_cyan]",
        border_style="bright_cyan", padding=(0, 0),
    ))

    _rule_sub()
    _sub_label("💰", "POSITION SIZING  ·  Risk Management  ·  Kelly Criterion", "bright_magenta")
    _con.print(Panel(
        _sizing(r, capital),
        title=f"[bold bright_magenta]  Position Sizing Calculator  ·  Portfolio: ₹{capital/1e5:.1f}L  [/bold bright_magenta]",
        border_style="bright_magenta", padding=(0, 0),
    ))

    _rule_sub()
    _sub_label("🧮", "8-FACTOR SIGNAL BREAKDOWN  ·  Score Decomposition", "cyan")
    _con.print(Panel(
        _factor_tbl(r),
        title="[bold cyan]  8-Factor Quantitative Signal Decomposition with Weighted Contributions  [/bold cyan]",
        border_style="cyan", padding=(0, 0),
    ))

    _rule_sub()
    _sub_label("🎯  +  📋", "PATTERN HITS  ·  CONVICTION SCORECARD", "bright_green")
    _con.print(Columns([_pattern_panel(r), _conviction_scorecard(r)], expand=True))

    _con.print()
    _con.print(Rule(characters="▁", style="dim green"))

# ══════════════════════════════════════════════════════════════════════════════
# §17  WATCHLIST DIGEST  (split into two tables to prevent column truncation)
# ══════════════════════════════════════════════════════════════════════════════

def render_watchlist(alerts: list, nifty: dict, bt: dict) -> None:
    _section_hdr("⭐", "Top 10 Watchlist Digest", colour="bright_yellow")

    nc = "bright_green" if nifty.get("trend",0) >= 0.3 else "red" if nifty.get("trend",0) <= -0.3 else "yellow"
    _con.print(
        f"  [bold {nc}]Nifty50  ₹{nifty.get('last',0):,.2f}[/bold {nc}]"
        f"  [dim]{nifty.get('label','N/A')}[/dim]"
        f"  RSI: [{'bright_green' if nifty.get('rsi',50)<40 else 'yellow'}]"
        f"{nifty.get('rsi',50):.1f}[/{'bright_green' if nifty.get('rsi',50)<40 else 'yellow'}]"
        f"  1M: [{'bright_green' if nifty.get('chg_1m',0)>=0 else 'red'}]"
        f"{nifty.get('chg_1m',0):+.2f}%[/{'bright_green' if nifty.get('chg_1m',0)>=0 else 'red'}]"
        f"  3M: [{'bright_green' if nifty.get('chg_3m',0)>=0 else 'red'}]"
        f"{nifty.get('chg_3m',0):+.2f}%[/{'bright_green' if nifty.get('chg_3m',0)>=0 else 'red'}]\n"
    )

    top = alerts[:10]

    # ── TABLE A: Identity + Signal Scores + Technical ─────────────────────
    _rule_minor("  Part A — Identity  ·  Grades  ·  Technical Indicators  ")
    ta = Table(box=rbox.DOUBLE_EDGE, header_style="bold bright_yellow",
               border_style="bright_yellow", show_lines=True, expand=True)
    for name, w, jus, sty in [
        ("#",              4,  "center", "dim"),
        ("Stock Symbol",  14,  "left",   "bold bright_white"),
        ("Grade",          7,  "center", "bold"),
        ("AI Score",       9,  "right",  "bold cyan"),
        ("Market %",       9,  "right",  "bold blue"),
        ("Pattern %",      9,  "right",  "bold green"),
        ("Comp Score",    10,  "right",  "bold yellow"),
        ("RSI (14)",       9,  "right",  "dim"),
        ("ADX",            7,  "right",  "dim"),
        ("Volume ×",       9,  "right",  "dim"),
        ("ATR %",          8,  "right",  "dim"),
        ("F&O",            5,  "center", "dim"),
        ("Price ₹",       12,  "right",  "bold bright_cyan"),
        ("Index Groups",  26,  "left",   "dim"),
    ]:
        ta.add_column(name, width=w, justify=jus, style=sty)

    for i, r in enumerate(top, 1):
        ai  = r["ai"]; mk_ = r["mkt"]
        ltr, lcol = _grade(ai["ai_pct"])
        sc_col = _sc(r["score"])
        ta.add_row(
            f"[bold]{i}[/bold]",
            f"[bold bright_white]{r['symbol']}[/bold bright_white]",
            f"[{lcol}]{ltr}[/{lcol}]",
            f"[bold cyan]{ai['ai_pct']:.1f}%[/bold cyan]",
            f"[bold blue]{mk_['pct']:.1f}%[/bold blue]",
            f"[bold green]{r['pat_conf']*100:.1f}%[/bold green]",
            f"[{sc_col}]{r['score']:+.4f}[/{sc_col}]",
            f"{r['rsi']:.1f}",
            f"{r['adx']:.1f}",
            f"{r['vol_ratio']:.2f}",
            f"{r['atr_pct']:.2f}%",
            "[bold bright_green]✅[/bold bright_green]" if r["is_fo"] else "[dim]—[/dim]",
            f"[bold bright_cyan]₹{r['last_close']:,.2f}[/bold bright_cyan]",
            r["indices"],
        )
    _con.print(ta)
    _con.print()

    # ── TABLE B: Trade Levels ─────────────────────────────────────────────
    _rule_minor("  Part B — Trade Levels  ·  Entry  ·  Targets  ·  Stop Losses  ·  R:R  ")
    tb = Table(box=rbox.DOUBLE_EDGE, header_style="bold bright_yellow",
               border_style="bright_yellow", show_lines=True, expand=True)
    for name, w, jus, sty in [
        ("#",              4,  "center", "dim"),
        ("Stock Symbol",  14,  "left",   "bold bright_white"),
        ("Grade",          7,  "center", "bold"),
        ("Top Signal (Full Text)",        44,  "left",   "bright_white"),
        ("ST Entry ₹",    13,  "right",  "bold cyan"),
        ("ST Target ₹",   13,  "right",  "bold bright_green"),
        ("ST Stop Loss ₹",13,  "right",  "bold red"),
        ("ST R:R",         9,  "center", "bold white"),
        ("ST % Upside",   10,  "right",  "bright_green"),
        ("LT Target ₹",   13,  "right",  "bold bright_green"),
        ("LT Stop Loss ₹",13,  "right",  "bold red"),
        ("LT R:R",         9,  "center", "bold white"),
        ("LT % Upside",   10,  "right",  "bright_green"),
    ]:
        tb.add_column(name, width=w, justify=jus, style=sty)

    for i, r in enumerate(top, 1):
        lvl  = r["levels"]; st = lvl["short_term"]; lt = lvl["long_term"]
        ltr, lcol = _grade(r["ai"]["ai_pct"])
        hits = r["hits"]
        # Full signal text — no truncation
        top_sig = hits[0][1] if hits else "No dominant signal"
        tb.add_row(
            f"[bold]{i}[/bold]",
            f"[bold bright_white]{r['symbol']}[/bold bright_white]",
            f"[{lcol}]{ltr}[/{lcol}]",
            top_sig,
            f"[bold cyan]₹{st['entry']:,.2f}[/bold cyan]",
            f"[bold bright_green]₹{st['tp']:,.2f}[/bold bright_green]",
            f"[bold red]₹{st['sl']:,.2f}[/bold red]",
            f"[bold]{st['rr_str']}[/bold]",
            _pct_m(st["entry"], st["tp"]),
            f"[bold bright_green]₹{lt['tp']:,.2f}[/bold bright_green]",
            f"[bold red]₹{lt['sl']:,.2f}[/bold red]",
            f"[bold]{lt['rr_str']}[/bold]",
            _pct_m(lt["entry"], lt["tp"]),
        )
    _con.print(tb)
    _con.print()

    # ── Backtest Performance ──────────────────────────────────────────────
    _rule_minor("  📊  Backtest Performance Summary  ")
    bt_ret=bt.get("ret",0); bt_sh=bt.get("sharpe",0); bt_dd=bt.get("maxdd",0)
    bt_wr=bt.get("winrate",0); bt_tr=bt.get("trades",0); bt_ar=bt.get("avg_ret",0)
    bt_ab=bt.get("avg_bars",0); bt_fin=bt.get("final",0)
    rc  = "bright_green" if bt_ret >= 0 else "red"
    sc_ = "bright_green" if bt_sh > 1 else "yellow" if bt_sh > 0 else "red"
    wc  = "bright_green" if bt_wr > 0.5 else "red"

    bt_t = Table(box=rbox.ROUNDED, expand=False, header_style="bold magenta",
                 border_style="magenta", padding=(0, 3), show_lines=True)
    for col in ["Metric", "Value", "Metric", "Value"]:
        bt_t.add_column(col, ratio=1)

    bt_t.add_row(
        "[dim]Total Return[/dim]",   f"[bold {rc}]{bt_ret:+.2%}[/bold {rc}]",
        "[dim]Win Rate[/dim]",       f"[bold {wc}]{bt_wr:.1%}[/bold {wc}]")
    bt_t.add_row(
        "[dim]Sharpe Ratio[/dim]",   f"[bold {sc_}]{bt_sh:.3f}[/bold {sc_}]",
        "[dim]Trade Count[/dim]",    f"[bold]{bt_tr}[/bold]")
    bt_t.add_row(
        "[dim]Max Drawdown[/dim]",   f"[bold red]{bt_dd:.2%}[/bold red]",
        "[dim]Avg Trade Return[/dim]",f"[{rc}]{bt_ar:+.2%}[/{rc}]")
    bt_t.add_row(
        "[dim]Avg Hold (bars)[/dim]",f"[bold]{bt_ab:.1f}[/bold]",
        "[dim]Final Equity[/dim]",   f"[bold bright_cyan]₹{bt_fin:,.2f}[/bold bright_cyan]")
    _con.print(Align.center(bt_t))

# ══════════════════════════════════════════════════════════════════════════════
# §18  FOOTER
# ══════════════════════════════════════════════════════════════════════════════

def render_footer(alerts: list, elapsed: float) -> None:
    _con.print()
    _con.print(Rule(characters="═", style="bright_cyan"))
    _con.print(
        "[dim]  ⚠️  DISCLAIMER: This output is generated by a quantitative model for research "
        "and educational purposes only. It does not constitute financial advice, investment "
        "recommendation, or solicitation to buy or sell any security. Past backtest performance "
        "is not indicative of future results. Always conduct your own due diligence and consult "
        "a SEBI-registered investment advisor before making any investment decisions.[/dim]"
    )
    _con.print(Rule(characters="═", style="bright_cyan"))
    _con.print(Align.center(
        f"[bold bright_cyan]  ✅  NSE Swing Trader v10.0  ·  {len(alerts)} bullish signal(s)  "
        f"·  {datetime.now().strftime('%d %b %Y  %H:%M IST')}  "
        f"·  Runtime: {elapsed:.1f}s  [/bold bright_cyan]"
    ))
    _con.print(Rule(characters="═", style="bright_cyan"))
    _con.print()


def plain_report(alerts: list, nifty: dict, bt: dict, cfg: Cfg) -> None:
    SEP = "=" * 110
    print(f"\n{SEP}")
    print(f"NSE SWING TRADER v10.0  |  {datetime.now().strftime('%d %b %Y %H:%M')}")
    print(f"Market: {nifty.get('label','N/A')} | Nifty ₹{nifty.get('last',0):,.2f} | "
          f"RSI: {nifty.get('rsi',50):.1f} | 1M: {nifty.get('chg_1m',0):+.2f}%")
    print(f"Backtest — Return: {bt.get('ret',0):+.2%}  "
          f"Sharpe: {bt.get('sharpe',0):.2f}  "
          f"MaxDD: {bt.get('maxdd',0):.2%}  "
          f"WinRate: {bt.get('winrate',0):.1%}  "
          f"Trades: {bt.get('trades',0)}")
    print(SEP)
    for i, r in enumerate(alerts[:cfg.top_n], 1):
        st   = r["levels"]["short_term"]; lt = r["levels"]["long_term"]
        hits = r["hits"]; ai = r["ai"]; mk = r["mkt"]
        print(f"\n[{i:>2}]  *** {r['symbol']} ***  "
              f"Close: ₹{r['last_close']:,.2f}  "
              f"AI: {ai['ai_pct']:.1f}%  "
              f"Market: {mk['pct']:.1f}%  "
              f"ADX: {r['adx']:.1f}  "
              f"RSI: {r['rsi']:.1f}  "
              f"Vol: {r['vol_ratio']:.2f}x")
        print(f"       Signal:    {hits[0][1] if hits else '—'}")
        print(f"       Short-Term: Entry ₹{st['entry']:,.2f}  →  Target ₹{st['tp']:,.2f}"
              f"  |  Stop ₹{st['sl']:,.2f}  |  R:R {st['rr_str']}  |  {st['window']}")
        print(f"       Long-Term:  Entry ₹{lt['entry']:,.2f}  →  Target ₹{lt['tp']:,.2f}"
              f"  |  Stop ₹{lt['sl']:,.2f}  |  R:R {lt['rr_str']}  |  {lt['window']}")
        print(f"       Reason:     {r['reason'][:115]}")
        print(f"       Indices:    {r['indices']}")
        print("-" * 110)

def run(cfg:Cfg)->tuple:
    cfg.output_dir.mkdir(parents=True,exist_ok=True)
    logging.basicConfig(level=logging.INFO,
                        format="%(asctime)s | %(levelname)-8s | %(message)s",
                        handlers=[logging.StreamHandler(sys.stdout),
                                  logging.FileHandler(cfg.output_dir/"nse_v10.log")])
    for nm in ("yfinance","peewee","urllib3","requests","charset_normalizer"):
        logging.getLogger(nm).setLevel(logging.CRITICAL)
    t0=time.time()

    if _HAS_RICH:
        _con.print()
        _con.print(Rule(characters="═",style="bright_cyan"))
        _con.print(Align.center(
            "[bold bright_white on #003366]"
            "   NSE SWING TRADER v10.0  ·  GOD-LEVEL BULLISH EDITION  "
            "·  FULLY STANDALONE   "
            "[/bold bright_white on #003366]"))
        _con.print(Align.center(
            f"[dim]  {datetime.now().strftime('%A, %d %B %Y  |  %H:%M IST')}  "
            f"|  Capital: ₹{cfg.capital/1e5:.0f}L  "
            f"|  Universe: {len(_ALL_SYMS)} symbols  [/dim]"))
        _con.print(Rule(characters="═",style="bright_cyan"))
        _con.print(f"\n[dim]📡 Fetching Nifty50 benchmark...[/dim]")

    nifty=nifty50_state(); nifty_trend=nifty.get("trend",0.0)

    if _HAS_RICH and nifty:
        threshold=cfg.bear_threshold if nifty_trend<=-0.5 else cfg.base_threshold
        _con.print(
            f"[bold blue]📊 Nifty50:[/bold blue]  ₹{nifty.get('last',0):.0f}  "
            f"{nifty.get('label','N/A')}  RSI:{nifty.get('rsi',50):.0f}  "
            f"1M:{nifty.get('chg_1m',0):+.1f}%  3M:{nifty.get('chg_3m',0):+.1f}%")
        if nifty_trend<=-0.5:
            _con.print(f"[bold red]⚠️  Bear market — threshold raised to {threshold:.2f}[/bold red]\n")
        else:
            _con.print()

    syms=cfg.symbols if cfg.symbols else _ALL_SYMS
    syms=sorted(set(syms)-_SKIP_SYMBOLS)
    if _HAS_RICH: _con.print(f"[dim]📋 Scanning {len(syms)} unique symbols...[/dim]\n")

    all_frames:list[pd.DataFrame]=[]; fund_cache:dict[str,dict]={}; ok=0; fail=0

    if cfg.use_sample:
        end=pd.Timestamp.today().normalize(); start=end-pd.Timedelta(days=260)
        demo=syms[:10]; all_frames.append(sample_ohlcv(demo,start,end)); ok=len(demo)
    elif _HAS_YF:
        def _fetch(sym):
            nonlocal ok,fail
            df=fetch_ohlcv(sym,cfg.live_period,cfg.live_interval)
            if not df.empty and len(df)>=cfg.min_bars:
                all_frames.append(df)
                if cfg.fetch_fundamentals: fund_cache[sym]=fetch_fundamentals(sym)
                ok+=1
            else: fail+=1

        if _HAS_RICH:
            with Progress(SpinnerColumn(),TextColumn("[progress.description]{task.description}"),
                          BarColumn(bar_width=24),TextColumn("{task.completed}/{task.total}"),
                          TimeElapsedColumn(),console=_con) as prog:
                task=prog.add_task("[cyan]🔴 Live scanning NSE...",total=len(syms))
                for sym in syms:
                    prog.update(task,description=f"[cyan]🔴 [bold]{sym:<14}[/bold]")
                    _fetch(sym); prog.advance(task)
        else:
            for i,sym in enumerate(syms,1):
                if i%20==0: LOG.info("Progress: %d/%d  ok:%d fail:%d",i,len(syms),ok,fail)
                _fetch(sym)
    elif cfg.prices_csv.exists():
        df=pd.read_csv(cfg.prices_csv); df["date"]=_norm_dates(df["date"])
        df["symbol"]=df["symbol"].str.upper(); all_frames.append(df); ok=df["symbol"].nunique()
    else:
        end=pd.Timestamp.today().normalize(); start=end-pd.Timedelta(days=260)
        all_frames.append(sample_ohlcv(_ALL_SYMS[:10],start,end)); ok=10

    if not all_frames: raise ValueError("No price data loaded.")

    prices=pd.concat(all_frames,ignore_index=True)
    prices=(prices.sort_values(["symbol","date"])
            .drop_duplicates(["date","symbol"],keep="last")
            .reset_index(drop=True))
    LOG.info("Loaded %d bars across %d symbols.",len(prices),prices["symbol"].nunique())

    # Pandas 2.2+ drops groupby key column in apply — use direct map instead
    def _get_fund(sym, field, default=0.0):
        return fund_cache.get(sym, {}).get(field) or default
    prices["_pe"]   = prices["symbol"].map(lambda s: _get_fund(s, "pe",    0.0))
    prices["_roe"]  = prices["symbol"].map(lambda s: _get_fund(s, "roe",   0.0))
    prices["_epsg"] = prices["symbol"].map(lambda s: _get_fund(s, "eps_g", 0.0))

    if _HAS_RICH: _con.print("[dim]⚙️  Computing 45+ indicators per symbol...[/dim]")
    feat=engineer_all(prices,cfg)
    if feat.empty: raise ValueError("Feature engineering returned empty frame.")

    feat=add_scores(feat,cfg,nifty_trend)

    if _HAS_RICH: _con.print("[dim]📊 Running backtest...[/dim]")
    bt=run_backtest(feat,cfg)

    alerts,rej=build_alerts(feat,nifty,fund_cache,cfg)
    elapsed=round(time.time()-t0,1)

    if _HAS_RICH:
        _con.print(
            f"\n[dim]🔍 Scan done in [bold]{elapsed}s[/bold]  |  "
            f"Fetched:[bold]{ok}[/bold]  Skipped:[yellow]{fail}[/yellow]  "
            f"Alerts:[bold bright_green]{len(alerts)}[/bold bright_green]  "
            f"Rejected:{dict(rej)}[/dim]\n")

    save_all(alerts,bt,nifty,cfg)

    if not _HAS_RICH:
        plain_report(alerts,nifty,bt,cfg); return alerts,bt

    if not alerts:
        _section_hdr("⚠️","No Signals Found",colour="yellow")
        _con.print("[bold yellow]  No bullish signals passed all quality gates.[/bold yellow]")
        _con.print("[dim]  Try: --threshold 0.18  or  --min-vol 1000000  or  --sample[/dim]")
        render_footer([],elapsed); return alerts,bt

    # ═══════════════════════════════════════════════════════════════════════
    # PRO DISPLAY
    # ═══════════════════════════════════════════════════════════════════════
    render_market_banner(nifty,bt,alerts)
    render_sector_heatmap(alerts)

    _section_hdr("🟢",f"Detailed Bullish Signal Cards  ({len(alerts)} stocks)",
                 sub=f"Capital ₹{cfg.capital/1e5:.0f}L  ·  Min R:R {cfg.min_rr}  ·  Threshold {cfg.base_threshold}",
                 colour="bright_green")
    for i,r in enumerate(alerts[:cfg.top_n],1):
        render_pro_card(r,rank=i,feat_df=feat,capital=cfg.capital)

    render_watchlist(alerts,nifty,bt)
    render_footer(alerts,elapsed)

    return alerts,bt

# ══════════════════════════════════════════════════════════════════════════════
# §20  CLI
# ══════════════════════════════════════════════════════════════════════════════

def main():
    p=argparse.ArgumentParser(
        description="NSE Swing Trader v10.0 — Fully Standalone (engine + pro display)",
        formatter_class=argparse.RawTextHelpFormatter)
    p.add_argument("--symbols",    type=str,   default="",
                   help="Comma-separated NSE symbols. Default: full universe.")
    p.add_argument("--group",      type=str,   default="",
                   help="Index group: 'NIFTY BANK', 'FO STOCKS', 'NIFTY IT', etc.")
    p.add_argument("--top-n",      type=int,   default=10)
    p.add_argument("--sample",     action="store_true",help="Use synthetic OHLCV (no internet).")
    p.add_argument("--prices-csv", type=Path,  default=Path("data/prices.csv"))
    p.add_argument("--output-dir", type=Path,  default=Path("nse_v10_output"))
    p.add_argument("--period",     type=str,   default="8mo")
    p.add_argument("--min-vol",    type=int,   default=1_500_000)
    p.add_argument("--min-tv",     type=float, default=5.0)
    p.add_argument("--min-rr",     type=float, default=1.5)
    p.add_argument("--threshold",  type=float, default=0.22)
    p.add_argument("--no-fund",    action="store_true",help="Skip fundamental fetch (faster).")
    p.add_argument("--capital",    type=float, default=1_000_000,
                   help="Portfolio capital for position sizing (default ₹10,00,000).")
    a=p.parse_args()

    cfg=Cfg()
    cfg.use_sample         = a.sample
    cfg.output_dir         = a.output_dir
    cfg.prices_csv         = a.prices_csv
    cfg.live_period        = a.period
    cfg.top_n              = a.top_n
    cfg.min_avg_vol        = a.min_vol
    cfg.min_traded_val_cr  = a.min_tv
    cfg.min_rr             = a.min_rr
    cfg.base_threshold     = a.threshold
    cfg.fetch_fundamentals = not a.no_fund
    cfg.capital            = a.capital

    if a.symbols:
        cfg.symbols=[s.strip().upper() for s in a.symbols.split(",") if s.strip()]
    elif a.group:
        gk=a.group.strip().upper()
        matched=[sl for grp,sl in _UNIVERSE.items() if gk in grp.upper()]
        if matched:
            cfg.symbols=sorted({s for sub in matched for s in sub}-_SKIP_SYMBOLS)
        else:
            print(f"Group '{a.group}' not found. Available: {list(_UNIVERSE.keys())}")
            sys.exit(1)

    run(cfg)


# ══════════════════════════════════════════════════════════════════════════════

# ══════════════════════════════════════════════════════════════════════════════
# §21  STREAMLIT DASHBOARD  —  TradingView-Style Interactive UI
# ══════════════════════════════════════════════════════════════════════════════

import streamlit.components.v1 as components

# ─────────────────────────────────────────────────────────────────────────────
# GLOBAL CSS  —  Bloomberg dark + TradingView aesthetic
# ─────────────────────────────────────────────────────────────────────────────
_CSS = """
<style>
@import url('https://fonts.googleapis.com/css2?family=JetBrains+Mono:wght@300;400;600;700&family=Syne:wght@400;700;800&display=swap');

:root {
    --bg:        #0b0e11;
    --bg1:       #131722;
    --bg2:       #1c2030;
    --bg3:       #252d3d;
    --border:    #2a3347;
    --border2:   #3d4f6b;
    --green:     #26a69a;
    --green2:    #00e676;
    --red:       #ef5350;
    --amber:     #f59e0b;
    --cyan:      #38bdf8;
    --blue:      #3b82f6;
    --purple:    #a855f7;
    --text:      #d1d4dc;
    --text2:     #787b86;
    --text3:     #434651;
}

/* ── Base ─────────────────────────────────────────────── */
.stApp,[data-testid="stAppViewContainer"]{
    background:var(--bg)!important;
    font-family:'JetBrains Mono',monospace!important;
    color:var(--text)!important;
}
[data-testid="stSidebar"]{
    background:var(--bg1)!important;
    border-right:1px solid var(--border)!important;
}
[data-testid="stSidebar"] *{ color:var(--text)!important; }

/* ── Metrics ──────────────────────────────────────────── */
[data-testid="metric-container"]{
    background:var(--bg1)!important;
    border:1px solid var(--border)!important;
    border-radius:4px!important;
    padding:12px 16px!important;
}
[data-testid="stMetricValue"]{
    font-family:'Syne',sans-serif!important;
    font-size:1.35rem!important;
    font-weight:800!important;
}
[data-testid="stMetricLabel"]{
    font-size:.65rem!important;
    letter-spacing:.1em!important;
    text-transform:uppercase!important;
    color:var(--text2)!important;
}

/* ── Tabs ─────────────────────────────────────────────── */
.stTabs [data-baseweb="tab-list"]{
    background:var(--bg1)!important;
    border-bottom:1px solid var(--border)!important;
    gap:0!important;
    padding:0 8px!important;
}
.stTabs [data-baseweb="tab"]{
    background:transparent!important;
    color:var(--text2)!important;
    font-family:'JetBrains Mono',monospace!important;
    font-size:.75rem!important;
    letter-spacing:.08em!important;
    border-radius:0!important;
    padding:10px 18px!important;
    border-bottom:2px solid transparent!important;
    transition:all .15s!important;
}
.stTabs [aria-selected="true"]{
    color:var(--cyan)!important;
    border-bottom:2px solid var(--cyan)!important;
}

/* ── Buttons ──────────────────────────────────────────── */
.stButton>button{
    background:var(--bg2)!important;
    border:1px solid var(--border2)!important;
    color:var(--cyan)!important;
    font-family:'JetBrains Mono',monospace!important;
    font-size:.75rem!important;
    border-radius:3px!important;
    padding:6px 14px!important;
    transition:all .15s!important;
    letter-spacing:.04em!important;
}
.stButton>button:hover{
    background:var(--border2)!important;
    color:#fff!important;
    border-color:var(--cyan)!important;
}

/* ── Inputs ───────────────────────────────────────────── */
.stSelectbox>div>div,.stMultiSelect>div>div{
    background:var(--bg2)!important;
    border-color:var(--border)!important;
    font-size:.8rem!important;
}
input[type="number"],.stNumberInput input{
    background:var(--bg2)!important;
    border-color:var(--border)!important;
    color:var(--text)!important;
}

/* ── Dataframe ────────────────────────────────────────── */
[data-testid="stDataFrame"]{
    border:1px solid var(--border)!important;
    border-radius:4px!important;
    font-size:.76rem!important;
}

/* ── Expander ─────────────────────────────────────────── */
.streamlit-expanderHeader{
    background:var(--bg2)!important;
    border:1px solid var(--border)!important;
    border-radius:3px!important;
    font-size:.78rem!important;
}

/* ── Custom components ────────────────────────────────── */
.tv-card{
    background:var(--bg1);
    border:1px solid var(--border);
    border-radius:4px;
    padding:16px 18px;
    margin-bottom:10px;
    transition:border-color .15s;
}
.tv-card:hover{ border-color:var(--border2); }
.tv-card-bull{ border-left:2px solid var(--green)!important; }
.tv-card-bear{ border-left:2px solid var(--red)!important; }
.tv-card-warn{ border-left:2px solid var(--amber)!important; }

.tv-badge-sym{
    display:inline-block;
    background:var(--green);
    color:#0b0e11;
    font-family:'Syne',sans-serif;
    font-weight:800;
    font-size:1rem;
    padding:3px 12px;
    border-radius:3px;
    letter-spacing:.05em;
}
.tv-badge-grade{
    display:inline-block;
    font-family:'Syne',sans-serif;
    font-weight:800;
    font-size:1.1rem;
    padding:3px 10px;
    border-radius:3px;
}
.tv-label{
    font-size:.62rem;
    letter-spacing:.1em;
    text-transform:uppercase;
    color:var(--text2);
    margin-bottom:3px;
}
.tv-val{
    font-family:'Syne',sans-serif;
    font-weight:700;
}
.tv-section{
    font-family:'Syne',sans-serif;
    font-weight:700;
    font-size:.9rem;
    letter-spacing:.1em;
    text-transform:uppercase;
    color:var(--cyan);
    padding:6px 0 5px;
    border-bottom:1px solid var(--border);
    margin-bottom:12px;
    margin-top:16px;
}
.tv-divider{ border:none; border-top:1px solid var(--border); margin:12px 0; }
.blink{animation:blink 1.4s step-start infinite;}
@keyframes blink{50%{opacity:0;}}
.pill-green{background:rgba(38,166,154,.18);color:#26a69a;padding:2px 8px;border-radius:2px;font-size:.68rem;font-weight:600;}
.pill-red  {background:rgba(239,83,80,.18); color:#ef5350;padding:2px 8px;border-radius:2px;font-size:.68rem;font-weight:600;}
.pill-cyan {background:rgba(56,189,248,.18);color:#38bdf8;padding:2px 8px;border-radius:2px;font-size:.68rem;font-weight:600;}
.pill-amber{background:rgba(245,158,11,.18);color:#f59e0b;padding:2px 8px;border-radius:2px;font-size:.68rem;font-weight:600;}
</style>
"""

# ─────────────────────────────────────────────────────────────────────────────
# TRADINGVIEW WIDGETS
# ─────────────────────────────────────────────────────────────────────────────

def tv_chart_widget(symbol: str, height: int = 520) -> str:
    """Full TradingView Advanced Chart widget — live NSE data, all indicators."""
    tv_sym = f"NSE:{symbol}"
    return f"""
    <div id="tv_chart_{symbol}" style="height:{height}px;border-radius:4px;overflow:hidden;border:1px solid #2a3347;"></div>
    <script type="text/javascript" src="https://s3.tradingview.com/tv.js"></script>
    <script type="text/javascript">
    new TradingView.widget({{
        "autosize": true,
        "symbol": "{tv_sym}",
        "interval": "D",
        "timezone": "Asia/Kolkata",
        "theme": "dark",
        "style": "1",
        "locale": "en",
        "toolbar_bg": "#131722",
        "enable_publishing": false,
        "withdateranges": true,
        "hide_side_toolbar": false,
        "allow_symbol_change": true,
        "watchlist": [],
        "details": true,
        "hotlist": false,
        "calendar": false,
        "show_popup_button": true,
        "popup_width": "1000",
        "popup_height": "650",
        "studies": [
            "MASimple@tv-basicstudies",
            "MASimple@tv-basicstudies",
            "RSI@tv-basicstudies",
            "MACD@tv-basicstudies",
            "Volume@tv-basicstudies"
        ],
        "studies_overrides": {{
            "moving average.length": 21,
            "moving average.plot.color": "#f59e0b",
            "moving average.plot.linewidth": 1.5
        }},
        "overrides": {{
            "mainSeriesProperties.candleStyle.upColor": "#26a69a",
            "mainSeriesProperties.candleStyle.downColor": "#ef5350",
            "mainSeriesProperties.candleStyle.borderUpColor": "#26a69a",
            "mainSeriesProperties.candleStyle.borderDownColor": "#ef5350",
            "mainSeriesProperties.candleStyle.wickUpColor": "#26a69a",
            "mainSeriesProperties.candleStyle.wickDownColor": "#ef5350",
            "paneProperties.background": "#0b0e11",
            "paneProperties.backgroundType": "solid",
            "paneProperties.gridLinesMode": "both",
            "paneProperties.vertGridProperties.color": "#2a3347",
            "paneProperties.horzGridProperties.color": "#2a3347",
            "scalesProperties.textColor": "#787b86",
            "scalesProperties.fontSize": 11
        }},
        "container_id": "tv_chart_{symbol}"
    }});
    </script>
    """

def tv_mini_chart(symbol: str, height: int = 180, nonce: str = "") -> str:
    """Compact mini sparkline-style TradingView chart."""
    tv_sym = f"NSE:{symbol}"
    return f"""
    <div style="border:1px solid #2a3347;border-radius:4px;overflow:hidden;">
    <!-- TradingView Widget BEGIN nonce={nonce} -->
    <div class="tradingview-widget-container" style="height:{height}px;">
      <div class="tradingview-widget-container__widget" style="height:100%;"></div>
      <script type="text/javascript" src="https://s3.tradingview.com/external-embedding/embed-widget-mini-symbol-overview.js" async>
      {{
        "symbol": "{tv_sym}",
        "width": "100%",
        "height": "{height}",
        "locale": "en",
        "dateRange": "3M",
        "colorTheme": "dark",
        "trendLineColor": "rgba(38, 166, 154, 1)",
        "underLineColor": "rgba(38, 166, 154, 0.1)",
        "underLineBottomColor": "rgba(41, 98, 255, 0)",
        "isTransparent": true,
        "autosize": false,
        "largeChartUrl": ""
      }}
      </script>
    </div>
    <!-- TradingView Widget END -->
    </div>
    """

def tv_ticker_tape(symbols: list) -> str:
    """Scrolling ticker tape at the top."""
    syms = [{"proName": f"NSE:{s}", "title": s} for s in symbols[:20]]
    import json
    syms_json = json.dumps(syms)
    return f"""
    <!-- TradingView Widget BEGIN -->
    <div class="tradingview-widget-container" style="margin-bottom:8px;">
      <div class="tradingview-widget-container__widget"></div>
      <script type="text/javascript" src="https://s3.tradingview.com/external-embedding/embed-widget-ticker-tape.js" async>
      {{
        "symbols": {syms_json},
        "showSymbolLogo": false,
        "isTransparent": true,
        "displayMode": "adaptive",
        "colorTheme": "dark",
        "locale": "en"
      }}
      </script>
    </div>
    <!-- TradingView Widget END -->
    """

def tv_market_overview(nifty_last: float, nifty_label: str) -> str:
    """TradingView market overview for Indian indices."""
    return """
    <!-- TradingView Widget BEGIN -->
    <div class="tradingview-widget-container" style="border:1px solid #2a3347;border-radius:4px;overflow:hidden;">
      <div class="tradingview-widget-container__widget"></div>
      <script type="text/javascript" src="https://s3.tradingview.com/external-embedding/embed-widget-market-overview.js" async>
      {
        "colorTheme": "dark",
        "dateRange": "1M",
        "showChart": true,
        "locale": "en",
        "largeChartUrl": "",
        "isTransparent": true,
        "showSymbolLogo": true,
        "showFloatingTooltip": true,
        "width": "100%",
        "height": "400",
        "tabs": [
          {
            "title": "Indian Indices",
            "symbols": [
              {"s": "BSE:SENSEX", "d": "Sensex"},
              {"s": "NSE:NIFTY", "d": "Nifty 50"},
              {"s": "NSE:BANKNIFTY", "d": "Bank Nifty"},
              {"s": "NSE:CNXIT", "d": "Nifty IT"},
              {"s": "NSE:CNXENERGY", "d": "Nifty Energy"},
              {"s": "NSE:CNXAUTO", "d": "Nifty Auto"}
            ],
            "originalTitle": "Indian Indices"
          },
          {
            "title": "F&O Leaders",
            "symbols": [
              {"s": "NSE:RELIANCE"},
              {"s": "NSE:HDFCBANK"},
              {"s": "NSE:ICICIBANK"},
              {"s": "NSE:TCS"},
              {"s": "NSE:INFY"},
              {"s": "NSE:SBIN"}
            ]
          }
        ]
      }
      </script>
    </div>
    <!-- TradingView Widget END -->
    """

def tv_heatmap() -> str:
    """TradingView stock heatmap for NSE."""
    return """
    <!-- TradingView Widget BEGIN -->
    <div class="tradingview-widget-container" style="border:1px solid #2a3347;border-radius:4px;overflow:hidden;">
      <div class="tradingview-widget-container__widget"></div>
      <script type="text/javascript" src="https://s3.tradingview.com/external-embedding/embed-widget-stock-heatmap.js" async>
      {
        "exchanges": [],
        "dataSource": "SENSEX",
        "grouping": "sector",
        "blockSize": "market_cap_basic",
        "blockColor": "change",
        "locale": "en",
        "symbolUrl": "",
        "colorTheme": "dark",
        "hasTopBar": true,
        "isDataSetEnabled": false,
        "isZoomEnabled": true,
        "hasSymbolTooltip": true,
        "isMonoSize": false,
        "width": "100%",
        "height": "480"
      }
      </script>
    </div>
    <!-- TradingView Widget END -->
    """

def tv_economic_calendar() -> str:
    """TradingView Economic Calendar widget."""
    return """
    <!-- TradingView Widget BEGIN -->
    <div class="tradingview-widget-container" style="border:1px solid #2a3347;border-radius:4px;overflow:hidden;">
      <div class="tradingview-widget-container__widget"></div>
      <script type="text/javascript" src="https://s3.tradingview.com/external-embedding/embed-widget-events.js" async>
      {
        "colorTheme": "dark",
        "isTransparent": true,
        "width": "100%",
        "height": "450",
        "locale": "en",
        "importanceFilter": "0,1",
        "countryFilter": "in"
      }
      </script>
    </div>
    <!-- TradingView Widget END -->
    """

def tv_symbol_info(symbol: str, nonce: str = "") -> str:
    """TradingView Symbol Info bar."""
    return f"""
    <!-- TradingView Widget BEGIN nonce={nonce} -->
    <div class="tradingview-widget-container" style="margin-bottom:10px;">
      <div class="tradingview-widget-container__widget"></div>
      <script type="text/javascript" src="https://s3.tradingview.com/external-embedding/embed-widget-symbol-info.js" async>
      {{
        "symbol": "NSE:{symbol}",
        "width": "100%",
        "locale": "en",
        "colorTheme": "dark",
        "isTransparent": true
      }}
      </script>
    </div>
    <!-- TradingView Widget END -->
    """

def tv_technical_analysis(symbol: str, nonce: str = "") -> str:
    """TradingView Technical Analysis (buy/sell gauge).
    Pass nonce=symbol to force re-render when symbol changes."""
    return f"""
    <!-- TradingView Widget BEGIN nonce={nonce} -->
    <div class="tradingview-widget-container" style="border:1px solid #2a3347;border-radius:4px;overflow:hidden;">
      <div class="tradingview-widget-container__widget"></div>
      <script type="text/javascript" src="https://s3.tradingview.com/external-embedding/embed-widget-technical-analysis.js" async>
      {{
        "interval": "1D",
        "width": "100%",
        "isTransparent": true,
        "height": "450",
        "symbol": "NSE:{symbol}",
        "showIntervalTabs": true,
        "displayMode": "multiple",
        "locale": "en",
        "colorTheme": "dark"
      }}
      </script>
    </div>
    <!-- TradingView Widget END -->
    """

def tv_financials(symbol: str) -> str:
    """TradingView Financials widget."""
    return f"""
    <!-- TradingView Widget BEGIN -->
    <div class="tradingview-widget-container" style="border:1px solid #2a3347;border-radius:4px;overflow:hidden;">
      <div class="tradingview-widget-container__widget"></div>
      <script type="text/javascript" src="https://s3.tradingview.com/external-embedding/embed-widget-financials.js" async>
      {{
        "isTransparent": true,
        "largeChartUrl": "",
        "displayMode": "regular",
        "width": "100%",
        "height": "830",
        "colorTheme": "dark",
        "symbol": "NSE:{symbol}",
        "locale": "en"
      }}
      </script>
    </div>
    <!-- TradingView Widget END -->
    """

def tv_news(symbol: str) -> str:
    """TradingView News widget for a symbol."""
    return f"""
    <!-- TradingView Widget BEGIN -->
    <div class="tradingview-widget-container" style="border:1px solid #2a3347;border-radius:4px;overflow:hidden;">
      <div class="tradingview-widget-container__widget"></div>
      <script type="text/javascript" src="https://s3.tradingview.com/external-embedding/embed-widget-timeline.js" async>
      {{
        "feedMode": "symbol",
        "isTransparent": true,
        "displayMode": "regular",
        "width": "100%",
        "height": "500",
        "colorTheme": "dark",
        "locale": "en",
        "symbol": "NSE:{symbol}"
      }}
      </script>
    </div>
    <!-- TradingView Widget END -->
    """

def tv_screener() -> str:
    """TradingView Stock Screener for NSE."""
    return """
    <!-- TradingView Widget BEGIN -->
    <div class="tradingview-widget-container" style="border:1px solid #2a3347;border-radius:4px;overflow:hidden;">
      <div class="tradingview-widget-container__widget"></div>
      <script type="text/javascript" src="https://s3.tradingview.com/external-embedding/embed-widget-screener.js" async>
      {
        "width": "100%",
        "height": "600",
        "defaultColumn": "overview",
        "defaultScreen": "most_capitalized",
        "market": "india",
        "showToolbar": true,
        "colorTheme": "dark",
        "locale": "en",
        "isTransparent": true
      }
      </script>
    </div>
    <!-- TradingView Widget END -->
    """

# ─────────────────────────────────────────────────────────────────────────────
# PLOTLY HELPERS  (no 8-digit hex — all rgba)
# ─────────────────────────────────────────────────────────────────────────────

def _d_grade(pct: float) -> tuple:
    if pct >= 88: return "A+","#26a69a"
    if pct >= 78: return "A", "#4db6ac"
    if pct >= 68: return "B+","#f59e0b"
    if pct >= 58: return "B", "#fbbf24"
    if pct >= 48: return "C+","#ef5350"
    return "C","#e53935"

def _d_rsi_col(r):
    if r < 30: return "#26a69a"
    if r < 45: return "#4db6ac"
    if r < 60: return "#f59e0b"
    if r < 75: return "#ffa726"
    return "#ef5350"

def _d_adx_col(a):
    if a >= 40: return "#26a69a"
    if a >= 28: return "#4db6ac"
    if a >= 20: return "#f59e0b"
    return "#ef5350"

def _d_vol_col(v):
    if v >= 2.0: return "#26a69a"
    if v >= 1.5: return "#4db6ac"
    if v >= 1.0: return "#f59e0b"
    return "#ef5350"

def _d_score_col(s):
    if s >= 0.35: return "#26a69a"
    if s >= 0.22: return "#4db6ac"
    if s >= 0.10: return "#f59e0b"
    return "#787b86"

def gauge_html(pct: float, label: str = "", width: int = 200) -> str:
    pct = max(0, min(pct, 100))
    col = "#26a69a" if pct >= 70 else "#f59e0b" if pct >= 45 else "#ef5350"
    f   = int(pct / 100 * width)
    return (f"<div style='margin:4px 0'>"
            f"<div class='tv-label' style='margin-bottom:3px'>{label}</div>"
            f"<div style='display:flex;align-items:center;gap:8px'>"
            f"<svg width='{width}' height='6' style='border-radius:3px;overflow:hidden'>"
            f"<rect width='{width}' height='6' fill='#2a3347'/>"
            f"<rect width='{f}' height='6' fill='{col}' rx='3'/></svg>"
            f"<span style='color:{col};font-size:.8rem;font-weight:700'>{pct:.1f}%</span>"
            f"</div></div>")

def fmt_inr(v):
    if v is None or (isinstance(v,float) and np.isnan(v)): return "N/A"
    return f"₹{v:,.2f}"

def fmt_cr(v):
    if v is None or (isinstance(v,float) and np.isnan(v)): return "N/A"
    return f"₹{v:,.2f} Cr"

def radar_fig(ai_dict: dict):
    cats = ["Trend","Momentum","Breakout","Pullback","Volume","Pattern","Fundamental","Sentiment"]
    keys = ["trend_s","mom_s","brk_s","trend_s","vol_s","pat_s","fund_s","sent_s"]
    vals = [(float(ai_dict.get(k,0))+1)/2*100 for k in keys]
    vc = vals + [vals[0]]; cc = cats + [cats[0]]
    fig = go.Figure()
    fig.add_trace(go.Scatterpolar(r=vc, theta=cc, fill="toself",
        fillcolor="rgba(56,189,248,0.10)", line=dict(color="#38bdf8",width=2), name="Score"))
    fig.update_layout(
        polar=dict(bgcolor="#131722",
            radialaxis=dict(visible=True,range=[0,100],tickfont=dict(size=8,color="#434651"),
                gridcolor="#2a3347",linecolor="#2a3347"),
            angularaxis=dict(tickfont=dict(size=9,color="#787b86"),
                gridcolor="#2a3347",linecolor="#2a3347")),
        paper_bgcolor="#131722",font=dict(family="JetBrains Mono"),
        showlegend=False,height=280,margin=dict(l=20,r=20,t=10,b=10))
    return fig

def factor_bar_fig(ai_dict: dict):
    factors=[("Trend",0.24,"trend_s"),("Momentum",0.16,"mom_s"),("Breakout",0.17,"brk_s"),
             ("Volume",0.10,"vol_s"),("Pattern",0.10,"pat_s"),("Fundamental",0.08,"fund_s"),
             ("Sentiment",0.04,"sent_s"),("Pullback",0.11,"trend_s")]
    names=[f[0] for f in factors]
    scores=[float(ai_dict.get(f[2],0)) for f in factors]
    colors=["#26a69a" if s>0.1 else "#ef5350" if s<-0.1 else "#f59e0b" for s in scores]
    fig=go.Figure(go.Bar(y=names,x=scores,orientation="h",marker_color=colors,
        text=[f"{s:+.3f}" for s in scores],textposition="outside",
        textfont=dict(size=10,color="#787b86")))
    fig.add_vline(x=0,line_color="#434651",line_width=1)
    fig.update_layout(height=280,paper_bgcolor="#131722",plot_bgcolor="#0b0e11",
        font=dict(family="JetBrains Mono",color="#787b86",size=10),
        xaxis=dict(range=[-1.1,1.1],showgrid=True,gridcolor="#2a3347",
                   zeroline=False,tickformat="+.1f"),
        yaxis=dict(showgrid=False),
        margin=dict(l=5,r=60,t=5,b=5),showlegend=False)
    return fig

def equity_curve_fig(bt_dict: dict):
    df=bt_dict.get("trades_df",pd.DataFrame())
    if df.empty: return go.Figure()
    df=df.copy()
    exit_col="exit" if "exit" in df.columns else "exit_date" if "exit_date" in df.columns else None
    if not exit_col: return go.Figure()
    df["_exit"]=pd.to_datetime(df[exit_col]); df=df.dropna(subset=["_exit"]).sort_values("_exit")
    df["cum_pnl"]=df["pnl"].cumsum()
    col="#26a69a" if df["cum_pnl"].iloc[-1]>=0 else "#ef5350"
    fig=go.Figure()
    fig.add_trace(go.Scatter(x=df["_exit"],y=df["cum_pnl"],mode="lines",
        line=dict(color=col,width=2),fill="tozeroy",
        fillcolor="rgba(38,166,154,0.08)" if col=="#26a69a" else "rgba(239,83,80,0.08)",
        name="P&L"))
    fig.add_hline(y=0,line_color="#434651",line_dash="dot",line_width=1)
    fig.update_layout(height=240,paper_bgcolor="#131722",plot_bgcolor="#0b0e11",
        font=dict(family="JetBrains Mono",color="#787b86",size=10),
        margin=dict(l=0,r=0,t=10,b=0),
        xaxis=dict(showgrid=True,gridcolor="#2a3347"),
        yaxis=dict(showgrid=True,gridcolor="#2a3347",tickprefix="₹"),
        showlegend=False)
    return fig

def multi_score_fig(alerts_list: list):
    top=alerts_list[:12]
    syms=[r["symbol"] for r in top]
    ai_v=[r["ai"]["ai_pct"] for r in top]
    mk_v=[r["mkt"]["pct"] for r in top]
    pt_v=[r["pat_conf"]*100 for r in top]
    fig=go.Figure()
    fig.add_trace(go.Bar(name="AI Score",x=syms,y=ai_v,marker_color="#38bdf8",
        text=[f"{v:.0f}%" for v in ai_v],textposition="outside",textfont=dict(size=9)))
    fig.add_trace(go.Bar(name="Market",x=syms,y=mk_v,marker_color="#3b82f6",
        text=[f"{v:.0f}%" for v in mk_v],textposition="outside",textfont=dict(size=9)))
    fig.add_trace(go.Bar(name="Pattern",x=syms,y=pt_v,marker_color="#26a69a",
        text=[f"{v:.0f}%" for v in pt_v],textposition="outside",textfont=dict(size=9)))
    fig.update_layout(barmode="group",height=320,
        paper_bgcolor="#131722",plot_bgcolor="#0b0e11",
        font=dict(family="JetBrains Mono",color="#787b86",size=10),
        legend=dict(bgcolor="rgba(0,0,0,0)",font=dict(size=9)),
        margin=dict(l=0,r=0,t=20,b=0),
        xaxis=dict(showgrid=False,tickfont=dict(size=9)),
        yaxis=dict(showgrid=True,gridcolor="#2a3347",range=[0,115]))
    return fig

def score_distribution_fig(alerts_list: list):
    scores=[r["score"] for r in alerts_list]
    if not scores: return go.Figure()
    fig=go.Figure()
    fig.add_trace(go.Histogram(x=scores,nbinsx=20,
        marker_color="#38bdf8",opacity=0.75,name="Score Distribution"))
    fig.add_vline(x=np.mean(scores),line_color="#f59e0b",line_dash="dash",
        annotation_text=f"Avg {np.mean(scores):.3f}",annotation_font_color="#f59e0b")
    fig.update_layout(height=220,paper_bgcolor="#131722",plot_bgcolor="#0b0e11",
        font=dict(family="JetBrains Mono",color="#787b86",size=10),
        margin=dict(l=0,r=0,t=10,b=0),
        xaxis=dict(showgrid=True,gridcolor="#2a3347",title="Composite Score"),
        yaxis=dict(showgrid=True,gridcolor="#2a3347"),showlegend=False)
    return fig

def rsi_vs_adx_fig(alerts_list: list):
    if not alerts_list: return go.Figure()
    syms=[r["symbol"] for r in alerts_list]
    rsi_v=[r["rsi"] for r in alerts_list]
    adx_v=[r["adx"] for r in alerts_list]
    ai_v=[r["ai"]["ai_pct"] for r in alerts_list]
    fig=go.Figure()
    fig.add_trace(go.Scatter(x=rsi_v,y=adx_v,mode="markers+text",
        marker=dict(size=[a/5 for a in ai_v],color=ai_v,
                    colorscale=[[0,"#ef5350"],[0.5,"#f59e0b"],[1,"#26a69a"]],
                    showscale=True,colorbar=dict(title="AI%",thickness=10,
                        tickfont=dict(size=8,color="#787b86"))),
        text=syms,textposition="top center",textfont=dict(size=8,color="#787b86"),
        hovertemplate="<b>%{text}</b><br>RSI: %{x:.1f}<br>ADX: %{y:.1f}<extra></extra>"))
    fig.add_vline(x=50,line_color="#434651",line_dash="dot",line_width=1)
    fig.add_hline(y=25,line_color="#434651",line_dash="dot",line_width=1)
    fig.update_layout(height=320,paper_bgcolor="#131722",plot_bgcolor="#0b0e11",
        font=dict(family="JetBrains Mono",color="#787b86",size=10),
        margin=dict(l=0,r=30,t=10,b=0),
        xaxis=dict(showgrid=True,gridcolor="#2a3347",title="RSI (14)",range=[0,100]),
        yaxis=dict(showgrid=True,gridcolor="#2a3347",title="ADX"),
        showlegend=False)
    return fig

def waterfall_fig(bt_dict: dict):
    df=bt_dict.get("trades_df",pd.DataFrame())
    if df.empty or len(df)<2: return go.Figure()
    df=df.copy().head(20)
    syms=[str(r.get("sym",r.get("symbol","?"))) for _,r in df.iterrows()]
    pnls=[float(r["pnl"]) for _,r in df.iterrows()]
    colors=["rgba(38,166,154,0.8)" if p>=0 else "rgba(239,83,80,0.8)" for p in pnls]
    fig=go.Figure(go.Bar(x=syms,y=pnls,marker_color=colors,
        text=[f"₹{p:+,.0f}" for p in pnls],textposition="outside",
        textfont=dict(size=8,color="#787b86")))
    fig.add_hline(y=0,line_color="#434651",line_width=1)
    fig.update_layout(height=250,paper_bgcolor="#131722",plot_bgcolor="#0b0e11",
        font=dict(family="JetBrains Mono",color="#787b86",size=9),
        margin=dict(l=0,r=0,t=10,b=40),
        xaxis=dict(showgrid=False,tickangle=-45,tickfont=dict(size=8)),
        yaxis=dict(showgrid=True,gridcolor="#2a3347",tickprefix="₹"),
        showlegend=False)
    return fig

def trade_scenario_html(title: str, lvl: dict, color: str, window: str) -> str:
    e=lvl["entry"]; tp=lvl["tp"]; sl=lvl["sl"]; rr=lvl["rr_str"]
    ri=lvl["risk"]; rw=lvl["reward"]
    up=(tp/e-1)*100 if e else 0; dn=(sl/e-1)*100 if e else 0
    return f"""
    <div class='tv-card' style='border-color:{color}66;text-align:center'>
      <div style='color:{color};font-weight:700;font-size:.82rem;margin-bottom:12px;letter-spacing:.06em'>{title}</div>
      <div class='tv-label'>Entry</div>
      <div style='font-family:Syne,sans-serif;font-weight:700;font-size:1.2rem;color:#38bdf8;margin-bottom:10px'>₹{e:,.2f}</div>
      <div style='display:flex;justify-content:space-around;margin-bottom:10px'>
        <div><div class='tv-label'>Target</div>
          <div style='font-weight:700;color:#26a69a'>₹{tp:,.2f}</div>
          <div style='font-size:.7rem;color:#26a69a'>{up:+.1f}%</div></div>
        <div><div class='tv-label'>Stop Loss</div>
          <div style='font-weight:700;color:#ef5350'>₹{sl:,.2f}</div>
          <div style='font-size:.7rem;color:#ef5350'>{dn:+.1f}%</div></div>
      </div>
      <div style='display:flex;justify-content:space-around;padding-top:8px;border-top:1px solid #2a3347'>
        <div><div class='tv-label'>R:R</div><div style='color:#d1d4dc;font-weight:700'>{rr}</div></div>
        <div><div class='tv-label'>Risk</div><div style='color:#ef5350'>₹{ri:.2f}</div></div>
        <div><div class='tv-label'>Reward</div><div style='color:#26a69a'>₹{rw:.2f}</div></div>
      </div>
      <div style='margin-top:8px;font-size:.68rem;color:#434651'>{window}</div>
    </div>"""

# ─────────────────────────────────────────────────────────────────────────────
# MAIN DASHBOARD
# ─────────────────────────────────────────────────────────────────────────────

@st.cache_data(ttl=1800, show_spinner=False)
def _sector_scan(group_key, threshold, min_vol, period, top_n, fo_only, use_smp=False):
    """Run full engine on a specific sector. Module-level for st.cache_data compatibility."""
    import pandas as pd  # ensure pd available in cached context
    """Run the full engine on a specific sector and return alerts."""
    import io as _io, contextlib as _cl, glob as _glob

    _cfg           = Cfg()
    _cfg.use_sample        = False
    _cfg.base_threshold    = max(threshold - 0.02, 0.08)
    _cfg.bear_threshold    = threshold + 0.04
    _cfg.min_avg_vol       = min(min_vol, 400_000)
    _cfg.min_traded_val_cr = 1.0
    _cfg.min_rr            = 1.1
    _cfg.min_categories    = 2
    _cfg.live_period       = period
    _cfg.top_n             = max(top_n * 2, 20)
    _cfg.fetch_fundamentals= False
    _cfg.output_dir        = Path("nse_v10_output")

    if group_key:
        gk = group_key.strip().upper()
        matched = [sl for grp, sl in _UNIVERSE.items() if gk in grp.upper()]
        syms = sorted({s for sub in matched for s in sub} - _SKIP_SYMBOLS)
    else:
        syms = _ALL_SYMS

    if fo_only:
        syms = [s for s in syms if s in _FO_SET]

    _cfg.symbols = syms

    _buf = _io.StringIO()
    with _cl.redirect_stdout(_buf), _cl.redirect_stderr(_buf):
        try:
            _alerts, _bt = run(_cfg)
            _nifty = nifty50_state()
        except Exception as _e:
            return [], {}, {}, str(_e)

    return _alerts, _bt, nifty50_state(), ""



# ─────────────────────────────────────────────────────────────────────────────
# POWER SCAN — High Volatility Bullish + Penny Golden Opportunity
# ─────────────────────────────────────────────────────────────────────────────

@st.cache_data(ttl=1800, show_spinner=False)
def _power_scan_volatile(threshold, period, top_n, use_smp=False):
    """Scan full NSE for high-momentum volatile stocks with strong bullish signals.
    Engine-level gates relaxed; post-filters select high-conviction momentum plays."""
    import pandas as pd
    import io as _io, contextlib as _cl

    cfg = Cfg()
    cfg.use_sample         = False
    cfg.base_threshold     = threshold
    cfg.bear_threshold     = threshold + 0.06
    cfg.min_avg_vol        = 400_000       # mid-caps included
    cfg.min_traded_val_cr  = 1.0
    cfg.min_rr             = 1.2
    cfg.min_atr_pct        = 0.015         # let engine pass all, filter after
    cfg.max_atr_pct        = 0.20
    cfg.min_price          = 30.0
    cfg.min_categories     = 2
    cfg.live_period        = period
    cfg.top_n              = max(top_n * 4, 40)  # scan wide
    cfg.fetch_fundamentals = False
    cfg.output_dir         = Path("nse_v10_output")
    cfg.symbols            = _ALL_SYMS

    buf = _io.StringIO()
    with _cl.redirect_stdout(buf), _cl.redirect_stderr(buf):
        try:
            alerts_out, bt_out = run(cfg)
        except Exception as e:
            return [], {}, {}, str(e)

    # Post-filter: volatile = ATR% > 2.5% AND ADX > 22 AND vol_ratio > 1.2
    volatile = [
        r for r in alerts_out
        if r["atr_pct"]    >= 2.5
        and r["adx"]       >= 22
        and r["vol_ratio"] >= 1.2
        and r["rsi"]       <= 72
    ]
    # Momentum score = ATR% × ADX × vol_ratio × AI%
    volatile.sort(key=lambda r: -(
        r["atr_pct"] * (r["adx"] / 30) * r["vol_ratio"] * r["ai"]["ai_pct"] / 1000
    ))
    return volatile[:top_n], bt_out, nifty50_state(), ""


@st.cache_data(ttl=1800, show_spinner=False)
def _power_scan_penny(threshold, period, top_n, use_smp=False):
    """Multi-bagger penny stock scanner — finds low-price NSE stocks with
    lifetime breakout setups, volume explosions, and strong AI conviction.
    Price range ₹10–₹500. Looks for: ST flip, breakout, vol surge, EMA alignment."""
    import pandas as pd
    import io as _io, contextlib as _cl

    cfg = Cfg()
    cfg.use_sample         = False
    cfg.base_threshold     = max(threshold - 0.06, 0.08)  # very loose: catch everything
    cfg.bear_threshold     = threshold + 0.02
    cfg.min_avg_vol        = 100_000       # penny stocks: allow very low volume
    cfg.min_traded_val_cr  = 0.3
    cfg.min_rr             = 1.1           # minimum R:R
    cfg.min_atr_pct        = 0.015
    cfg.max_atr_pct        = 0.35          # penny stocks can be very volatile
    cfg.min_categories     = 1             # even 1 category is ok for penny
    cfg.min_price          = 10.0          # true penny starts at ₹10
    cfg.live_period        = period
    cfg.top_n              = max(top_n * 5, 50)  # scan very wide
    cfg.fetch_fundamentals = False
    cfg.output_dir         = Path("nse_v10_output")
    cfg.symbols            = _ALL_SYMS

    buf = _io.StringIO()
    with _cl.redirect_stdout(buf), _cl.redirect_stderr(buf):
        try:
            alerts_out, bt_out = run(cfg)
        except Exception as e:
            return [], {}, {}, str(e)

    # Filter for multi-bagger penny stocks — price ₹10–₹500
    penny_golden = []
    for r in alerts_out:
        price = r["last_close"]
        if price > 500 or price < 10:   # out of penny range
            continue
        hits = r["hits"]
        cats = {h[2] for h in hits}
        top_confs = [h[0] for h in hits[:3]]
        avg_top_conf = sum(top_confs) / max(len(top_confs), 1)

        # Golden criteria: ST flip OR breakout + high AI + multi-category
        has_breakout  = any("Breakout" in h[2] for h in hits)
        has_volume    = any("Volume" in h[2] for h in hits)
        has_stf       = r.get("st_flip", 0)
        has_trend     = any("Trend" in h[2] for h in hits)
        golden_score  = (
            (3.0 if has_stf else 0)
            + (2.5 if has_breakout else 0)
            + (1.5 if has_volume else 0)
            + (1.0 if has_trend else 0)
            + r["ai"]["ai_pct"] / 20
            + r["n_cats"] * 0.5
            + avg_top_conf * 2
        )

        # Must have at least 2 of: breakout, volume surge, ST flip, strong trend
        signals_count = sum([has_breakout, has_volume, bool(has_stf), has_trend])
        if signals_count < 1:  # at least 1 strong signal
            continue

        r = dict(r)  # copy
        r["golden_score"]  = round(golden_score, 2)
        r["has_breakout"]  = has_breakout
        r["has_volume_surge"] = has_volume
        r["has_stf"]       = bool(has_stf)
        r["price_category"]= (
            "💎 Ultra Penny (< ₹50)"   if price < 50   else
            "🔶 Penny (₹50–₹150)"      if price < 150  else
            "🔷 Small Cap (₹150–₹300)" if price < 300  else
            "📘 Mid Value (₹300–₹500)"
        )
        penny_golden.append(r)

    penny_golden.sort(key=lambda r: -r["golden_score"])
    return penny_golden[:top_n], bt_out, nifty50_state(), ""


@st.cache_data(ttl=900, show_spinner=False)
def _nse_52w_lows_today() -> list:
    """
    Fetch today's 52-week low stocks from NSE India public API.
    Returns list of NSE symbols confirmed at 52W low today.
    Falls back to empty list gracefully if NSE unreachable.
    """
    import requests as _rq, time as _t2
    confirmed = []
    try:
        hdrs = {
            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36",
            "Accept":     "application/json, text/plain, */*",
            "Referer":    "https://www.nseindia.com/",
        }
        sess = _rq.Session()
        sess.get("https://www.nseindia.com/", headers=hdrs, timeout=8)
        _t2.sleep(0.4)
        for url in [
            "https://www.nseindia.com/api/liveEquity-derivatives?index=securities52WeekLow",
        ]:
            try:
                r = sess.get(url, headers=hdrs, timeout=12)
                if r.status_code == 200:
                    js = r.json()
                    for row in js.get("data", []):
                        sym = str(row.get("symbol") or row.get("Symbol") or "").strip().upper()
                        if sym and sym not in _SKIP_SYMBOLS:
                            confirmed.append(sym)
                    if confirmed:
                        break
            except Exception:
                continue
    except Exception:
        pass
    return list(dict.fromkeys(confirmed))


def _safe_fetch_52w(sym: str) -> dict:
    """
    Fetch 1-year OHLCV + full fundamentals for one NSE symbol.
    Returns populated dict or empty dict on any failure.
    FIXED: proper timezone handling, robust field extraction.
    """
    import io as _io2, contextlib as _cl2
    empty = {}
    try:
        buf = _io2.StringIO()
        with _cl2.redirect_stdout(buf), _cl2.redirect_stderr(buf):
            tk   = yf.Ticker(yf_ticker(sym))
            hist = tk.history(period="1y", auto_adjust=True)
            info = {}
            try:
                info = tk.info or {}
            except Exception:
                info = {}

        if hist is None or hist.empty or len(hist) < 30:
            return empty

        # ── FIX: safe timezone stripping ──────────────────────────────────
        idx = pd.to_datetime(hist.index)
        if hasattr(idx, "tz") and idx.tz is not None:
            idx = idx.tz_convert(None)   # tz-aware → tz-naive (CORRECT)
        else:
            try:
                idx = idx.tz_localize(None)  # tz-naive stays tz-naive
            except Exception:
                pass
        hist.index = idx

        # Ensure column names are normalised
        hist.columns = [str(c).strip() for c in hist.columns]
        col_map = {}
        for c in hist.columns:
            cl = c.lower()
            if cl == "close":  col_map[c] = "close"
            elif cl == "open": col_map[c] = "open"
            elif cl == "high": col_map[c] = "high"
            elif cl == "low":  col_map[c] = "low"
            elif cl in ("volume","vol"): col_map[c] = "volume"
        hist = hist.rename(columns=col_map)

        need = {"close","high","low","volume"}
        if not need.issubset(set(hist.columns)):
            return empty

        closes = hist["close"].dropna()
        highs  = hist["high"].dropna()
        lows   = hist["low"].dropna()
        vols   = hist["volume"].fillna(0)

        if len(closes) < 30 or float(closes.iloc[-1]) <= 0:
            return empty

        c_now  = float(closes.iloc[-1])
        c_prev = float(closes.iloc[-2])  if len(closes) >= 2 else c_now
        c_5d   = float(closes.iloc[-6])  if len(closes) >= 6 else c_now
        c_20d  = float(closes.iloc[-21]) if len(closes) >= 21 else c_now

        # ── 52-Week range from full 1y history ────────────────────────────
        w52l_h = float(lows.min())
        w52h_h = float(highs.max())

        # Supplement with yfinance info keys
        def _gf(key, cast=float):
            v = info.get(key)
            try:
                if v is None or str(v) in ("nan","None","inf","-inf",""): return None
                return cast(v)
            except Exception:
                return None

        w52l_i = _gf("fiftyTwoWeekLow")
        w52h_i = _gf("fiftyTwoWeekHigh")
        w52l   = min(w52l_h, w52l_i) if w52l_i and w52l_i > 0 else w52l_h
        w52h   = max(w52h_h, w52h_i) if w52h_i and w52h_i > 0 else w52h_h

        if w52l <= 0 or w52h <= w52l:
            return empty

        pct_above_low  = round((c_now / w52l - 1) * 100, 2)
        pct_below_high = round((c_now / w52h - 1) * 100, 2)
        range_52w_pct  = round((w52h / w52l - 1) * 100, 1)

        # ── Recency: when did this stock last touch 52W low? ──────────────
        # Check last 10 bars
        low_bars = []
        for i in range(1, min(11, len(lows) + 1)):
            if float(lows.iloc[-i]) <= w52l * 1.015:   # within 1.5% of 52W low
                low_bars.append(i)

        is_new_52w_low  = (c_now <= w52l * 1.005)
        touched_today   = (1 in low_bars)
        touched_5d      = bool(low_bars and min(low_bars) <= 5)
        touched_10d     = bool(low_bars and min(low_bars) <= 10)
        days_since_low  = min(low_bars) if low_bars else 999

        chg_1d  = round((c_now / c_prev - 1) * 100, 2) if c_prev > 0 else 0
        chg_5d  = round((c_now / c_5d  - 1) * 100, 2) if c_5d  > 0 else 0
        chg_20d = round((c_now / c_20d - 1) * 100, 2) if c_20d > 0 else 0

        # ── Volume ─────────────────────────────────────────────────────────
        avg20    = float(vols.iloc[-20:].mean()) if len(vols) >= 20 else float(vols.mean())
        avg5     = float(vols.iloc[-5:].mean())  if len(vols) >= 5  else avg20
        vol_now  = float(vols.iloc[-1])
        vol_r    = round(vol_now / avg20, 2)     if avg20 > 0 else 1.0
        vol_t    = round(avg5   / avg20, 2)      if avg20 > 0 else 1.0
        traded_cr= round(c_now  * avg20  / 1e7,  2)

        # ── RSI(14) ────────────────────────────────────────────────────────
        diff   = closes.diff()
        gain   = diff.clip(lower=0).ewm(alpha=1/14, adjust=False, min_periods=14).mean()
        loss   = (-diff.clip(upper=0)).ewm(alpha=1/14, adjust=False, min_periods=14).mean()
        rs     = gain / loss.replace(0, float("nan"))
        rsi_v  = round(float((100 - 100 / (1 + rs)).fillna(50).iloc[-1]), 1)

        # ── EMA 21 / 50 ────────────────────────────────────────────────────
        ema21 = round(float(closes.ewm(span=21, adjust=False).mean().iloc[-1]), 2)
        ema50 = round(float(closes.ewm(span=50, adjust=False).mean().iloc[-1]), 2)

        # ── ATR(14) ────────────────────────────────────────────────────────
        pc     = closes.shift(1)
        tr     = pd.concat([highs - lows, (highs - pc).abs(), (lows - pc).abs()], axis=1).max(axis=1)
        atr_v  = round(float(tr.ewm(alpha=1/14, adjust=False, min_periods=14).mean().iloc[-1]), 2)
        atr_p  = round(atr_v / c_now * 100, 2) if c_now > 0 else 0

        # ── Candlestick (last 2 bars) ──────────────────────────────────────
        candle = ""
        if len(hist) >= 3:
            try:
                _o  = float(hist.get("open", closes).iloc[-1])
                _c  = c_now
                _h  = float(highs.iloc[-1])
                _l  = float(lows.iloc[-1])
                _po = float(hist.get("open", closes).iloc[-2])
                _pc = c_prev
                body    = abs(_c - _o)
                lo_wick = min(_c, _o) - _l
                hi_wick = _h - max(_c, _o)
                rng_    = max(_h - _l, 0.001)
                if lo_wick >= 2 * body and hi_wick <= 0.5 * body and body / rng_ > 0.05:
                    candle = "🔨 Hammer"
                elif _c > _o and _pc < _po and _o <= _pc and _c >= _po:
                    candle = "🕯️ Bullish Engulfing"
                elif body / rng_ < 0.06:
                    candle = "⚖️ Doji"
                elif _c > _o and _c > (_po + _pc) / 2:
                    candle = "🌅 Bullish Close"
            except Exception:
                pass

        # ── Fundamentals (robust, field-by-field) ─────────────────────────
        pe        = _gf("trailingPE")  or _gf("forwardPE")
        roe       = _gf("returnOnEquity")
        eps_g     = _gf("earningsGrowth") or _gf("earningsQuarterlyGrowth")
        rev_g     = _gf("revenueGrowth")
        de        = _gf("debtToEquity")
        pb        = _gf("priceToBook")
        peg       = _gf("pegRatio")
        curr_r    = _gf("currentRatio")
        profit_m  = _gf("profitMargins")
        oper_m    = _gf("operatingMargins")
        div_y     = _gf("dividendYield")
        beta      = _gf("beta")
        target_p  = _gf("targetMeanPrice")
        mcap_raw  = _gf("marketCap") or 0
        mcap_cr   = round(mcap_raw / 1e7, 1) if mcap_raw and mcap_raw > 0 else None
        num_an    = int(_gf("numberOfAnalystOpinions", int) or 0)
        analyst   = str(info.get("recommendationKey") or "").lower().replace("_","")
        sector    = str(info.get("sector")   or "N/A")
        industry  = str(info.get("industry") or "N/A")

        return dict(
            symbol        = sym,
            c_now         = c_now,
            c_prev        = c_prev,
            w52l          = round(w52l, 2),
            w52h          = round(w52h, 2),
            range_52w_pct = range_52w_pct,
            pct_above_low = pct_above_low,
            pct_below_high= pct_below_high,
            is_new_52w_low= is_new_52w_low,
            touched_today = touched_today,
            touched_5d    = touched_5d,
            touched_10d   = touched_10d,
            days_since_low= days_since_low,
            chg_1d=chg_1d, chg_5d=chg_5d, chg_20d=chg_20d,
            pe=pe, roe=roe, eps_g=eps_g, rev_g=rev_g,
            de=de, pb=pb, peg=peg, curr_r=curr_r,
            profit_m=profit_m, oper_m=oper_m,
            mcap_cr=mcap_cr, div_y=div_y, beta=beta,
            sector=sector, industry=industry,
            analyst=analyst, target_price=target_p, num_analyst=num_an,
            vol_ratio=vol_r, vol_trend=vol_t,
            avg_vol=int(avg20), traded_cr=traded_cr,
            rsi=rsi_v, ema21=ema21, ema50=ema50,
            atr=atr_v, atr_pct=atr_p,
            candle=candle,
            is_fo=(sym in _FO_SET),
            indices=symbol_tags(sym),
        )
    except Exception:
        return empty


@st.cache_data(ttl=3600, show_spinner=False)
def _power_scan_52w_recovery(period, top_n, use_smp=False):
    """
    NSE 52W LOW RECOVERY SCANNER — FULL NSE UNIVERSE
    ════════════════════════════════════════════════════════════════
    ① NSE public API → confirmed 52W low symbols today (priority)
    ② yfinance 1y OHLCV + full fundamentals for ALL ~200 symbols
    ③ Hard gate: price within 20% of 52W low
    ④ Fresh 52W Low Detection: touched today / last 5d / new low
    ⑤ Fundamental Score (0–20): P/E · ROE · D/E · Margins · MCap
    ⑥ Growth Score (0–20): EPS · Revenue · Analyst target · PEG
    ⑦ Recency Bonus (0–10): NSE confirmed > today > 5d > 10d > near
    ⑧ Composite Opportunity Score = Fund×2.5 + Growth×2.0 + Recency×2.0
       + RSI×1.5 + Volume×1.2 + Candle×0.8
    ════════════════════════════════════════════════════════════════
    """
    import time as _t3

    nifty    = nifty50_state()
    nse_syms = set(_nse_52w_lows_today())
    all_syms = _ALL_SYMS

    # Priority: NSE-confirmed first, then rest
    scan_order = [s for s in all_syms if s in nse_syms] + \
                 [s for s in all_syms if s not in nse_syms]

    fetched = []
    for sym in scan_order:
        d = _safe_fetch_52w(sym)
        if d:                          # non-empty dict = success
            d["nse_confirmed"] = (sym in nse_syms)
            fetched.append(d)
        _t3.sleep(0.05)

    picks = []

    for d in fetched:
        sym  = d["symbol"]        # ← FIX: always read sym from d, never from outer loop
        c    = d["c_now"]
        w52l = d["w52l"]
        w52h = d["w52h"]

        # ── Hard gate: within 20% of 52W low, range ≥ 10% ───────────────
        if d["pct_above_low"] > 20:
            continue
        if d["range_52w_pct"] < 10:
            continue

        # ═══════════════════════════════════════════════════════════════
        # FUNDAMENTAL SCORE  (0 – 20 pts)
        # ═══════════════════════════════════════════════════════════════
        fs = 0
        ff = []   # fund flags (human readable)

        # P/E Valuation
        pe = d["pe"]
        if pe:
            if   0 < pe < 12:  fs += 4; ff.append(f"P/E {pe:.1f} — Deeply cheap")
            elif pe < 20:      fs += 3; ff.append(f"P/E {pe:.1f} — Attractive")
            elif pe < 30:      fs += 2; ff.append(f"P/E {pe:.1f} — Fair")
            elif pe < 50:      fs += 1; ff.append(f"P/E {pe:.1f} — Slightly high")

        # Return on Equity
        roe = d["roe"]
        if roe:
            if   roe > 0.25:   fs += 4; ff.append(f"ROE {roe*100:.1f}% — World-class")
            elif roe > 0.18:   fs += 3; ff.append(f"ROE {roe*100:.1f}% — Excellent")
            elif roe > 0.12:   fs += 2; ff.append(f"ROE {roe*100:.1f}% — Good")
            elif roe > 0.06:   fs += 1; ff.append(f"ROE {roe*100:.1f}% — Acceptable")

        # Debt / Equity
        de = d["de"]
        if de is not None:
            if   de < 0.3:     fs += 3; ff.append(f"D/E {de:.2f} — Almost debt-free")
            elif de < 0.8:     fs += 2; ff.append(f"D/E {de:.2f} — Low debt")
            elif de < 1.5:     fs += 1; ff.append(f"D/E {de:.2f} — Manageable")
            elif de > 4.0:     fs -= 2; ff.append(f"D/E {de:.2f} — Heavy debt ⚠️")

        # Net Profit Margin
        pm = d["profit_m"]
        if pm:
            if   pm > 0.20:    fs += 2; ff.append(f"Net margin {pm*100:.1f}% — Exceptional")
            elif pm > 0.10:    fs += 1; ff.append(f"Net margin {pm*100:.1f}% — Healthy")
            elif pm < 0:       fs -= 1; ff.append(f"Net margin negative ⚠️")

        # Market Cap
        mc = d["mcap_cr"]
        if mc:
            if   mc > 10000:   fs += 2; ff.append(f"Large Cap ₹{mc:,.0f} Cr (index quality)")
            elif mc > 3000:    fs += 2; ff.append(f"Mid-Large ₹{mc:,.0f} Cr")
            elif mc > 500:     fs += 1; ff.append(f"Mid Cap ₹{mc:,.0f} Cr")

        # Current Ratio (liquidity)
        cr = d["curr_r"]
        if cr:
            if   cr > 2.0:     fs += 1; ff.append(f"Current ratio {cr:.1f} — Very liquid")
            elif cr > 1.2:     ff.append(f"Current ratio {cr:.1f} — Liquid")

        # F&O eligible (institutional quality)
        if d["is_fo"]:
            fs += 1; ff.append("F&O listed — institutional grade")

        # Dividend
        if d["div_y"] and d["div_y"] > 0.015:
            fs += 1; ff.append(f"Dividend {d['div_y']*100:.1f}% yield")

        # ═══════════════════════════════════════════════════════════════
        # GROWTH SCORE  (0 – 20 pts)
        # ═══════════════════════════════════════════════════════════════
        gs = 0
        gf = []   # growth flags

        # EPS Growth
        eg = d["eps_g"]
        if eg:
            if   eg > 0.30:    gs += 5; gf.append(f"EPS growth {eg*100:.1f}% — Accelerating")
            elif eg > 0.15:    gs += 4; gf.append(f"EPS growth {eg*100:.1f}% — Strong")
            elif eg > 0.05:    gs += 2; gf.append(f"EPS growth {eg*100:.1f}% — Positive")
            elif eg < -0.20:   gs -= 2; gf.append(f"EPS declining {eg*100:.1f}% ⚠️")

        # Revenue Growth
        rg = d["rev_g"]
        if rg:
            if   rg > 0.25:    gs += 4; gf.append(f"Revenue growth {rg*100:.1f}% — Rapid")
            elif rg > 0.12:    gs += 3; gf.append(f"Revenue growth {rg*100:.1f}% — Healthy")
            elif rg > 0.04:    gs += 1; gf.append(f"Revenue growth {rg*100:.1f}%")

        # Analyst consensus
        an = d["analyst"]
        na = d["num_analyst"]
        if   "strongbuy" in an or "strongbuy" == an:
            gs += 3; gf.append(f"Analyst: STRONG BUY ({na} analysts)")
        elif an == "buy":
            gs += 2; gf.append(f"Analyst: BUY ({na} analysts)")
        elif an == "hold":
            gs += 1; gf.append(f"Analyst: HOLD")

        # Analyst target price upside
        tp = d["target_price"]
        if tp and c > 0:
            aup = (tp / c - 1) * 100
            if   aup > 40:     gs += 4; gf.append(f"Analyst target ₹{tp:.0f} (+{aup:.0f}%) — Huge")
            elif aup > 20:     gs += 3; gf.append(f"Analyst target ₹{tp:.0f} (+{aup:.0f}%)")
            elif aup > 10:     gs += 2; gf.append(f"Analyst target ₹{tp:.0f} (+{aup:.0f}%)")
            elif aup > 0:      gs += 1; gf.append(f"Analyst target ₹{tp:.0f} (+{aup:.0f}%)")
        else:
            aup = None

        # PEG ratio
        pg = d["peg"]
        if pg:
            if   0 < pg < 0.8:  gs += 3; gf.append(f"PEG {pg:.2f} — Growth deeply discounted")
            elif pg < 1.2:      gs += 2; gf.append(f"PEG {pg:.2f} — Growth at fair price")
            elif pg < 2.0:      gs += 1; gf.append(f"PEG {pg:.2f}")

        # ═══════════════════════════════════════════════════════════════
        # RECENCY SCORE  (0 – 10 pts)
        # ═══════════════════════════════════════════════════════════════
        nse_ok = d.get("nse_confirmed", False)
        rs = 0
        rl = ""

        if nse_ok and (d["is_new_52w_low"] or d["touched_today"]):
            rs = 10; rl = "🔴 NSE CONFIRMED — AT 52W LOW TODAY"
        elif nse_ok:
            rs = 8;  rl = "🟠 NSE CONFIRMED — Recent 52W Low"
        elif d["is_new_52w_low"] and d["touched_today"]:
            rs = 9;  rl = "🆕 NEW 52W LOW — Fresh all-time base"
        elif d["touched_today"]:
            rs = 7;  rl = "🔴 AT 52W LOW TODAY"
        elif d["touched_5d"]:
            rs = 6;  rl = f"🟠 Hit 52W Low {d['days_since_low']} days ago"
        elif d["touched_10d"]:
            rs = 4;  rl = f"🟡 Hit 52W Low {d['days_since_low']} days ago"
        elif d["pct_above_low"] <= 5:
            rs = 3;  rl = f"🟢 Within 5% of 52W Low"
        elif d["pct_above_low"] <= 10:
            rs = 2;  rl = f"📊 Within 10% of 52W Low"
        else:
            rs = 1;  rl = f"📉 {d['pct_above_low']:.1f}% above 52W Low"

        # ═══════════════════════════════════════════════════════════════
        # RSI BONUS  (0 – 5 pts)
        # ═══════════════════════════════════════════════════════════════
        rv  = d["rsi"]
        if   rv < 20:  rb = 5; rz = f"CRITICALLY OVERSOLD ({rv:.0f})"
        elif rv < 25:  rb = 4; rz = f"DEEPLY OVERSOLD ({rv:.0f})"
        elif rv < 30:  rb = 3; rz = f"Oversold ({rv:.0f})"
        elif rv < 40:  rb = 2; rz = f"Near oversold ({rv:.0f})"
        elif rv < 50:  rb = 1; rz = f"Recovering ({rv:.0f})"
        else:          rb = 0; rz = f"Neutral ({rv:.0f})"

        # ═══════════════════════════════════════════════════════════════
        # VOLUME / ACCUMULATION BONUS  (0 – 4 pts)
        # ═══════════════════════════════════════════════════════════════
        vr  = d["vol_ratio"]
        vb  = 0; has_acc = False
        if   vr >= 2.5:  vb = 4; has_acc = True; ff.append(f"🔊 Surge {vr:.1f}× — Institutional")
        elif vr >= 1.8:  vb = 3; has_acc = True; ff.append(f"📈 Volume {vr:.1f}× — Smart money")
        elif vr >= 1.3:  vb = 2; has_acc = True; ff.append(f"Volume elevated {vr:.1f}×")
        elif vr >= 1.1:  vb = 1; has_acc = True; ff.append(f"Volume slightly up {vr:.1f}×")
        if d["vol_trend"] >= 1.3:
            vb = min(vb + 1, 4); ff.append("5-day volume rising")

        # ═══════════════════════════════════════════════════════════════
        # CANDLE BONUS  (0 – 3 pts)
        # ═══════════════════════════════════════════════════════════════
        cv  = d.get("candle", "")
        cb  = 3 if "Engulfing" in cv or "Hammer" in cv else 2 if "Bullish" in cv else 1 if cv else 0

        # ═══════════════════════════════════════════════════════════════
        # COMPOSITE OPPORTUNITY SCORE
        # ═══════════════════════════════════════════════════════════════
        opp = (
            fs  * 2.5
            + gs  * 2.0
            + rs  * 2.0
            + rb  * 1.5
            + vb  * 1.2
            + cb  * 0.8
        )

        # Skip if fundamentals are completely absent AND no strong recency signal
        if fs == 0 and gs == 0 and rs < 5:
            continue

        # ── Price targets ──────────────────────────────────────────────────
        t50  = round(w52l + (w52h - w52l) * 0.50, 2)
        t75  = round(w52l + (w52h - w52l) * 0.75, 2)
        up50 = round((t50 / c - 1) * 100, 1) if c > 0 else 0
        up75 = round((t75 / c - 1) * 100, 1) if c > 0 else 0
        uph  = round((w52h / c - 1) * 100, 1) if c > 0 else 0

        # ── Category ────────────────────────────────────────────────────────
        if rs >= 8 and fs >= 8 and gs >= 5:
            cat = "💎 JACKPOT — Fresh 52W Low + Strong Fund + Growth"
            col = "#26a69a"
        elif rs >= 6 and fs >= 7:
            cat = "🔥 PRIME — Recent 52W Low + Quality"
            col = "#4db6ac"
        elif d["nse_confirmed"] or d["touched_today"] or d["touched_5d"]:
            cat = "🟠 FRESH LOW — Just Touched 52W Low"
            col = "#f59e0b"
        elif fs >= 10 and gs >= 6:
            cat = "🔵 DEEP VALUE — Exceptional Fundamentals"
            col = "#38bdf8"
        elif gs >= 8:
            cat = "🚀 GROWTH GEM — High Growth + Discounted"
            col = "#a855f7"
        else:
            cat = "📋 VALUE WATCH"
            col = "#787b86"

        picks.append(dict(
            # Identification
            symbol      = sym,
            last_close  = c,
            sector      = d["sector"],
            industry    = d["industry"],
            indices     = d["indices"],
            is_fo       = d["is_fo"],
            # 52W metrics
            w52l        = w52l,
            w52h        = w52h,
            range_52w_pct = d["range_52w_pct"],
            pct_from_low  = d["pct_above_low"],
            pct_from_high = d["pct_below_high"],
            discount_pct  = round(abs(d["pct_below_high"]), 1),
            is_new_52w_low= d["is_new_52w_low"],
            touched_today = d["touched_today"],
            touched_5d    = d["touched_5d"],
            touched_10d   = d["touched_10d"],
            days_since_low= d["days_since_low"],
            nse_confirmed = nse_ok,
            recency_label = rl,
            chg_1d  = d["chg_1d"],
            chg_5d  = d["chg_5d"],
            chg_20d = d["chg_20d"],
            # Scores
            opp_score    = round(opp, 2),
            fund_score   = fs,
            growth_score = gs,
            recency_score= rs,
            rsi_bonus    = rb,
            vol_bonus    = vb,
            # Fundamentals (raw)
            pe    = d["pe"],    roe   = d["roe"],
            eps_g = d["eps_g"], rev_g = d["rev_g"],
            de    = d["de"],    pb    = d["pb"],
            peg   = d["peg"],   mcap_cr = d["mcap_cr"],
            profit_m = d["profit_m"], oper_m = d["oper_m"],
            beta  = d["beta"],  div_y = d["div_y"],
            analyst = d["analyst"], target_price = d["target_price"],
            num_analyst = d["num_analyst"], analyst_upside = aup,
            # Technical
            rsi      = rv,       rsi_zone      = rz,
            vol_ratio= vr,       vol_trend     = d["vol_trend"],
            avg_vol  = d["avg_vol"], traded_cr = d["traded_cr"],
            has_accumulation = has_acc,
            ema21    = d["ema21"], ema50 = d["ema50"],
            atr      = d["atr"],   atr_pct = d["atr_pct"],
            candle_signal = cv,
            # Recovery projections
            target_50  = t50,  target_75  = t75,  target_high = w52h,
            upside_50  = up50, upside_75  = up75, upside_to_high = uph,
            target_recovery   = t75,
            upside_to_target  = up75,
            # Flags
            fund_flags   = ff,
            growth_flags = gf,
            opp_category = cat,
            opp_color    = col,
            # Shims for display compatibility
            levels = dict(
                short_term = dict(
                    entry=c, tp=t75, sl=round(w52l * 0.97, 2),
                    risk=round(c - w52l * 0.97, 2),
                    reward=round(t75 - c, 2),
                    rr=round((t75 - c) / max(c - w52l * 0.97, 0.01), 2),
                    rr_str=f"1:{round((t75-c)/max(c-w52l*0.97,0.01),1)}",
                    window="3–9 months"),
                long_term = dict(
                    entry=c, tp=round(w52h, 2), sl=round(w52l * 0.95, 2),
                    risk=round(c - w52l * 0.95, 2),
                    reward=round(w52h - c, 2),
                    rr=round((w52h - c) / max(c - w52l * 0.95, 0.01), 2),
                    rr_str=f"1:{round((w52h-c)/max(c-w52l*0.95,0.01),1)}",
                    window="9–24 months"),
            ),
            ai = dict(
                ai_pct  = min(99, max(1, round((fs + gs + rs) / 0.42, 1))),
                trend_s = 0.0,
                mom_s   = 0.2 if rv < 40 else 0.0,
                brk_s   = 0.0,
                vol_s   = 0.4 if has_acc else 0.0,
                fund_s  = min(1.0, fs / 14.0),
                sent_s  = 0.0,
                pat_s   = 0.0,
            ),
            mkt  = dict(pct=50.0, label="Value Recovery", align="Long-term hold"),
            hits = [(0.8, cv, "Candlestick")] if cv else [],
            pat_conf = 0.45,
            n_cats   = 1,
            score    = round(opp / 60, 4),
            reason   = (f"{rl} | Fund:{fs} Growth:{gs} | "
                        + " | ".join(ff[:3]) + " | " + " | ".join(gf[:2])),
            scan_ts  = datetime.now().strftime("%Y-%m-%d %H:%M"),
            vol_z    = 0.0,
            ema9     = c,
            traded_val_cr = d["traded_cr"],
            st_flip  = 0,
            adx      = 25.0,
        ))

    picks.sort(key=lambda x: -x["opp_score"])
    return picks[:top_n], {}, nifty, ""


# ══════════════════════════════════════════════════════════════════════════════
# SIGNAL TRACKER — persistent JSON store for all generated signals
# ══════════════════════════════════════════════════════════════════════════════

_TRACKER_FILE   = Path("nse_v10_output") / "signal_tracker.json"
_TRACKER_CUTOFF = "2026-03-20"   # Track signals from this date onwards

# ─────────────────────────────────────────────────────────────────────────────
# PERSISTENCE: GitHub Gist (cloud) → local file (fallback)
# ─────────────────────────────────────────────────────────────────────────────
# On Streamlit Cloud the filesystem is ephemeral — data is lost on restart.
# Solution: store tracker JSON in a GitHub Gist via the GitHub API.
# The Gist ID and Personal Access Token are read from st.secrets (set once
# in the Streamlit Cloud dashboard under Settings → Secrets).
#
# st.secrets required keys:
#   [gist]
#   token   = "ghp_xxxxxxxxxxxxxxxxxxxx"   ← GitHub PAT (repo + gist scope)
#   gist_id = "abcdef1234567890abcdef"     ← ID of a Gist you created
#
# If secrets are absent the app silently falls back to local file storage
# (works fine for local development).
# ─────────────────────────────────────────────────────────────────────────────

def _gist_token():
    """Read token from st.secrets — tries nested [gist] section AND flat keys."""
    try:
        # Format 1 (correct): [gist] section → token = "..."
        v = st.secrets.get("gist", {}).get("token")
        if v: return str(v).strip()
    except Exception:
        pass
    try:
        # Format 2 (flat): token = "..." at root level
        v = st.secrets.get("token")
        if v: return str(v).strip()
    except Exception:
        pass
    return None

def _gist_id():
    """Read gist_id from st.secrets — tries nested [gist] section AND flat keys."""
    try:
        v = st.secrets.get("gist", {}).get("gist_id")
        if v: return str(v).strip()
    except Exception:
        pass
    try:
        v = st.secrets.get("gist_id")
        if v: return str(v).strip()
    except Exception:
        pass
    return None

_GIST_FILENAME = "nse_signal_tracker.json"

def _gist_load() -> list | None:
    """Fetch tracker JSON from GitHub Gist. Returns None on failure."""
    import requests as _rq
    token = _gist_token(); gid = _gist_id()
    if not token or not gid:
        return None
    try:
        r = _rq.get(
            f"https://api.github.com/gists/{gid}",
            headers={"Authorization": f"token {token}",
                     "Accept": "application/vnd.github.v3+json"},
            timeout=10,
        )
        if r.status_code == 200:
            files = r.json().get("files", {})
            f     = files.get(_GIST_FILENAME)
            if f:
                raw_url = f.get("raw_url")
                if raw_url:
                    raw = _rq.get(raw_url, timeout=10)
                    if raw.status_code == 200:
                        data = raw.json()
                        return data if isinstance(data, list) else []
    except Exception:
        pass
    return None

def _gist_save(records: list) -> bool:
    """Push tracker JSON to GitHub Gist. Returns True on success."""
    import requests as _rq
    token = _gist_token(); gid = _gist_id()
    if not token or not gid:
        return False
    try:
        payload = {
            "files": {
                _GIST_FILENAME: {
                    "content": json.dumps(records, indent=2, default=str)
                }
            }
        }
        r = _rq.patch(
            f"https://api.github.com/gists/{gid}",
            headers={"Authorization": f"token {token}",
                     "Accept": "application/vnd.github.v3+json"},
            json=payload,
            timeout=15,
        )
        return r.status_code == 200
    except Exception:
        return False

def _tracker_load() -> list:
    """
    Load tracked signals.
    Priority: (1) Gist cloud storage  (2) local JSON file  (3) empty list
    """
    # 1. Try Gist (Streamlit Cloud deployment)
    gist_data = _gist_load()
    if gist_data is not None:
        # Mirror to local file so local dev also has the data
        try:
            _TRACKER_FILE.parent.mkdir(parents=True, exist_ok=True)
            with open(_TRACKER_FILE, "w") as f:
                json.dump(gist_data, f, indent=2, default=str)
        except Exception:
            pass
        return gist_data

    # 2. Fall back to local file
    try:
        _TRACKER_FILE.parent.mkdir(parents=True, exist_ok=True)
        if _TRACKER_FILE.exists():
            with open(_TRACKER_FILE, "r") as f:
                data = json.load(f)
            return data if isinstance(data, list) else []
    except Exception:
        pass
    return []

def _tracker_save(records: list) -> None:
    """
    Save tracked signals.
    Writes to Gist (if configured) AND local file (always).
    """
    # 1. Save to Gist first (cloud persistence)
    _gist_save(records)

    # 2. Always write local file (for local dev + as cache)
    try:
        _TRACKER_FILE.parent.mkdir(parents=True, exist_ok=True)
        with open(_TRACKER_FILE, "w") as f:
            json.dump(records, f, indent=2, default=str)
    except Exception:
        pass

def _tracker_ingest(alerts: list) -> int:
    """
    Called every time a scan completes. Adds new signals to the tracker
    (only signals on or after _TRACKER_CUTOFF, deduped by symbol+scan_date).
    Returns count of newly added records.
    """
    if not alerts:
        return 0
    existing = _tracker_load()
    existing_keys = {(r["symbol"], r["scan_date"]) for r in existing}
    added = 0
    today_str = datetime.now().strftime("%Y-%m-%d")
    # Only ingest if today >= cutoff
    if today_str < _TRACKER_CUTOFF:
        return 0
    for r in alerts:
        sym       = r["symbol"]
        scan_date = datetime.now().strftime("%Y-%m-%d")
        scan_time = datetime.now().strftime("%H:%M")
        if (sym, scan_date) in existing_keys:
            continue  # already recorded today
        stl = r["levels"]["short_term"]
        ltl = r["levels"]["long_term"]
        record = {
            "symbol":        sym,
            "scan_date":     scan_date,
            "scan_time":     scan_time,
            "entry_price":   r["last_close"],
            "st_target":     stl["tp"],
            "st_sl":         stl["sl"],
            "lt_target":     ltl["tp"],
            "lt_sl":         ltl["sl"],
            "st_rr":         stl["rr_str"],
            "lt_rr":         ltl["rr_str"],
            "rsi":           r["rsi"],
            "adx":           r["adx"],
            "atr_pct":       r["atr_pct"],
            "vol_ratio":     r["vol_ratio"],
            "ai_score":      r["ai"]["ai_pct"],
            "comp_score":    r["score"],
            "top_signal":    r["hits"][0][1] if r["hits"] else "",
            "n_signals":     len(r["hits"]),
            "sector":        r.get("sector", "N/A"),
            "indices":       r.get("indices", ""),
            "is_fo":         r.get("is_fo", False),
            "st_flip":       r.get("st_flip", 0),
            "status":        "OPEN",          # OPEN / TARGET_HIT / SL_HIT / CLOSED
            "close_date":    None,
            "close_price":   None,
            "close_reason":  None,
        }
        existing.append(record)
        existing_keys.add((sym, scan_date))
        added += 1
    if added > 0:
        _tracker_save(existing)
    return added


def _tracker_enrich(sym: str, entry_price: float, entry_date: str) -> dict:
    """
    Given just symbol + entry_price + entry_date, fetch live data and compute
    ALL signal fields: ST/LT targets, SL, R:R, RSI, ADX, volume, AI score, etc.
    Returns a complete tracker record ready to save.
    """
    import io as _io3, contextlib as _cl3

    try:
        buf = _io3.StringIO()
        with _cl3.redirect_stdout(buf), _cl3.redirect_stderr(buf):
            tk   = yf.Ticker(yf_ticker(sym))
            hist = tk.history(period="6mo", auto_adjust=True)
            info = {}
            try:
                info = tk.info or {}
            except Exception:
                pass

        if hist is None or hist.empty or len(hist) < 20:
            # Minimal fallback with just what we know
            return _tracker_minimal(sym, entry_price, entry_date)

        # Normalise index timezone
        idx = pd.to_datetime(hist.index)
        if hasattr(idx, "tz") and idx.tz is not None:
            idx = idx.tz_convert(None)
        hist.index = idx
        hist.columns = [str(c).lower().strip() for c in hist.columns]

        closes = hist["close"].dropna()
        highs  = hist["high"].dropna()
        lows   = hist["low"].dropna()
        vols   = hist["volume"].fillna(0)

        if len(closes) < 20:
            return _tracker_minimal(sym, entry_price, entry_date)

        c_now = float(closes.iloc[-1])

        # ── ATR(14) for stop/target levels ────────────────────────────────
        pc_  = closes.shift(1)
        tr_  = pd.concat([highs - lows, (highs-pc_).abs(), (lows-pc_).abs()], axis=1).max(axis=1)
        atr  = float(tr_.ewm(alpha=1/14, adjust=False, min_periods=14).mean().iloc[-1])
        atr_pct = round(atr / c_now * 100, 2) if c_now > 0 else 0

        # ── ST / LT Targets and Stop Losses ──────────────────────────────
        # Use entry_price (not current price) as the reference for levels
        ep   = entry_price if entry_price > 0 else c_now
        # ATR-based levels from entry price
        st_sl  = round(ep - 1.0 * atr, 2)
        st_tp  = round(ep + 1.8 * atr, 2)
        lt_sl  = round(ep - 1.5 * atr, 2)
        lt_tp  = round(ep + 3.5 * atr, 2)
        # R:R
        st_rr  = round((st_tp - ep) / max(ep - st_sl, 0.01), 1)
        lt_rr  = round((lt_tp - ep) / max(ep - lt_sl, 0.01), 1)

        # ── RSI(14) ───────────────────────────────────────────────────────
        diff   = closes.diff()
        gain   = diff.clip(lower=0).ewm(alpha=1/14, adjust=False, min_periods=14).mean()
        loss   = (-diff.clip(upper=0)).ewm(alpha=1/14, adjust=False, min_periods=14).mean()
        rsi_v  = round(float((100 - 100/(1+gain/loss.replace(0,float("nan")))).fillna(50).iloc[-1]), 1)

        # ── ADX(14) ───────────────────────────────────────────────────────
        hd   = highs.diff(); ld = -lows.diff()
        pdm  = pd.Series([(h if h>l and h>0 else 0) for h,l in zip(hd,ld)], index=closes.index)
        mdm  = pd.Series([(l if l>h and l>0 else 0) for h,l in zip(hd,ld)], index=closes.index)
        atr14= tr_.ewm(alpha=1/14, adjust=False, min_periods=14).mean().replace(0, float("nan"))
        pdi  = pdm.ewm(alpha=1/14, adjust=False, min_periods=14).mean() / atr14 * 100
        mdi  = mdm.ewm(alpha=1/14, adjust=False, min_periods=14).mean() / atr14 * 100
        dx   = ((pdi-mdi).abs() / (pdi+mdi).replace(0,float("nan")) * 100)
        adx  = round(float(dx.ewm(alpha=1/14, adjust=False, min_periods=14).mean().fillna(20).iloc[-1]), 1)

        # ── Volume ────────────────────────────────────────────────────────
        avg_vol = float(vols.iloc[-20:].mean()) if len(vols) >= 20 else float(vols.mean())
        vol_r   = round(float(vols.iloc[-1]) / avg_vol, 2) if avg_vol > 0 else 1.0

        # ── EMA alignment ────────────────────────────────────────────────
        ema9   = float(closes.ewm(span=9,  adjust=False).mean().iloc[-1])
        ema21  = float(closes.ewm(span=21, adjust=False).mean().iloc[-1])
        ema50  = float(closes.ewm(span=50, adjust=False).mean().iloc[-1])
        ema_bull = (c_now > ema9 > ema21 > ema50)

        # ── Fundamentals ──────────────────────────────────────────────────
        def _gv(key):
            v = info.get(key)
            try:
                if v is None or str(v) in ("nan","None",""): return None
                return float(v)
            except Exception: return None

        pe      = _gv("trailingPE") or _gv("forwardPE")
        roe     = _gv("returnOnEquity")
        de      = _gv("debtToEquity")
        eps_g   = _gv("earningsGrowth")
        rev_g   = _gv("revenueGrowth")
        mcap_r  = float(info.get("marketCap") or 0)
        mcap_cr = round(mcap_r / 1e7, 1) if mcap_r > 0 else None
        sector  = str(info.get("sector")   or "N/A")
        industry= str(info.get("industry") or "N/A")
        analyst = str(info.get("recommendationKey") or "").lower()
        target_p= _gv("targetMeanPrice")

        # ── Simple AI score estimate ──────────────────────────────────────
        ai_score = 0
        if ema_bull:             ai_score += 25
        if rsi_v < 30:           ai_score += 20
        elif rsi_v < 45:         ai_score += 12
        if adx > 30:             ai_score += 15
        elif adx > 20:           ai_score += 8
        if vol_r >= 1.5:         ai_score += 10
        if pe and 0 < pe < 25:   ai_score += 10
        if roe and roe > 0.15:   ai_score += 10
        if analyst in ("buy","strongbuy"): ai_score += 10
        ai_score = min(ai_score, 99)

        # ── Top signal text ───────────────────────────────────────────────
        sigs = []
        if ema_bull:            sigs.append("EMA Bull Stack")
        if rsi_v < 30:          sigs.append(f"RSI Oversold {rsi_v:.0f}")
        elif rsi_v < 45:        sigs.append(f"RSI Near Oversold {rsi_v:.0f}")
        if adx > 30:            sigs.append(f"ADX Strong {adx:.0f}")
        if vol_r >= 1.5:        sigs.append(f"Vol Surge {vol_r:.1f}x")
        top_sig = " · ".join(sigs[:3]) or "Manual entry"

        w52l = float(info.get("fiftyTwoWeekLow")  or lows.min())
        w52h = float(info.get("fiftyTwoWeekHigh") or highs.max())

        return {
            "symbol":       sym,
            "scan_date":    entry_date,
            "scan_time":    "Manual+Backtest",
            "entry_price":  round(ep, 2),
            "current_price":round(c_now, 2),
            "st_target":    st_tp,
            "st_sl":        st_sl,
            "lt_target":    lt_tp,
            "lt_sl":        lt_sl,
            "st_rr":        f"1:{st_rr}",
            "lt_rr":        f"1:{lt_rr}",
            "atr":          round(atr, 2),
            "atr_pct":      atr_pct,
            "rsi":          rsi_v,
            "adx":          adx,
            "vol_ratio":    vol_r,
            "avg_vol":      int(avg_vol),
            "ema9":         round(ema9, 2),
            "ema21":        round(ema21, 2),
            "ema50":        round(ema50, 2),
            "ema_bull":     ema_bull,
            "ai_score":     ai_score,
            "comp_score":   round(ai_score / 100, 4),
            "top_signal":   top_sig,
            "n_signals":    len(sigs),
            "pe":           round(pe, 1) if pe else None,
            "roe":          round(roe * 100, 1) if roe else None,
            "de":           round(de, 2) if de else None,
            "eps_g":        round(eps_g * 100, 1) if eps_g else None,
            "rev_g":        round(rev_g * 100, 1) if rev_g else None,
            "mcap_cr":      mcap_cr,
            "sector":       sector,
            "industry":     industry,
            "analyst":      analyst.upper(),
            "analyst_target": round(target_p, 2) if target_p else None,
            "w52l":         round(w52l, 2),
            "w52h":         round(w52h, 2),
            "indices":      symbol_tags(sym),
            "is_fo":        sym in _FO_SET,
            "st_flip":      0,
            "status":       "OPEN",
            "close_date":   None,
            "close_price":  None,
            "close_reason": None,
            "added_method": "manual",
        }

    except Exception as _e:
        return _tracker_minimal(sym, entry_price, entry_date)


def _tracker_minimal(sym: str, entry_price: float, entry_date: str) -> dict:
    """Fallback when data fetch fails — stores only what user provided."""
    ep = entry_price if entry_price > 0 else 0
    return {
        "symbol":       sym,
        "scan_date":    entry_date,
        "scan_time":    "Manual",
        "entry_price":  round(ep, 2),
        "current_price":None,
        "st_target":    round(ep * 1.09, 2) if ep > 0 else None,
        "st_sl":        round(ep * 0.96, 2) if ep > 0 else None,
        "lt_target":    round(ep * 1.18, 2) if ep > 0 else None,
        "lt_sl":        round(ep * 0.93, 2) if ep > 0 else None,
        "st_rr":        "1:2.25",
        "lt_rr":        "1:4.5",
        "atr":          None, "atr_pct":None,
        "rsi":          None, "adx":None, "vol_ratio":None,
        "avg_vol":      None, "ema9":None, "ema21":None, "ema50":None,
        "ema_bull":     None,
        "ai_score":     0,    "comp_score":0,
        "top_signal":   "Manually tracked",
        "n_signals":    0,
        "pe":None, "roe":None, "de":None, "eps_g":None, "rev_g":None,
        "mcap_cr":None, "sector":"N/A", "industry":"N/A",
        "analyst":"",   "analyst_target":None,
        "w52l":None,    "w52h":None,
        "indices":      symbol_tags(sym),
        "is_fo":        sym in _FO_SET,
        "st_flip":      0,
        "status":       "OPEN",
        "close_date":   None,
        "close_price":  None,
        "close_reason": None,
        "added_method": "manual-minimal",
    }


def run_dashboard(alerts_in, bt_in, nifty_in, feat_df_in):
    # Unique-key counter — every chart/component gets _uid() as its key
    _counter = [0]
    def _uid(prefix="el"):
        _counter[0] += 1
        return f"{prefix}_{_counter[0]:04d}"

    alerts  = alerts_in  or []
    bt      = bt_in      or {}
    nifty   = nifty_in   or {}

    st.markdown(_CSS, unsafe_allow_html=True)

    # ── Nifty values ──────────────────────────────────────────────────────
    trend=nifty.get("trend",0); last=nifty.get("last",0)
    rsi_n=nifty.get("rsi",50);  chg1m=nifty.get("chg_1m",0)
    chg3m=nifty.get("chg_3m",0); lbl=nifty.get("label","N/A")
    e9=nifty.get("ema9",0); e21=nifty.get("ema21",0); e50=nifty.get("ema50",0)
    nt_col="#26a69a" if trend>=0.5 else "#ef5350" if trend<=-0.5 else "#f59e0b"

    avg_ai=sum(r["ai"]["ai_pct"] for r in alerts)/max(len(alerts),1)
    n_fo=sum(1 for r in alerts if r.get("is_fo"))
    bt_ret=bt.get("ret",0); bt_sh=bt.get("sharpe",0)
    bt_dd=bt.get("maxdd",0); bt_wr=bt.get("winrate",0); bt_tr=bt.get("trades",0)

    # ── Ticker tape ───────────────────────────────────────────────────────
    if alerts:
        components.html(tv_ticker_tape([r["symbol"] for r in alerts]), height=55, scrolling=False)

    # ── Header ────────────────────────────────────────────────────────────
    st.markdown(f"""
    <div class='tv-card' style='margin-bottom:16px;padding:14px 20px'>
      <div style='display:flex;align-items:center;justify-content:space-between;flex-wrap:wrap;gap:10px'>
        <div>
          <div style='font-family:Syne,sans-serif;font-weight:800;font-size:1.4rem;color:#38bdf8;letter-spacing:.08em'>
            📈 NSE SWING TRADER PRO
          </div>
          <div class='tv-label'>{datetime.now().strftime("%d %b %Y  %H:%M IST")}
            &nbsp;·&nbsp; <span class='blink' style='color:#26a69a'>●</span> LIVE
          </div>
        </div>
        <div style='display:flex;gap:22px;flex-wrap:wrap'>
          <div style='text-align:center'>
            <div class='tv-label'>Nifty50</div>
            <div style='font-family:Syne,sans-serif;font-weight:800;font-size:1.3rem;color:#38bdf8'>₹{last:,.2f}</div>
            <div style='font-size:.72rem;color:{nt_col}'>{lbl}</div>
          </div>
          <div style='text-align:center'>
            <div class='tv-label'>1M</div>
            <div style='font-family:Syne,sans-serif;font-weight:700;font-size:1.1rem;color:{"#26a69a" if chg1m>=0 else "#ef5350"}'>{chg1m:+.2f}%</div>
          </div>
          <div style='text-align:center'>
            <div class='tv-label'>Signals</div>
            <div style='font-family:Syne,sans-serif;font-weight:800;font-size:1.3rem;color:#26a69a'>{len(alerts)}</div>
          </div>
          <div style='text-align:center'>
            <div class='tv-label'>F&O</div>
            <div style='font-family:Syne,sans-serif;font-weight:800;font-size:1.3rem;color:#38bdf8'>{n_fo}</div>
          </div>
        </div>
      </div>
    </div>""", unsafe_allow_html=True)

    # ── KPI row ───────────────────────────────────────────────────────────
    kc = st.columns(7)
    kc[0].metric("🟢 Signals",    f"{len(alerts)}", f"F&O: {n_fo}")
    kc[1].metric("🤖 Avg AI",     f"{avg_ai:.1f}%", f"Top: {alerts[0]['ai']['ai_pct']:.1f}%" if alerts else "—")
    kc[2].metric("📊 BT Return",  f"{bt_ret:+.2%}")
    kc[3].metric("📐 Sharpe",     f"{bt_sh:.3f}")
    kc[4].metric("📉 Max DD",     f"{abs(bt_dd):.2%}")
    kc[5].metric("🎯 Win Rate",   f"{bt_wr:.1%}",   f"{bt_tr} trades")
    kc[6].metric("📡 Nifty RSI",  f"{rsi_n:.1f}")

    if not alerts:
        st.warning("⚠️  No signals. Try --threshold 0.18 or lower --min-vol.")
        return

    st.markdown("<br>", unsafe_allow_html=True)

    # ── Main tabs ─────────────────────────────────────────────────────────
    tabs = st.tabs([
        "🌐  Overview",
        "🏭  Sector Intelligence",
        "🎯  Sector Scan",
        "🔥  Power Scan",
        "📋  Signal Cards",
        "📊  Live Charts",
        "🔬  Analysis",
        "📈  Backtest",
        "📡  Signal Tracker",
        "📉  Strategy Analytics",
        "🔍  Screener",
        "📰  News & Calendar",
    ])

    # ══════════════════════════════════════════════════════════════════════
    # TAB 0 — OVERVIEW
    # ══════════════════════════════════════════════════════════════════════
    with tabs[0]:
        # ── Market + Scan summary ─────────────────────────────────────────
        ov_m, ov_r = st.columns([2, 1.2])

        with ov_m:
            st.markdown("<div class='tv-section'>🌐 Indian Market Overview</div>",
                        unsafe_allow_html=True)
            components.html(tv_market_overview(last, lbl), height=420, scrolling=False)

            # ── Grouped Sector Heatmap ────────────────────────────────────
            st.markdown("<div class='tv-section' style='margin-top:20px'>🗺️ Sector Signal Heatmap — All Groups</div>",
                        unsafe_allow_html=True)

            # Build sector stats from alerts
            _ov_gs: dict = {}
            for _r in alerts:
                for _t in _r["indices"].split(" · "):
                    _t = _t.strip()
                    if not _t or _t == "—": continue
                    _d = _ov_gs.setdefault(_t, {"count":0,"ai":[],"syms":[],"stflip":0})
                    _d["count"] += 1; _d["ai"].append(_r["ai"]["ai_pct"])
                    _d["syms"].append(_r["symbol"])
                    if _r.get("st_flip"): _d["stflip"] += 1

            _OV_GF = {"N50L":"Nifty 50 Leaders","N50":"Nifty 50","NN50":"Nifty Next 50",
                      "MC100":"Nifty Midcap 100","SC250":"Nifty Smallcap 250",
                      "BNK":"Nifty Bank","IT":"Nifty IT","NRG":"Nifty Energy",
                      "AUTO":"Nifty Auto","INFRA":"Nifty Infra","F&O":"F&O Stocks"}
            _OV_GI = {"N50L":"👑","N50":"📊","NN50":"🔵","MC100":"🟡","SC250":"🟠",
                      "BNK":"🏦","IT":"💻","NRG":"⚡","AUTO":"🚗","INFRA":"🏗️","F&O":"🔰"}
            _OV_ORDER = ["BNK","IT","NRG","AUTO","INFRA","N50L","N50","NN50","MC100","SC250","F&O"]

            # Group into Large Cap / Sectoral / Broad
            _OV_GROUPS = {
                "🏦 Sectoral Indices (Theme-Based)": ["BNK","IT","NRG","AUTO","INFRA"],
                "👑 Large Cap Indices":              ["N50L","N50","NN50"],
                "📊 Broad Market Indices":           ["MC100","SC250","F&O"],
            }

            for _grp_name, _grp_tags in _OV_GROUPS.items():
                # Check if any sector in this group has signals
                _grp_has = any(t in _ov_gs for t in _grp_tags)
                with st.expander(_grp_name + (" ✅" if _grp_has else " — No signals"), expanded=_grp_has):
                    for _tag in _grp_tags:
                        _ico  = _OV_GI.get(_tag,"📊")
                        _full = _OV_GF.get(_tag, _tag)
                        if _tag not in _ov_gs:
                            st.markdown(
                                f"<div style='display:flex;align-items:center;gap:10px;padding:8px 12px;"
                                f"background:#0b0e11;border-radius:4px;margin-bottom:5px;opacity:.45'>"
                                f"<span style='font-size:1rem'>{_ico}</span>"
                                f"<span style='font-size:.82rem;color:#434651'>{_full}</span>"
                                f"<span style='font-size:.72rem;color:#434651;margin-left:auto'>No signals</span>"
                                f"</div>",
                                unsafe_allow_html=True,
                            )
                            continue

                        _d   = _ov_gs[_tag]; _cnt = _d["count"]; _avg = sum(_d["ai"]) / _cnt
                        _stf = _d["stflip"]
                        if _avg >= 78 and _cnt >= 3:   _hc, _verdict = "#26a69a","🔥 HOT"
                        elif _avg >= 65 or _cnt >= 3:  _hc, _verdict = "#4db6ac","🟢 BULL"
                        elif _avg >= 50 or _cnt >= 2:  _hc, _verdict = "#f59e0b","🟡 NEUTRAL"
                        else:                           _hc, _verdict = "#ef5350","🔴 WEAK"

                        _bar_w = max(4, int(min(_cnt/max(len(alerts)*0.25,1),1)*200))
                        _syms_html = "  ".join(
                            f"<span style='background:{_hc}22;color:{_hc};padding:1px 7px;"
                            f"border-radius:2px;font-size:.72rem;font-weight:600'>{s}</span>"
                            for s in _d["syms"][:6]
                        ) + (f"<span style='color:#434651;font-size:.7rem'> +{len(_d['syms'])-6}</span>"
                             if len(_d["syms"]) > 6 else "")

                        st.markdown(
                            f"<div style='padding:10px 14px;margin-bottom:6px;background:#131722;"
                            f"border-radius:4px;border-left:3px solid {_hc}'>"
                            # Row 1: icon + name + verdict + count + stflip
                            f"<div style='display:flex;align-items:center;gap:10px;margin-bottom:7px'>"
                            f"<span style='font-size:1.1rem'>{_ico}</span>"
                            f"<span style='font-family:Syne,sans-serif;font-weight:700;font-size:.9rem;color:#d1d4dc'>{_full}</span>"
                            f"<span style='font-size:.72rem;color:{_hc};font-weight:700;margin-left:4px'>{_verdict}</span>"
                            f"<span style='font-size:.7rem;color:#787b86;margin-left:auto'>{_cnt} signals</span>"
                            + (f"<span style='background:#a855f722;color:#a855f7;padding:1px 7px;border-radius:2px;font-size:.68rem'>⚡ {_stf} ST Flip</span>" if _stf else "")
                            + f"</div>"
                            # Row 2: heat bar + avg AI
                            f"<div style='display:flex;align-items:center;gap:10px;margin-bottom:8px'>"
                            f"<div style='flex:1;height:5px;background:#1c2030;border-radius:3px'>"
                            f"<div style='height:5px;width:{_bar_w}px;max-width:100%;background:{_hc};border-radius:3px'></div></div>"
                            f"<span style='font-size:.78rem;font-weight:700;color:{_hc};min-width:46px'>AI {_avg:.0f}%</span>"
                            f"<div style='min-width:80px;text-align:right'>"
                            f"{gauge_html(_avg,'',80)}"
                            f"</div></div>"
                            # Row 3: symbol pills
                            f"<div style='line-height:1.8'>{_syms_html}</div>"
                            f"</div>",
                            unsafe_allow_html=True,
                        )

            # TradingView NSE heatmap
            st.markdown("<div class='tv-section' style='margin-top:16px'>🌐 TradingView Live NSE Heatmap</div>",
                        unsafe_allow_html=True)
            components.html(tv_heatmap(), height=480, scrolling=False)

        with ov_r:
            st.markdown("<div class='tv-section'>⭐ Top 8 Signals</div>", unsafe_allow_html=True)
            for r in alerts[:8]:
                ai_p   = r["ai"]["ai_pct"]
                ltr, gc = _d_grade(ai_p)
                stl    = r["levels"]["short_term"]
                up     = (stl["tp"] / stl["entry"] - 1) * 100
                fo_tag = "<span class='pill-cyan'>F&amp;O</span>" if r["is_fo"] else ""
                gauge  = gauge_html(ai_p, "AI Score", 180)
                html_card = (
                    "<div class='tv-card tv-card-bull' style='padding:10px 14px;margin-bottom:8px'>"
                    "<div style='display:flex;justify-content:space-between;align-items:center'>"
                    "<div style='display:flex;align-items:center;gap:8px'>"
                    f"<span class='tv-badge-sym' style='font-size:.85rem;padding:2px 10px'>{r['symbol']}</span>"
                    f"<span class='tv-badge-grade' style='background:{gc}22;color:{gc};font-size:.9rem;padding:2px 8px'>{ltr}</span>"
                    f"{fo_tag}"
                    "</div>"
                    "<div style='text-align:right'>"
                    f"<div style='font-family:Syne,sans-serif;font-weight:700;color:#38bdf8'>&#8377;{r['last_close']:,.2f}</div>"
                    f"<div style='font-size:.7rem;color:#26a69a'>&#9650; {up:+.1f}% to target</div>"
                    "</div></div>"
                    f"<div style='margin-top:5px'>{gauge}</div>"
                    f"<div style='font-size:.68rem;color:#434651;margin-top:3px'>{r['indices']}</div>"
                    "</div>"
                )
                st.markdown(html_card, unsafe_allow_html=True)

            st.markdown("<div class='tv-section' style='margin-top:16px'>📊 Score Distribution</div>",
                        unsafe_allow_html=True)
            st.plotly_chart(score_distribution_fig(alerts), use_container_width=True,
                            config={"displayModeBar": False}, key=_uid("pc"))


    # TAB 1 — SECTOR INTELLIGENCE
    # ══════════════════════════════════════════════════════════════════════
    with tabs[1]:

        # ── SECTOR DEFINITIONS WITH INTEL ────────────────────────────────
        SECTOR_INTEL = {
            "BNK": {
                "name": "Nifty Bank",
                "icon": "🏦",
                "description": "The backbone of Indian equity markets. Banks drive credit growth, NIM expansion, and are highly sensitive to RBI rate decisions and liquidity conditions.",
                "key_drivers": ["RBI rate policy & liquidity","Credit growth (YoY)","NPA & provisioning levels","CASA ratio trends","Net Interest Margins (NIM)"],
                "watch_macro": ["RBI Monetary Policy Committee dates","CPI / WPI inflation prints","10-year G-Sec yield","FII flow into financials","Q results: HDFC Bank, ICICI Bank, SBI"],
                "bull_catalyst": ["Rate cut cycle begins","Credit growth above 14%","NPA ratios declining","Strong CASA growth"],
                "bear_catalyst": ["Rate hikes / prolonged pause","NPA spike","Liquidity tightening","Slowdown in credit"],
                "symbols_all": ["HDFCBANK","ICICIBANK","SBIN","AXISBANK","KOTAKBANK","INDUSINDBK","BANDHANBNK","FEDERALBNK","AUBANK","IDFCFIRSTB","PNB","BANKBARODA","CANBK","INDIANB","UNIONBANK","RBLBANK"],
            },
            "IT": {
                "name": "Nifty IT",
                "icon": "💻",
                "description": "Export-driven sector closely tied to US/Europe tech spending. Revenue in USD makes it a natural hedge against INR depreciation. Margin-sensitive to wage inflation.",
                "key_drivers": ["US IT budgets & deal wins","USD/INR exchange rate","Attrition & wage costs","AI/Cloud adoption pace","Employee utilisation rates"],
                "watch_macro": ["US GDP & unemployment data","Fed rate decisions (USD impact)","Quarterly guidance from Accenture/IBM","TCS & Infosys deal announcements","H-1B visa policy"],
                "bull_catalyst": ["USD strengthens vs INR","US tech spending recovery","Large deal wins","AI-driven revenue acceleration"],
                "bear_catalyst": ["USD weakens","US recession fears","Budget cuts at enterprise clients","Margin pressure from hikes"],
                "symbols_all": ["TCS","INFY","HCLTECH","WIPRO","TECHM","LTM","MPHASIS","COFORGE","PERSISTENT","KPITTECH","TATAELXSI","OFSS","NAUKRI"],
            },
            "NRG": {
                "name": "Nifty Energy",
                "icon": "⚡",
                "description": "Blend of legacy oil & gas and fast-growing renewable energy. Government capex push in solar/wind provides long tailwind while crude prices drive PSU upstream earnings.",
                "key_drivers": ["Global crude oil prices (Brent)","Renewable capacity additions (GW)","Government PLI & subsidy policies","Refining margins (GRM)","Power demand growth"],
                "watch_macro": ["OPEC+ production decisions","INR/USD (crude import cost)","Union Budget energy allocations","REC/IREDA bond issuances","State electricity tariff revisions"],
                "bull_catalyst": ["Crude above $80 (upstream)","Renewable capacity targets raised","Rate cuts boost infra capex","Power demand surge"],
                "bear_catalyst": ["Crude collapse below $65","Subsidy rollback","Interest rate spike hurts capex","Slow renewable execution"],
                "symbols_all": ["RELIANCE","ONGC","NTPC","POWERGRID","TATAPOWER","ADANIGREEN","ADANIPOWER","JSWENERGY","NHPC","IREDA","SUZLON","INOXWIND","WAAREEENER","TORNTPOWER","BPCL","IOC","HINDPETRO","OIL","GAIL","PETRONET","COALINDIA"],
            },
            "AUTO": {
                "name": "Nifty Auto",
                "icon": "🚗",
                "description": "Cyclical sector driven by consumer sentiment, rural income, and the EV transition. Two-wheelers proxy for rural demand; PVs proxy for urban affluence. EV disruption is structural.",
                "key_drivers": ["Monthly auto sales volumes","Rural income & Kharif/Rabi crop output","Fuel prices (petrol/diesel)","EV penetration rate","Input costs (steel, aluminium, semiconductors)"],
                "watch_macro": ["SIAM monthly sales data (2nd of every month)","Monsoon progress (rural sentiment)","Commodity price index","PLI incentives for EV","Budget: excise duty on vehicles"],
                "bull_catalyst": ["Strong festive season sales","Rural recovery post monsoon","EV subsidy expansion","Commodity deflation"],
                "bear_catalyst": ["Weak monsoon → rural slump","Fuel price spike","Commodity cost push","EV disruption to ICE OEMs"],
                "symbols_all": ["MARUTI","M&M","TATAMOTORS","BAJAJ-AUTO","EICHERMOT","HEROMOTOCO","TVSMOTOR","MOTHERSON","BOSCHLTD","TIINDIA","SONACOMS","UNOMINDA","EXIDEIND","BHARATFORG"],
            },
            "INFRA": {
                "name": "Nifty Infra",
                "icon": "🏗️",
                "description": "Government capex-driven sector. National infrastructure pipeline (NIP), roads, railways, urban metro, ports, and housing drive order books. Long cycle, high visibility.",
                "key_drivers": ["Union Budget capex allocation","NHAI & railway order awards","Order book growth & execution","Interest rates (high leverage)","Working capital cycle"],
                "watch_macro": ["Monthly infrastructure output index","NHAI toll collection data","L&T order inflows (industry proxy)","RBI rate trajectory","State capex budgets"],
                "bull_catalyst": ["Budget capex increase","Rate cut cycle","Large order wins","Pre-election government spending"],
                "bear_catalyst": ["Fiscal consolidation → capex cut","Rate hike → higher borrowing cost","Project delays / land acquisition issues","Commodity cost overruns"],
                "symbols_all": ["LT","ADANIPORTS","POWERGRID","NTPC","COALINDIA","BHEL","SIEMENS","ABB","HAVELLS","POLYCAB","KEI","CUMMINSIND","RVNL","NBCC","HUDCO","IRFC","GMRAIRPORT","CONCOR","DELHIVERY","DLF","LODHA"],
            },
            "N50L": {"name":"Nifty 50 Leaders","icon":"👑","description":"Largest-cap blue chips. Lead the broader index. FII ownership is highest here; flows drive sharp moves. Safe haven in risk-off; laggards in risk-on.","key_drivers":["FII net flows","Index rebalancing","Macro GDP outlook","Earnings season consensus"],"watch_macro":["FII/DII daily data","NSE/BSE index changes","RBI policy","US Fed FOMC"],"bull_catalyst":["FII inflows","GDP upgrade","Broad market bull run"],"bear_catalyst":["FII outflows","Global risk-off","Rupee depreciation"],"symbols_all":["RELIANCE","HDFCBANK","ICICIBANK","TCS","INFY","SBIN","BHARTIARTL","LT","AXISBANK","KOTAKBANK","ITC","HINDUNILVR","BAJFINANCE","SUNPHARMA","TITAN","MARUTI","M&M","NTPC","POWERGRID","ADANIPORTS"]},
            "N50":  {"name":"Nifty 50","icon":"📊","description":"India's benchmark index. 50 largest listed companies. Performance benchmark for most funds. Tracks economic cycles closely.","key_drivers":["Broad macro","Earnings growth","Global risk appetite","Domestic flows"],"watch_macro":["PCE / CPI global","RBI bi-monthly policy","Quarterly GDP prints","Q earnings season"],"bull_catalyst":["Earnings upgrade cycle","DII SIP inflows","Global bull market"],"bear_catalyst":["Earnings downgrade","Global recession","FII selling"],"symbols_all":["RELIANCE","HDFCBANK","BHARTIARTL","SBIN","ICICIBANK","TCS","BAJFINANCE","INFY","HINDUNILVR","LT","SUNPHARMA","MARUTI"]},
            "NN50": {"name":"Nifty Next 50","icon":"🔵","description":"Mid-large caps poised for Nifty 50 inclusion. Higher beta than Nifty 50; outperform in bull markets, underperform in corrections.","key_drivers":["Nifty 50 inclusion probability","Index rebalancing flows","Sector rotation"],"watch_macro":["NSE index reconstitution","Mid-cap flows","Earnings visibility"],"bull_catalyst":["Nifty 50 upgrade","Sector bull run","Liquidity expansion"],"bear_catalyst":["Mid-cap selloff","Index exclusion risk","Liquidity tightening"],"symbols_all":["LICI","ADANIGREEN","HAL","SIEMENS","GODREJCP","PIDILITIND","DMART","MARICO","BRITANNIA","HAVELLS","GAIL"]},
            "MC100":{"name":"Nifty Midcap 100","icon":"🟡","description":"High-growth mid-sized companies. Highest earnings growth potential in bull markets. Volatile in corrections but multi-bagger potential in 2–3 year horizon.","key_drivers":["Earnings growth momentum","Domestic institutional flows","Sector-specific tailwinds","Promoter stake changes"],"watch_macro":["BSE Midcap PE vs historical","DII/SIP flows","Smallcap-midcap rotation"],"bull_catalyst":["DII inflows","Domestic consumption boom","Sector re-rating"],"bear_catalyst":["Liquidity crunch","Earnings miss","Mid-cap PE de-rating"],"symbols_all":["TVSMOTOR","CHOLAFIN","MUTHOOTFIN","LUPIN","AUROPHARMA","DIVISLAB","ALKEM","TORNTPHARM","AUBANK","FEDERALBNK","RBLBANK","SRF","ASTRAL"]},
            "SC250":{"name":"Nifty Smallcap 250","icon":"🟠","description":"High-risk, high-reward. Driven by domestic retail flows and SIP money. Illiquid in corrections. Disproportionate upside in broad bull runs.","key_drivers":["Retail investor sentiment","SIP inflows","Stock-specific catalysts","Promoter quality"],"watch_macro":["AMFI SIP data","VIX index","Smallcap vs largecap PE spread","Margin trading data"],"bull_catalyst":["SIP surge","Retail confidence","Momentum factor","Bullish macro"],"bear_catalyst":["VIX spike","Retail panic selling","Liquidity dry-up","Earnings disappoint"],"symbols_all":["IREDA","RVNL","IRFC","NHPC","HUDCO","SJVN","NBCC","PNBHOUSING","MANAPPURAM","INOXWIND","WAAREEENER","OIL","COLPAL","EMAMILTD"]},
            "F&O": {"name":"F&O Stocks","icon":"🔰","description":"Futures & Options eligible stocks — most liquid NSE stocks. Used for hedging, leverage, and directional bets. Option OI data provides sentiment clues.","key_drivers":["Open Interest build-up / unwinding","PCR (Put-Call Ratio)","Max Pain level","IV (Implied Volatility)","Rollover data"],"watch_macro":["Weekly F&O expiry (Thursday)","PCR of Nifty/BankNifty","NSE F&O ban list","VIX trend"],"bull_catalyst":["PCR rising (more puts written)","OI build on calls unwinding","IV compression"],"bear_catalyst":["PCR falling below 0.7","High IV spike","OI build on puts"],"symbols_all":["RELIANCE","HDFCBANK","ICICIBANK","SBIN","TCS","INFY","AXISBANK","KOTAKBANK","LT","BAJFINANCE","WIPRO","HCLTECH"]},
        }

        # ── Build sector stats from alerts ───────────────────────────────
        all_symbols_seen = {r["symbol"] for r in alerts}

        sector_stats: dict = {}
        for r_s in alerts:
            for tag in r_s["indices"].split(" · "):
                tag = tag.strip()
                if not tag or tag == "—": continue
                d = sector_stats.setdefault(tag, {
                    "bullish": [], "all_ai": [], "all_sc": [],
                    "all_rsi": [], "all_adx": [], "all_vol": [],
                })
                d["bullish"].append(r_s)
                d["all_ai"].append(r_s["ai"]["ai_pct"])
                d["all_sc"].append(r_s["score"])
                d["all_rsi"].append(r_s["rsi"])
                d["all_adx"].append(r_s["adx"])
                d["all_vol"].append(r_s["vol_ratio"])

        # Compute sector bullishness scores
        def _sec_bull(d):
            if not d["all_ai"]: return 0
            n = len(d["all_ai"])
            return (
                0.35 * (sum(d["all_ai"]) / n / 100)
                + 0.25 * min(sum(d["all_sc"]) / n / 0.5, 1.0)
                + 0.20 * min(sum(d["all_adx"]) / n / 40, 1.0)
                + 0.10 * min(sum(d["all_vol"]) / n / 2.5, 1.0)
                + 0.10 * max((70 - sum(d["all_rsi"]) / n) / 40, 0)
            )

        TAG_ORDER = ["BNK","IT","NRG","AUTO","INFRA","N50L","N50","NN50","MC100","SC250","F&O"]

        # ── MARKET PULSE HEADER ───────────────────────────────────────────
        bullish_sectors = [t for t in TAG_ORDER if t in sector_stats and _sec_bull(sector_stats[t]) >= 0.55]
        neutral_sectors = [t for t in TAG_ORDER if t in sector_stats and 0.35 <= _sec_bull(sector_stats[t]) < 0.55]
        weak_sectors    = [t for t in TAG_ORDER if t in sector_stats and _sec_bull(sector_stats[t]) < 0.35]
        empty_sectors   = [t for t in TAG_ORDER if t not in sector_stats]

        overall_bias = "#26a69a" if len(bullish_sectors) >= 4 else "#f59e0b" if len(bullish_sectors) >= 2 else "#ef5350"
        overall_txt  = "🐂 BROAD MARKET BULLISH" if len(bullish_sectors) >= 4 else "🔄 MIXED MARKET" if len(bullish_sectors) >= 2 else "🐻 DEFENSIVE — MARKET WEAK"

        st.markdown(f"""
        <div class='tv-card' style='padding:16px 20px;border-color:{overall_bias}99;margin-bottom:16px'>
          <div style='font-family:Syne,sans-serif;font-weight:800;font-size:1.3rem;color:{overall_bias};margin-bottom:10px'>
            {overall_txt}
          </div>
          <div style='display:flex;gap:28px;flex-wrap:wrap'>
            <div>
              <div class='tv-label'>Bullish Sectors</div>
              <div style='font-family:Syne,sans-serif;font-weight:800;font-size:1.5rem;color:#26a69a'>{len(bullish_sectors)}</div>
              <div style='font-size:.72rem;color:#26a69a'>{'  ·  '.join(SECTOR_INTEL.get(t,{}).get('icon','📊')+' '+t for t in bullish_sectors) or '—'}</div>
            </div>
            <div>
              <div class='tv-label'>Neutral Sectors</div>
              <div style='font-family:Syne,sans-serif;font-weight:800;font-size:1.5rem;color:#f59e0b'>{len(neutral_sectors)}</div>
              <div style='font-size:.72rem;color:#f59e0b'>{'  ·  '.join(SECTOR_INTEL.get(t,{}).get('icon','📊')+' '+t for t in neutral_sectors) or '—'}</div>
            </div>
            <div>
              <div class='tv-label'>Weak / No Signals</div>
              <div style='font-family:Syne,sans-serif;font-weight:800;font-size:1.5rem;color:#ef5350'>{len(weak_sectors) + len(empty_sectors)}</div>
              <div style='font-size:.72rem;color:#ef5350'>{'  ·  '.join(SECTOR_INTEL.get(t,{}).get('icon','📊')+' '+t for t in weak_sectors+empty_sectors) or '—'}</div>
            </div>
            <div>
              <div class='tv-label'>Total Bullish Stocks</div>
              <div style='font-family:Syne,sans-serif;font-weight:800;font-size:1.5rem;color:#38bdf8'>{len(alerts)}</div>
            </div>
          </div>
        </div>""", unsafe_allow_html=True)

        # ── SECTOR FILTER ─────────────────────────────────────────────────
        filter_c1, filter_c2 = st.columns([2, 1])
        with filter_c1:
            view_filter = st.radio(
                "Show", ["🟢 Bullish Only", "🟡 Neutral Too", "📋 All Sectors"],
                horizontal=True, key="sec_filter"
            )
        with filter_c2:
            expand_all = st.toggle("Expand All", value=False, key="sec_expand")

        if view_filter == "🟢 Bullish Only":
            display_order = bullish_sectors + neutral_sectors
        elif view_filter == "🟡 Neutral Too":
            display_order = bullish_sectors + neutral_sectors + weak_sectors
        else:
            display_order = [t for t in TAG_ORDER if t in sector_stats] + empty_sectors

        # ── PER-SECTOR DETAILED CARDS ─────────────────────────────────────
        for tag in display_order:
            intel = SECTOR_INTEL.get(tag, {
                "name": tag, "icon": "📊", "description": "",
                "key_drivers": [], "watch_macro": [], "bull_catalyst": [], "bear_catalyst": [],
                "symbols_all": [],
            })

            has_data  = tag in sector_stats
            d         = sector_stats.get(tag, {"bullish":[],"all_ai":[],"all_sc":[],"all_rsi":[],"all_adx":[],"all_vol":[]})
            bull_pct  = _sec_bull(d) if has_data else 0
            n_bull    = len(d["bullish"])
            avg_ai    = sum(d["all_ai"]) / max(n_bull, 1)
            avg_adx   = sum(d["all_adx"]) / max(n_bull, 1)
            avg_rsi   = sum(d["all_rsi"]) / max(n_bull, 1)
            avg_vol   = sum(d["all_vol"]) / max(n_bull, 1)
            avg_sc    = sum(d["all_sc"])  / max(n_bull, 1)

            sec_col   = "#26a69a" if bull_pct >= 0.55 else "#f59e0b" if bull_pct >= 0.35 else "#ef5350"
            sec_label = "🟢 BULLISH" if bull_pct >= 0.55 else "🟡 NEUTRAL" if bull_pct >= 0.35 else "🔴 WEAK"
            sec_label = "⬜ NO SIGNALS" if not has_data else sec_label

            # Identify which sector symbols are bearish (in universe but NOT in alerts)
            all_sym_in_sector = set(intel.get("symbols_all", [])) - _SKIP_SYMBOLS
            bullish_syms  = {r["symbol"] for r in d["bullish"]}
            bearish_syms  = all_sym_in_sector - all_symbols_seen          # not scanned or no signal
            neutral_syms  = (all_symbols_seen & all_sym_in_sector) - bullish_syms  # scanned but no signal

            expander_title = (
                f"{intel['icon']}  {intel['name']}  ·  {sec_label}"
                f"{'  ·  Bullishness: ' + str(round(bull_pct*100)) + '%' if has_data else ''}"
                f"{'  ·  ' + str(n_bull) + ' bullish stocks' if n_bull else ''}"
            )

            with st.expander(expander_title, expanded=(expand_all or bull_pct >= 0.55)):

                # ── ROW A: Description + Key Drivers + Macro Watch ────────
                rA1, rA2, rA3 = st.columns([2, 1.2, 1.2])

                with rA1:
                    st.markdown(
                        f"<div class='tv-card' style='border-left:3px solid {sec_col};height:100%'>"
                        f"<div style='font-family:Syne,sans-serif;font-weight:700;font-size:.95rem;"
                        f"color:{sec_col};margin-bottom:8px'>{intel['icon']} {intel['name']}</div>"
                        f"<div style='font-size:.82rem;color:#787b86;line-height:1.7'>{intel['description']}</div>"
                        f"<div style='margin-top:10px;padding-top:8px;border-top:1px solid #2a3347;"
                        f"display:flex;gap:16px;flex-wrap:wrap'>"
                        f"<span style='font-size:.72rem'>"
                        f"<b style='color:{sec_col}'>Bullishness:</b> <span style='color:#d1d4dc'>{bull_pct*100:.0f}%</span></span>"
                        f"<span style='font-size:.72rem'>"
                        f"<b style='color:#38bdf8'>Bullish Stocks:</b> <span style='color:#26a69a'>{n_bull}</span></span>"
                        f"<span style='font-size:.72rem'>"
                        f"<b style='color:#f59e0b'>Neutral:</b> <span style='color:#d1d4dc'>{len(neutral_syms)}</span></span>"
                        f"</div></div>",
                        unsafe_allow_html=True,
                    )

                with rA2:
                    drivers_html = "".join(
                        f"<div style='display:flex;align-items:flex-start;gap:6px;margin-bottom:5px'>"
                        f"<span style='color:#38bdf8;font-size:.75rem;margin-top:1px'>▸</span>"
                        f"<span style='font-size:.78rem;color:#787b86'>{d_}</span></div>"
                        for d_ in intel.get("key_drivers", [])
                    )
                    st.markdown(
                        f"<div class='tv-card' style='height:100%'>"
                        f"<div class='tv-label' style='margin-bottom:8px'>📌 KEY DRIVERS</div>"
                        f"{drivers_html}</div>",
                        unsafe_allow_html=True,
                    )

                with rA3:
                    macro_html = "".join(
                        f"<div style='display:flex;align-items:flex-start;gap:6px;margin-bottom:5px'>"
                        f"<span style='color:#f59e0b;font-size:.75rem;margin-top:1px'>◆</span>"
                        f"<span style='font-size:.78rem;color:#787b86'>{m_}</span></div>"
                        for m_ in intel.get("watch_macro", [])
                    )
                    st.markdown(
                        f"<div class='tv-card' style='height:100%'>"
                        f"<div class='tv-label' style='margin-bottom:8px'>👁️ WATCH LIST</div>"
                        f"{macro_html}</div>",
                        unsafe_allow_html=True,
                    )

                # ── ROW B: Bull Catalysts + Bear Catalysts ─────────────────
                rB1, rB2 = st.columns(2)
                with rB1:
                    bull_c_html = "".join(
                        f"<div style='display:flex;align-items:flex-start;gap:6px;margin-bottom:5px'>"
                        f"<span style='color:#26a69a;font-size:.8rem'>✅</span>"
                        f"<span style='font-size:.8rem;color:#d1d4dc'>{c_}</span></div>"
                        for c_ in intel.get("bull_catalyst", [])
                    )
                    st.markdown(
                        f"<div class='tv-card' style='border-left:2px solid #26a69a'>"
                        f"<div class='tv-label' style='margin-bottom:8px;color:#26a69a'>🟢 BULLISH CATALYSTS</div>"
                        f"{bull_c_html}</div>",
                        unsafe_allow_html=True,
                    )
                with rB2:
                    bear_c_html = "".join(
                        f"<div style='display:flex;align-items:flex-start;gap:6px;margin-bottom:5px'>"
                        f"<span style='color:#ef5350;font-size:.8rem'>⚠️</span>"
                        f"<span style='font-size:.8rem;color:#d1d4dc'>{c_}</span></div>"
                        for c_ in intel.get("bear_catalyst", [])
                    )
                    st.markdown(
                        f"<div class='tv-card' style='border-left:2px solid #ef5350'>"
                        f"<div class='tv-label' style='margin-bottom:8px;color:#ef5350'>🔴 BEARISH RISKS</div>"
                        f"{bear_c_html}</div>",
                        unsafe_allow_html=True,
                    )

                # ── ROW C: Sector KPIs ─────────────────────────────────────
                if has_data:
                    k1,k2,k3,k4,k5,k6 = st.columns(6)
                    k1.metric("Bullishness",   f"{bull_pct*100:.1f}%")
                    k2.metric("Avg AI Score",  f"{avg_ai:.1f}%")
                    k3.metric("Avg ADX",       f"{avg_adx:.1f}")
                    k4.metric("Avg RSI",       f"{avg_rsi:.1f}")
                    k5.metric("Avg Volume",    f"{avg_vol:.2f}×")
                    k6.metric("Avg Score",     f"{avg_sc:+.3f}")

                # ── ROW D: BULLISH STOCKS ─────────────────────────────────
                if d["bullish"]:
                    st.markdown(
                        f"<div style='font-family:Syne,sans-serif;font-weight:700;font-size:.82rem;"
                        f"letter-spacing:.1em;color:#26a69a;text-transform:uppercase;"
                        f"padding:8px 0 6px;border-bottom:1px solid #1c2030;margin:10px 0 8px'>"
                        f"🟢 BULLISH STOCKS ({n_bull}) — Currently Showing Buy Signals</div>",
                        unsafe_allow_html=True,
                    )

                    sorted_bull = sorted(d["bullish"], key=lambda x: -x["ai"]["ai_pct"])

                    # Show 3-column grid of bullish stock cards
                    cols_per_row = 3
                    rows_b = [sorted_bull[i:i+cols_per_row] for i in range(0, len(sorted_bull), cols_per_row)]
                    for row_b in rows_b:
                        rcols = st.columns(cols_per_row)
                        for ci, r_b in enumerate(row_b):
                            ai_b    = r_b["ai"]["ai_pct"]
                            ltr_b, gc_b = _d_grade(ai_b)
                            stl_b   = r_b["levels"]["short_term"]
                            ltl_b   = r_b["levels"]["long_term"]
                            up_b    = (stl_b["tp"] / stl_b["entry"] - 1) * 100
                            hits_b  = r_b["hits"]
                            sig_b   = hits_b[0][1] if hits_b else "—"
                            fo_b    = "🔰 F&O" if r_b["is_fo"] else ""
                            stf_b   = "⚡ ST FLIP" if r_b.get("st_flip") else ""
                            gb_b    = gauge_html(ai_b, "", 160)
                            rsi_cb  = _d_rsi_col(r_b["rsi"])
                            adx_cb  = _d_adx_col(r_b["adx"])
                            rcols[ci].markdown(
                                f"<div class='tv-card tv-card-bull' style='padding:12px;border-color:{gc_b}55'>"
                                # Header
                                f"<div style='display:flex;justify-content:space-between;align-items:center;margin-bottom:8px'>"
                                f"<span class='tv-badge-sym' style='font-size:.9rem;padding:3px 10px'>{r_b['symbol']}</span>"
                                f"<span class='tv-badge-grade' style='background:{gc_b}22;color:{gc_b};font-size:.9rem'>{ltr_b}</span>"
                                f"</div>"
                                # Price
                                f"<div style='font-family:Syne,sans-serif;font-weight:800;font-size:1.1rem;color:#38bdf8;margin-bottom:4px'>&#8377;{r_b['last_close']:,.2f}</div>"
                                # Tags
                                f"<div style='display:flex;gap:5px;margin-bottom:8px;flex-wrap:wrap'>"
                                + (f"<span class='pill-cyan' style='font-size:.62rem'>{fo_b}</span>" if fo_b else "")
                                + (f"<span class='pill-amber' style='font-size:.62rem'>{stf_b}</span>" if stf_b else "")
                                + f"</div>"
                                # Metrics grid
                                f"<div style='display:grid;grid-template-columns:1fr 1fr;gap:4px;margin-bottom:8px;font-size:.72rem'>"
                                f"<div><span style='color:#434651'>AI:</span> <b style='color:{gc_b}'>{ai_b:.0f}%</b></div>"
                                f"<div><span style='color:#434651'>Score:</span> <b style='color:#d1d4dc'>{r_b['score']:+.3f}</b></div>"
                                f"<div><span style='color:#434651'>RSI:</span> <b style='color:{rsi_cb}'>{r_b['rsi']:.1f}</b></div>"
                                f"<div><span style='color:#434651'>ADX:</span> <b style='color:{adx_cb}'>{r_b['adx']:.1f}</b></div>"
                                f"<div><span style='color:#434651'>Vol:</span> <b style='color:#d1d4dc'>{r_b['vol_ratio']:.2f}&times;</b></div>"
                                f"<div><span style='color:#434651'>ATR:</span> <b style='color:#d1d4dc'>{r_b['atr_pct']:.1f}%</b></div>"
                                f"</div>"
                                # Trade levels
                                f"<div style='background:#1c2030;border-radius:3px;padding:7px;margin-bottom:8px;font-size:.72rem'>"
                                f"<div style='display:flex;justify-content:space-between;margin-bottom:3px'>"
                                f"<span style='color:#434651'>Entry</span><b style='color:#38bdf8'>&#8377;{stl_b['entry']:,.2f}</b></div>"
                                f"<div style='display:flex;justify-content:space-between;margin-bottom:3px'>"
                                f"<span style='color:#434651'>ST Target</span><b style='color:#26a69a'>&#8377;{stl_b['tp']:,.2f} ({up_b:+.1f}%)</b></div>"
                                f"<div style='display:flex;justify-content:space-between;margin-bottom:3px'>"
                                f"<span style='color:#434651'>Stop Loss</span><b style='color:#ef5350'>&#8377;{stl_b['sl']:,.2f}</b></div>"
                                f"<div style='display:flex;justify-content:space-between;margin-bottom:3px'>"
                                f"<span style='color:#434651'>R:R</span><b style='color:#d1d4dc'>{stl_b['rr_str']}</b></div>"
                                f"<div style='display:flex;justify-content:space-between'>"
                                f"<span style='color:#434651'>LT Target</span><b style='color:#4db6ac'>&#8377;{ltl_b['tp']:,.2f}</b></div>"
                                f"</div>"
                                # Gauge + signal
                                f"{gb_b}"
                                f"<div style='font-size:.68rem;color:#434651;margin-top:5px;line-height:1.4'>&#127919; {sig_b[:55]}</div>"
                                "</div>",
                                unsafe_allow_html=True,
                            )

                # ── ROW E: NEUTRAL / NO-SIGNAL STOCKS ─────────────────────
                if neutral_syms:
                    neutral_sorted = sorted(list(neutral_syms))
                    neutral_pills  = "".join(
                        f"<span style='background:#1c2030;color:#787b86;padding:3px 9px;border-radius:3px;"
                        f"font-size:.75rem;margin:2px 3px;display:inline-block'>{sym}</span>"
                        for sym in neutral_sorted
                    )
                    st.markdown(
                        f"<div style='font-family:Syne,sans-serif;font-weight:700;font-size:.82rem;"
                        f"letter-spacing:.1em;color:#f59e0b;text-transform:uppercase;"
                        f"padding:8px 0 6px;border-bottom:1px solid #1c2030;margin:10px 0 8px'>"
                        f"🟡 NEUTRAL / WATCHING ({len(neutral_syms)}) — Scanned but no buy signal yet</div>"
                        f"<div style='margin-bottom:8px'>{neutral_pills}</div>",
                        unsafe_allow_html=True,
                    )

                # ── ROW F: BEARISH / NOT SIGNALLING STOCKS ─────────────────
                # Bearish = in sector universe but NOT in our scanned universe (or below all gates)
                if bearish_syms:
                    bear_sorted = sorted(list(bearish_syms))[:20]
                    bear_pills  = "".join(
                        f"<span style='background:#1c203080;color:#434651;padding:3px 9px;border-radius:3px;"
                        f"font-size:.75rem;margin:2px 3px;display:inline-block;text-decoration:line-through'>{sym}</span>"
                        for sym in bear_sorted
                    )
                    st.markdown(
                        f"<div style='font-family:Syne,sans-serif;font-weight:700;font-size:.82rem;"
                        f"letter-spacing:.1em;color:#ef5350;text-transform:uppercase;"
                        f"padding:8px 0 6px;border-bottom:1px solid #1c2030;margin:10px 0 8px'>"
                        f"🔴 BEARISH / NO SIGNAL ({len(bearish_syms)}) — Did not pass quality gates</div>"
                        f"<div style='margin-bottom:6px'>{bear_pills}"
                        + (f"<span style='font-size:.7rem;color:#434651;margin-left:6px'>+{len(bearish_syms)-20} more</span>"
                           if len(bearish_syms) > 20 else "") +
                        f"</div>",
                        unsafe_allow_html=True,
                    )

                # ── ROW G: Live TradingView chart for top pick ──────────────
                if d["bullish"]:
                    best_sym = sorted(d["bullish"], key=lambda x: -x["ai"]["ai_pct"])[0]["symbol"]
                    st.markdown(
                        f"<div class='tv-label' style='margin:10px 0 4px'>"
                        f"📊 Live Chart — {best_sym} (Top Pick in {intel['name']})</div>",
                        unsafe_allow_html=True,
                    )
                    components.html(
                        tv_mini_chart(best_sym, height=200, nonce=f"sec_{tag}_{best_sym}"),
                        height=204, scrolling=False,
                    )

                st.markdown("<div style='margin-bottom:8px'></div>", unsafe_allow_html=True)

        # ── EMPTY SECTORS ─────────────────────────────────────────────────
        if empty_sectors and view_filter == "📋 All Sectors":
            for tag in empty_sectors:
                intel = SECTOR_INTEL.get(tag, {"name": tag, "icon": "📊"})
                st.markdown(
                    f"<div style='background:#0b0e11;border:1px solid #1c2030;border-radius:4px;"
                    f"padding:12px 16px;margin-bottom:6px;opacity:.5'>"
                    f"<span style='color:#434651;font-size:.82rem'>{intel['icon']} {intel['name']}</span>"
                    f"<span style='color:#434651;font-size:.72rem;margin-left:12px'>No signals in this scan</span>"
                    f"</div>",
                    unsafe_allow_html=True,
                )


    # ══════════════════════════════════════════════════════════════════════
    # TAB 2 — SECTOR SCAN  (engine-backed per-sector signal discovery)
    # ══════════════════════════════════════════════════════════════════════
    with tabs[2]:

        st.markdown("<div class='tv-section'>🎯 Sector-Wise Swing Trade Signal Scan</div>",
                    unsafe_allow_html=True)

        # ── Controls ──────────────────────────────────────────────────────
        sc_ctrl1, sc_ctrl2, sc_ctrl3, sc_ctrl4 = st.columns([1.5, 1, 1, 1])
        with sc_ctrl1:
            scan_sector = st.selectbox(
                "Select Sector to Scan",
                ["All Sectors"] + [
                    "🏦 Nifty Bank",   "💻 Nifty IT",    "⚡ Nifty Energy",
                    "🚗 Nifty Auto",   "🏗️ Nifty Infra",  "👑 Nifty 50 Leaders",
                    "📊 Nifty 50",     "🔵 Nifty Next 50","🟡 Nifty Midcap 100",
                    "🟠 Nifty Smallcap","🔰 F&O Stocks",
                ],
                key="scan_sec_sel",
            )
        with sc_ctrl2:
            scan_threshold = st.slider("Signal Threshold", 0.08, 0.30, 0.14, 0.01,
                                       key="scan_thresh", help="0.12–0.16 recommended")
        with sc_ctrl3:
            scan_top_n = st.slider("Max Results", 5, 30, 15, 5, key="scan_top_n")
        with sc_ctrl4:
            scan_period = st.select_slider(
                "Lookback",
                options=["3mo","4mo","6mo","8mo","1y"],
                value="6mo", key="scan_period"
            )

        sc_c1, sc_c2 = st.columns([2, 1])
        with sc_c1:
            scan_min_vol = st.select_slider(
                "Min Avg Daily Volume",
                options=[100_000, 200_000, 300_000, 500_000, 750_000, 1_000_000],
                value=300_000,
                format_func=lambda x: f"{x/1e5:.1f}L shares",
                key="scan_min_vol",
            )
        with sc_c2:
            scan_fo_only = st.toggle("F&O Eligible Only", value=False, key="scan_fo")

        run_scan_btn = st.button(
            "🚀  Run Sector Scan",
            use_container_width=True,
            key="run_sector_scan",
            help="Fetches live data and runs the full signal engine on the selected sector",
        )

        # ── Map selection to universe ─────────────────────────────────────
        SCAN_MAP = {
            "All Sectors":        None,
            "🏦 Nifty Bank":      "NIFTY BANK",
            "💻 Nifty IT":        "NIFTY IT",
            "⚡ Nifty Energy":    "NIFTY ENERGY",
            "🚗 Nifty Auto":      "NIFTY AUTO",
            "🏗️ Nifty Infra":     "NIFTY INFRA",
            "👑 Nifty 50 Leaders":"NIFTY 50 LEADERS",
            "📊 Nifty 50":        "NIFTY 50",
            "🔵 Nifty Next 50":   "NIFTY NEXT 50",
            "🟡 Nifty Midcap 100":"NIFTY MIDCAP 100",
            "🟠 Nifty Smallcap":  "NIFTY SMALLCAP 250",
            "🔰 F&O Stocks":      "FO STOCKS",
        }

        # ── Cached scan function ──────────────────────────────────────────
        # _sector_scan is defined at module level above run_dashboard
        # ── State management ──────────────────────────────────────────────
        scan_cache_key = (
            scan_sector, scan_threshold, scan_min_vol,
            scan_period, scan_top_n, scan_fo_only,
        )

        if run_scan_btn or ("sector_scan_results" not in st.session_state
                            or st.session_state.get("sector_scan_key") != scan_cache_key):
            if run_scan_btn:
                with st.spinner(f"🔴 Scanning {scan_sector} — fetching live NSE data..."):
                    gk = SCAN_MAP.get(scan_sector)
                    sc_alerts, sc_bt, sc_nifty, sc_err = _sector_scan(
                        gk, scan_threshold, scan_min_vol,
                        scan_period, scan_top_n, scan_fo_only,
                        False,
                    )
                st.session_state["sector_scan_results"] = (sc_alerts, sc_bt, sc_nifty, sc_err)
                st.session_state["sector_scan_key"]     = scan_cache_key
            else:
                sc_alerts, sc_bt, sc_nifty, sc_err = [], {}, {}, ""
        else:
            sc_alerts, sc_bt, sc_nifty, sc_err = st.session_state["sector_scan_results"]

        if sc_err:
            st.error(f"Scan error: {sc_err}")

        if not sc_alerts and not run_scan_btn:
            st.info("👆  Configure parameters above and click **🚀 Run Sector Scan** to discover the best swing trade setups in your chosen sector.", icon="🎯")

        elif sc_alerts:
            sc_nifty_trend = sc_nifty.get("trend", 0)
            sc_avg_ai = sum(r["ai"]["ai_pct"] for r in sc_alerts) / max(len(sc_alerts), 1)
            sc_n_fo   = sum(1 for r in sc_alerts if r.get("is_fo"))
            sc_n_stf  = sum(1 for r in sc_alerts if r.get("st_flip"))

            # ── SCAN RESULT SUMMARY ───────────────────────────────────────
            mkt_col  = "#26a69a" if sc_nifty_trend >= 0.3 else "#ef5350" if sc_nifty_trend <= -0.3 else "#f59e0b"
            st.markdown(
                f"<div class='tv-card' style='padding:14px 20px;border-color:{mkt_col}88;margin-bottom:16px'>"
                f"<div style='display:flex;justify-content:space-between;align-items:center;flex-wrap:wrap;gap:12px'>"
                f"<div>"
                f"<div style='font-family:Syne,sans-serif;font-weight:800;font-size:1.1rem;color:{mkt_col}'>"
                f"🎯 {scan_sector} — Scan Complete</div>"
                f"<div class='tv-label'>Threshold {scan_threshold:.2f} · {scan_period} lookback · Vol &gt;{scan_min_vol/1e5:.1f}L</div>"
                f"</div>"
                f"<div style='display:flex;gap:20px;flex-wrap:wrap'>"
                f"<div style='text-align:center'><div class='tv-label'>Signals Found</div>"
                f"<div style='font-family:Syne,sans-serif;font-weight:800;font-size:1.4rem;color:#26a69a'>{len(sc_alerts)}</div></div>"
                f"<div style='text-align:center'><div class='tv-label'>F&O Eligible</div>"
                f"<div style='font-family:Syne,sans-serif;font-weight:800;font-size:1.4rem;color:#38bdf8'>{sc_n_fo}</div></div>"
                f"<div style='text-align:center'><div class='tv-label'>Avg AI Score</div>"
                f"<div style='font-family:Syne,sans-serif;font-weight:800;font-size:1.4rem;color:#f59e0b'>{sc_avg_ai:.1f}%</div></div>"
                f"<div style='text-align:center'><div class='tv-label'>ST Flips</div>"
                f"<div style='font-family:Syne,sans-serif;font-weight:800;font-size:1.4rem;color:#a855f7'>{sc_n_stf}</div></div>"
                f"<div style='text-align:center'><div class='tv-label'>Nifty</div>"
                f"<div style='font-family:Syne,sans-serif;font-weight:800;font-size:1.4rem;color:{mkt_col}'>{sc_nifty.get('label','N/A')}</div></div>"
                f"</div></div></div>",
                unsafe_allow_html=True,
            )

            # ── QUICK SORT + FILTER ───────────────────────────────────────
            qf1, qf2, qf3 = st.columns([1, 1, 1])
            with qf1:
                sort_mode = st.selectbox(
                    "Sort by",
                    ["AI Score", "Composite Score", "Volume", "ADX Strength",
                     "RSI (Lowest First)", "R:R Ratio"],
                    key="sc_sort",
                    label_visibility="collapsed",
                )
            with qf2:
                min_grade = st.selectbox(
                    "Min Grade",
                    ["All Grades", "B+ and above", "A and above", "A+ only"],
                    key="sc_grade_filter",
                    label_visibility="collapsed",
                )
            with qf3:
                show_stf_only = st.toggle("SuperTrend Flips Only", value=False, key="sc_stf_filter")

            # Apply sort
            sort_fn = {
                "AI Score":          lambda r: -r["ai"]["ai_pct"],
                "Composite Score":   lambda r: -r["score"],
                "Volume":            lambda r: -r["vol_ratio"],
                "ADX Strength":      lambda r: -r["adx"],
                "RSI (Lowest First)":lambda r:  r["rsi"],
                "R:R Ratio":         lambda r: -r["levels"]["short_term"]["rr"],
            }
            # Apply grade filter
            grade_min = {"All Grades":0,"B+ and above":68,"A and above":78,"A+ only":88}
            min_pct = grade_min.get(min_grade, 0)

            filtered = [
                r for r in sc_alerts
                if r["ai"]["ai_pct"] >= min_pct
                and (not show_stf_only or r.get("st_flip"))
            ]
            filtered.sort(key=sort_fn.get(sort_mode, sort_fn["AI Score"]))

            st.markdown(
                f"<div class='tv-label' style='margin:4px 0 12px'>"
                f"Showing <b style='color:#38bdf8'>{len(filtered)}</b> of {len(sc_alerts)} signals</div>",
                unsafe_allow_html=True,
            )

            # ── SIGNAL PANEL — COMPACT SUMMARY TABLE ─────────────────────
            st.markdown("<div class='tv-section'>📊 Signal Summary Table</div>",
                        unsafe_allow_html=True)

            tbl_rows = []
            for r_t in filtered:
                ltr_t, _ = _d_grade(r_t["ai"]["ai_pct"])
                stl_t    = r_t["levels"]["short_term"]
                ltl_t    = r_t["levels"]["long_term"]
                up_t     = (stl_t["tp"] / stl_t["entry"] - 1) * 100
                lt_up_t  = (ltl_t["tp"] / ltl_t["entry"] - 1) * 100
                hits_t   = r_t["hits"]
                tbl_rows.append({
                    "Symbol":    r_t["symbol"],
                    "Grade":     ltr_t,
                    "AI%":       round(r_t["ai"]["ai_pct"], 1),
                    "Score":     round(r_t["score"], 4),
                    "RSI":       round(r_t["rsi"], 1),
                    "ADX":       round(r_t["adx"], 1),
                    "Vol×":      round(r_t["vol_ratio"], 2),
                    "ATR%":      round(r_t["atr_pct"], 2),
                    "F&O":       "✅" if r_t["is_fo"] else "—",
                    "ST⚡":      "⚡" if r_t.get("st_flip") else "—",
                    "Price ₹":   r_t["last_close"],
                    "ST Entry":  stl_t["entry"],
                    "ST Tgt":    stl_t["tp"],
                    "ST SL":     stl_t["sl"],
                    "ST R:R":    stl_t["rr_str"],
                    "ST %Up":    round(up_t, 1),
                    "LT Tgt":    ltl_t["tp"],
                    "LT R:R":    ltl_t["rr_str"],
                    "LT %Up":    round(lt_up_t, 1),
                    "Top Signal":hits_t[0][1] if hits_t else "—",
                    "Indices":   r_t["indices"],
                })

            tbl_df = pd.DataFrame(tbl_rows)
            st.dataframe(
                tbl_df,
                use_container_width=True,
                height=360,
                hide_index=True,
                column_config={
                    "AI%":       st.column_config.ProgressColumn("AI%",      min_value=0,max_value=100, format="%.1f%%"),
                    "Price ₹":   st.column_config.NumberColumn("Price ₹",    format="₹%.2f"),
                    "ST Entry":  st.column_config.NumberColumn("ST Entry",   format="₹%.2f"),
                    "ST Tgt":    st.column_config.NumberColumn("ST Target",  format="₹%.2f"),
                    "ST SL":     st.column_config.NumberColumn("ST SL",      format="₹%.2f"),
                    "LT Tgt":    st.column_config.NumberColumn("LT Target",  format="₹%.2f"),
                    "Score":     st.column_config.NumberColumn("Score",      format="%.4f"),
                    "ST %Up":    st.column_config.NumberColumn("ST %Up",     format="+%.1f%%"),
                    "LT %Up":    st.column_config.NumberColumn("LT %Up",     format="+%.1f%%"),
                },
            )

            # CSV export
            csv_sc = tbl_df.to_csv(index=False).encode()
            st.download_button(
                "⬇️  Download Signal List CSV",
                data=csv_sc,
                file_name=f"sector_scan_{scan_sector.replace(' ','_')}_{datetime.now().strftime('%Y%m%d_%H%M')}.csv",
                mime="text/csv",
                key="sc_csv_dl",
            )

            st.markdown("<hr style='border-color:#2a3347;margin:16px 0'>",
                        unsafe_allow_html=True)

            # ── DETAILED SIGNAL CARDS ─────────────────────────────────────
            st.markdown(
                f"<div class='tv-section'>📋 Detailed Signal Cards — Top {min(len(filtered), scan_top_n)} Swing Setups</div>",
                unsafe_allow_html=True,
            )

            for rank_sc, r_sc in enumerate(filtered[:scan_top_n], 1):
                ai_sc  = r_sc["ai"]; mk_sc = r_sc["mkt"]
                stl_sc = r_sc["levels"]["short_term"]
                ltl_sc = r_sc["levels"]["long_term"]
                hits_sc= r_sc["hits"]
                ai_p_sc= ai_sc["ai_pct"]
                ltr_sc, gc_sc = _d_grade(ai_p_sc)
                up_sc  = (stl_sc["tp"] / stl_sc["entry"] - 1) * 100
                lt_up_sc= (ltl_sc["tp"] / ltl_sc["entry"] - 1) * 100
                stf_sc = r_sc.get("st_flip", 0)
                sig_sc = hits_sc[0][1] if hits_sc else "—"

                # Pre-compute for HTML
                _rsi_c  = _d_rsi_col(r_sc["rsi"])
                _adx_c  = _d_adx_col(r_sc["adx"])
                _vol_c  = _d_vol_col(r_sc["vol_ratio"])
                _fo_tag = "<span class='pill-cyan' style='font-size:.68rem'>F&O ✅</span>" if r_sc["is_fo"] else ""
                _stf_tag= "<span class='pill-amber' style='font-size:.68rem'>⚡ ST FLIP</span>" if stf_sc else ""
                _gauge_sc = gauge_html(ai_p_sc, "", 220)

                with st.expander(
                    f"#{rank_sc}  {r_sc['symbol']}  ·  Grade: {ltr_sc}  ·  "
                    f"AI: {ai_p_sc:.1f}%  ·  ₹{r_sc['last_close']:,.2f}  ·  "
                    f"ST Target: ₹{stl_sc['tp']:,.2f} ({up_sc:+.1f}%)  ·  {r_sc['indices']}",
                    expanded=(rank_sc <= 3),
                ):
                    # ── TOP: Symbol header ─────────────────────────────────
                    st.markdown(
                        f"<div class='tv-card tv-card-bull' style='border-color:{gc_sc}55;padding:14px 18px'>"
                        f"<div style='display:flex;justify-content:space-between;align-items:center;flex-wrap:wrap;gap:12px'>"
                        f"<div style='display:flex;align-items:center;gap:12px'>"
                        f"<span class='tv-badge-sym' style='font-size:1.2rem;padding:5px 16px'>{r_sc['symbol']}</span>"
                        f"<span class='tv-badge-grade' style='background:{gc_sc}22;color:{gc_sc};font-size:1.3rem'>{ltr_sc}</span>"
                        f"{_fo_tag}{_stf_tag}"
                        f"<div>"
                        f"<div class='tv-label'>{r_sc.get('sector','N/A')}</div>"
                        f"<div style='font-size:.7rem;color:#434651'>{r_sc['indices']}</div>"
                        f"</div></div>"
                        f"<div style='display:flex;gap:20px;flex-wrap:wrap'>"
                        f"<div style='text-align:center'><div class='tv-label'>Price</div>"
                        f"<div style='font-family:Syne,sans-serif;font-weight:800;font-size:1.6rem;color:#38bdf8'>&#8377;{r_sc['last_close']:,.2f}</div></div>"
                        f"<div style='text-align:center'><div class='tv-label'>AI Score</div>"
                        f"<div style='font-family:Syne,sans-serif;font-weight:700;font-size:1.3rem;color:{gc_sc}'>{ai_p_sc:.1f}%</div></div>"
                        f"<div style='text-align:center'><div class='tv-label'>RSI</div>"
                        f"<div style='font-family:Syne,sans-serif;font-weight:700;font-size:1.3rem;color:{_rsi_c}'>{r_sc['rsi']:.1f}</div></div>"
                        f"<div style='text-align:center'><div class='tv-label'>ADX</div>"
                        f"<div style='font-family:Syne,sans-serif;font-weight:700;font-size:1.3rem;color:{_adx_c}'>{r_sc['adx']:.1f}</div></div>"
                        f"<div style='text-align:center'><div class='tv-label'>Volume</div>"
                        f"<div style='font-family:Syne,sans-serif;font-weight:700;font-size:1.3rem;color:{_vol_c}'>{r_sc['vol_ratio']:.2f}&times;</div></div>"
                        f"</div></div></div>",
                        unsafe_allow_html=True,
                    )

                    # ── CONFIDENCE ROW ─────────────────────────────────────
                    cc_a, cc_b, cc_c = st.columns(3)
                    for _cw, _pct, _lbl, _clr in [
                        (cc_a, ai_p_sc,             "🤖 AI Confidence",     gc_sc),
                        (cc_b, mk_sc["pct"],         "📊 Market Confidence", "#26a69a" if mk_sc["pct"]>=65 else "#f59e0b" if mk_sc["pct"]>=40 else "#ef5350"),
                        (cc_c, r_sc["pat_conf"]*100, "🎯 Pattern Strength",  "#26a69a" if r_sc["pat_conf"]*100>=60 else "#f59e0b" if r_sc["pat_conf"]*100>=40 else "#ef5350"),
                    ]:
                        _gb = gauge_html(_pct, "", 200)
                        _cw.markdown(
                            f"<div class='tv-card' style='text-align:center;padding:12px'>"
                            f"<div class='tv-label'>{_lbl}</div>"
                            f"<div style='font-family:Syne,sans-serif;font-weight:800;font-size:1.7rem;color:{_clr};margin:6px 0'>{_pct:.1f}%</div>"
                            f"{_gb}</div>",
                            unsafe_allow_html=True,
                        )

                    # ── TRADE LEVELS ───────────────────────────────────────
                    st.markdown("<div class='tv-section' style='margin-top:6px'>📐 Trade Setup</div>",
                                unsafe_allow_html=True)
                    tl_a, tl_b, tl_c = st.columns(3)

                    def _sc_trade_col(col_w, title, entry, target, sl, rr_str, up_pct, window, border_col):
                        risk  = round(entry - sl,    2) if entry > sl    else round(sl - entry, 2)
                        reward= round(target - entry, 2) if target > entry else round(entry - target, 2)
                        col_w.markdown(
                            f"<div class='tv-card' style='border-left:3px solid {border_col};padding:12px;text-align:center'>"
                            f"<div style='color:{border_col};font-size:.78rem;font-weight:700;margin-bottom:10px'>{title}</div>"
                            f"<div class='tv-label'>Entry</div>"
                            f"<div style='font-family:Syne,sans-serif;font-weight:700;font-size:1.1rem;color:#38bdf8;margin-bottom:8px'>&#8377;{entry:,.2f}</div>"
                            f"<div style='display:flex;justify-content:space-around;margin-bottom:8px'>"
                            f"<div><div class='tv-label'>Target</div><div style='font-weight:700;color:#26a69a'>&#8377;{target:,.2f}</div>"
                            f"<div style='font-size:.68rem;color:#26a69a'>{up_pct:+.1f}%</div></div>"
                            f"<div><div class='tv-label'>Stop Loss</div><div style='font-weight:700;color:#ef5350'>&#8377;{sl:,.2f}</div>"
                            f"<div style='font-size:.68rem;color:#ef5350'>{(sl/entry-1)*100:+.1f}%</div></div>"
                            f"</div>"
                            f"<div style='display:flex;justify-content:space-around;padding-top:6px;border-top:1px solid #2a3347'>"
                            f"<div><div class='tv-label'>R:R</div><div style='color:#d1d4dc;font-weight:700'>{rr_str}</div></div>"
                            f"<div><div class='tv-label'>Risk ₹</div><div style='color:#ef5350'>&#8377;{risk:.2f}</div></div>"
                            f"<div><div class='tv-label'>Reward ₹</div><div style='color:#26a69a'>&#8377;{reward:.2f}</div></div>"
                            f"</div>"
                            f"<div style='margin-top:6px;font-size:.65rem;color:#434651'>{window}</div>"
                            f"</div>",
                            unsafe_allow_html=True,
                        )

                    _sc_trade_col(tl_a, "⚡ Aggressive (ST)", stl_sc["entry"], stl_sc["tp"], stl_sc["sl"], stl_sc["rr_str"], up_sc,    "2–5 trading days",  "#f59e0b")
                    _sc_trade_col(tl_b, "📅 Swing (LT)",      ltl_sc["entry"], ltl_sc["tp"], ltl_sc["sl"], ltl_sc["rr_str"], lt_up_sc, "10–20 trading days","#26a69a")
                    _dip_e  = round(r_sc["last_close"] * 0.98, 2)
                    _dip_sl = round(_dip_e - r_sc["atr"] * 1.2, 2)
                    _dip_tp = round(_dip_e + r_sc["atr"] * 2.5, 2)
                    _dip_rr = round((_dip_tp - _dip_e) / max(_dip_e - _dip_sl, 0.01), 2)
                    _sc_trade_col(tl_c, "📌 Limit/Dip Entry", _dip_e, _dip_tp, _dip_sl, f"1:{_dip_rr}", (_dip_tp/_dip_e-1)*100, "Limit at −2%", "#3b82f6")

                    # ── 8-FACTOR SCORES ────────────────────────────────────
                    st.markdown("<div class='tv-section' style='margin-top:4px'>🧮 8-Factor Score Breakdown</div>",
                                unsafe_allow_html=True)
                    fa_cols = st.columns(8)
                    for fi, (factor_name, factor_key, factor_wt) in enumerate([
                        ("Trend",       "trend_s", 0.24), ("Momentum", "mom_s",  0.16),
                        ("Breakout",    "brk_s",   0.17), ("Volume",   "vol_s",  0.10),
                        ("Pattern",     "pat_s",   0.10), ("Fund",     "fund_s", 0.08),
                        ("Sentiment",   "sent_s",  0.04), ("Pullback", "trend_s",0.11),
                    ]):
                        _fsc  = float(ai_sc.get(factor_key, 0))
                        _fc   = "#26a69a" if _fsc > 0.1 else "#ef5350" if _fsc < -0.1 else "#f59e0b"
                        _farr = "▲" if _fsc > 0.1 else "▼" if _fsc < -0.1 else "◆"
                        fa_cols[fi].markdown(
                            f"<div style='background:#1c2030;border-radius:3px;padding:8px 6px;text-align:center'>"
                            f"<div class='tv-label'>{factor_name}</div>"
                            f"<div style='font-family:Syne,sans-serif;font-weight:700;color:{_fc};font-size:.9rem'>{_fsc:+.2f}</div>"
                            f"<div style='font-size:.75rem;color:{_fc}'>{_farr} {factor_wt:.0%}</div>"
                            f"</div>",
                            unsafe_allow_html=True,
                        )

                    # ── PATTERN HITS + AI GAUGE + RESEARCH NOTE ───────────
                    ph_col, rn_col = st.columns([1, 1.4])

                    with ph_col:
                        st.markdown("<div class='tv-section' style='margin-top:6px'>🎯 Pattern Signals</div>",
                                    unsafe_allow_html=True)
                        CAT_C_SC = {
                            "Trend":"#38bdf8","Momentum":"#f59e0b","Candlestick":"#a855f7",
                            "Breakout":"#26a69a","Volume":"#3b82f6","Volatility":"#787b86",
                            "Price Action":"#d1d4dc","Structure":"#67e8f9",
                        }
                        for _sc_h, _lb_h, _cat_h in hits_sc[:7]:
                            _cc_h = CAT_C_SC.get(_cat_h, "#787b86")
                            _bw_h = int(_sc_h * 140)
                            st.markdown(
                                f"<div style='display:flex;align-items:center;gap:8px;padding:5px 8px;"
                                f"margin-bottom:4px;background:#1c2030;border-radius:3px;"
                                f"border-left:2px solid {_cc_h}'>"
                                f"<span style='font-size:.68rem;color:{_cc_h};min-width:64px'>{_cat_h}</span>"
                                f"<div style='height:4px;width:{_bw_h}px;background:{_cc_h};border-radius:2px'></div>"
                                f"<span style='font-size:.75rem;color:#d1d4dc;flex:1'>{_lb_h[:38]}</span>"
                                f"<span style='font-size:.7rem;color:{_cc_h};font-weight:700'>{_sc_h*100:.0f}%</span>"
                                f"</div>",
                                unsafe_allow_html=True,
                            )
                        st.markdown(
                            f"<div style='font-size:.7rem;color:#434651;padding:4px 0'>"
                            f"Total: {len(hits_sc)} signals across {r_sc['n_cats']} categories</div>",
                            unsafe_allow_html=True,
                        )

                    with rn_col:
                        st.markdown("<div class='tv-section' style='margin-top:6px'>📝 Research Note</div>",
                                    unsafe_allow_html=True)
                        st.markdown(
                            f"<div class='tv-card' style='font-size:.82rem;color:#787b86;line-height:1.75'>"
                            f"{r_sc['reason'].replace('  •  ', '<br><span style=\"color:#38bdf8\">→</span> ')}"
                            f"<div style='margin-top:10px;padding-top:8px;border-top:1px solid #2a3347;"
                            f"font-size:.72rem;color:#434651;display:grid;grid-template-columns:1fr 1fr;gap:4px'>"
                            f"<span>EMA 9/21/50: <b style='color:#d1d4dc'>{r_sc['ema9']:.1f} / {r_sc['ema21']:.1f} / {r_sc['ema50']:.1f}</b></span>"
                            f"<span>ATR: <b style='color:#d1d4dc'>{r_sc['atr_pct']:.2f}%</b></span>"
                            f"<span>Vol Z-Score: <b style='color:#d1d4dc'>{r_sc.get('vol_z',0):.2f}&sigma;</b></span>"
                            f"<span>Traded: <b style='color:#d1d4dc'>&#8377;{r_sc['traded_val_cr']:.2f} Cr/d</b></span>"
                            f"<span>Composite: <b style='color:#d1d4dc'>{r_sc['score']:+.4f}</b></span>"
                            f"<span>Market: <b style='color:#d1d4dc'>{mk_sc['label']}</b></span>"
                            f"</div></div>",
                            unsafe_allow_html=True,
                        )
                        st.markdown("<div style='margin-top:8px'></div>", unsafe_allow_html=True)
                        st.markdown(_gauge_sc, unsafe_allow_html=True)

                    # ── TradingView mini chart ─────────────────────────────
                    st.markdown(
                        f"<div class='tv-label' style='margin:8px 0 3px'>📊 Live Chart — {r_sc['symbol']}</div>",
                        unsafe_allow_html=True,
                    )
                    components.html(
                        tv_mini_chart(r_sc["symbol"], height=190,
                                      nonce=f"scscan_{r_sc['symbol']}_{rank_sc}"),
                        height=194, scrolling=False,
                    )

    # TAB 3 — POWER SCAN
    # ══════════════════════════════════════════════════════════════════════
    with tabs[3]:
        st.markdown("""
        <div style='padding:14px 20px;background:linear-gradient(135deg,#ef535018,#a855f718,#38bdf818,#0b0e11);
                    border-left:4px solid #ef5350;border-radius:6px;margin-bottom:16px'>
          <div style='font-family:Syne,sans-serif;font-weight:800;font-size:1.25rem;color:#ef5350;letter-spacing:.06em'>
            🔥 POWER SCAN — Live NSE Universe
          </div>
          <div style='font-size:.76rem;color:#787b86;margin-top:3px'>
            ⚡ High-Volatility Momentum Movers &nbsp;·&nbsp; 💎 Multi-Bagger Penny Opportunities
            &nbsp;·&nbsp; Full engine scan across ~200 NSE symbols &nbsp;·&nbsp; Click any result for deep analysis
          </div>
        </div>""", unsafe_allow_html=True)

        # ── Controls ──────────────────────────────────────────────────────
        psc1, psc2, psc3 = st.columns(3)
        with psc1:
            ps_threshold = st.slider("Signal Threshold", 0.08, 0.30, 0.14, 0.01, key="ps_thresh",
                                     help="Lower = more signals. 0.14 recommended for power scan.")
        with psc2:
            ps_period = st.select_slider("Lookback Period",
                                          ["3mo","4mo","6mo","8mo","1y"], value="6mo", key="ps_period")
        with psc3:
            ps_top_n = st.slider("Results to Show", 5, 20, 10, 1, key="ps_top_n")

        run_ps_btn = st.button("🚀  Run Power Scan — Full NSE Universe",
                               use_container_width=True, key="run_power_scan")

        ps_tab1, ps_tab2, ps_tab3 = st.tabs(["⚡ High-Volatility Momentum", "💎 Multi-Bagger Penny Picks", "🔍 52W Low Recovery Gems"])

        # ── State ─────────────────────────────────────────────────────────
        ps_key = (ps_threshold, ps_period, ps_top_n)
        if run_ps_btn or ("ps_results" not in st.session_state
                          or st.session_state.get("ps_key") != ps_key):
            if run_ps_btn:
                with st.spinner("⚡ Scanning high-volatility movers across full NSE universe..."):
                    _ps_vol, _ps_vbt, _ps_vnifty, _ps_verr = _power_scan_volatile(
                        ps_threshold, ps_period, ps_top_n, False)
                with st.spinner("💎 Discovering multi-bagger penny opportunities..."):
                    _ps_pny, _ps_pbt, _ps_pnifty, _ps_perr = _power_scan_penny(
                        ps_threshold, ps_period, ps_top_n, False)
                with st.spinner("🔍 Finding fundamentally strong stocks near 52W lows..."):
                    _ps_rec, _ps_rbt, _ps_rnifty, _ps_rerr = _power_scan_52w_recovery(
                        ps_period, ps_top_n, False)
                st.session_state["ps_results"] = (_ps_vol, _ps_pny, _ps_vbt, _ps_pbt, _ps_rec, _ps_rbt)
                st.session_state["ps_key"]     = ps_key
            else:
                _ps_vol = _ps_pny = _ps_rec = []; _ps_vbt = _ps_pbt = _ps_rbt = {}
        else:
            _ps_vol, _ps_pny, _ps_vbt, _ps_pbt, _ps_rec, _ps_rbt = st.session_state["ps_results"]

        if not _ps_vol and not _ps_pny and not _ps_rec and not run_ps_btn:
            st.info(
                "👆 Click **🚀 Run Power Scan** to discover momentum movers, "
                "multi-bagger penny stocks, and 52W low recovery gems.\n\n"
                "⏱️ Live scan takes **4–10 minutes** (includes fundamentals for recovery scan). "
                "Results cached for 1 hour.",
                icon="🔥",
            )

        # ══════════════════════════════════════════════════════════════════
        # HELPER: full interactive stock detail panel
        # ══════════════════════════════════════════════════════════════════
        def _render_stock_detail(r_d, source_label="Power Scan"):
            """Render the complete deep-dive panel for one stock."""
            ai_d   = r_d["ai"];  mk_d = r_d["mkt"]
            stl_d  = r_d["levels"]["short_term"]
            ltl_d  = r_d["levels"]["long_term"]
            hits_d = r_d["hits"]
            ai_p_d = ai_d["ai_pct"]
            ltr_d, gc_d = _d_grade(ai_p_d)
            up_d   = (stl_d["tp"] / stl_d["entry"] - 1) * 100
            lt_up_d= (ltl_d["tp"] / ltl_d["entry"] - 1) * 100
            stf_d  = r_d.get("st_flip", 0) or r_d.get("has_stf", False)

            _rsi_cd = _d_rsi_col(r_d["rsi"])
            _adx_cd = _d_adx_col(r_d["adx"])
            _vol_cd = _d_vol_col(r_d["vol_ratio"])

            # ── Header ────────────────────────────────────────────────────
            _fo_h  = "<span class='pill-cyan' style='font-size:.7rem'>F&O ✅</span>" if r_d.get("is_fo") else ""
            _stf_h = "<span class='pill-amber' style='font-size:.7rem'>⚡ ST FLIP</span>" if stf_d else ""
            _brk_h = "<span style='background:#26a69a22;color:#26a69a;padding:2px 8px;border-radius:2px;font-size:.7rem'>🚀 BREAKOUT</span>" if r_d.get("has_breakout") else ""
            _vol_h = "<span style='background:#3b82f622;color:#3b82f6;padding:2px 8px;border-radius:2px;font-size:.7rem'>🔊 VOL SURGE</span>" if r_d.get("has_volume_surge") else ""

            st.markdown(
                f"<div class='tv-card tv-card-bull' style='border-color:{gc_d}66;padding:16px 20px'>"
                f"<div style='display:flex;justify-content:space-between;align-items:center;flex-wrap:wrap;gap:14px'>"
                f"<div style='display:flex;align-items:center;gap:12px;flex-wrap:wrap'>"
                f"<span class='tv-badge-sym' style='font-size:1.3rem;padding:5px 18px'>{r_d['symbol']}</span>"
                f"<span class='tv-badge-grade' style='background:{gc_d}22;color:{gc_d};font-size:1.3rem'>{ltr_d}</span>"
                f"{_fo_h}{_stf_h}{_brk_h}{_vol_h}"
                f"<div><div class='tv-label'>{r_d.get('sector','N/A')}</div>"
                f"<div style='font-size:.7rem;color:#434651'>{r_d['indices']}</div></div>"
                f"</div>"
                f"<div style='display:flex;gap:18px;flex-wrap:wrap'>"
                f"<div style='text-align:center'><div class='tv-label'>Price</div>"
                f"<div style='font-family:Syne,sans-serif;font-weight:800;font-size:1.7rem;color:#38bdf8'>&#8377;{r_d['last_close']:,.2f}</div></div>"
                f"<div style='text-align:center'><div class='tv-label'>AI Score</div>"
                f"<div style='font-family:Syne,sans-serif;font-weight:700;font-size:1.3rem;color:{gc_d}'>{ai_p_d:.1f}%</div></div>"
                f"<div style='text-align:center'><div class='tv-label'>RSI</div>"
                f"<div style='font-family:Syne,sans-serif;font-weight:700;font-size:1.3rem;color:{_rsi_cd}'>{r_d['rsi']:.1f}</div></div>"
                f"<div style='text-align:center'><div class='tv-label'>ADX</div>"
                f"<div style='font-family:Syne,sans-serif;font-weight:700;font-size:1.3rem;color:{_adx_cd}'>{r_d['adx']:.1f}</div></div>"
                f"<div style='text-align:center'><div class='tv-label'>ATR%</div>"
                f"<div style='font-family:Syne,sans-serif;font-weight:700;font-size:1.3rem;color:#f59e0b'>{r_d['atr_pct']:.2f}%</div></div>"
                f"<div style='text-align:center'><div class='tv-label'>Vol</div>"
                f"<div style='font-family:Syne,sans-serif;font-weight:700;font-size:1.3rem;color:{_vol_cd}'>{r_d['vol_ratio']:.2f}&times;</div></div>"
                f"</div></div></div>",
                unsafe_allow_html=True,
            )

            # ── Section A: WHY SELECTED ───────────────────────────────────
            st.markdown("""
            <div style='margin:14px 0 10px;padding:8px 16px;background:linear-gradient(90deg,#38bdf815,#0b0e11);
                        border-left:3px solid #38bdf8;border-radius:3px'>
              <span style='font-family:Syne,sans-serif;font-weight:700;font-size:.85rem;color:#38bdf8;letter-spacing:.08em'>
                🎯 WHY THIS STOCK IS SELECTED
              </span>
            </div>""", unsafe_allow_html=True)

            ws1, ws2 = st.columns([1.2, 1])
            with ws1:
                # Pattern hits grouped
                CAT_C_D = {"Trend":"#38bdf8","Momentum":"#f59e0b","Candlestick":"#a855f7",
                           "Breakout":"#26a69a","Volume":"#3b82f6","Volatility":"#ef5350",
                           "Price Action":"#d1d4dc","Structure":"#67e8f9"}
                for sh,slb,scat in hits_d[:8]:
                    scc = CAT_C_D.get(scat,"#787b86")
                    bw  = int(sh*180)
                    st.markdown(
                        f"<div style='display:flex;align-items:center;gap:9px;padding:6px 10px;"
                        f"margin-bottom:4px;background:#131722;border-radius:3px;border-left:2px solid {scc}'>"
                        f"<span style='font-size:.65rem;color:{scc};font-weight:600;min-width:80px'>{scat}</span>"
                        f"<div style='height:4px;width:{bw}px;background:{scc};border-radius:2px;flex-shrink:0'></div>"
                        f"<span style='font-size:.8rem;color:#d1d4dc;flex:1'>{slb}</span>"
                        f"<b style='font-size:.72rem;color:{scc}'>{sh*100:.0f}%</b></div>",
                        unsafe_allow_html=True,
                    )
                st.markdown(
                    f"<div style='font-size:.72rem;color:#434651;margin-top:5px'>"
                    f"Total: {len(hits_d)} signals across {r_d['n_cats']} categories &nbsp;·&nbsp; "
                    f"Pattern confidence: {r_d['pat_conf']*100:.1f}%</div>",
                    unsafe_allow_html=True,
                )
            with ws2:
                st.markdown(
                    f"<div class='tv-card' style='border-left:3px solid #38bdf8;padding:12px;font-size:.82rem;color:#787b86;line-height:1.75'>"
                    f"<div class='tv-label' style='margin-bottom:6px'>📝 Engine Research Note</div>"
                    f"{r_d['reason'].replace('  •  ','<br><span style=\"color:#38bdf8\">→</span> ')}"
                    f"<div style='margin-top:8px;padding-top:7px;border-top:1px solid #2a3347;font-size:.7rem;color:#434651'>"
                    f"EMA9/21/50: {r_d['ema9']:.1f} / {r_d['ema21']:.1f} / {r_d['ema50']:.1f} &nbsp;·&nbsp; "
                    f"Score: {r_d['score']:+.4f} &nbsp;·&nbsp; ATR ₹{r_d['atr']:.2f}"
                    f"</div></div>",
                    unsafe_allow_html=True,
                )

            # ── Section B: CONFIDENCE ─────────────────────────────────────
            st.markdown("""
            <div style='margin:14px 0 10px;padding:8px 16px;background:linear-gradient(90deg,#a855f715,#0b0e11);
                        border-left:3px solid #a855f7;border-radius:3px'>
              <span style='font-family:Syne,sans-serif;font-weight:700;font-size:.85rem;color:#a855f7;letter-spacing:.08em'>
                🤖 AI CONFIDENCE ANALYSIS
              </span>
            </div>""", unsafe_allow_html=True)

            c1,c2,c3,c4 = st.columns(4)
            for cw,pct_,lbl_,clr_ in [
                (c1, ai_p_d,              "🤖 AI Model",    gc_d),
                (c2, mk_d["pct"],         "📊 Market",      "#26a69a" if mk_d["pct"]>=65 else "#f59e0b" if mk_d["pct"]>=40 else "#ef5350"),
                (c3, r_d["pat_conf"]*100, "🎯 Patterns",    "#26a69a" if r_d["pat_conf"]*100>=55 else "#f59e0b"),
                (c4, (ai_d.get("trend_s",0)+1)/2*100, "📈 Trend", "#26a69a" if ai_d.get("trend_s",0)>0.3 else "#f59e0b"),
            ]:
                _gb = gauge_html(pct_,"",160)
                cw.markdown(
                    f"<div style='background:#131722;border-radius:4px;padding:10px;text-align:center;border-top:2px solid {clr_}'>"
                    f"<div class='tv-label'>{lbl_}</div>"
                    f"<div style='font-family:Syne,sans-serif;font-weight:800;font-size:1.3rem;color:{clr_};margin:6px 0'>{pct_:.1f}%</div>"
                    f"{_gb}</div>",
                    unsafe_allow_html=True,
                )

            # 8-factor mini bars
            st.markdown("<div style='display:flex;gap:6px;margin-top:8px;flex-wrap:wrap'>", unsafe_allow_html=True)
            for fn,fk in [("Trend","trend_s"),("Mom","mom_s"),("Brk","brk_s"),("Vol","vol_s"),("Pat","pat_s"),("Fund","fund_s"),("Sent","sent_s")]:
                fv = float(ai_d.get(fk,0))
                fc_ = "#26a69a" if fv>0.1 else "#ef5350" if fv<-0.1 else "#f59e0b"
                st.markdown(
                    f"<span style='background:#1c2030;border-radius:3px;padding:4px 8px;font-size:.7rem;"
                    f"color:{fc_}'>{fn}: <b>{fv:+.2f}</b></span>",
                    unsafe_allow_html=True,
                )
            st.markdown("</div>", unsafe_allow_html=True)

            # ── Section C: TRADE PLAN ─────────────────────────────────────
            st.markdown("""
            <div style='margin:14px 0 10px;padding:8px 16px;background:linear-gradient(90deg,#26a69a15,#0b0e11);
                        border-left:3px solid #26a69a;border-radius:3px'>
              <span style='font-family:Syne,sans-serif;font-weight:700;font-size:.85rem;color:#26a69a;letter-spacing:.08em'>
                📐 COMPLETE TRADE PLAN
              </span>
            </div>""", unsafe_allow_html=True)

            tp1,tp2,tp3 = st.columns(3)
            _dip_e=round(r_d["last_close"]*0.98,2); _dip_sl=round(_dip_e-r_d["atr"]*1.2,2)
            _dip_tp=round(_dip_e+r_d["atr"]*2.5,2); _dip_rr=round((_dip_tp-_dip_e)/max(_dip_e-_dip_sl,0.01),2)
            for tc_,tit_,e_,tp_,sl_,rr_,up_,win_,brd_ in [
                (tp1,"⚡ Aggressive (ST)",stl_d["entry"],stl_d["tp"],stl_d["sl"],stl_d["rr_str"],up_d,"2–5 days","#f59e0b"),
                (tp2,"📅 Swing (LT)",     ltl_d["entry"],ltl_d["tp"],ltl_d["sl"],ltl_d["rr_str"],lt_up_d,"10–20 days","#26a69a"),
                (tp3,"📌 Limit/Dip",      _dip_e,_dip_tp,_dip_sl,f"1:{_dip_rr}",(_dip_tp/_dip_e-1)*100,"Limit at −2%","#3b82f6"),
            ]:
                tc_.markdown(
                    f"<div style='background:#131722;border-left:3px solid {brd_};border-radius:4px;padding:12px;text-align:center'>"
                    f"<div style='color:{brd_};font-size:.75rem;font-weight:700;margin-bottom:8px'>{tit_}</div>"
                    f"<div class='tv-label'>Entry</div>"
                    f"<div style='font-family:Syne,sans-serif;font-weight:700;font-size:1.1rem;color:#38bdf8;margin:5px 0'>&#8377;{e_:,.2f}</div>"
                    f"<div style='display:flex;justify-content:space-around;margin-bottom:7px'>"
                    f"<div><div class='tv-label'>Target</div><div style='font-weight:700;color:#26a69a'>&#8377;{tp_:,.2f}</div><div style='font-size:.68rem;color:#26a69a'>{up_:+.1f}%</div></div>"
                    f"<div><div class='tv-label'>Stop</div><div style='font-weight:700;color:#ef5350'>&#8377;{sl_:,.2f}</div><div style='font-size:.68rem;color:#ef5350'>{(sl_/e_-1)*100:+.1f}%</div></div>"
                    f"</div>"
                    f"<div style='border-top:1px solid #2a3347;padding-top:6px;font-size:.75rem;color:#d1d4dc'>"
                    f"R:R {rr_} &nbsp;·&nbsp; {win_}</div></div>",
                    unsafe_allow_html=True,
                )

            # ── Section D: POSITION SIZING ────────────────────────────────
            st.markdown("""
            <div style='margin:14px 0 10px;padding:8px 16px;background:linear-gradient(90deg,#f59e0b15,#0b0e11);
                        border-left:3px solid #f59e0b;border-radius:3px'>
              <span style='font-family:Syne,sans-serif;font-weight:700;font-size:.85rem;color:#f59e0b;letter-spacing:.08em'>
                💰 POSITION SIZING CALCULATOR
              </span>
            </div>""", unsafe_allow_html=True)

            cap_ = st.session_state.get("capital_val", 1_000_000)
            sl_r_ = stl_d["risk"]
            kf_   = min(max(0.55-(1-0.55)/1.5,0),0.25)
            ps_c1,ps_c2,ps_c3,ps_c4 = st.columns(4)
            for psc_,plbl_,qty_ in [
                (ps_c1,"1% Risk Rule",  max(1,int(cap_*0.01/sl_r_)) if sl_r_ else 0),
                (ps_c2,"2% Risk Rule",  max(1,int(cap_*0.02/sl_r_)) if sl_r_ else 0),
                (ps_c3,f"Half-Kelly",   max(1,int(cap_*kf_/r_d["last_close"])) if r_d["last_close"] else 0),
                (ps_c4,"Fixed 20%",     int(cap_*0.20/r_d["last_close"]) if r_d["last_close"] else 0),
            ]:
                inv_=qty_*r_d["last_close"]; ml_=qty_*sl_r_; tp__=qty_*stl_d["reward"]; pct_=inv_/cap_*100
                pc_="#26a69a" if pct_<=20 else "#f59e0b" if pct_<=30 else "#ef5350"
                psc_.markdown(
                    f"<div style='background:#131722;border-radius:4px;padding:10px;text-align:center;border-bottom:2px solid {pc_}'>"
                    f"<div class='tv-label'>{plbl_}</div>"
                    f"<div style='font-family:Syne,sans-serif;font-weight:800;font-size:1.3rem;color:#38bdf8;margin:5px 0'>{qty_:,}</div>"
                    f"<div style='font-size:.72rem;line-height:1.6'>"
                    f"<div style='display:flex;justify-content:space-between'><span style='color:#434651'>Capital</span><span>&#8377;{inv_:,.0f}</span></div>"
                    f"<div style='display:flex;justify-content:space-between'><span style='color:#434651'>Risk</span><span style='color:#ef5350'>&#8377;{ml_:,.0f}</span></div>"
                    f"<div style='display:flex;justify-content:space-between'><span style='color:#434651'>Reward</span><span style='color:#26a69a'>&#8377;{tp__:,.0f}</span></div>"
                    f"<div style='display:flex;justify-content:space-between'><span style='color:#434651'>Portfolio%</span><span style='color:{pc_}'>{pct_:.1f}%</span></div>"
                    f"</div></div>",
                    unsafe_allow_html=True,
                )

            # ── Section E: FUTURE GROWTH PREDICTION ──────────────────────
            st.markdown("""
            <div style='margin:14px 0 10px;padding:8px 16px;background:linear-gradient(90deg,#6366f115,#0b0e11);
                        border-left:3px solid #6366f1;border-radius:3px'>
              <span style='font-family:Syne,sans-serif;font-weight:700;font-size:.85rem;color:#6366f1;letter-spacing:.08em'>
                🔮 GROWTH SCENARIO PROJECTIONS
              </span>
              <span style='font-size:.68rem;color:#434651;margin-left:8px'>Based on current momentum &amp; historical ATR patterns</span>
            </div>""", unsafe_allow_html=True)

            _price  = r_d["last_close"]
            _atr    = r_d["atr"]
            _ai     = ai_p_d / 100
            _adx    = r_d["adx"]
            _vol    = r_d["vol_ratio"]
            _stf_   = 1 if stf_d else 0
            _brk_   = 1 if r_d.get("has_breakout") else 0

            # Momentum multiplier
            _mom_mult = 1.0 + (_ai * 0.5) + (_adx / 100) + (_vol / 10) + (_stf_ * 0.3) + (_brk_ * 0.2)
            # Scenario projections
            _sc_base   = round(_price * (1 + _atr / _price * 3 * _mom_mult * 0.8), 2)
            _sc_bull   = round(_price * (1 + _atr / _price * 5 * _mom_mult), 2)
            _sc_super  = round(_price * (1 + _atr / _price * 9 * _mom_mult * 1.3), 2)
            _sc_bear   = round(_price * (1 - _atr / _price * 2), 2)
            _sc_crash  = round(_price * (1 - _atr / _price * 4), 2)

            sc_cols = st.columns(5)
            for scc_,slbl_,sprice_,sicon_,sbrd_ in [
                (sc_cols[0],"🐻 Bear Case",    _sc_crash, "📉", "#ef5350"),
                (sc_cols[1],"⚠️ Soft Stop",    _sc_bear,  "↘️",  "#f59e0b"),
                (sc_cols[2],"📊 Base Case",    _sc_base,  "→",  "#787b86"),
                (sc_cols[3],"🚀 Bull Case",    _sc_bull,  "📈", "#26a69a"),
                (sc_cols[4],"🔥 Super Bull",   _sc_super, "⭐", "#a855f7"),
            ]:
                _pct_ = (sprice_/_price-1)*100
                _col_ = "#26a69a" if _pct_>0 else "#ef5350"
                scc_.markdown(
                    f"<div style='background:#131722;border-radius:4px;padding:10px;text-align:center;border-top:3px solid {sbrd_}'>"
                    f"<div class='tv-label'>{slbl_}</div>"
                    f"<div style='font-size:1.2rem;margin:4px 0'>{sicon_}</div>"
                    f"<div style='font-family:Syne,sans-serif;font-weight:700;color:{_col_};font-size:.95rem'>&#8377;{sprice_:,.2f}</div>"
                    f"<div style='font-size:.72rem;color:{_col_};font-weight:600'>{_pct_:+.1f}%</div>"
                    f"</div>",
                    unsafe_allow_html=True,
                )
            st.markdown(
                f"<div style='font-size:.68rem;color:#434651;text-align:center;margin-top:4px'>"
                f"⚠️ Projections are momentum-model estimates only, not financial advice. "
                f"Momentum multiplier: {_mom_mult:.2f}× &nbsp;·&nbsp; "
                f"Based on ATR={_atr:.2f}, ADX={_adx:.1f}, Vol={_vol:.2f}×"
                f"</div>",
                unsafe_allow_html=True,
            )

            # ── Section F: LIVE CHART ─────────────────────────────────────
            st.markdown("""
            <div style='margin:14px 0 8px;padding:8px 16px;background:linear-gradient(90deg,#ef535015,#0b0e11);
                        border-left:3px solid #ef5350;border-radius:3px'>
              <span style='font-family:Syne,sans-serif;font-weight:700;font-size:.85rem;color:#ef5350;letter-spacing:.08em'>
                📊 LIVE TRADINGVIEW CHART
              </span>
            </div>""", unsafe_allow_html=True)
            ch1_, ch2_ = st.columns([2,1])
            with ch1_:
                components.html(
                    tv_chart_widget(r_d["symbol"], height=400),
                    height=415, scrolling=False,
                )
            with ch2_:
                components.html(
                    tv_technical_analysis(r_d["symbol"], nonce=f"ps_{r_d['symbol']}"),
                    height=415, scrolling=False,
                )

        # ══════════════════════════════════════════════════════════════════
        # SUB-TAB 1: HIGH VOLATILITY MOVERS
        # ══════════════════════════════════════════════════════════════════
        with ps_tab1:
            if not _ps_vol:
                if run_ps_btn:
                    st.warning("⚠️ No high-volatility signals found. Try lowering threshold to 0.12.", icon="⚡")
                else:
                    st.info("Run Power Scan to see momentum movers.", icon="⚡")
            else:
                # Summary strip
                _avg_atr_v = sum(r["atr_pct"] for r in _ps_vol)/len(_ps_vol)
                _avg_ai_v  = sum(r["ai"]["ai_pct"] for r in _ps_vol)/len(_ps_vol)
                _n_stf_v   = sum(1 for r in _ps_vol if r.get("st_flip"))
                st.markdown(
                    f"<div style='display:flex;gap:10px;margin-bottom:14px;flex-wrap:wrap'>"
                    + "".join(
                        f"<div style='background:#131722;border-radius:4px;padding:10px 16px;text-align:center;border-top:2px solid {c_}'>"
                        f"<div class='tv-label'>{l_}</div>"
                        f"<div style='font-family:Syne,sans-serif;font-weight:800;font-size:1.3rem;color:{c_}'>{v_}</div></div>"
                        for l_,v_,c_ in [
                            ("Signals Found",str(len(_ps_vol)),"#ef5350"),
                            ("Avg ATR%",f"{_avg_atr_v:.1f}%","#f59e0b"),
                            ("Avg AI%",f"{_avg_ai_v:.1f}%","#38bdf8"),
                            ("ST Flips ⚡",str(_n_stf_v),"#a855f7"),
                        ]
                    )
                    + f"</div>",
                    unsafe_allow_html=True,
                )

                # Sort
                v_s1, v_s2 = st.columns([2,1])
                with v_s1:
                    v_sort = st.selectbox("Sort by", [
                        "Momentum Score (ATR×ADX×Vol×AI)",
                        "ATR% Highest", "AI Score", "ADX Strength",
                        "Volume Surge", "RSI Oversold",
                    ], key="v_sort", label_visibility="collapsed")
                with v_s2:
                    v_stf_only = st.toggle("ST Flips Only", False, key="v_stf")

                _vsf = {"Momentum Score (ATR×ADX×Vol×AI)": lambda r:-(r["atr_pct"]*(r["adx"]/30)*r["vol_ratio"]*r["ai"]["ai_pct"]/1000),
                        "ATR% Highest":lambda r:-r["atr_pct"],"AI Score":lambda r:-r["ai"]["ai_pct"],
                        "ADX Strength":lambda r:-r["adx"],"Volume Surge":lambda r:-r["vol_ratio"],"RSI Oversold":lambda r:r["rsi"]}
                _vlist = sorted([r for r in _ps_vol if (not v_stf_only or r.get("st_flip"))],
                                key=_vsf.get(v_sort,_vsf["Momentum Score (ATR×ADX×Vol×AI)"]))

                # Quick summary table
                _vrows=[]
                for rv in _vlist:
                    _lt,_=_d_grade(rv["ai"]["ai_pct"]); _sl=rv["levels"]["short_term"]; _ll=rv["levels"]["long_term"]
                    _vrows.append({"#":_vlist.index(rv)+1,"Symbol":rv["symbol"],"Grade":_lt,
                        "AI%":round(rv["ai"]["ai_pct"],1),"ATR%":round(rv["atr_pct"],2),
                        "ADX":round(rv["adx"],1),"RSI":round(rv["rsi"],1),"Vol×":round(rv["vol_ratio"],2),
                        "⚡":("⚡" if rv.get("st_flip") else "—"),"F&O":("✅" if rv.get("is_fo") else "—"),
                        "Price ₹":rv["last_close"],"ST Tgt":_sl["tp"],"R:R":_sl["rr_str"],
                        "Up%":round((_sl["tp"]/_sl["entry"]-1)*100,1),"LT Tgt":_ll["tp"],
                        "Top Signal":rv["hits"][0][1][:35] if rv["hits"] else "—"})
                _vdf = pd.DataFrame(_vrows)
                st.dataframe(_vdf, use_container_width=True, height=280, hide_index=True,
                    column_config={
                        "AI%":   st.column_config.ProgressColumn("AI%",min_value=0,max_value=100,format="%.1f%%"),
                        "Price ₹":st.column_config.NumberColumn("Price ₹",format="₹%.2f"),
                        "ST Tgt":st.column_config.NumberColumn("ST Tgt",format="₹%.2f"),
                        "LT Tgt":st.column_config.NumberColumn("LT Tgt",format="₹%.2f"),
                        "Up%":  st.column_config.NumberColumn("Upside%",format="+%.1f%%"),
                    })
                _vcsv=_vdf.to_csv(index=False).encode()
                st.download_button("⬇️ Download CSV",data=_vcsv,
                    file_name=f"volatile_{datetime.now().strftime('%Y%m%d_%H%M')}.csv",
                    mime="text/csv",key="v_csv")

                st.markdown("<hr style='border-color:#2a3347;margin:14px 0'>", unsafe_allow_html=True)
                st.markdown(f"<div class='tv-section'>⚡ Interactive Signal Cards — Click to Expand Full Analysis</div>",
                            unsafe_allow_html=True)

                for rank_v, rv in enumerate(_vlist[:ps_top_n], 1):
                    _ai_v=rv["ai"]["ai_pct"]; _ltr_v,_gc_v=_d_grade(_ai_v)
                    _stl_v=rv["levels"]["short_term"]; _up_v=(_stl_v["tp"]/_stl_v["entry"]-1)*100
                    with st.expander(
                        f"#{rank_v}  {rv['symbol']}  ·  ATR {rv['atr_pct']:.1f}%  ·  "
                        f"AI {_ai_v:.1f}%  ·  ADX {rv['adx']:.1f}  ·  "
                        f"₹{rv['last_close']:,.2f}  →  ₹{_stl_v['tp']:,.2f} ({_up_v:+.1f}%)"
                        + ("  ·  ⚡ ST FLIP" if rv.get("st_flip") else ""),
                        expanded=(rank_v == 1),
                    ):
                        _render_stock_detail(rv, "Volatile Scan")

        # ══════════════════════════════════════════════════════════════════
        # SUB-TAB 2: MULTI-BAGGER PENNY PICKS
        # ══════════════════════════════════════════════════════════════════
        with ps_tab2:
            if not _ps_pny:
                if run_ps_btn:
                    st.warning("⚠️ No penny multi-bagger signals found. Try lowering threshold to 0.10.", icon="💎")
                else:
                    st.info("Run Power Scan to discover multi-bagger penny stocks.", icon="💎")
            else:
                st.markdown("""
                <div style='padding:10px 16px;background:#ef535015;border-left:3px solid #ef5350;border-radius:3px;margin-bottom:12px'>
                  <div style='font-size:.8rem;color:#ef5350;font-weight:700'>⚠️ EXTREME RISK WARNING</div>
                  <div style='font-size:.72rem;color:#787b86;margin-top:2px'>
                    Penny stocks carry extreme risk. Max 0.5–1% portfolio per trade. Stop-losses are mandatory.
                    This is technical analysis only — not financial advice. Always do independent research.
                  </div>
                </div>""", unsafe_allow_html=True)

                _avg_ai_p=sum(r["ai"]["ai_pct"] for r in _ps_pny)/len(_ps_pny)
                _n_stf_p=sum(1 for r in _ps_pny if r.get("has_stf"))
                _n_brk_p=sum(1 for r in _ps_pny if r.get("has_breakout"))
                _categories_p={}
                for r in _ps_pny:
                    c_=r["price_category"]; _categories_p[c_]=_categories_p.get(c_,0)+1

                st.markdown(
                    "<div style='display:flex;gap:10px;margin-bottom:12px;flex-wrap:wrap'>"
                    + "".join(
                        f"<div style='background:#131722;border-radius:4px;padding:10px 16px;text-align:center;border-top:2px solid {c_}'>"
                        f"<div class='tv-label'>{l_}</div>"
                        f"<div style='font-family:Syne,sans-serif;font-weight:800;font-size:1.3rem;color:{c_}'>{v_}</div></div>"
                        for l_,v_,c_ in [
                            ("Golden Picks",str(len(_ps_pny)),"#a855f7"),
                            ("Avg AI Score",f"{_avg_ai_p:.1f}%","#38bdf8"),
                            ("ST Flips ⚡",str(_n_stf_p),"#f59e0b"),
                            ("Breakouts 🚀",str(_n_brk_p),"#26a69a"),
                            ("Top Score",f"{_ps_pny[0]['golden_score']:.1f}⭐","#67e8f9"),
                        ]
                    )
                    + "</div>",
                    unsafe_allow_html=True,
                )

                # Category pills
                _cat_pills="".join(
                    f"<span style='background:#1c2030;color:#d1d4dc;padding:3px 10px;border-radius:2px;font-size:.76rem;margin:2px'>{c_} — {n_}</span>"
                    for c_,n_ in sorted(_categories_p.items())
                )
                st.markdown(f"<div style='margin-bottom:10px'>{_cat_pills}</div>", unsafe_allow_html=True)

                # Sort + filter
                pp1,pp2,pp3 = st.columns([2,1,1])
                with pp1:
                    p_sort=st.selectbox("Sort by",[
                        "Golden Score","AI Score","Price Low First","ST Flip First","Breakout First"
                    ],key="p_sort",label_visibility="collapsed")
                with pp2:
                    p_max_price=st.number_input("Max Price ₹",min_value=10,max_value=500,value=300,step=50,key="p_maxp",label_visibility="collapsed")
                with pp3:
                    p_ultra=st.toggle("Under ₹100 Only",False,key="p_ultra")

                _psf={"Golden Score":lambda r:-r["golden_score"],"AI Score":lambda r:-r["ai"]["ai_pct"],
                      "Price Low First":lambda r:r["last_close"],"ST Flip First":lambda r:-(r.get("has_stf",0)*100+r["ai"]["ai_pct"]),
                      "Breakout First":lambda r:-(r.get("has_breakout",0)*100+r["ai"]["ai_pct"])}
                _plist=sorted([r for r in _ps_pny if r["last_close"]<=p_max_price and (not p_ultra or r["last_close"]<100)],
                              key=_psf.get(p_sort,_psf["Golden Score"]))

                # Summary table
                _prows=[]
                for rp in _plist:
                    _lt,_=_d_grade(rp["ai"]["ai_pct"]); _sl=rp["levels"]["short_term"]; _ll=rp["levels"]["long_term"]
                    _up_p2=(_sl["tp"]/_sl["entry"]-1)*100
                    _prows.append({"Symbol":rp["symbol"],"Category":rp["price_category"],
                        "Price ₹":rp["last_close"],"Golden⭐":round(rp["golden_score"],1),
                        "AI%":round(rp["ai"]["ai_pct"],1),"Grade":_lt,
                        "Breakout":("🚀" if rp.get("has_breakout") else "—"),
                        "Vol Surge":("🔊" if rp.get("has_volume_surge") else "—"),
                        "ST Flip":("⚡" if rp.get("has_stf") else "—"),
                        "ADX":round(rp["adx"],1),"RSI":round(rp["rsi"],1),
                        "ST Tgt":_sl["tp"],"R:R":_sl["rr_str"],"Up%":round(_up_p2,1),
                        "LT Tgt":_ll["tp"],"Top Signal":rp["hits"][0][1][:35] if rp["hits"] else "—"})
                _pdf=pd.DataFrame(_prows)
                st.dataframe(_pdf,use_container_width=True,height=300,hide_index=True,
                    column_config={
                        "AI%":     st.column_config.ProgressColumn("AI%",min_value=0,max_value=100,format="%.1f%%"),
                        "Golden⭐":st.column_config.NumberColumn("Golden⭐",format="%.1f"),
                        "Price ₹": st.column_config.NumberColumn("Price ₹",format="₹%.2f"),
                        "ST Tgt":  st.column_config.NumberColumn("ST Tgt",format="₹%.2f"),
                        "LT Tgt":  st.column_config.NumberColumn("LT Tgt",format="₹%.2f"),
                        "Up%":     st.column_config.NumberColumn("Upside%",format="+%.1f%%"),
                    })
                _pcsv=_pdf.to_csv(index=False).encode()
                st.download_button("⬇️ Download Penny CSV",data=_pcsv,
                    file_name=f"penny_multibagger_{datetime.now().strftime('%Y%m%d_%H%M')}.csv",
                    mime="text/csv",key="p_csv")

                st.markdown("<hr style='border-color:#2a3347;margin:14px 0'>", unsafe_allow_html=True)
                st.markdown(f"<div class='tv-section'>💎 Interactive Multi-Bagger Cards — Click to Expand Full Analysis</div>",
                            unsafe_allow_html=True)

                for rank_p, rp in enumerate(_plist[:ps_top_n], 1):
                    _ai_p=rp["ai"]["ai_pct"]; _ltr_p,_gc_p=_d_grade(_ai_p)
                    _stl_p=rp["levels"]["short_term"]; _up_p=(_stl_p["tp"]/_stl_p["entry"]-1)*100
                    _gold=rp["golden_score"]; _stars="⭐"*min(5,max(1,int(_gold/3)))
                    with st.expander(
                        f"#{rank_p}  {rp['symbol']}  ·  {rp['price_category']}  ·  "
                        f"₹{rp['last_close']:.2f}  ·  Golden {_gold:.1f} {_stars}  ·  AI {_ai_p:.1f}%"
                        + ("  ·  ⚡ ST FLIP" if rp.get("has_stf") else "")
                        + ("  ·  🚀 BREAKOUT" if rp.get("has_breakout") else ""),
                        expanded=(rank_p == 1),
                    ):
                        # Golden score header
                        g1,g2,g3,g4,g5 = st.columns(5)
                        for gc_w_,gl_,gv_,gc_c_ in [
                            (g1,"Golden Score",f"{_gold:.1f} {_stars}","#a855f7"),
                            (g2,"AI Score",f"{_ai_p:.1f}%",_gc_p),
                            (g3,"Breakout","🚀 YES" if rp.get("has_breakout") else "—","#26a69a" if rp.get("has_breakout") else "#434651"),
                            (g4,"Vol Surge","🔊 YES" if rp.get("has_volume_surge") else "—","#3b82f6" if rp.get("has_volume_surge") else "#434651"),
                            (g5,"ST Flip","⚡ YES" if rp.get("has_stf") else "—","#f59e0b" if rp.get("has_stf") else "#434651"),
                        ]:
                            gc_w_.markdown(
                                f"<div style='background:#131722;border-radius:4px;padding:10px;text-align:center;border-top:3px solid {gc_c_}'>"
                                f"<div class='tv-label'>{gl_}</div>"
                                f"<div style='font-family:Syne,sans-serif;font-weight:700;color:{gc_c_};font-size:.95rem;margin:6px 0'>{gv_}</div>"
                                f"</div>",
                                unsafe_allow_html=True,
                            )
                        _render_stock_detail(rp, "Penny Multi-Bagger")


# ══════════════════════════════════════════════════════════════════
        # SUB-TAB 3: 52W LOW RECOVERY GEMS (UPGRADED)
        # ══════════════════════════════════════════════════════════════════
        with ps_tab3:
            st.markdown("""
            <div style='padding:14px 20px;background:linear-gradient(135deg,#26a69a20,#38bdf818,#a855f718,#0b0e11);
                        border-left:4px solid #26a69a;border-radius:6px;margin-bottom:16px'>
              <div style='font-family:Syne,sans-serif;font-weight:800;font-size:1.2rem;color:#26a69a'>
                🔍 52-Week Low Recovery Gems — Full NSE Universe Direct Scan
              </div>
              <div style='font-size:.76rem;color:#787b86;margin-top:4px;line-height:1.7'>
                Scans <b style='color:#d1d4dc'>all ~200 NSE symbols directly via yfinance</b> — no engine pre-filter.
                Finds stocks that <b style='color:#26a69a'>recently touched their 52-week low</b> (last 1–5 days)
                OR are currently at their lowest price, with <b style='color:#38bdf8'>strong fundamentals</b>
                (P/E, ROE, Revenue Growth, Low Debt) and <b style='color:#a855f7'>high future growth potential</b>.
                <br><b style='color:#f59e0b'>⚠️ This scan fetches fundamentals — takes 8–15 minutes. Worth it.</b>
              </div>
            </div>""", unsafe_allow_html=True)

            if not _ps_rec:
                if run_ps_btn:
                    st.warning("⚠️ No 52W recovery gems found. Check internet connection — this scan requires fetching fundamentals from yfinance.", icon="🔍")
                else:
                    st.info(
                        "Run Power Scan to discover **fundamentally strong stocks near or AT 52-week lows**.\n\n"
                        "🔍 This scan:\n"
                        "- Directly fetches ALL ~200 NSE symbols\n"
                        "- Finds stocks that hit 52W low in last **1–5 trading days**\n"
                        "- Scores each on P/E, ROE, EPS Growth, Revenue Growth, Debt, Analyst targets\n"
                        "- Ranks by composite Opportunity Score",
                        icon="🔍",
                    )
            else:
                # ── Summary KPIs ──────────────────────────────────────────
                _n_jackpot  = sum(1 for r in _ps_rec if "JACKPOT" in r.get("opp_category",""))
                _n_prime    = sum(1 for r in _ps_rec if "PRIME" in r.get("opp_category",""))
                _n_fresh    = sum(1 for r in _ps_rec if r.get("at_52w_low_now") or r.get("touched_recent"))
                _avg_disc   = sum(r.get("discount_pct",0) for r in _ps_rec) / max(len(_ps_rec),1)
                _avg_upside = sum(r.get("upside_to_high",0) for r in _ps_rec) / max(len(_ps_rec),1)

                st.markdown(
                    "<div style='display:grid;grid-template-columns:repeat(6,1fr);gap:10px;margin-bottom:16px'>"
                    + "".join(
                        f"<div style='background:#131722;border-radius:4px;padding:12px;text-align:center;border-top:3px solid {c_}'>"
                        f"<div class='tv-label'>{l_}</div>"
                        f"<div style='font-family:Syne,sans-serif;font-weight:800;font-size:1.4rem;color:{c_}'>{v_}</div></div>"
                        for l_,v_,c_ in [
                            ("Total Gems",        str(len(_ps_rec)),    "#38bdf8"),
                            ("💎 Jackpot",         str(_n_jackpot),      "#26a69a"),
                            ("🔥 Prime",           str(_n_prime),        "#4db6ac"),
                            ("🔴 At/Near 52W Low", str(_n_fresh),        "#f59e0b"),
                            ("Avg Discount",       f"{_avg_disc:.1f}%",  "#a855f7"),
                            ("Avg Upside",         f"+{_avg_upside:.1f}%","#26a69a"),
                        ]
                    )
                    + "</div>",
                    unsafe_allow_html=True,
                )

                # ── Controls ──────────────────────────────────────────────
                rc1, rc2, rc3, rc4 = st.columns([2,1,1,1])
                with rc1:
                    rec_sort = st.selectbox("Sort by", [
                        "Opportunity Score (Default)",
                        "At 52W Low Right Now",
                        "Deepest Discount",
                        "Best Fundamentals",
                        "Highest Growth",
                        "Highest Upside to 52W High",
                        "RSI Most Oversold",
                        "Volume Surge",
                    ], key="rec_sort", label_visibility="collapsed")
                with rc2:
                    rec_fresh = st.toggle("Recently Hit Low Only", False, key="rec_fresh",
                                         help="Only stocks that hit 52W low in last 5 days")
                with rc3:
                    rec_cat = st.selectbox("Category", [
                        "All","💎 Jackpot","🔥 Prime","🟠 Fresh Low","🔵 Deep Value","🚀 Growth"
                    ], key="rec_cat", label_visibility="collapsed")
                with rc4:
                    rec_fo = st.toggle("F&O Only", False, key="rec_fo")

                _rsf = {
                    "Opportunity Score (Default)": lambda r: -r.get("opp_score",0),
                    "At 52W Low Right Now":        lambda r: (0 if r.get("at_52w_low_now") else 1 if r.get("touched_recent") else 2, r.get("pct_from_low",99)),
                    "Deepest Discount":            lambda r:  r.get("pct_from_low",99),
                    "Best Fundamentals":           lambda r: -r.get("fund_score",0),
                    "Highest Growth":              lambda r: -r.get("growth_score",0),
                    "Highest Upside to 52W High":  lambda r: -r.get("upside_to_high",0),
                    "RSI Most Oversold":           lambda r:  r.get("rsi",50),
                    "Volume Surge":                lambda r: -r.get("vol_ratio",1),
                }
                _cat_map = {
                    "💎 Jackpot":"JACKPOT","🔥 Prime":"PRIME",
                    "🟠 Fresh Low":"FRESH","🔵 Deep Value":"DEEP VALUE","🚀 Growth":"GROWTH"
                }
                _rec_list = [
                    r for r in _ps_rec
                    if (rec_cat == "All" or _cat_map.get(rec_cat,"") in r.get("opp_category",""))
                    and (not rec_fresh or r.get("at_52w_low_now") or r.get("touched_recent"))
                    and (not rec_fo or r.get("is_fo"))
                ]
                _rec_list.sort(key=_rsf.get(rec_sort, _rsf["Opportunity Score (Default)"]))

                st.markdown(
                    f"<div class='tv-label' style='margin:4px 0 10px'>"
                    f"Showing <b style='color:#38bdf8'>{len(_rec_list)}</b> stocks | "
                    f"Sorted by: <b style='color:#d1d4dc'>{rec_sort}</b></div>",
                    unsafe_allow_html=True,
                )

                # ── Summary Table ─────────────────────────────────────────
                st.markdown("<div class='tv-section'>📊 Recovery Gems — Full Detail Table</div>",
                            unsafe_allow_html=True)
                _rec_rows = []
                for rr in _rec_list:
                    _rec_rows.append({
                        "Symbol":         rr["symbol"],
                        "Category":       rr.get("opp_category","")[:25],
                        "Recency":        rr.get("recency_label",""),
                        "Price ₹":        rr["last_close"],
                        "52W Low ₹":      rr.get("w52l",""),
                        "52W High ₹":     rr.get("w52h",""),
                        "From Low%":      f"+{rr.get('pct_from_low',0):.1f}%",
                        "Discount%":      f"{rr.get('pct_from_high',0):.1f}%",
                        "Target(75%)₹":  rr.get("target_recovery",""),
                        "Upside 75%":    f"+{rr.get('upside_to_target',0):.1f}%",
                        "Upside 52WH":   f"+{rr.get('upside_to_high',0):.1f}%",
                        "Opp Score":     round(rr.get("opp_score",0),1),
                        "Fund Score":    rr.get("fund_score",0),
                        "Growth Score":  rr.get("growth_score",0),
                        "P/E":           round(rr["pe"],1) if rr.get("pe") else "N/A",
                        "ROE%":          f"{rr['roe']*100:.1f}%" if rr.get("roe") else "N/A",
                        "EPS Growth%":   f"+{rr['eps_g']*100:.1f}%" if rr.get("eps_g") and rr["eps_g"]>0 else (f"{rr['eps_g']*100:.1f}%" if rr.get("eps_g") else "N/A"),
                        "Rev Growth%":   f"+{rr['rev_g']*100:.1f}%" if rr.get("rev_g") and rr["rev_g"]>0 else "N/A",
                        "D/E":           f"{rr['de']:.2f}" if rr.get("de") else "N/A",
                        "MCap Cr":       rr.get("mcap_cr",""),
                        "RSI":           rr.get("rsi",""),
                        "RSI Zone":      rr.get("rsi_zone",""),
                        "Vol Ratio":     rr.get("vol_ratio",""),
                        "Analyst":       rr.get("analyst","").upper(),
                        "Analyst Tgt":   f"₹{rr['target_price']:.0f} (+{rr['analyst_upside']:.0f}%)" if rr.get("target_price") and rr.get("analyst_upside") else "N/A",
                        "F&O":           "✅" if rr.get("is_fo") else "—",
                        "Candle":        rr.get("candle_signal",""),
                        "Sector":        rr.get("sector",""),
                    })
                _rec_df = pd.DataFrame(_rec_rows)
                st.dataframe(_rec_df, use_container_width=True, height=380, hide_index=True,
                    column_config={
                        "Price ₹":    st.column_config.NumberColumn("Price ₹",   format="₹%.2f"),
                        "52W Low ₹":  st.column_config.NumberColumn("52W Low",   format="₹%.2f"),
                        "52W High ₹": st.column_config.NumberColumn("52W High",  format="₹%.2f"),
                        "Target(75%)₹":st.column_config.NumberColumn("Target 75%",format="₹%.2f"),
                        "Opp Score":  st.column_config.NumberColumn("Opp Score", format="%.1f"),
                        "MCap Cr":    st.column_config.NumberColumn("MCap Cr",   format="₹%.0f Cr"),
                        "Vol Ratio":  st.column_config.NumberColumn("Vol×",      format="%.2f"),
                        "RSI":        st.column_config.NumberColumn("RSI",       format="%.1f"),
                    })

                # Downloads
                _dd1, _dd2 = st.columns(2)
                with _dd1:
                    st.download_button("⬇️ CSV Download", data=_rec_df.to_csv(index=False).encode(),
                        file_name=f"52W_Recovery_{__import__('datetime').datetime.now().strftime('%Y%m%d_%H%M')}.csv",
                        mime="text/csv", use_container_width=True, key="rec_csv")
                with _dd2:
                    try:
                        import io as _rio, openpyxl as _rxl
                        from openpyxl.styles import PatternFill, Font, Alignment
                        from openpyxl.utils import get_column_letter
                        _rwb = _rxl.Workbook(); _rws = _rwb.active; _rws.title = "52W Recovery Gems"
                        _hdrs = ["Symbol","Category","Recency","Price","52W Low","52W High",
                                 "From Low%","Discount%","Target 75%","Upside 75%","Upside 52WH",
                                 "Opp Score","Fund Score","Growth Score","P/E","ROE%",
                                 "EPS Growth","Rev Growth","D/E","MCap Cr",
                                 "RSI","RSI Zone","Vol Ratio","Analyst","Analyst Target",
                                 "Candle Signal","F&O","Sector","Industry",
                                 "Fund Flags","Growth Flags",
                                 "ST Entry","ST Target","ST SL","LT Target","Hold Period"]
                        _rws.append(_hdrs)
                        for ci in range(1, len(_hdrs)+1):
                            _rws.cell(row=1, column=ci).fill = PatternFill("solid", fgColor="0D1117")
                            _rws.cell(row=1, column=ci).font = Font(bold=True, color="26A69A", size=9)
                            _rws.cell(row=1, column=ci).alignment = Alignment(horizontal="center")
                        _rws.row_dimensions[1].height = 26
                        for rr in _rec_list:
                            _stl = rr["levels"]["short_term"]; _ltl = rr["levels"]["long_term"]
                            _rws.append([
                                rr["symbol"], rr.get("opp_category","")[:35],
                                rr.get("recency_label",""),
                                rr["last_close"], rr.get("w52l",""), rr.get("w52h",""),
                                f"+{rr.get('pct_from_low',0):.1f}%",
                                f"{rr.get('pct_from_high',0):.1f}%",
                                rr.get("target_recovery",""),
                                f"+{rr.get('upside_to_target',0):.1f}%",
                                f"+{rr.get('upside_to_high',0):.1f}%",
                                round(rr.get("opp_score",0),2),
                                rr.get("fund_score",0), rr.get("growth_score",0),
                                round(rr["pe"],1) if rr.get("pe") else "N/A",
                                f"{rr['roe']*100:.1f}%" if rr.get("roe") else "N/A",
                                f"+{rr['eps_g']*100:.1f}%" if rr.get("eps_g") and rr["eps_g"]>0 else "N/A",
                                f"+{rr['rev_g']*100:.1f}%" if rr.get("rev_g") and rr["rev_g"]>0 else "N/A",
                                f"{rr['de']:.2f}" if rr.get("de") else "N/A",
                                rr.get("mcap_cr","N/A"),
                                round(rr.get("rsi",0),1), rr.get("rsi_zone",""),
                                round(rr.get("vol_ratio",1),2),
                                rr.get("analyst","").upper(),
                                f"₹{rr['target_price']:.0f} (+{rr['analyst_upside']:.0f}%)" if rr.get("target_price") else "N/A",
                                rr.get("candle_signal",""),
                                "YES" if rr.get("is_fo") else "NO",
                                rr.get("sector",""), rr.get("industry",""),
                                " | ".join(rr.get("fund_flags",[])),
                                " | ".join(rr.get("growth_flags",[])),
                                _stl["entry"], _stl["tp"], _stl["sl"],
                                _ltl["tp"], "3–18 months (recovery)",
                            ])
                            _ri = _rws.max_row
                            _fc = "0D2B1E" if "JACKPOT" in rr.get("opp_category","") else "0D1A2B" if "PRIME" in rr.get("opp_category","") else "0B0E11"
                            _fnc = "26A69A" if "JACKPOT" in rr.get("opp_category","") else "38BDF8" if "PRIME" in rr.get("opp_category","") else "D1D4DC"
                            for ci in range(1, len(_hdrs)+1):
                                _rws.cell(row=_ri,column=ci).fill = PatternFill("solid", fgColor=_fc)
                                _rws.cell(row=_ri,column=ci).font = Font(color=_fnc, size=9)
                                _rws.cell(row=_ri,column=ci).alignment = Alignment(horizontal="left")
                            _rws.row_dimensions[_ri].height = 14
                        _cws = [12,28,22,9,10,10,10,9,10,10,10,10,8,8,8,8,10,10,8,10,7,14,8,10,18,18,5,16,18,40,30,9,10,9,10,20]
                        for ci,w in enumerate(_cws[:len(_hdrs)],1):
                            _rws.column_dimensions[get_column_letter(ci)].width = w
                        _rws.freeze_panes = "A2"
                        _rbuf = _rio.BytesIO(); _rwb.save(_rbuf); _rbuf.seek(0)
                        st.download_button("⬇️ Excel Download", data=_rbuf.getvalue(),
                            file_name=f"52W_Recovery_{__import__('datetime').datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            use_container_width=True, key="rec_xl")
                    except Exception as _xe:
                        st.caption(f"Excel: {_xe}")

                st.markdown("<hr style='border-color:#2a3347;margin:14px 0'>", unsafe_allow_html=True)
                st.markdown("<div class='tv-section'>🔍 Interactive Recovery Cards — Click to Deep Dive</div>",
                            unsafe_allow_html=True)

                for rank_r, rr in enumerate(_rec_list[:ps_top_n], 1):
                    _col_r  = rr.get("opp_color","#38bdf8")
                    _cat_r  = rr.get("opp_category","")
                    _w52l_r = rr.get("w52l",0) or 0
                    _w52h_r = rr.get("w52h",0) or 0
                    _c_r    = rr["last_close"]
                    _uph_r  = rr.get("upside_to_high",0)
                    _upt_r  = rr.get("upside_to_target",0)
                    _tgt_r  = rr.get("target_recovery",0)
                    _fsc_r  = rr.get("fund_score",0)
                    _gsc_r  = rr.get("growth_score",0)
                    _osc_r  = rr.get("opp_score",0)
                    _rec_lbl= rr.get("recency_label","")
                    _rsi_zr = rr.get("rsi_zone","")
                    _ffl_r  = rr.get("fund_flags",[])
                    _gfl_r  = rr.get("growth_flags",[])
                    _stl_r  = rr["levels"]["short_term"]
                    _ltl_r  = rr["levels"]["long_term"]
                    _atr_r  = rr.get("atr",0); _atrp_r = rr.get("atr_pct",0)

                    with st.expander(
                        f"#{rank_r}  {rr['symbol']}  ·  {_cat_r[:32]}  ·  "
                        f"₹{_c_r:,.2f}  ·  {_rec_lbl}  ·  "
                        f"Fund:{_fsc_r}  Growth:{_gsc_r}  ·  "
                        f"Upside +{_uph_r:.1f}%"
                        + ("  ·  ✅F&O" if rr.get("is_fo") else ""),
                        expanded=(rank_r <= 2),
                    ):
                        # ── Top header ────────────────────────────────────
                        st.markdown(
                            f"<div class='tv-card' style='border-color:{_col_r}88;border-left:4px solid {_col_r};padding:16px 20px'>"
                            f"<div style='display:flex;justify-content:space-between;align-items:center;flex-wrap:wrap;gap:14px'>"
                            f"<div style='display:flex;align-items:center;gap:12px;flex-wrap:wrap'>"
                            f"<span class='tv-badge-sym' style='font-size:1.2rem;padding:5px 16px'>{rr['symbol']}</span>"
                            f"<span style='background:{_col_r}22;color:{_col_r};padding:3px 10px;border-radius:3px;font-size:.75rem;font-weight:700'>{_cat_r[:35]}</span>"
                            + (f"<span style='background:#f59e0b22;color:#f59e0b;padding:3px 10px;border-radius:3px;font-size:.72rem;font-weight:700'>{_rec_lbl}</span>" if _rec_lbl else "")
                            + (f"<span class='pill-cyan' style='font-size:.7rem'>F&O ✅</span>" if rr.get("is_fo") else "")
                            + f"<div><div class='tv-label'>{rr.get('sector','N/A')}</div>"
                            f"<div style='font-size:.7rem;color:#434651'>{rr.get('industry','N/A')} · {rr['indices']}</div></div></div>"
                            f"<div style='display:flex;gap:16px;flex-wrap:wrap'>"
                            f"<div style='text-align:center'><div class='tv-label'>Current Price</div>"
                            f"<div style='font-family:Syne,sans-serif;font-weight:800;font-size:1.6rem;color:#38bdf8'>&#8377;{_c_r:,.2f}</div></div>"
                            f"<div style='text-align:center'><div class='tv-label'>52W Low</div>"
                            f"<div style='font-family:Syne,sans-serif;font-weight:700;color:#26a69a'>&#8377;{_w52l_r:,.2f}</div>"
                            f"<div style='font-size:.68rem;color:#26a69a'>+{rr.get('pct_from_low',0):.1f}% above</div></div>"
                            f"<div style='text-align:center'><div class='tv-label'>52W High</div>"
                            f"<div style='font-family:Syne,sans-serif;font-weight:700;color:#ef5350'>&#8377;{_w52h_r:,.2f}</div>"
                            f"<div style='font-size:.68rem;color:#ef5350'>{rr.get('pct_from_high',0):.1f}%</div></div>"
                            f"<div style='text-align:center'><div class='tv-label'>Recovery Target</div>"
                            f"<div style='font-family:Syne,sans-serif;font-weight:800;color:#26a69a;font-size:1.2rem'>+{_upt_r:.1f}%</div>"
                            f"<div style='font-size:.68rem;color:#26a69a'>&#8377;{_tgt_r:,.2f}</div></div>"
                            f"<div style='text-align:center'><div class='tv-label'>Opp Score</div>"
                            f"<div style='font-family:Syne,sans-serif;font-weight:800;color:{_col_r};font-size:1.3rem'>{_osc_r:.1f}</div></div>"
                            f"</div></div></div>",
                            unsafe_allow_html=True,
                        )

                        # ── 52W Range visual ──────────────────────────────
                        if _w52l_r and _w52h_r and _w52l_r < _w52h_r:
                            _rng   = _w52h_r - _w52l_r
                            _cpct  = min((_c_r - _w52l_r) / _rng * 100, 100)
                            _tpct  = min((_tgt_r - _w52l_r) / _rng * 100, 100) if _tgt_r else 75
                            st.markdown(
                                f"<div style='background:#131722;border-radius:4px;padding:12px 16px;margin-bottom:12px'>"
                                f"<div style='font-size:.72rem;color:{_col_r};font-weight:700;text-transform:uppercase;letter-spacing:.08em;margin-bottom:8px'>📊 52-WEEK PRICE RANGE MAP</div>"
                                f"<div style='display:flex;justify-content:space-between;font-size:.7rem;color:#787b86;margin-bottom:5px'>"
                                f"<span>&#8377;{_w52l_r:,.2f} (52W Low)</span>"
                                f"<span style='color:#38bdf8;font-weight:600'>&#8594; You are here: &#8377;{_c_r:,.2f} ({_cpct:.0f}%)</span>"
                                f"<span>&#8377;{_w52h_r:,.2f} (52W High)</span></div>"
                                f"<div style='position:relative;height:18px;background:#1c2030;border-radius:9px;overflow:visible'>"
                                f"<div style='position:absolute;left:0;height:100%;width:{_cpct:.1f}%;background:linear-gradient(90deg,#ef535040,#f59e0b40);border-radius:9px'></div>"
                                f"<div style='position:absolute;left:{_cpct:.1f}%;height:100%;width:{max(0,_tpct-_cpct):.1f}%;background:{_col_r}55;border-radius:0 9px 9px 0'></div>"
                                f"<div style='position:absolute;left:calc({_cpct:.1f}% - 3px);top:-3px;width:6px;height:24px;background:#38bdf8;border-radius:3px;box-shadow:0 0 6px #38bdf8'></div>"
                                f"<div style='position:absolute;left:calc({_tpct:.1f}% - 1px);top:0;width:2px;height:100%;background:{_col_r};opacity:.8'></div>"
                                f"</div>"
                                f"<div style='display:flex;justify-content:space-between;margin-top:5px;font-size:.68rem'>"
                                f"<span style='color:#ef5350'>🔴 Panic selling zone</span>"
                                f"<span style='color:#38bdf8'>📍 Current</span>"
                                f"<span style='color:{_col_r}'>🎯 75% Recovery Target</span>"
                                f"<span style='color:#787b86'>52W High</span></div>"
                                f"</div>",
                                unsafe_allow_html=True,
                            )

                        # ── 3-column analysis ─────────────────────────────
                        rca, rcb, rcc = st.columns(3)

                        with rca:
                            # Fundamental Analysis
                            st.markdown(f"""
                            <div style='margin:0 0 8px;padding:7px 12px;background:linear-gradient(90deg,#26a69a15,#0b0e11);
                                        border-left:3px solid #26a69a;border-radius:3px'>
                              <span style='font-size:.75rem;font-weight:700;color:#26a69a;letter-spacing:.08em;text-transform:uppercase'>
                                🏦 FUNDAMENTAL QUALITY — {_fsc_r}/12
                              </span>
                            </div>""", unsafe_allow_html=True)
                            _ffl_html = "".join(
                                f"<div style='display:flex;gap:8px;padding:5px 0;border-bottom:1px solid #1c2030'>"
                                f"<span style='color:#26a69a;flex-shrink:0'>✅</span>"
                                f"<span style='font-size:.78rem;color:#d1d4dc'>{fl}</span></div>"
                                for fl in _ffl_r
                            )
                            _ffl_html = _ffl_html or f"<div style='color:#434651;font-size:.78rem'>No fundamental flags</div>"
                            st.markdown(
                                f"<div style='background:#131722;border-radius:4px;padding:12px;margin-bottom:8px'>"
                                f"{_ffl_html}</div>",
                                unsafe_allow_html=True,
                            )
                            # Quick metrics
                            for lbl_,val_,col_ in [
                                ("RSI", f"{rr.get('rsi',0):.1f} — {_rsi_zr}", _d_rsi_col(rr.get("rsi",50))),
                                ("Volume", f"{rr.get('vol_ratio',1):.2f}× (trend {rr.get('vol_trend',1):.2f}×)", _d_vol_col(rr.get("vol_ratio",1))),
                                ("Candle", rr.get("candle_signal","—") or "—", "#a855f7"),
                            ]:
                                st.markdown(
                                    f"<div style='display:flex;justify-content:space-between;padding:5px 8px;margin-bottom:3px;background:#131722;border-radius:3px'>"
                                    f"<span style='font-size:.72rem;color:#434651'>{lbl_}</span>"
                                    f"<span style='font-size:.76rem;color:{col_};font-weight:600'>{val_}</span></div>",
                                    unsafe_allow_html=True,
                                )

                        with rcb:
                            # Growth potential
                            st.markdown(f"""
                            <div style='margin:0 0 8px;padding:7px 12px;background:linear-gradient(90deg,#a855f715,#0b0e11);
                                        border-left:3px solid #a855f7;border-radius:3px'>
                              <span style='font-size:.75rem;font-weight:700;color:#a855f7;letter-spacing:.08em;text-transform:uppercase'>
                                🚀 FUTURE GROWTH POTENTIAL — {_gsc_r}/12
                              </span>
                            </div>""", unsafe_allow_html=True)
                            _gfl_html = "".join(
                                f"<div style='display:flex;gap:8px;padding:5px 0;border-bottom:1px solid #1c2030'>"
                                f"<span style='color:#a855f7;flex-shrink:0'>🚀</span>"
                                f"<span style='font-size:.78rem;color:#d1d4dc'>{fl}</span></div>"
                                for fl in _gfl_r
                            )
                            _gfl_html = _gfl_html or f"<div style='color:#434651;font-size:.78rem'>Growth data not available</div>"
                            st.markdown(
                                f"<div style='background:#131722;border-radius:4px;padding:12px;margin-bottom:8px'>"
                                f"{_gfl_html}</div>",
                                unsafe_allow_html=True,
                            )
                            # Recovery scenarios
                            for _slbl, _stgt, _sup, _scol in [
                                ("Conservative (50%)", round(_w52l_r+(_w52h_r-_w52l_r)*0.5,2),
                                 round((_w52l_r+(_w52h_r-_w52l_r)*0.5)/_c_r*100-100,1) if _c_r>0 else 0, "#f59e0b"),
                                ("Base (75% recovery)", _tgt_r, _upt_r, "#26a69a"),
                                ("Bull (52W High)",     _w52h_r, _uph_r, "#a855f7"),
                            ]:
                                st.markdown(
                                    f"<div style='background:#131722;border-left:3px solid {_scol};border-radius:4px;padding:7px 10px;margin-bottom:4px;display:flex;justify-content:space-between;align-items:center'>"
                                    f"<span style='font-size:.72rem;color:{_scol};font-weight:600'>{_slbl}</span>"
                                    f"<span style='font-family:Syne,sans-serif;font-weight:700;color:{_scol}'>&#8377;{_stgt:,.2f} <span style='font-size:.8rem'>(+{_sup:.1f}%)</span></span>"
                                    f"</div>",
                                    unsafe_allow_html=True,
                                )

                        with rcc:
                            # Entry plan
                            st.markdown("""
                            <div style='margin:0 0 8px;padding:7px 12px;background:linear-gradient(90deg,#38bdf815,#0b0e11);
                                        border-left:3px solid #38bdf8;border-radius:3px'>
                              <span style='font-size:.75rem;font-weight:700;color:#38bdf8;letter-spacing:.08em;text-transform:uppercase'>
                                💡 SMART ENTRY PLAN
                              </span>
                            </div>""", unsafe_allow_html=True)
                            _fresh_r = rr.get("at_52w_low_now") or rr.get("touched_recent")
                            st.markdown(
                                f"<div style='background:#131722;border-radius:4px;padding:12px;font-size:.78rem'>"
                                f"<div style='color:#38bdf8;font-weight:700;margin-bottom:8px'>"
                                f"{'🎯 OPTIMAL ENTRY — at 52W Low now!' if _fresh_r else '📍 ACCUMULATION ZONE — near 52W Low'}</div>"
                                f"<div style='display:flex;justify-content:space-between;margin-bottom:5px'>"
                                f"<span style='color:#434651'>Entry Zone</span>"
                                f"<span style='color:#38bdf8;font-weight:600'>&#8377;{_stl_r['entry']:,.2f}</span></div>"
                                f"<div style='display:flex;justify-content:space-between;margin-bottom:5px'>"
                                f"<span style='color:#434651'>ST Target (75%)</span>"
                                f"<span style='color:#26a69a;font-weight:600'>&#8377;{_stl_r['tp']:,.2f}</span></div>"
                                f"<div style='display:flex;justify-content:space-between;margin-bottom:5px'>"
                                f"<span style='color:#434651'>Hard Stop</span>"
                                f"<span style='color:#ef5350;font-weight:600'>&#8377;{_stl_r['sl']:,.2f}</span></div>"
                                f"<div style='display:flex;justify-content:space-between;margin-bottom:8px'>"
                                f"<span style='color:#434651'>LT Target (52WH)</span>"
                                f"<span style='color:#a855f7;font-weight:600'>&#8377;{_ltl_r['tp']:,.2f}</span></div>"
                                f"<div style='background:#1c2030;border-radius:3px;padding:8px;color:#787b86;line-height:1.7'>"
                                f"{'<b style=\"color:#d1d4dc\">Stagger buy</b>: 33% now → 33% on confirmation → 33% on breakout<br>' if not _fresh_r else '<b style=\"color:#d1d4dc\">Start full position</b> — at 52W Low is optimal entry<br>'}"
                                f"<b style='color:#d1d4dc'>Hold:</b> 3–18 months &nbsp;·&nbsp; "
                                f"<b style='color:#d1d4dc'>Max position:</b> 2–3% portfolio<br>"
                                f"<b style='color:#ef5350'>Exit if</b> breaks below &#8377;{_w52l_r*0.97:,.2f}"
                                f"</div></div>",
                                unsafe_allow_html=True,
                            )
                            # ATR context
                            st.markdown(
                                f"<div style='background:#ef535015;border-left:3px solid #ef5350;border-radius:4px;padding:8px 10px;margin-top:6px;font-size:.72rem;color:#787b86'>"
                                f"<div style='color:#ef5350;font-weight:700;margin-bottom:3px'>⚠️ RISK</div>"
                                f"ATR: ₹{_atr_r:.2f} ({_atrp_r:.1f}%) &nbsp;·&nbsp; "
                                f"Recovery plays need <b style='color:#d1d4dc'>patience 3–18 months</b>. "
                                f"If 52W low breaks decisively → exit immediately.</div>",
                                unsafe_allow_html=True,
                            )

                        # Live chart
                        st.markdown(f"<div class='tv-label' style='margin:10px 0 4px'>📊 Live Chart — {rr['symbol']}</div>",
                                    unsafe_allow_html=True)
                        components.html(
                            tv_mini_chart(rr["symbol"], height=180, nonce=f"rec_{rr['symbol']}_{rank_r}"),
                            height=184, scrolling=False,
                        )


    # TAB 4 — SIGNAL CARDS
    # ══════════════════════════════════════════════════════════════════════
    with tabs[4]:

        # ── Download bar ─────────────────────────────────────────────────
        st.markdown("<div class='tv-section'>📥 Download All Signals</div>", unsafe_allow_html=True)
        _dl1, _dl2, _dl3 = st.columns([2, 1, 1])
        with _dl1:
            st.markdown(
                f"<div style='font-size:.82rem;color:#787b86;padding:6px 0'>"
                f"<b style='color:#d1d4dc'>{len(alerts)} bullish signals</b> — "
                f"download as Excel (support zones, hold windows, all details) or CSV</div>",
                unsafe_allow_html=True,
            )
        with _dl2:
            try:
                import io as _io_xl, openpyxl as _opx
                from openpyxl.styles import PatternFill, Font, Alignment
                from openpyxl.utils import get_column_letter

                wb = _opx.Workbook()
                ws = wb.active
                ws.title = "NSE Signals"
                headers = [
                    "Scan Time","Symbol","Grade","AI%","Market%","Pattern%","Score",
                    "Price","RSI","ADX","Vol Ratio","ATR%","Vol Z",
                    "Sector","Indices","F&O","ST Flip",
                    "S1 Support","S2 Support","R1 Resist","R2 Resist",
                    "EMA9","EMA21","EMA50","EMA200","52W Low","52W High",
                    "ST Entry","ST Target","ST SL","ST RR","ST Up%","ST Window",
                    "LT Entry","LT Target","LT SL","LT RR","LT Up%","LT Window",
                    "Hold Rec","Hold Min Days","Hold Max Days","Hold Basis",
                    "Top Signal","All Signals","Signal Cats",
                    "P/E","ROE%","MCap Cr","Beta","Traded Cr/d","Research Note",
                ]
                ws.append(headers)
                hfill = PatternFill("solid", fgColor="0D1117")
                hfont = Font(bold=True, color="38BDF8", size=9)
                haln  = Alignment(horizontal="center", vertical="center", wrap_text=True)
                for ci in range(1, len(headers)+1):
                    cell = ws.cell(row=1, column=ci)
                    cell.fill = hfill; cell.font = hfont; cell.alignment = haln
                ws.row_dimensions[1].height = 28

                gfill = PatternFill("solid", fgColor="0D2B1E")
                afill = PatternFill("solid", fgColor="2B1F00")
                dfill = PatternFill("solid", fgColor="0B0E11")
                gfont = Font(color="26A69A", size=9)
                rfont = Font(color="EF5350", size=9)
                bfont = Font(color="38BDF8", size=9)
                wfont = Font(color="D1D4DC", size=9)
                pfont = Font(color="A855F7", size=9)
                saln  = Alignment(horizontal="left", vertical="center", wrap_text=False)
                naln  = Alignment(horizontal="left", vertical="top", wrap_text=True)

                for r_xl in alerts:
                    aip = r_xl["ai"]["ai_pct"]
                    ltr_xl, _ = _d_grade(aip)
                    stl_xl = r_xl["levels"]["short_term"]
                    ltl_xl = r_xl["levels"]["long_term"]
                    c_xl   = r_xl["last_close"]; atr_xl = r_xl["atr"]
                    e9x    = r_xl["ema9"]; e21x = r_xl["ema21"]
                    e50x   = r_xl["ema50"]; e200x= r_xl["ema200"]
                    hits_xl= r_xl["hits"]
                    esups  = sorted([x for x in [e9x,e21x,e50x] if 0<x<c_xl], reverse=True)
                    s1x    = round(max(c_xl-atr_xl, esups[0] if esups else c_xl-atr_xl), 2)
                    s2x    = round(c_xl - 2*atr_xl, 2)
                    r1x    = round(c_xl + atr_xl, 2)
                    r2x    = round(c_xl + 2*atr_xl, 2)
                    adxx   = r_xl["adx"]; rsix = r_xl["rsi"]
                    stfx   = r_xl.get("st_flip",0); atpx = r_xl["atr_pct"]
                    if stfx and adxx>=30:   hr,hmin,hmax,hb="Momentum 3-7d",3,7,"ST flip+ADX≥30"
                    elif rsix<40 and adxx>=25: hr,hmin,hmax,hb="Bounce 5-10d",5,10,"RSI<40 recovery"
                    elif atpx>=4.0:          hr,hmin,hmax,hb="Volatile 1-3d",1,3,"High ATR"
                    elif adxx>=35 and c_xl>e21x>e50x: hr,hmin,hmax,hb="Swing 10-20d",10,20,"Strong trend"
                    elif any("Breakout" in h[2] for h in hits_xl): hr,hmin,hmax,hb="Breakout 7-15d",7,15,"Breakout confirmed"
                    else:                    hr,hmin,hmax,hb="Standard 5-12d",5,12,"Normal bullish"
                    stup = (stl_xl["tp"]/stl_xl["entry"]-1)*100
                    ltup = (ltl_xl["tp"]/ltl_xl["entry"]-1)*100
                    ws.append([
                        r_xl["scan_ts"], r_xl["symbol"], ltr_xl,
                        round(aip,1), round(r_xl["mkt"]["pct"],1),
                        round(r_xl["pat_conf"]*100,1), round(r_xl["score"],4),
                        c_xl, round(rsix,1), round(adxx,1),
                        round(r_xl["vol_ratio"],2), round(atpx,2), round(r_xl.get("vol_z",0),2),
                        r_xl.get("sector","N/A"), r_xl["indices"],
                        "YES" if r_xl["is_fo"] else "NO",
                        "YES" if stfx else "NO",
                        s1x, s2x, r1x, r2x,
                        e9x, e21x, e50x, e200x if e200x else "N/A",
                        r_xl.get("w52l","N/A"), r_xl.get("w52h","N/A"),
                        stl_xl["entry"], stl_xl["tp"], stl_xl["sl"],
                        stl_xl["rr_str"], round(stup,1), stl_xl.get("window","2-5 days"),
                        ltl_xl["entry"], ltl_xl["tp"], ltl_xl["sl"],
                        ltl_xl["rr_str"], round(ltup,1), ltl_xl.get("window","10-20 days"),
                        hr, hmin, hmax, hb,
                        hits_xl[0][1] if hits_xl else "-",
                        " | ".join(h[1] for h in hits_xl[:6]),
                        ", ".join(sorted({h[2] for h in hits_xl})),
                        round(r_xl.get("pe") or 0, 1) or "N/A",
                        round((r_xl.get("roe") or 0)*100, 1) or "N/A",
                        round((r_xl.get("mcap") or 0)/1e7, 1) or "N/A",
                        round(r_xl.get("beta") or 0, 2) or "N/A",
                        round(r_xl["traded_val_cr"],2),
                        r_xl["reason"][:400],
                    ])
                    ri = ws.max_row
                    fill_ = gfill if aip>=75 else afill if aip>=60 else dfill
                    for ci in range(1, len(headers)+1):
                        cell = ws.cell(row=ri, column=ci)
                        cell.fill = fill_
                        cell.font = bfont if ci in (2,28,34) else gfont if ci in (4,29,35) else rfont if ci in (30,36) else pfont if ci==40 else wfont
                        cell.alignment = naln if ci==len(headers) else saln
                    ws.row_dimensions[ri].height = 14

                col_ws = [16,10,6,8,8,8,10,9,8,8,9,7,7,16,22,5,5,10,10,10,10,9,9,9,9,10,10,10,10,10,8,8,12,10,10,10,8,8,12,26,8,8,30,30,55,22,8,8,10,6,10,55]
                for ci,w in enumerate(col_ws[:len(headers)],1):
                    ws.column_dimensions[get_column_letter(ci)].width = w
                ws.freeze_panes = "A2"

                ws2 = wb.create_sheet("Support & Resistance")
                ws2.append(["Symbol","Price","S2","S1","Current","R1","R2","EMA9","EMA21","EMA50","52W Low","52W High","Zone Width%"])
                for r_xl in alerts:
                    c_xl=r_xl["last_close"]; a_=r_xl["atr"]
                    e9_=r_xl["ema9"]; e21_=r_xl["ema21"]; e50_=r_xl["ema50"]
                    es_=[x for x in [e9_,e21_,e50_] if 0<x<c_xl]
                    s1_=round(max(c_xl-a_, max(es_) if es_ else c_xl-a_),2)
                    s2_=round(c_xl-2*a_,2); r1_=round(c_xl+a_,2); r2_=round(c_xl+2*a_,2)
                    ws2.append([r_xl["symbol"],c_xl,s2_,s1_,c_xl,r1_,r2_,e9_,e21_,e50_,
                                r_xl.get("w52l","N/A"),r_xl.get("w52h","N/A"),round((r2_-s2_)/c_xl*100,1)])

                ws3 = wb.create_sheet("Holding Guide")
                ws3.append(["Symbol","Hold Rec","Min Days","Max Days","Hold Basis","ST Window","LT Window","ADX","RSI","ATR%","ST Flip"])
                for r_xl in alerts:
                    ad_=r_xl["adx"]; rs_=r_xl["rsi"]; sf_=r_xl.get("st_flip",0); ap_=r_xl["atr_pct"]
                    c_=r_xl["last_close"]; e21_=r_xl["ema21"]; e50_=r_xl["ema50"]
                    h_=r_xl["hits"]
                    if sf_ and ad_>=30:    hhr,hm,hx,hb="Momentum 3-7d",3,7,"ST flip+ADX>=30"
                    elif rs_<40:            hhr,hm,hx,hb="Bounce 5-10d",5,10,"RSI<40"
                    elif ap_>=4.0:          hhr,hm,hx,hb="Volatile 1-3d",1,3,"High ATR"
                    elif ad_>=35:           hhr,hm,hx,hb="Swing 10-20d",10,20,"Strong trend"
                    elif any("Breakout" in h[2] for h in h_): hhr,hm,hx,hb="Breakout 7-15d",7,15,"Breakout"
                    else:                   hhr,hm,hx,hb="Standard 5-12d",5,12,"Normal"
                    ws3.append([r_xl["symbol"],hhr,hm,hx,hb,
                                r_xl["levels"]["short_term"].get("window","2-5d"),
                                r_xl["levels"]["long_term"].get("window","10-20d"),
                                round(ad_,1),round(rs_,1),round(ap_,2),"YES" if sf_ else "NO"])

                buf_xl = _io_xl.BytesIO()
                wb.save(buf_xl); buf_xl.seek(0)
                st.download_button("⬇️ Excel (Full)", data=buf_xl.getvalue(),
                    file_name=f"NSE_Signals_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True, key="xl_dl")
            except ImportError:
                st.warning("pip install openpyxl")
            except Exception as _xl_e:
                st.error(f"Excel: {_xl_e}")

        with _dl3:
            _crows = []
            for r_c in alerts:
                stc=r_c["levels"]["short_term"]; ltc=r_c["levels"]["long_term"]
                lc_,_=_d_grade(r_c["ai"]["ai_pct"])
                c_c=r_c["last_close"]; a_c=r_c["atr"]
                e9c=r_c["ema9"]; e21c=r_c["ema21"]; e50c=r_c["ema50"]
                es_c=sorted([x for x in [e9c,e21c,e50c] if 0<x<c_c],reverse=True)
                s1c=round(max(c_c-a_c,es_c[0] if es_c else c_c-a_c),2); s2c=round(c_c-2*a_c,2)
                ad_c=r_c["adx"]; rs_c=r_c["rsi"]; sf_c=r_c.get("st_flip",0); ap_c=r_c["atr_pct"]
                if sf_c and ad_c>=30: hr_c="Momentum 3-7d"
                elif rs_c<40: hr_c="Bounce 5-10d"
                elif ap_c>=4: hr_c="Volatile 1-3d"
                elif ad_c>=35: hr_c="Swing 10-20d"
                elif any("Breakout" in h[2] for h in r_c["hits"]): hr_c="Breakout 7-15d"
                else: hr_c="Standard 5-12d"
                _crows.append({
                    "Symbol":r_c["symbol"],"Grade":lc_,"AI%":round(r_c["ai"]["ai_pct"],1),
                    "Score":round(r_c["score"],4),"Price":c_c,"RSI":round(rs_c,1),"ADX":round(ad_c,1),
                    "Vol":round(r_c["vol_ratio"],2),"ATR%":round(ap_c,2),
                    "S1":s1c,"S2":s2c,"R1":round(c_c+a_c,2),
                    "EMA9":e9c,"EMA21":e21c,"EMA50":e50c,
                    "ST_Entry":stc["entry"],"ST_Tgt":stc["tp"],"ST_SL":stc["sl"],"ST_RR":stc["rr_str"],
                    "LT_Tgt":ltc["tp"],"LT_RR":ltc["rr_str"],
                    "Hold":hr_c,"Sector":r_c.get("sector","N/A"),
                    "FO":"Y" if r_c["is_fo"] else "N","STF":"Y" if sf_c else "N",
                    "Signal":r_c["hits"][0][1] if r_c["hits"] else "-","Indices":r_c["indices"],
                })
            st.download_button("⬇️ CSV", data=pd.DataFrame(_crows).to_csv(index=False).encode(),
                file_name=f"NSE_{datetime.now().strftime('%Y%m%d_%H%M')}.csv",
                mime="text/csv", use_container_width=True, key="csv_all")

        st.markdown("<hr style='border-color:#2a3347;margin:12px 0'>", unsafe_allow_html=True)

        # ── Signal list ───────────────────────────────────────────────────
        st.markdown("<div class='tv-section'>🟢 All Bullish Signals — Click to Expand</div>",
                    unsafe_allow_html=True)

        for idx_r, r_li in enumerate(alerts):
            ai_li = r_li["ai"]["ai_pct"]
            ltr_li, gc_li = _d_grade(ai_li)
            stl_li = r_li["levels"]["short_term"]
            ltl_li = r_li["levels"]["long_term"]
            c_li   = r_li["last_close"]; atr_li = r_li["atr"]
            up_li  = (stl_li["tp"]/stl_li["entry"]-1)*100
            e9l    = r_li["ema9"]; e21l = r_li["ema21"]; e50l = r_li["ema50"]
            esl    = sorted([x for x in [e9l,e21l,e50l] if 0<x<c_li], reverse=True)
            s1l    = round(max(c_li-atr_li, esl[0] if esl else c_li-atr_li), 2)
            s2l    = round(c_li - 2*atr_li, 2)
            r1l    = round(c_li + atr_li, 2)
            adxl   = r_li["adx"]; rsil = r_li["rsi"]
            stfl   = r_li.get("st_flip",0); atpl = r_li["atr_pct"]
            if stfl and adxl>=30:   hrl,hcl="Momentum 3-7d","#f59e0b"
            elif rsil<40:            hrl,hcl="Bounce 5-10d","#38bdf8"
            elif atpl>=4:            hrl,hcl="Volatile 1-3d","#ef5350"
            elif adxl>=35:           hrl,hcl="Swing 10-20d","#26a69a"
            elif any("Breakout" in h[2] for h in r_li["hits"]): hrl,hcl="Breakout 7-15d","#26a69a"
            else:                    hrl,hcl="Standard 5-12d","#787b86"

            with st.expander(
                f"#{idx_r+1}  {r_li['symbol']}  Grade:{ltr_li}  AI:{ai_li:.1f}%  "
                f"₹{c_li:,.2f}  →₹{stl_li['tp']:,.2f}({up_li:+.1f}%)  "
                f"{hrl}  {r_li['indices']}"
                + ("  ✅F&O" if r_li["is_fo"] else "")
                + ("  ⚡ST" if stfl else ""),
                expanded=False,
            ):
                ea, eb, ec_ = st.columns(3)
                with ea:
                    ec1,ec2 = st.columns(2)
                    ec1.metric("AI Score", f"{ai_li:.1f}%")
                    ec2.metric("ADX", f"{adxl:.1f}")
                    ec3,ec4 = st.columns(2)
                    ec3.metric("RSI", f"{rsil:.1f}")
                    ec4.metric("Vol", f"{r_li['vol_ratio']:.2f}×")
                with eb:
                    st.markdown(
                        f"<div style='background:#131722;border-left:3px solid #26a69a;"
                        f"border-radius:4px;padding:10px 12px;font-size:.78rem'>"
                        f"<div style='color:#26a69a;font-weight:700;margin-bottom:6px'>🛡️ SUPPORT / RESISTANCE</div>"
                        f"<div style='display:flex;justify-content:space-between;margin-bottom:3px'>"
                        f"<span style='color:#434651'>S2 (Strong)</span><span style='color:#ef5350;font-weight:600'>₹{s2l:,.2f}</span></div>"
                        f"<div style='display:flex;justify-content:space-between;margin-bottom:3px'>"
                        f"<span style='color:#434651'>S1 (Immed.)</span><span style='color:#f59e0b;font-weight:600'>₹{s1l:,.2f}</span></div>"
                        f"<div style='display:flex;justify-content:space-between;padding:3px 0;border-top:1px solid #2a3347;border-bottom:1px solid #2a3347;margin-bottom:3px'>"
                        f"<span style='color:#787b86'>Current</span><span style='color:#38bdf8;font-weight:700'>₹{c_li:,.2f}</span></div>"
                        f"<div style='display:flex;justify-content:space-between;margin-bottom:3px'>"
                        f"<span style='color:#434651'>R1 (Resist.)</span><span style='color:#26a69a;font-weight:600'>₹{r1l:,.2f}</span></div>"
                        f"<div style='display:flex;justify-content:space-between'>"
                        f"<span style='color:#434651'>R2 / Target</span><span style='color:#26a69a;font-weight:600'>₹{stl_li['tp']:,.2f}</span></div>"
                        f"<div style='margin-top:6px;font-size:.7rem;color:#434651'>"
                        f"EMA9:{e9l:.0f} · EMA21:{e21l:.0f} · EMA50:{e50l:.0f}</div>"
                        f"</div>",
                        unsafe_allow_html=True,
                    )
                with ec_:
                    st.markdown(
                        f"<div style='background:#131722;border-left:3px solid {hcl};"
                        f"border-radius:4px;padding:10px 12px;font-size:.78rem'>"
                        f"<div style='color:{hcl};font-weight:700;margin-bottom:6px'>⏱️ HOLD WINDOW</div>"
                        f"<div style='font-family:Syne,sans-serif;font-weight:700;color:{hcl};font-size:.95rem;margin-bottom:6px'>{hrl}</div>"
                        f"<div style='color:#787b86;line-height:1.6;margin-bottom:6px'>"
                        f"{'ST flip momentum' if stfl else 'RSI recovery' if rsil<40 else 'High ATR quick' if atpl>=4 else 'Trend swing' if adxl>=35 else 'Breakout hold' if any('Breakout' in h[2] for h in r_li['hits']) else 'Standard swing'}"
                        f"</div>"
                        f"<div style='display:flex;justify-content:space-between;margin-bottom:3px'>"
                        f"<span style='color:#434651'>ST Target</span><span style='color:#26a69a'>₹{stl_li['tp']:,.2f}</span></div>"
                        f"<div style='display:flex;justify-content:space-between;margin-bottom:3px'>"
                        f"<span style='color:#434651'>SL</span><span style='color:#ef5350'>₹{stl_li['sl']:,.2f}</span></div>"
                        f"<div style='display:flex;justify-content:space-between'>"
                        f"<span style='color:#434651'>R:R</span><span style='color:#d1d4dc'>{stl_li['rr_str']}</span></div>"
                        f"</div>",
                        unsafe_allow_html=True,
                    )
                if st.button(f"📊 Full Analysis — {r_li['symbol']}", key=f"open_{idx_r}"):
                    st.session_state["card_sym"] = r_li["symbol"]
                    st.rerun()

        # ── Detailed Card ─────────────────────────────────────────────────
        st.markdown("<hr style='border-color:#2a3347;margin:20px 0'>", unsafe_allow_html=True)
        st.markdown("<div class='tv-section'>📋 Detailed Signal Card</div>", unsafe_allow_html=True)

        fc1,fc2,fc3,fc4 = st.columns([2,1,1,1])
        with fc1:
            sel_sym = st.selectbox("Select Stock", [r["symbol"] for r in alerts],
                format_func=lambda s: f"{s}  —  AI: {next(r['ai']['ai_pct'] for r in alerts if r['symbol']==s):.1f}%",
                key="card_sym")
        with fc2:
            chart_interval = st.selectbox("Chart Interval", ["D","W","60","15","5"], key="ci")
        with fc3:
            show_tv_ta = st.toggle("TV Analysis", value=True, key="tv_ta")
        with fc4:
            show_tv_fin = st.toggle("Financials", value=False, key="tv_fin")

        r  = next(x for x in alerts if x["symbol"] == sel_sym)
        ai = r["ai"]; mk = r["mkt"]
        stl = r["levels"]["short_term"]; ltl = r["levels"]["long_term"]
        hits = r["hits"]; ai_p = ai["ai_pct"]; mk_p = mk["pct"]; pt_p = r["pat_conf"]*100
        ltr, gc = _d_grade(ai_p)
        c_d  = r["last_close"]; atr_d = r["atr"]
        e9d  = r["ema9"]; e21d = r["ema21"]; e50d = r["ema50"]

        # Support zones
        esd  = sorted([x for x in [e9d,e21d,e50d] if 0<x<c_d], reverse=True)
        s1d  = round(max(c_d-atr_d, esd[0] if esd else c_d-atr_d), 2)
        s2d  = round(c_d - 2*atr_d, 2)
        r1d  = round(c_d + atr_d, 2)
        r2d  = round(c_d + 2*atr_d, 2)
        zwd  = round((r2d-s2d)/c_d*100, 1)

        # Hold window
        adxd = r["adx"]; rsid = r["rsi"]; stfd = r.get("st_flip",0); atpd = r["atr_pct"]
        if stfd and adxd>=30:    hrd,hmid,hmxd,hbd,hrcd = "Momentum Play",3,7,"ST flip + ADX≥30","#f59e0b"
        elif rsid<40 and adxd>=25: hrd,hmid,hmxd,hbd,hrcd = "Oversold Bounce",5,10,"RSI<40 recovery","#38bdf8"
        elif atpd>=4.0:            hrd,hmid,hmxd,hbd,hrcd = "Volatile Quick",1,3,"ATR≥4% — tight SL","#ef5350"
        elif adxd>=35 and c_d>e21d>e50d: hrd,hmid,hmxd,hbd,hrcd = "Swing Trade",10,20,"Trend + EMA stack","#26a69a"
        elif any("Breakout" in h[2] for h in hits): hrd,hmid,hmxd,hbd,hrcd = "Breakout Hold",7,15,"Confirmed breakout","#26a69a"
        else:                      hrd,hmid,hmxd,hbd,hrcd = "Standard Swing",5,12,"Normal bullish","#787b86"

        components.html(tv_symbol_info(sel_sym, nonce=sel_sym), height=80, scrolling=False)

        # Header strip
        _rcc = _d_rsi_col(r["rsi"]); _acc = _d_adx_col(r["adx"]); _vcc = _d_vol_col(r["vol_ratio"])
        _fot = "<span class='pill-cyan' style='font-size:.75rem'>F&amp;O ✅</span>" if r["is_fo"] else ""
        _stt = "<span class='pill-amber' style='font-size:.75rem'>⚡ ST FLIP</span>" if stfd else ""
        st.markdown(
            f"<div class='tv-card tv-card-bull' style='border-color:{gc}66'>"
            f"<div style='display:flex;align-items:center;justify-content:space-between;flex-wrap:wrap;gap:14px'>"
            f"<div style='display:flex;align-items:center;gap:14px'>"
            f"<span class='tv-badge-sym' style='font-size:1.3rem;padding:5px 16px'>{r['symbol']}</span>"
            f"<span class='tv-badge-grade' style='background:{gc}22;color:{gc};font-size:1.4rem'>{ltr}</span>"
            f"{_fot}{_stt}"
            f"<div><div class='tv-label'>{r.get('sector','N/A')}</div>"
            f"<div style='font-size:.72rem;color:#434651'>{r['indices']}</div></div></div>"
            f"<div style='display:flex;gap:22px;flex-wrap:wrap'>"
            f"<div style='text-align:center'><div class='tv-label'>Price</div>"
            f"<div style='font-family:Syne,sans-serif;font-weight:800;font-size:1.8rem;color:#38bdf8'>&#8377;{r['last_close']:,.2f}</div></div>"
            f"<div style='text-align:center'><div class='tv-label'>RSI</div><div style='font-family:Syne,sans-serif;font-weight:700;font-size:1.3rem;color:{_rcc}'>{r['rsi']:.1f}</div></div>"
            f"<div style='text-align:center'><div class='tv-label'>ADX</div><div style='font-family:Syne,sans-serif;font-weight:700;font-size:1.3rem;color:{_acc}'>{r['adx']:.1f}</div></div>"
            f"<div style='text-align:center'><div class='tv-label'>Volume</div><div style='font-family:Syne,sans-serif;font-weight:700;font-size:1.3rem;color:{_vcc}'>{r['vol_ratio']:.2f}&times;</div></div>"
            f"<div style='text-align:center'><div class='tv-label'>Hold</div>"
            f"<div style='font-family:Syne,sans-serif;font-weight:700;color:{hrcd}'>{hmid}–{hmxd} days</div></div>"
            f"</div></div></div>",
            unsafe_allow_html=True,
        )

        # Confidence row
        cc1,cc2,cc3 = st.columns(3)
        for _cw,_pct,_lbl,_extra,_ch in [
            (cc1,ai_p,"🤖 AI Confidence",f"T:{ai['trend_s']:+.2f}  M:{ai['mom_s']:+.2f}  B:{ai['brk_s']:+.2f}",gc),
            (cc2,mk_p,"📊 Market",f"{mk['label']} · {mk['align']}","#26a69a" if mk_p>=65 else "#f59e0b" if mk_p>=40 else "#ef5350"),
            (cc3,pt_p,"🎯 Pattern",f"{len(hits)} signals · {r['n_cats']} categories","#26a69a" if pt_p>=60 else "#f59e0b"),
        ]:
            _cw.markdown(
                f"<div class='tv-card' style='text-align:center'>"
                f"<div class='tv-label'>{_lbl}</div>"
                f"<div style='font-family:Syne,sans-serif;font-weight:800;font-size:1.9rem;color:{_ch};margin:8px 0'>{_pct:.1f}%</div>"
                f"{gauge_html(_pct,'',220)}"
                f"<div style='font-size:.68rem;color:#434651;margin-top:5px'>{_extra}</div></div>",
                unsafe_allow_html=True,
            )

        # Support zones + Trade plan + Hold guide
        st.markdown("<div class='tv-section'>🛡️ Support Zones · 📐 Trade Plan · ⏱️ Holding Guide</div>",
                    unsafe_allow_html=True)
        sz1, sz2, sz3 = st.columns(3)

        with sz1:
            st.markdown(
                f"<div style='background:#131722;border-left:4px solid #26a69a;border-radius:4px;padding:14px 16px'>"
                f"<div style='color:#26a69a;font-weight:700;font-size:.8rem;letter-spacing:.08em;text-transform:uppercase;margin-bottom:12px'>🛡️ SUPPORT &amp; RESISTANCE ZONES</div>"
                f"<div style='display:flex;justify-content:space-between;align-items:center;padding:7px 0;border-bottom:1px solid #1c2030'>"
                f"<span style='font-size:.78rem;color:#787b86'>S2 — Strong Support</span>"
                f"<span style='font-family:Syne,sans-serif;font-weight:700;color:#ef5350'>₹{s2d:,.2f}</span></div>"
                f"<div style='display:flex;justify-content:space-between;align-items:center;padding:7px 0;border-bottom:1px solid #1c2030'>"
                f"<span style='font-size:.78rem;color:#787b86'>S1 — Immediate Support</span>"
                f"<span style='font-family:Syne,sans-serif;font-weight:700;color:#f59e0b'>₹{s1d:,.2f}</span></div>"
                f"<div style='display:flex;justify-content:space-between;align-items:center;padding:7px 0;border-bottom:1px solid #2a3347;background:#1c203080;margin:0 -4px;padding:8px 4px'>"
                f"<span style='font-size:.82rem;color:#d1d4dc;font-weight:600'>📍 Current Price</span>"
                f"<span style='font-family:Syne,sans-serif;font-weight:800;color:#38bdf8;font-size:1.1rem'>₹{c_d:,.2f}</span></div>"
                f"<div style='display:flex;justify-content:space-between;align-items:center;padding:7px 0;border-bottom:1px solid #1c2030'>"
                f"<span style='font-size:.78rem;color:#787b86'>R1 — First Resistance</span>"
                f"<span style='font-family:Syne,sans-serif;font-weight:700;color:#26a69a'>₹{r1d:,.2f}</span></div>"
                f"<div style='display:flex;justify-content:space-between;align-items:center;padding:7px 0;border-bottom:1px solid #1c2030'>"
                f"<span style='font-size:.78rem;color:#787b86'>R2 — ST Target</span>"
                f"<span style='font-family:Syne,sans-serif;font-weight:700;color:#26a69a'>₹{stl['tp']:,.2f}</span></div>"
                f"<div style='margin-top:10px;padding-top:8px;border-top:1px solid #2a3347'>"
                f"<div style='font-size:.72rem;color:#434651;margin-bottom:4px'>EMA Supports</div>"
                f"<div style='display:flex;gap:8px;flex-wrap:wrap'>"
                f"<span style='background:#38bdf822;color:#38bdf8;padding:2px 8px;border-radius:2px;font-size:.72rem'>EMA9: ₹{e9d:.1f}</span>"
                f"<span style='background:#f59e0b22;color:#f59e0b;padding:2px 8px;border-radius:2px;font-size:.72rem'>EMA21: ₹{e21d:.1f}</span>"
                f"<span style='background:#a855f722;color:#a855f7;padding:2px 8px;border-radius:2px;font-size:.72rem'>EMA50: ₹{e50d:.1f}</span>"
                f"</div>"
                f"<div style='font-size:.7rem;color:#434651;margin-top:6px'>Zone width: {zwd:.1f}% of price &nbsp;·&nbsp; "
                f"52W: ₹{r.get('w52l','N/A')} – ₹{r.get('w52h','N/A')}</div>"
                f"</div></div>",
                unsafe_allow_html=True,
            )

        with sz2:
            dip_e=round(c_d*0.98,2); dip_sl=round(dip_e-atr_d*1.2,2)
            dip_tp=round(dip_e+atr_d*2.5,2); dip_ri=round(dip_e-dip_sl,2)
            dip_rw=round(dip_tp-dip_e,2); dip_rr=round(dip_rw/dip_ri,2) if dip_ri>0 else 0
            st.markdown(trade_scenario_html("⚡ Aggressive (ST)", stl, "#f59e0b","2–5 trading days"), unsafe_allow_html=True)
            st.markdown("<div style='margin:5px 0'></div>", unsafe_allow_html=True)
            st.markdown(trade_scenario_html("📅 Swing (LT)", ltl, "#26a69a","10–20 trading days"), unsafe_allow_html=True)
            st.markdown("<div style='margin:5px 0'></div>", unsafe_allow_html=True)
            st.markdown(trade_scenario_html("📌 Limit/Dip (−2%)",
                dict(entry=dip_e,tp=dip_tp,sl=dip_sl,rr_str=f"1:{dip_rr}",risk=dip_ri,reward=dip_rw),
                "#3b82f6","Limit at −2%"), unsafe_allow_html=True)

        with sz3:
            st.markdown(
                f"<div style='background:#131722;border-left:4px solid {hrcd};border-radius:4px;padding:14px 16px'>"
                f"<div style='color:{hrcd};font-weight:700;font-size:.8rem;letter-spacing:.08em;text-transform:uppercase;margin-bottom:10px'>⏱️ HOLDING GUIDE</div>"
                f"<div style='font-family:Syne,sans-serif;font-weight:800;font-size:1.1rem;color:{hrcd};margin-bottom:10px'>{hrd}</div>"
                f"<div style='display:flex;gap:10px;margin-bottom:12px'>"
                f"<div style='background:#1c2030;border-radius:3px;padding:10px;text-align:center;flex:1'>"
                f"<div class='tv-label'>Min Days</div>"
                f"<div style='font-family:Syne,sans-serif;font-weight:700;font-size:1.4rem;color:{hrcd}'>{hmid}</div></div>"
                f"<div style='background:#1c2030;border-radius:3px;padding:10px;text-align:center;flex:1'>"
                f"<div class='tv-label'>Max Days</div>"
                f"<div style='font-family:Syne,sans-serif;font-weight:700;font-size:1.4rem;color:{hrcd}'>{hmxd}</div></div>"
                f"</div>"
                f"<div style='font-size:.78rem;color:#787b86;line-height:1.7;margin-bottom:10px'>{hbd}</div>"
                f"<div style='background:#1c2030;border-radius:3px;padding:10px;font-size:.75rem'>"
                f"<div style='display:flex;justify-content:space-between;margin-bottom:4px'>"
                f"<span style='color:#434651'>Exit on SL hit</span><span style='color:#ef5350'>₹{stl['sl']:,.2f}</span></div>"
                f"<div style='display:flex;justify-content:space-between;margin-bottom:4px'>"
                f"<span style='color:#434651'>Target exit</span><span style='color:#26a69a'>₹{stl['tp']:,.2f}</span></div>"
                f"<div style='display:flex;justify-content:space-between;margin-bottom:4px'>"
                f"<span style='color:#434651'>Max hold</span><span style='color:{hrcd}'>{hmxd} days</span></div>"
                f"<div style='display:flex;justify-content:space-between'>"
                f"<span style='color:#434651'>ATR</span><span style='color:#f59e0b'>₹{atr_d:.2f} ({atpd:.1f}%)</span></div>"
                f"</div></div>",
                unsafe_allow_html=True,
            )

        # Position sizing
        st.markdown("<div class='tv-section'>💰 Position Sizing</div>", unsafe_allow_html=True)
        capital = st.session_state.get("capital_val", 1_000_000)
        sl_risk = stl["risk"]; kf = min(max(0.55-(1-0.55)/1.5,0),0.25)
        psc_ = st.columns(4)
        for ix_,(lbl_sz,qty) in enumerate([
            ("1% Risk",  max(1,int(capital*0.01/sl_risk)) if sl_risk else 0),
            ("2% Risk",  max(1,int(capital*0.02/sl_risk)) if sl_risk else 0),
            ("Half-Kelly",max(1,int(capital*kf/r["last_close"])) if r["last_close"] else 0),
            ("Fixed 20%",int(capital*0.20/r["last_close"]) if r["last_close"] else 0),
        ]):
            inv=qty*r["last_close"]; ml=qty*sl_risk; tp__=qty*stl["reward"]; pp=inv/capital*100
            pcc="#26a69a" if pp<=20 else "#f59e0b" if pp<=30 else "#ef5350"
            psc_[ix_].markdown(
                f"<div class='tv-card' style='text-align:center'>"
                f"<div class='tv-label'>{lbl_sz}</div>"
                f"<div style='font-family:Syne,sans-serif;font-weight:800;font-size:1.4rem;color:#38bdf8;margin:6px 0'>{qty:,}</div>"
                f"<div style='font-size:.74rem;line-height:1.6'>"
                f"<div style='display:flex;justify-content:space-between'><span class='tv-label'>Capital</span><span>₹{inv:,.0f}</span></div>"
                f"<div style='display:flex;justify-content:space-between'><span class='tv-label'>Max Loss</span><span style='color:#ef5350'>₹{ml:,.0f}</span></div>"
                f"<div style='display:flex;justify-content:space-between'><span class='tv-label'>Target P&L</span><span style='color:#26a69a'>₹{tp__:,.0f}</span></div>"
                f"<div style='display:flex;justify-content:space-between'><span class='tv-label'>Portfolio%</span><span style='color:{pcc}'>{pp:.1f}%</span></div>"
                f"</div></div>",
                unsafe_allow_html=True,
            )

        # Fundamentals + Research
        st.markdown("<div class='tv-section'>🏦 Fundamentals &amp; Research</div>", unsafe_allow_html=True)
        fn_c, no_c = st.columns([1,2])
        pe=r.get("pe"); roe=r.get("roe"); mcap=r.get("mcap"); beta=r.get("beta")
        w52h=r.get("w52h"); w52l=r.get("w52l")
        pec="#26a69a" if pe and pe<20 else "#f59e0b" if pe and pe<35 else "#ef5350"
        rec="#26a69a" if roe and roe>0.18 else "#f59e0b" if roe and roe>0.10 else "#ef5350"
        fn_c.markdown(
            "<div class='tv-card'><div style='display:grid;gap:6px;font-size:.82rem'>"
            + "".join(f"<div style='display:flex;justify-content:space-between;border-bottom:1px solid #2a3347;padding-bottom:5px'><span class='tv-label'>{lb2}</span><span style='color:{vc};font-weight:600'>{vv}</span></div>"
            for lb2,vc,vv in [
                ("P/E",pec,f"{pe:.1f}" if pe else "N/A"),
                ("ROE",rec,f"{roe*100:.1f}%" if roe else "N/A"),
                ("Market Cap","#d1d4dc",fmt_cr(mcap)),
                ("Beta","#d1d4dc",f"{beta:.2f}" if beta else "N/A"),
                ("52W High","#26a69a",fmt_inr(w52h)),
                ("52W Low","#f59e0b",fmt_inr(w52l)),
                ("Traded Val","#38bdf8",f"₹{r['traded_val_cr']:.2f} Cr/d"),
            ])
            + "</div></div>", unsafe_allow_html=True,
        )
        _score_color = _d_score_col(r["score"])
        _note_html = (
            "<div class='tv-card' style='line-height:1.75;font-size:.82rem;color:#787b86'>"
            "<div class='tv-label' style='margin-bottom:8px'>&#128221; Research Summary</div>"
            + r["reason"].replace("  •  ", "<br>&rarr; ")
            + "<div style='margin-top:10px;padding-top:8px;border-top:1px solid #2a3347;font-size:.7rem;color:#434651'>"
            + f"Score: <span style='color:{_score_color};font-weight:700'>{r['score']:+.4f}</span>"
            + f" &nbsp;&middot;&nbsp; ATR: {r['atr_pct']:.2f}%"
            + f" &nbsp;&middot;&nbsp; Vol Z: {r.get('vol_z',0):.2f}&sigma;"
            + f" &nbsp;&middot;&nbsp; &#8377;{r['traded_val_cr']:.2f} Cr/d"
            + "</div></div>"
        )
        no_c.markdown(_note_html, unsafe_allow_html=True)
        # Pattern hits
        st.markdown("<div class='tv-section'>🎯 Pattern Hits</div>", unsafe_allow_html=True)
        CAT_C = {"Trend":"#38bdf8","Momentum":"#f59e0b","Candlestick":"#a855f7",
                 "Breakout":"#26a69a","Volume":"#3b82f6","Volatility":"#787b86",
                 "Price Action":"#d1d4dc","Structure":"#67e8f9"}
        cm: dict = defaultdict(list)
        for sc,lb,cat in hits: cm[cat].append((sc,lb))
        for cat, items in cm.items():
            cc_=CAT_C.get(cat,"#787b86")
            rh="".join([f"<div style='display:flex;align-items:center;gap:10px;margin-bottom:6px'><div style='width:{int(sc*120)}px;height:4px;background:{cc_};border-radius:2px'></div><span style='font-size:.8rem'>{lb}</span><span style='font-size:.7rem;color:{cc_};margin-left:auto'>{sc*100:.0f}%</span></div>" for sc,lb in items])
            with st.expander(f"  {cat}  ({len(items)})", expanded=len(items)>=2):
                st.markdown(f"<div style='padding:4px 0'>{rh}</div>", unsafe_allow_html=True)

        if show_tv_ta:
            st.markdown("<div class='tv-section'>🧮 TradingView Technical Analysis</div>", unsafe_allow_html=True)
            components.html(tv_technical_analysis(sel_sym, nonce=sel_sym), height=470, scrolling=False)
        if show_tv_fin:
            st.markdown("<div class='tv-section'>💹 Financial Statements</div>", unsafe_allow_html=True)
            components.html(tv_financials(sel_sym), height=850, scrolling=False)


    # TAB 5 — LIVE CHARTS (TradingView)
    # ══════════════════════════════════════════════════════════════════════
    with tabs[5]:
        ch_c1, ch_c2 = st.columns([3, 1])
        with ch_c1:
            chart_sym = st.selectbox("Select Stock for Chart",
                [r["symbol"] for r in alerts], key="tv_chart_sym")
        with ch_c2:
            chart_h = st.select_slider("Chart Height",
                options=[400,480,560,640,720], value=560, key="chart_h")

        # Main TradingView chart
        r_chart = next(x for x in alerts if x["symbol"] == chart_sym)
        st.markdown("<div class='tv-section'>📊 TradingView Live Chart</div>", unsafe_allow_html=True)

        # Signal level overlay info
        stl_c = r_chart["levels"]["short_term"]; ltl_c = r_chart["levels"]["long_term"]
        info_c1, info_c2, info_c3, info_c4, info_c5 = st.columns(5)
        info_c1.metric("Entry",     f"₹{stl_c['entry']:,.2f}")
        info_c2.metric("ST Target", f"₹{stl_c['tp']:,.2f}",    f"{(stl_c['tp']/stl_c['entry']-1)*100:+.1f}%")
        info_c3.metric("ST SL",     f"₹{stl_c['sl']:,.2f}",    f"{(stl_c['sl']/stl_c['entry']-1)*100:+.1f}%")
        info_c4.metric("LT Target", f"₹{ltl_c['tp']:,.2f}",    f"{(ltl_c['tp']/ltl_c['entry']-1)*100:+.1f}%")
        info_c5.metric("R:R (ST)",  stl_c["rr_str"])

        components.html(tv_chart_widget(chart_sym, height=chart_h), height=chart_h+10, scrolling=False)

        # ── Mini charts grid ───────────────────────────────────────────────
        st.markdown("<div class='tv-section' style='margin-top:20px'>📊 All Signal Mini Charts</div>", unsafe_allow_html=True)
        grid_cols = st.columns(3)
        for i, r_g in enumerate(alerts[:9]):
            with grid_cols[i % 3]:
                ai_g=r_g["ai"]["ai_pct"]; ltr_g,gc_g=_d_grade(ai_g)
                st_g=r_g["levels"]["short_term"]
                st.markdown(f"""
                <div class='tv-card tv-card-bull' style='padding:8px 12px;margin-bottom:4px'>
                  <div style='display:flex;justify-content:space-between;align-items:center;margin-bottom:4px'>
                    <span class='tv-badge-sym' style='font-size:.8rem;padding:2px 8px'>{r_g["symbol"]}</span>
                    <span style='font-family:Syne,sans-serif;font-weight:700;color:#38bdf8'>₹{r_g["last_close"]:,.2f}</span>
                  </div>
                  <div style='display:flex;justify-content:space-between;font-size:.7rem;color:#434651'>
                    <span>AI: <span style='color:{gc_g}'>{ai_g:.0f}%</span></span>
                    <span>Tgt: <span style='color:#26a69a'>₹{st_g["tp"]:,.0f}</span></span>
                    <span>{st_g["rr_str"]}</span>
                  </div>
                </div>""", unsafe_allow_html=True)
                components.html(tv_mini_chart(r_g["symbol"], height=160, nonce=r_g["symbol"]), height=164, scrolling=False)

    # ══════════════════════════════════════════════════════════════════════
    # TAB 6 — ANALYSIS
    # ══════════════════════════════════════════════════════════════════════
    with tabs[6]:
        # ── Stock + View selector ─────────────────────────────────────────
        anc1, anc2 = st.columns([2, 2])
        with anc1:
            an_sym = st.selectbox(
                "Stock for Analysis",
                [r["symbol"] for r in alerts], key="an_sym",
                format_func=lambda s: f"{s}  —  AI: {next(r['ai']['ai_pct'] for r in alerts if r['symbol']==s):.1f}%"
            )
        with anc2:
            an_view = st.radio(
                "View",
                ["📊 Basic Analysis", "🔬 Deep Analysis", "📺 TradingView TA"],
                horizontal=True, key="an_view"
            )

        r_an   = next(x for x in alerts if x["symbol"] == an_sym)
        ai_an  = r_an["ai"]; mk_an  = r_an["mkt"]
        stl_an = r_an["levels"]["short_term"]; ltl_an = r_an["levels"]["long_term"]
        ltr_an, gc_an = _d_grade(ai_an["ai_pct"])

        # ── Quick KPI strip ───────────────────────────────────────────────
        _kpi_rsi_col = _d_rsi_col(r_an["rsi"])
        _kpi_adx_col = _d_adx_col(r_an["adx"])
        _kpi_vol_col = _d_vol_col(r_an["vol_ratio"])
        _kpi_html = (
            f"<div class='tv-card tv-card-bull' style='padding:12px 18px;border-color:{gc_an}55;margin-bottom:4px'>"
            f"<div style='display:flex;gap:24px;flex-wrap:wrap;align-items:center'>"
            f"<span class='tv-badge-sym' style='font-size:.95rem;padding:3px 12px'>{an_sym}</span>"
            f"<span class='tv-badge-grade' style='background:{gc_an}22;color:{gc_an};font-size:1.1rem'>{ltr_an}</span>"
            f"<div style='text-align:center'><div class='tv-label'>Price</div>"
            f"<div style='font-family:Syne,sans-serif;font-weight:700;color:#38bdf8'>&#8377;{r_an['last_close']:,.2f}</div></div>"
            f"<div style='text-align:center'><div class='tv-label'>AI Score</div>"
            f"<div style='font-family:Syne,sans-serif;font-weight:700;color:{gc_an}'>{ai_an['ai_pct']:.1f}%</div></div>"
            f"<div style='text-align:center'><div class='tv-label'>RSI</div>"
            f"<div style='font-family:Syne,sans-serif;font-weight:700;color:{_kpi_rsi_col}'>{r_an['rsi']:.1f}</div></div>"
            f"<div style='text-align:center'><div class='tv-label'>ADX</div>"
            f"<div style='font-family:Syne,sans-serif;font-weight:700;color:{_kpi_adx_col}'>{r_an['adx']:.1f}</div></div>"
            f"<div style='text-align:center'><div class='tv-label'>Volume</div>"
            f"<div style='font-family:Syne,sans-serif;font-weight:700;color:{_kpi_vol_col}'>{r_an['vol_ratio']:.2f}&times;</div></div>"
            f"<div style='text-align:center'><div class='tv-label'>ATR %</div>"
            f"<div style='font-family:Syne,sans-serif;font-weight:700;color:#f59e0b'>{r_an['atr_pct']:.2f}%</div></div>"
            f"<div style='text-align:center'><div class='tv-label'>ST Target</div>"
            f"<div style='font-family:Syne,sans-serif;font-weight:700;color:#26a69a'>&#8377;{stl_an['tp']:,.2f}</div></div>"
            f"<div style='text-align:center'><div class='tv-label'>R:R</div>"
            f"<div style='font-family:Syne,sans-serif;font-weight:700;color:#d1d4dc'>{stl_an['rr_str']}</div></div>"
            f"</div></div>"
        )
        st.markdown(_kpi_html, unsafe_allow_html=True)

        # ══════════════════════════════════════════════════════════════════
        # VIEW 1 — BASIC ANALYSIS (Technical + Fundamental clearly separated)
        # ══════════════════════════════════════════════════════════════════
        if an_view == "📊 Basic Analysis":

            # ────────────────────────────────────────────────────────────────
            # SECTION A: TECHNICAL ANALYSIS
            # ────────────────────────────────────────────────────────────────
            st.markdown("""
            <div style='margin:18px 0 12px;padding:10px 18px;
                        background:linear-gradient(90deg,#38bdf820,#0b0e11);
                        border-left:4px solid #38bdf8;border-radius:4px'>
              <div style='font-family:Syne,sans-serif;font-weight:800;font-size:1.05rem;
                          color:#38bdf8;letter-spacing:.1em'>
                📊 TECHNICAL ANALYSIS
              </div>
              <div style='font-size:.72rem;color:#787b86;margin-top:2px'>
                Price action, indicators, momentum and trend analysis
              </div>
            </div>""", unsafe_allow_html=True)

            # ── T1: Trend & EMA Analysis ──────────────────────────────────
            st.markdown("<div style='font-size:.75rem;font-weight:700;color:#38bdf8;letter-spacing:.08em;text-transform:uppercase;padding:6px 0;border-bottom:1px solid #1c2030;margin-bottom:10px'>📈 Trend & EMA Structure</div>", unsafe_allow_html=True)

            e9=r_an["ema9"]; e21=r_an["ema21"]; e50=r_an["ema50"]; e200=r_an.get("ema200",0)
            c_=r_an["last_close"]
            ema_stack = "🟢 Full Bull (9>21>50>200)" if c_>e9>e21>e50>e200 and e200>0 else \
                        "🟢 Bull Stack (9>21>50)" if c_>e9>e21>e50 else \
                        "🟡 Partial Bull (9>21)" if c_>e9>e21 else \
                        "🔴 Bearish Stack"
            ema_col_  = "#26a69a" if "Bull" in ema_stack else "#ef5350"

            te1,te2,te3,te4 = st.columns(4)
            for tc, lbl_, val_, ref_, is_bull in [
                (te1,"EMA 9",  f"₹{e9:.2f}",  f"{'Above' if c_>e9 else 'Below'} price", c_>e9),
                (te2,"EMA 21", f"₹{e21:.2f}", f"{'Above' if c_>e21 else 'Below'} price",c_>e21),
                (te3,"EMA 50", f"₹{e50:.2f}", f"{'Above' if c_>e50 else 'Below'} price",c_>e50),
                (te4,"EMA 200",f"₹{e200:.2f}" if e200 else "N/A","Long-term trend",c_>e200),
            ]:
                _col = "#26a69a" if is_bull else "#ef5350"
                tc.markdown(
                    f"<div style='background:#131722;border-radius:4px;padding:10px 12px;"
                    f"border-left:3px solid {_col};text-align:center'>"
                    f"<div class='tv-label'>{lbl_}</div>"
                    f"<div style='font-family:Syne,sans-serif;font-weight:700;font-size:1rem;color:#d1d4dc;margin:4px 0'>{val_}</div>"
                    f"<div style='font-size:.72rem;color:{_col}'>{'✅' if is_bull else '❌'} {ref_}</div>"
                    f"</div>",
                    unsafe_allow_html=True,
                )

            st.markdown(
                f"<div style='background:#131722;border-radius:4px;padding:10px 16px;margin:8px 0;display:flex;gap:14px;align-items:center'>"
                f"<span style='font-size:.82rem;color:#787b86'>EMA Stack:</span>"
                f"<span style='font-weight:700;color:{ema_col_}'>{ema_stack}</span>"
                f"<span style='margin-left:auto;font-size:.75rem;color:#787b86'>SuperTrend:</span>"
                f"<span style='font-weight:700;color:{'#a855f7' if r_an.get('st_flip') else '#26a69a'}'>{'⚡ JUST FLIPPED BULLISH' if r_an.get('st_flip') else '🟢 Bullish Mode'}</span>"
                f"</div>",
                unsafe_allow_html=True,
            )

            # ── T2: Momentum Indicators ───────────────────────────────────
            st.markdown("<div style='font-size:.75rem;font-weight:700;color:#38bdf8;letter-spacing:.08em;text-transform:uppercase;padding:6px 0;border-bottom:1px solid #1c2030;margin:14px 0 10px'>⚡ Momentum Indicators</div>", unsafe_allow_html=True)

            rsi_=r_an["rsi"]; adx_=r_an["adx"]; macd_=r_an["macd_h"]; vol_=r_an["vol_ratio"]

            mom_data = [
                ("RSI (14)", rsi_, f"{rsi_:.1f}",
                 "OVERSOLD 🟢" if rsi_<30 else "Near Oversold" if rsi_<42 else "Neutral" if rsi_<58 else "Elevated ⚠️" if rsi_<72 else "OVERBOUGHT 🔴",
                 _d_rsi_col(rsi_)),
                ("ADX", adx_, f"{adx_:.1f}",
                 "Very Strong 💪" if adx_>=40 else "Strong" if adx_>=28 else "Moderate" if adx_>=20 else "Weak ⚠️",
                 _d_adx_col(adx_)),
                ("MACD Hist", None, f"{macd_:+.4f}",
                 "Positive — Bullish ✅" if macd_>0 else "Negative — Bearish",
                 "#26a69a" if macd_>0 else "#ef5350"),
                ("Volume Ratio", None, f"{vol_:.2f}×",
                 "SURGE 🔊" if vol_>=2.5 else "High 📈" if vol_>=1.5 else "Average" if vol_>=0.8 else "Low ⬇️",
                 _d_vol_col(vol_)),
                ("ATR %", None, f"{r_an['atr_pct']:.2f}%",
                 "Ideal ✓" if 1.5<r_an['atr_pct']<5 else "High ⚠️",
                 "#26a69a" if 1.5<r_an['atr_pct']<5 else "#f59e0b"),
            ]

            mc1,mc2,mc3,mc4,mc5 = st.columns(5)
            for mc, (lbl_m, num_m, val_m, status_m, col_m) in zip([mc1,mc2,mc3,mc4,mc5], mom_data):
                bar_w = int(min(max((num_m if num_m else 50)/100,0),1)*100) if num_m is not None else 50
                mc.markdown(
                    f"<div style='background:#131722;border-radius:4px;padding:10px 10px;border-bottom:3px solid {col_m}'>"
                    f"<div class='tv-label'>{lbl_m}</div>"
                    f"<div style='font-family:Syne,sans-serif;font-weight:700;font-size:1.1rem;color:{col_m};margin:5px 0'>{val_m}</div>"
                    f"<div style='font-size:.68rem;color:#787b86'>{status_m}</div>"
                    f"</div>",
                    unsafe_allow_html=True,
                )

            # ── T3: Confidence Gauges ─────────────────────────────────────
            st.markdown("<div style='font-size:.75rem;font-weight:700;color:#38bdf8;letter-spacing:.08em;text-transform:uppercase;padding:6px 0;border-bottom:1px solid #1c2030;margin:14px 0 10px'>🎯 Signal Confidence</div>", unsafe_allow_html=True)

            cg1,cg2,cg3 = st.columns(3)
            for gc_w, pct_, lbl_g, clr_g, extra_g in [
                (cg1, ai_an["ai_pct"],          "🤖 AI Model Score",      gc_an,
                 f"T:{ai_an['trend_s']:+.2f}  M:{ai_an['mom_s']:+.2f}  B:{ai_an['brk_s']:+.2f}  V:{ai_an['vol_s']:+.2f}"),
                (cg2, mk_an["pct"],              "📊 Market Confidence",
                 "#26a69a" if mk_an["pct"]>=65 else "#f59e0b" if mk_an["pct"]>=40 else "#ef5350",
                 f"{mk_an['label']}  ·  {mk_an['align']}"),
                (cg3, r_an["pat_conf"]*100,      "🎯 Pattern Confidence",
                 "#26a69a" if r_an["pat_conf"]*100>=60 else "#f59e0b" if r_an["pat_conf"]*100>=40 else "#ef5350",
                 f"{len(r_an['hits'])} signals  ·  {r_an['n_cats']} categories"),
            ]:
                _gb = gauge_html(pct_, "", 220)
                gc_w.markdown(
                    f"<div class='tv-card' style='text-align:center;padding:14px 12px;border-top:3px solid {clr_g}'>"
                    f"<div class='tv-label'>{lbl_g}</div>"
                    f"<div style='font-family:Syne,sans-serif;font-weight:800;font-size:1.8rem;color:{clr_g};margin:8px 0'>{pct_:.1f}%</div>"
                    f"{_gb}"
                    f"<div style='font-size:.68rem;color:#434651;margin-top:6px'>{extra_g}</div>"
                    f"</div>",
                    unsafe_allow_html=True,
                )

            # ── T4: Trade Setup ───────────────────────────────────────────
            st.markdown("<div style='font-size:.75rem;font-weight:700;color:#38bdf8;letter-spacing:.08em;text-transform:uppercase;padding:6px 0;border-bottom:1px solid #1c2030;margin:14px 0 10px'>📐 Trade Setup</div>", unsafe_allow_html=True)
            ts1,ts2,ts3 = st.columns(3)
            _dip_e2=round(c_*0.98,2); _dip_sl2=round(_dip_e2-r_an["atr"]*1.2,2)
            _dip_tp2=round(_dip_e2+r_an["atr"]*2.5,2); _dip_rr2=round((_dip_tp2-_dip_e2)/max(_dip_e2-_dip_sl2,0.01),2)
            for ts_col,title_,entry_,tp_,sl_,rr_,wind_,brd_ in [
                (ts1,"⚡ Aggressive (Short-Term)",stl_an["entry"],stl_an["tp"],stl_an["sl"],stl_an["rr_str"],"2–5 days","#f59e0b"),
                (ts2,"📅 Swing (Long-Term)",       ltl_an["entry"],ltl_an["tp"],ltl_an["sl"],ltl_an["rr_str"],"10–20 days","#26a69a"),
                (ts3,"📌 Limit / Dip Entry",       _dip_e2,_dip_tp2,_dip_sl2,f"1:{_dip_rr2}","Limit at −2%","#3b82f6"),
            ]:
                _up_ = (tp_/entry_-1)*100; _dn_ = (sl_/entry_-1)*100
                ts_col.markdown(
                    f"<div class='tv-card' style='border-top:3px solid {brd_};text-align:center;padding:12px'>"
                    f"<div style='color:{brd_};font-size:.78rem;font-weight:700;margin-bottom:10px'>{title_}</div>"
                    f"<div class='tv-label'>Entry</div>"
                    f"<div style='font-family:Syne,sans-serif;font-weight:700;font-size:1.15rem;color:#38bdf8;margin-bottom:8px'>&#8377;{entry_:,.2f}</div>"
                    f"<div style='display:flex;justify-content:space-around;margin-bottom:8px'>"
                    f"<div><div class='tv-label'>Target</div><div style='font-weight:700;color:#26a69a'>&#8377;{tp_:,.2f}</div>"
                    f"<div style='font-size:.68rem;color:#26a69a'>{_up_:+.1f}%</div></div>"
                    f"<div><div class='tv-label'>Stop</div><div style='font-weight:700;color:#ef5350'>&#8377;{sl_:,.2f}</div>"
                    f"<div style='font-size:.68rem;color:#ef5350'>{_dn_:+.1f}%</div></div>"
                    f"</div>"
                    f"<div style='display:flex;justify-content:space-around;border-top:1px solid #2a3347;padding-top:7px'>"
                    f"<div><div class='tv-label'>R:R</div><div style='font-weight:700;color:#d1d4dc'>{rr_}</div></div>"
                    f"<div><div class='tv-label'>Window</div><div style='font-size:.72rem;color:#787b86'>{wind_}</div></div>"
                    f"</div></div>",
                    unsafe_allow_html=True,
                )

            # ── T5: Pattern Hits ──────────────────────────────────────────
            st.markdown("<div style='font-size:.75rem;font-weight:700;color:#38bdf8;letter-spacing:.08em;text-transform:uppercase;padding:6px 0;border-bottom:1px solid #1c2030;margin:14px 0 10px'>🎯 Detected Patterns & Signals</div>", unsafe_allow_html=True)
            hits_an = r_an["hits"]
            CAT_C_AN = {"Trend":"#38bdf8","Momentum":"#f59e0b","Candlestick":"#a855f7",
                        "Breakout":"#26a69a","Volume":"#3b82f6","Volatility":"#787b86",
                        "Price Action":"#d1d4dc","Structure":"#67e8f9"}
            for sc_h, lb_h, cat_h in hits_an[:8]:
                _cc_ = CAT_C_AN.get(cat_h,"#787b86"); _bw_ = int(sc_h*200)
                st.markdown(
                    f"<div style='display:flex;align-items:center;gap:10px;padding:7px 12px;"
                    f"margin-bottom:5px;background:#131722;border-radius:4px;border-left:3px solid {_cc_}'>"
                    f"<span style='font-size:.68rem;font-weight:600;color:{_cc_};min-width:90px'>{cat_h}</span>"
                    f"<div style='height:4px;width:{_bw_}px;background:{_cc_};border-radius:2px;min-width:4px'></div>"
                    f"<span style='font-size:.82rem;color:#d1d4dc;flex:1'>{lb_h}</span>"
                    f"<span style='font-size:.75rem;color:{_cc_};font-weight:700;min-width:34px;text-align:right'>{sc_h*100:.0f}%</span>"
                    f"</div>",
                    unsafe_allow_html=True,
                )

            # ────────────────────────────────────────────────────────────────
            # SECTION B: FUNDAMENTAL ANALYSIS
            # ────────────────────────────────────────────────────────────────
            st.markdown("""
            <div style='margin:24px 0 12px;padding:10px 18px;
                        background:linear-gradient(90deg,#a855f720,#0b0e11);
                        border-left:4px solid #a855f7;border-radius:4px'>
              <div style='font-family:Syne,sans-serif;font-weight:800;font-size:1.05rem;
                          color:#a855f7;letter-spacing:.1em'>
                🏦 FUNDAMENTAL ANALYSIS
              </div>
              <div style='font-size:.72rem;color:#787b86;margin-top:2px'>
                Financial health, valuation, ownership and business quality
              </div>
            </div>""", unsafe_allow_html=True)

            pe_an=r_an.get("pe"); roe_an=r_an.get("roe"); mcap_an=r_an.get("mcap")
            beta_an=r_an.get("beta"); w52h_an=r_an.get("w52h"); w52l_an=r_an.get("w52l")
            pb_an=r_an.get("pb")

            # ── F1: Valuation Metrics ─────────────────────────────────────
            st.markdown("<div style='font-size:.75rem;font-weight:700;color:#a855f7;letter-spacing:.08em;text-transform:uppercase;padding:6px 0;border-bottom:1px solid #1c2030;margin-bottom:10px'>💰 Valuation & Size</div>", unsafe_allow_html=True)

            fv1,fv2,fv3,fv4,fv5 = st.columns(5)
            val_cards = [
                (fv1,"P/E Ratio",   f"{pe_an:.1f}" if pe_an else "N/A",
                 "Cheap" if pe_an and pe_an<15 else "Fair" if pe_an and pe_an<28 else "Expensive" if pe_an and pe_an<50 else "N/A",
                 "#26a69a" if pe_an and pe_an<20 else "#f59e0b" if pe_an and pe_an<35 else "#ef5350"),
                (fv2,"P/B Ratio",   f"{pb_an:.2f}" if pb_an else "N/A",
                 "Cheap" if pb_an and pb_an<1.5 else "Fair" if pb_an and pb_an<4 else "Expensive" if pb_an else "N/A",
                 "#26a69a" if pb_an and pb_an<2 else "#f59e0b" if pb_an and pb_an<5 else "#ef5350"),
                (fv3,"Market Cap",  fmt_cr(mcap_an) if mcap_an else "N/A",
                 "Large Cap" if mcap_an and mcap_an>2000 else "Mid Cap" if mcap_an and mcap_an>500 else "Small Cap" if mcap_an else "N/A",
                 "#38bdf8"),
                (fv4,"Beta",        f"{beta_an:.2f}" if beta_an else "N/A",
                 "Low risk" if beta_an and beta_an<0.8 else "Moderate" if beta_an and beta_an<1.3 else "High risk" if beta_an else "N/A",
                 "#26a69a" if beta_an and beta_an<0.8 else "#f59e0b" if beta_an and beta_an<1.3 else "#ef5350"),
                (fv5,"ROE",         f"{roe_an*100:.1f}%" if roe_an else "N/A",
                 "Excellent" if roe_an and roe_an>0.20 else "Good" if roe_an and roe_an>0.12 else "Weak" if roe_an else "N/A",
                 "#26a69a" if roe_an and roe_an>0.18 else "#f59e0b" if roe_an and roe_an>0.10 else "#ef5350"),
            ]
            for (vcol,vlbl,vval,vstatus,vcol_) in val_cards:
                vcol.markdown(
                    f"<div style='background:#131722;border-radius:4px;padding:10px 10px;border-bottom:3px solid {vcol_}'>"
                    f"<div class='tv-label'>{vlbl}</div>"
                    f"<div style='font-family:Syne,sans-serif;font-weight:700;font-size:1.05rem;color:{vcol_};margin:5px 0'>{vval}</div>"
                    f"<div style='font-size:.68rem;color:#787b86'>{vstatus}</div>"
                    f"</div>",
                    unsafe_allow_html=True,
                )

            # ── F2: 52-Week Price Range ───────────────────────────────────
            st.markdown("<div style='font-size:.75rem;font-weight:700;color:#a855f7;letter-spacing:.08em;text-transform:uppercase;padding:6px 0;border-bottom:1px solid #1c2030;margin:14px 0 10px'>📅 52-Week Price Range</div>", unsafe_allow_html=True)

            if w52h_an and w52l_an and c_:
                _rng   = w52h_an - w52l_an
                _pos   = (c_ - w52l_an) / _rng * 100 if _rng > 0 else 50
                _pct_from_low  = (c_/w52l_an - 1)*100
                _pct_from_high = (c_/w52h_an - 1)*100
                pr1,pr2,pr3 = st.columns([1,2,1])
                pr1.markdown(
                    f"<div style='background:#131722;border-radius:4px;padding:10px 14px;text-align:center'>"
                    f"<div class='tv-label'>52W Low</div>"
                    f"<div style='font-family:Syne,sans-serif;font-weight:700;color:#26a69a'>&#8377;{w52l_an:,.2f}</div>"
                    f"<div style='font-size:.68rem;color:#26a69a'>{_pct_from_low:+.1f}% from here</div></div>",
                    unsafe_allow_html=True,
                )
                _pos_col = "#26a69a" if _pos < 40 else "#f59e0b" if _pos < 70 else "#ef5350"
                pr2.markdown(
                    f"<div style='background:#131722;border-radius:4px;padding:14px 16px'>"
                    f"<div style='display:flex;justify-content:space-between;margin-bottom:6px'>"
                    f"<span style='font-size:.7rem;color:#787b86'>Current: <b style='color:#38bdf8'>&#8377;{c_:,.2f}</b></span>"
                    f"<span style='font-size:.7rem;color:{_pos_col};font-weight:700'>{_pos:.0f}% of range</span>"
                    f"</div>"
                    f"<div style='background:#2a3347;border-radius:4px;height:10px;position:relative'>"
                    f"<div style='background:{_pos_col};border-radius:4px;height:10px;width:{_pos:.0f}%'></div>"
                    f"<div style='position:absolute;left:{_pos:.0f}%;top:-4px;transform:translateX(-50%);width:3px;height:18px;background:#fff;border-radius:2px'></div>"
                    f"</div>"
                    f"<div style='font-size:.68rem;color:#434651;margin-top:4px;text-align:center'>"
                    f"{'Near 52W Low — potential accumulation zone 🟢' if _pos<25 else '52W High zone — momentum play ⚠️' if _pos>75 else 'Mid-range — balanced risk/reward'}"
                    f"</div></div>",
                    unsafe_allow_html=True,
                )
                pr3.markdown(
                    f"<div style='background:#131722;border-radius:4px;padding:10px 14px;text-align:center'>"
                    f"<div class='tv-label'>52W High</div>"
                    f"<div style='font-family:Syne,sans-serif;font-weight:700;color:#ef5350'>&#8377;{w52h_an:,.2f}</div>"
                    f"<div style='font-size:.68rem;color:#ef5350'>{_pct_from_high:+.1f}% from here</div></div>",
                    unsafe_allow_html=True,
                )
            else:
                st.info("52-week range data not available — enable fundamentals fetch.")

            # ── F3: Traded Volume & Liquidity ─────────────────────────────
            st.markdown("<div style='font-size:.75rem;font-weight:700;color:#a855f7;letter-spacing:.08em;text-transform:uppercase;padding:6px 0;border-bottom:1px solid #1c2030;margin:14px 0 10px'>💧 Liquidity & Traded Data</div>", unsafe_allow_html=True)

            liq1,liq2,liq3,liq4 = st.columns(4)
            liq_data = [
                (liq1,"Avg Daily Volume",  f"{r_an['avg_vol']/1e5:.1f}L shares",
                 "✅ Liquid" if r_an["avg_vol"]>=1_500_000 else "⚠️ Low liquidity","#26a69a"),
                (liq2,"Median Traded Value",f"₹{r_an['traded_val_cr']:.2f} Cr/day",
                 "✅ High liquidity" if r_an["traded_val_cr"]>=5 else "⚠️ Limited","#26a69a"),
                (liq3,"Vol Z-Score",       f"{r_an.get('vol_z',0):.2f}σ",
                 "Unusual activity 🔊" if abs(r_an.get('vol_z',0))>2 else "Normal range","#f59e0b"),
                (liq4,"F&O Eligible",      "YES ✅" if r_an["is_fo"] else "NO",
                 "Can hedge with options" if r_an["is_fo"] else "Cash segment only",
                 "#38bdf8" if r_an["is_fo"] else "#787b86"),
            ]
            for (lc,ll,lv,ls,lcl) in liq_data:
                lc.markdown(
                    f"<div style='background:#131722;border-radius:4px;padding:10px 10px;border-left:3px solid {lcl}'>"
                    f"<div class='tv-label'>{ll}</div>"
                    f"<div style='font-family:Syne,sans-serif;font-weight:700;color:{lcl};margin:5px 0'>{lv}</div>"
                    f"<div style='font-size:.68rem;color:#787b86'>{ls}</div>"
                    f"</div>",
                    unsafe_allow_html=True,
                )

            # ── F4: Research Note ─────────────────────────────────────────
            st.markdown("<div style='font-size:.75rem;font-weight:700;color:#a855f7;letter-spacing:.08em;text-transform:uppercase;padding:6px 0;border-bottom:1px solid #1c2030;margin:14px 0 10px'>📝 Integrated Research Note</div>", unsafe_allow_html=True)
            st.markdown(
                f"<div class='tv-card' style='border-left:3px solid #a855f7;font-size:.83rem;color:#787b86;line-height:1.8'>"
                f"<div style='font-size:.72rem;color:#a855f7;font-weight:600;margin-bottom:8px'>"
                f"SECTOR: {r_an.get('sector','N/A')}  ·  INDUSTRY: {r_an.get('industry','N/A')}  ·  INDICES: {r_an['indices']}</div>"
                f"{r_an['reason'].replace('  •  ', '<br><span style=\"color:#a855f7\">→</span> ')}"
                f"</div>",
                unsafe_allow_html=True,
            )

        # ══════════════════════════════════════════════════════════════════
        # VIEW 2 — DEEP ANALYSIS
        # ══════════════════════════════════════════════════════════════════
        elif an_view == "🔬 Deep Analysis":
            st.markdown("<div class='tv-section'>🧮 8-Factor Score Decomposition</div>", unsafe_allow_html=True)
            af1, af2 = st.columns([1.5, 1])
            with af1:
                st.plotly_chart(factor_bar_fig(r_an["ai"]), use_container_width=True,
                                config={"displayModeBar": False}, key=_uid("pc"))
            with af2:
                st.plotly_chart(radar_fig(r_an["ai"]), use_container_width=True,
                                config={"displayModeBar": False}, key=_uid("pc"))

            st.markdown("<div class='tv-section'>📊 All Signals — Multi-Stock Comparison</div>", unsafe_allow_html=True)
            st.plotly_chart(multi_score_fig(alerts), use_container_width=True,
                            config={"displayModeBar": False}, key=_uid("pc"))

            an2c1, an2c2 = st.columns(2)
            with an2c1:
                st.markdown("<div class='tv-section'>🎯 RSI vs ADX — Bubble Size = AI Score</div>", unsafe_allow_html=True)
                st.plotly_chart(rsi_vs_adx_fig(alerts), use_container_width=True,
                                config={"displayModeBar": False}, key=_uid("pc"))
            with an2c2:
                st.markdown("<div class='tv-section'>📐 Score Distribution</div>", unsafe_allow_html=True)
                st.plotly_chart(score_distribution_fig(alerts), use_container_width=True,
                                config={"displayModeBar": False}, key=_uid("pc"))

            st.markdown("<div class='tv-section'>📋 Factor Detail Table</div>", unsafe_allow_html=True)
            W_deep = {"trend_s":0.24,"mom_s":0.16,"brk_s":0.17,"vol_s":0.10,
                      "pat_s":0.10,"fund_s":0.08,"sent_s":0.04}
            factor_labels = {
                "trend_s":"📈 Trend","mom_s":"⚡ Momentum","brk_s":"🚀 Breakout",
                "vol_s":"🔊 Volume","pat_s":"🎯 Pattern","fund_s":"🏦 Fundamental",
                "sent_s":"📡 Sentiment",
            }
            factor_df_rows = []
            for key, label in factor_labels.items():
                sc_f = float(ai_an.get(key, 0))
                wt_f = W_deep.get(key, 0)
                direction = "▲▲ Strong Bull" if sc_f > 0.5 else "▲ Bullish" if sc_f > 0.2 else \
                            "◆ Neutral" if sc_f > -0.2 else "▽ Bearish"
                factor_df_rows.append({
                    "Factor": label, "Score": round(sc_f, 3),
                    "Weight": f"{wt_f:.0%}",
                    "Contribution": round(sc_f * wt_f, 4),
                    "Direction": direction,
                })
            st.dataframe(
                pd.DataFrame(factor_df_rows),
                use_container_width=True, hide_index=True,
                column_config={
                    "Score": st.column_config.NumberColumn("Score", format="%.3f"),
                    "Contribution": st.column_config.NumberColumn("Contribution", format="%.4f"),
                }
            )

        # ══════════════════════════════════════════════════════════════════
        # VIEW 3 — TRADINGVIEW TA
        # ══════════════════════════════════════════════════════════════════
        else:
            st.markdown("<div class='tv-section'>📺 TradingView Technical Analysis — Live Data</div>", unsafe_allow_html=True)
            st.info(f"📊 Showing live TradingView analysis for **NSE:{an_sym}** — switch stock above to update", icon="ℹ️")

            tv_interval = st.select_slider(
                "TradingView Interval",
                options=["1", "5", "15", "60", "240", "1D", "1W", "1M"],
                value="1D", key="tv_interval_sel"
            )

            tv_ta_html = f"""
            <!-- TradingView Widget BEGIN nonce={an_sym}_{tv_interval} -->
            <div class="tradingview-widget-container" style="border:1px solid #2a3347;border-radius:4px;overflow:hidden;">
              <div class="tradingview-widget-container__widget"></div>
              <script type="text/javascript" src="https://s3.tradingview.com/external-embedding/embed-widget-technical-analysis.js" async>
              {{"interval":"{tv_interval}","width":"100%","isTransparent":true,"height":"500",
               "symbol":"NSE:{an_sym}","showIntervalTabs":true,"displayMode":"multiple",
               "locale":"en","colorTheme":"dark"}}
              </script>
            </div>"""
            components.html(tv_ta_html, height=520, scrolling=False)

            st.markdown("<div class='tv-section' style='margin-top:16px'>📊 Symbol Info + Financials</div>", unsafe_allow_html=True)
            sf1, sf2 = st.columns([1, 1.5])
            with sf1:
                components.html(
                    f"""<!-- nonce={an_sym} -->
                    <div class="tradingview-widget-container">
                      <div class="tradingview-widget-container__widget"></div>
                      <script type="text/javascript" src="https://s3.tradingview.com/external-embedding/embed-widget-symbol-info.js" async>
                      {{"symbol":"NSE:{an_sym}","width":"100%","locale":"en","colorTheme":"dark","isTransparent":true}}
                      </script>
                    </div>""",
                    height=180, scrolling=False
                )
                pe_v=r_an.get("pe"); roe_v=r_an.get("roe"); mcap_v=r_an.get("mcap")
                beta_v=r_an.get("beta"); w52h_v=r_an.get("w52h"); w52l_v=r_an.get("w52l")
                for lbl_f,col_f,val_f in [
                    ("P/E","#26a69a" if pe_v and pe_v<20 else "#f59e0b",f"{pe_v:.1f}" if pe_v else "N/A"),
                    ("ROE","#26a69a" if roe_v and roe_v>0.18 else "#f59e0b",f"{roe_v*100:.1f}%" if roe_v else "N/A"),
                    ("Mkt Cap","#d1d4dc",fmt_cr(mcap_v)),
                    ("Beta","#d1d4dc",f"{beta_v:.2f}" if beta_v else "N/A"),
                    ("52W High","#26a69a",fmt_inr(w52h_v)),
                    ("52W Low","#f59e0b",fmt_inr(w52l_v)),
                    ("Traded Val","#38bdf8",f"₹{r_an['traded_val_cr']:.2f} Cr/d"),
                ]:
                    st.markdown(
                        f"<div style='display:flex;justify-content:space-between;padding:5px 10px;"
                        f"margin-bottom:4px;background:#1c2030;border-radius:3px'>"
                        f"<span style='font-size:.75rem;color:#787b86'>{lbl_f}</span>"
                        f"<span style='font-size:.8rem;color:{col_f};font-weight:600'>{val_f}</span>"
                        f"</div>", unsafe_allow_html=True,
                    )
            with sf2:
                components.html(
                    f"""<!-- nonce={an_sym}_fin -->
                    <div class="tradingview-widget-container" style="border:1px solid #2a3347;border-radius:4px;overflow:hidden;">
                      <div class="tradingview-widget-container__widget"></div>
                      <script type="text/javascript" src="https://s3.tradingview.com/external-embedding/embed-widget-financials.js" async>
                      {{"isTransparent":true,"displayMode":"regular","width":"100%","height":"600",
                       "colorTheme":"dark","symbol":"NSE:{an_sym}","locale":"en"}}
                      </script>
                    </div>""",
                    height=620, scrolling=False
                )


    # TAB 7 — BACKTEST + TRADE LOG
    # ══════════════════════════════════════════════════════════════════════
    with tabs[7]:
        # ── What is Trade Log explanation ─────────────────────────────────
        st.markdown("""
        <div style='padding:12px 18px;background:linear-gradient(90deg,#38bdf815,#0b0e11);
                    border-left:4px solid #38bdf8;border-radius:4px;margin-bottom:16px'>
          <div style='font-family:Syne,sans-serif;font-weight:700;font-size:.95rem;color:#38bdf8'>
            📋 What is the Trade Log?
          </div>
          <div style='font-size:.78rem;color:#787b86;margin-top:5px;line-height:1.7'>
            The Trade Log is the <b style='color:#d1d4dc'>simulated backtest result</b> of running our signal engine
            on historical price data. For each stock that generated a LONG signal, the engine:<br>
            ① <b style='color:#26a69a'>Enters</b> at next-day open price + slippage (5 bps)<br>
            ② <b style='color:#ef5350'>Exits</b> at stop-loss (−4%), take-profit (+9%), max hold (12 bars), or signal reversal<br>
            ③ Deducts <b style='color:#f59e0b'>12 bps transaction cost</b> per trade (brokerage + taxes)<br>
            This is <b style='color:#a855f7'>NOT live trading</b> — it shows how signals performed historically
            on the scanned period. Use it to gauge signal quality before placing real orders.
          </div>
        </div>""", unsafe_allow_html=True)

        # ── KPI metrics ───────────────────────────────────────────────────
        st.markdown("<div class='tv-section'>📈 Backtest Performance — {}</div>".format(
            f"Threshold {cfg_for_display.get('threshold','N/A') if 'cfg_for_display' in dir() else '—'}"
            if False else "Simulated Historical Performance"
        ), unsafe_allow_html=True)

        bt_kpi1,bt_kpi2,bt_kpi3,bt_kpi4,bt_kpi5,bt_kpi6,bt_kpi7 = st.columns(7)
        _bt_fin = bt.get("final", 0)
        _bt_cap = bt.get("final",0) / max(1+bt_ret, 0.001) if bt_ret != 0 else 1_000_000
        bt_kpi1.metric("📊 Total Return",     f"{bt_ret:+.2%}",  delta_color="normal")
        bt_kpi2.metric("📐 Sharpe Ratio",      f"{bt_sh:.3f}",   f"{'Good' if bt_sh>1 else 'OK' if bt_sh>0.5 else 'Poor'}")
        bt_kpi3.metric("📉 Max Drawdown",      f"{abs(bt_dd):.2%}", delta_color="inverse")
        bt_kpi4.metric("🎯 Win Rate",          f"{bt_wr:.1%}",   f"{bt_tr} trades")
        bt_kpi5.metric("📦 Avg Bars Held",     f"{bt.get('avg_bars',0):.1f}", "bars per trade")
        bt_kpi6.metric("💰 Avg Trade Return",  f"{bt.get('avg_ret',0):+.2%}")
        bt_kpi7.metric("🏦 Final Capital",     f"₹{_bt_fin:,.0f}" if _bt_fin else "—")

        # ── Charts row ────────────────────────────────────────────────────
        bc1, bc2 = st.columns([2.2, 1])
        with bc1:
            st.markdown("<div class='tv-section'>📊 Cumulative P&L Equity Curve</div>", unsafe_allow_html=True)
            ef = equity_curve_fig(bt)
            if ef.data:
                st.plotly_chart(ef, use_container_width=True, config={"displayModeBar":False}, key=_uid("pc"))
            else:
                st.info("No closed trades yet — run a scan first.", icon="📊")
        with bc2:
            st.markdown("<div class='tv-section'>💸 Per-Trade P&L (Last 20)</div>", unsafe_allow_html=True)
            st.plotly_chart(waterfall_fig(bt), use_container_width=True, config={"displayModeBar":False}, key=_uid("pc"))

        # ── Trade Log ─────────────────────────────────────────────────────
        trd_df = bt.get("trades_df", pd.DataFrame())
        if trd_df.empty:
            st.info("No trades recorded — run a scan with live data to populate the trade log.", icon="📋")
        else:
            st.markdown("<div class='tv-section'>📋 Live Trade Log — All Simulated Trades</div>",
                        unsafe_allow_html=True)

            # Controls
            tl1,tl2,tl3 = st.columns(3)
            with tl1:
                tl_filter = st.selectbox("Filter", ["All Trades","Winners Only","Losers Only","Stop Loss Exits","Take Profit Exits"], key="tl_filter", label_visibility="collapsed")
            with tl2:
                tl_sort = st.selectbox("Sort", ["Exit Date (Latest First)","P&L (Best First)","P&L (Worst First)","Return % (Best)","Bars Held (Shortest)"], key="tl_sort", label_visibility="collapsed")
            with tl3:
                tl_sym = st.text_input("Search Symbol", placeholder="e.g. RELIANCE", key="tl_sym", label_visibility="collapsed")

            disp = trd_df.copy()
            # Rename columns
            col_map = {"sym":"Symbol","entry":"Entry Date","exit":"Exit Date",
                       "ep":"Entry ₹","xp":"Exit ₹","bars":"Bars","reason":"Exit Reason",
                       "pnl":"P&L ₹","ret":"Return%"}
            disp = disp.rename(columns={k:v for k,v in col_map.items() if k in disp.columns})

            # Apply symbol search
            if tl_sym.strip() and "Symbol" in disp.columns:
                disp = disp[disp["Symbol"].str.contains(tl_sym.strip().upper(), na=False)]

            # Apply filter
            if "P&L ₹" in disp.columns:
                if tl_filter == "Winners Only":    disp = disp[disp["P&L ₹"] > 0]
                elif tl_filter == "Losers Only":   disp = disp[disp["P&L ₹"] < 0]
                elif tl_filter == "Stop Loss Exits" and "Exit Reason" in disp.columns:
                    disp = disp[disp["Exit Reason"] == "stop_loss"]
                elif tl_filter == "Take Profit Exits" and "Exit Reason" in disp.columns:
                    disp = disp[disp["Exit Reason"] == "take_profit"]

            # Apply sort
            if "P&L ₹" in disp.columns:
                _sort_map = {
                    "Exit Date (Latest First)": ("Exit Date", False),
                    "P&L (Best First)":         ("P&L ₹", False),
                    "P&L (Worst First)":        ("P&L ₹", True),
                    "Return % (Best)":          ("Return%", False),
                    "Bars Held (Shortest)":     ("Bars", True),
                }
                _sc,_asc = _sort_map.get(tl_sort, ("Exit Date", False))
                if _sc in disp.columns:
                    disp = disp.sort_values(_sc, ascending=_asc)

            # Compute formatted columns
            if "P&L ₹" in disp.columns:
                disp["P&L"] = disp["P&L ₹"].apply(lambda v: f"₹{v:+,.2f}")
                disp["Return"] = disp["Return%"].apply(lambda v: f"{v:+.2%}") if "Return%" in disp.columns else "—"

            # Summary cards
            _n_win   = int((trd_df.get("pnl", pd.Series()) > 0).sum()) if "pnl" in trd_df else 0
            _n_loss  = int((trd_df.get("pnl", pd.Series()) < 0).sum()) if "pnl" in trd_df else 0
            _tot_pnl = float(trd_df.get("pnl", pd.Series()).fillna(0).sum()) if "pnl" in trd_df else 0
            _avg_win = float(trd_df[trd_df.get("pnl",pd.Series()).fillna(0)>0]["pnl"].fillna(0).mean()) if _n_win > 0 else 0
            _avg_los = float(trd_df[trd_df.get("pnl",pd.Series()).fillna(0)<0]["pnl"].fillna(0).mean()) if _n_loss > 0 else 0
            _profit_factor = _safe_float(abs(_avg_win * _n_win / (_avg_los * _n_loss)), 0) if _n_loss > 0 and _avg_los != 0 else 0

            st.markdown(
                "<div style='display:flex;gap:10px;margin-bottom:12px;flex-wrap:wrap'>"
                + "".join(
                    f"<div style='background:#131722;border-radius:4px;padding:9px 14px;text-align:center;border-left:3px solid {c_}'>"
                    f"<div class='tv-label'>{l_}</div>"
                    f"<div style='font-family:Syne,sans-serif;font-weight:700;color:{c_};font-size:.95rem'>{v_}</div></div>"
                    for l_,v_,c_ in [
                        ("Showing",        f"{len(disp)} trades",    "#787b86"),
                        ("Winners 🟢",     str(_n_win),               "#26a69a"),
                        ("Losers 🔴",      str(_n_loss),              "#ef5350"),
                        ("Total P&L",      f"₹{_tot_pnl:+,.0f}",    "#26a69a" if _tot_pnl>=0 else "#ef5350"),
                        ("Avg Win",        f"₹{_avg_win:+,.0f}",    "#26a69a"),
                        ("Avg Loss",       f"₹{_avg_los:+,.0f}",    "#ef5350"),
                        ("Profit Factor",  f"{_safe_float(_profit_factor,0):.2f}",  "#38bdf8" if _safe_float(_profit_factor,0)>=1.5 else "#f59e0b"),
                    ]
                )
                + "</div>",
                unsafe_allow_html=True,
            )

            # Display table
            show_cols = [c for c in ["Symbol","Entry Date","Exit Date","Entry ₹","Exit ₹","Bars","Exit Reason","P&L","Return"] if c in disp.columns]
            st.dataframe(
                disp[show_cols],
                use_container_width=True,
                height=400,
                hide_index=True,
                column_config={
                    "Entry ₹":    st.column_config.NumberColumn("Entry ₹",  format="₹%.2f"),
                    "Exit ₹":     st.column_config.NumberColumn("Exit ₹",   format="₹%.2f"),
                    "Exit Reason":st.column_config.TextColumn("Exit Reason", help="stop_loss=hit SL | take_profit=hit TP | signal_exit=signal reversed | max_hold=time exit | eop=end of period"),
                },
            )

            # Legend
            st.markdown("""
            <div style='font-size:.72rem;color:#434651;margin-top:8px;line-height:1.8'>
              <b style='color:#787b86'>Exit Reason Legend:</b>
              &nbsp; <span style='color:#ef5350'>stop_loss</span> = Stop-loss triggered (−4% default)
              &nbsp;·&nbsp; <span style='color:#26a69a'>take_profit</span> = Take-profit hit (+9% default)
              &nbsp;·&nbsp; <span style='color:#f59e0b'>signal_exit</span> = Signal reversed (stock no longer bullish)
              &nbsp;·&nbsp; <span style='color:#787b86'>max_hold</span> = Maximum holding period reached (12 bars)
              &nbsp;·&nbsp; <span style='color:#434651'>eop</span> = End of backtest period (force closed)
            </div>""", unsafe_allow_html=True)

            # Download
            _tl_csv = disp[show_cols].to_csv(index=False).encode()
            st.download_button("⬇️ Download Trade Log CSV", data=_tl_csv,
                file_name=f"trade_log_{datetime.now().strftime('%Y%m%d_%H%M')}.csv",
                mime="text/csv", key="tl_csv")

    # ══════════════════════════════════════════════════════════════════════
    # ══════════════════════════════════════════════════════════════════════
    # TAB 8 — SIGNAL TRACKER (From March 20, 2026)
    # ══════════════════════════════════════════════════════════════════════
    with tabs[8]:

        st.markdown("""
        <div style='padding:14px 20px;background:linear-gradient(90deg,#a855f720,#38bdf810,#0b0e11);
                    border-left:4px solid #a855f7;border-radius:6px;margin-bottom:16px'>
          <div style='font-family:Syne,sans-serif;font-weight:800;font-size:1.1rem;color:#a855f7'>
            📡 Signal Tracker — All Signals Since March 20, 2026
          </div>
          <div style='font-size:.76rem;color:#787b86;margin-top:4px;line-height:1.7'>
            Every scan auto-saves signals here. <b style='color:#d1d4dc'>Add any stock manually</b> using
            just Symbol + Date + Entry Price — the system <b style='color:#26a69a'>auto-computes all targets,
            stop loss, RSI, ADX, AI score and fundamentals</b> via live data.
            Status auto-updates: <span style='color:#26a69a'>🚀 TARGET HIT</span> ·
            <span style='color:#ef5350'>🔴 SL HIT</span> · <span style='color:#38bdf8'>🔵 OPEN</span>
          </div>
        </div>""", unsafe_allow_html=True)

        # ── Load tracked signals ──────────────────────────────────────────
        _tracked = _tracker_load()

        # ── Import / Export panel (cloud data migration) ─────────────────
        with st.expander("📤 Import / Export — Backup & Restore Tracker Data", expanded=False):
            _ie1, _ie2 = st.columns(2)

            with _ie1:
                st.markdown(
                    "<div style='font-size:.78rem;color:#787b86;padding:4px 0 8px'>"
                    "<b style='color:#d1d4dc'>Export</b> — download all your tracked signals "
                    "as JSON. Use this to back up before redeployment or move data between devices."
                    "</div>",
                    unsafe_allow_html=True,
                )
                if _tracked:
                    _export_bytes = json.dumps(_tracked, indent=2, default=str).encode()
                    st.download_button(
                        "⬇️ Download tracker.json",
                        data=_export_bytes,
                        file_name=f"tracker_backup_{datetime.now().strftime('%Y%m%d_%H%M')}.json",
                        mime="application/json",
                        use_container_width=True,
                        key="export_json",
                    )
                    # Connection status
                    _has_gist = bool(_gist_token() and _gist_id())
                    st.markdown(
                        f"<div style='font-size:.7rem;margin-top:6px;padding:5px 8px;"
                        f"background:{'#26a69a15' if _has_gist else '#ef535015'};"
                        f"border-radius:3px;color:{'#26a69a' if _has_gist else '#ef5350'}'>"
                        f"{'✅ Gist storage active — data persists across restarts' if _has_gist else '⚠️ No Gist configured — data will be lost on restart. See setup guide below.'}"
                        f"</div>",
                        unsafe_allow_html=True,
                    )
                else:
                    st.info("No signals to export yet.", icon="📭")

            with _ie2:
                st.markdown(
                    "<div style='font-size:.78rem;color:#787b86;padding:4px 0 8px'>"
                    "<b style='color:#d1d4dc'>Import</b> — upload a previously exported "
                    "<code>tracker.json</code> to restore your signals. This <b>merges</b> "
                    "with existing data (deduplicates by symbol + date)."
                    "</div>",
                    unsafe_allow_html=True,
                )
                _up_file = st.file_uploader(
                    "Upload tracker.json", type=["json"],
                    key="import_json", label_visibility="collapsed",
                )
                if _up_file:
                    try:
                        _import_data = json.load(_up_file)
                        if not isinstance(_import_data, list):
                            st.error("Invalid file — must be a JSON array.")
                        else:
                            _existing_keys = {(r["symbol"], r.get("scan_date","")) for r in _tracked}
                            _added = 0
                            for _ir in _import_data:
                                _ik = (_ir.get("symbol",""), _ir.get("scan_date",""))
                                if _ik not in _existing_keys:
                                    _tracked.append(_ir)
                                    _existing_keys.add(_ik)
                                    _added += 1
                            if _added > 0:
                                _tracker_save(_tracked)
                                st.success(f"✅ Imported {_added} new signals ({len(_import_data) - _added} duplicates skipped).")
                                st.rerun()
                            else:
                                st.info("All signals in the file are already in the tracker.")
                    except Exception as _ie_err:
                        st.error(f"Error reading file: {_ie_err}")

            # Gist setup guide
            if not (_gist_token() and _gist_id()):
                st.markdown("""---
                <div style='font-size:.76rem;color:#787b86;line-height:1.8'>
                <b style='color:#f59e0b'>⚠️ Gist not configured — your tracker data will be lost when the app restarts.</b><br>
                To enable permanent storage, follow these 3 steps:
                </div>""", unsafe_allow_html=True)
                st.markdown("""
**Step 1 — Create a GitHub Gist:**
1. Go to [gist.github.com](https://gist.github.com)
2. Create a new **secret** gist
3. Filename: `nse_signal_tracker.json`, Content: `[]`
4. Click **Create secret gist**
5. Copy the **Gist ID** from the URL: `gist.github.com/username/`**`THIS_PART`**

**Step 2 — Create a GitHub Personal Access Token (PAT):**
1. GitHub → Settings → Developer settings → Personal access tokens → Tokens (classic)
2. Click **Generate new token (classic)**
3. Tick **`gist`** scope only
4. Click **Generate token** and copy it immediately

**Step 3 — Add to Streamlit Cloud Secrets:**
1. Streamlit Cloud → your app → **Settings** → **Secrets**
2. Paste exactly:
```toml
[gist]
token   = "ghp_your_token_here"
gist_id = "your_gist_id_here"
```
3. Click **Save** — app restarts and data is now permanent ✅
""")

        # ═══════════════════════════════════════════════════════════════════
        # ADD / REMOVE PANEL
        # ═══════════════════════════════════════════════════════════════════
        with st.expander("➕ Add Stock to Tracker   |   ❌ Remove Stock", expanded=True):
            _ap_tab, _rp_tab = st.tabs(["➕ Add Stock", "❌ Remove Stock"])

            # ── ADD STOCK ─────────────────────────────────────────────────
            with _ap_tab:
                st.markdown("""
                <div style='font-size:.78rem;color:#787b86;padding:6px 0 10px;line-height:1.7'>
                  Enter <b style='color:#d1d4dc'>Symbol · Date · Entry Price</b> only.
                  All other fields (ST Target, Stop Loss, LT Target, R:R, RSI, ADX,
                  EMA stack, Volume, AI Score, Fundamentals) are <b style='color:#26a69a'>auto-fetched
                  and computed from live market data</b>.
                </div>""", unsafe_allow_html=True)

                _ac1, _ac2, _ac3 = st.columns([2, 1.5, 1.5])
                with _ac1:
                    _add_sym   = st.text_input(
                        "NSE Symbol", placeholder="e.g. RELIANCE, HDFCBANK, TCS",
                        key="add_sym", help="Exact NSE symbol — will be validated against universe"
                    ).strip().upper()
                with _ac2:
                    _add_date  = st.date_input(
                        "Entry / Signal Date", key="add_date",
                        value=__import__('datetime').date(2026, 3, 20),
                        help="Date when you received the signal or bought the stock"
                    )
                with _ac3:
                    _add_price = st.number_input(
                        "Entry Price ₹", min_value=0.01, value=100.0,
                        step=0.5, format="%.2f", key="add_price",
                        help="Price at which you entered / plan to enter"
                    )

                _ab1, _ab2 = st.columns([1, 2])
                with _ab1:
                    _do_add = st.button(
                        "🚀 Add & Auto-Analyse", use_container_width=True, key="do_add",
                        help="Fetches live data and computes all signal fields automatically"
                    )
                with _ab2:
                    st.markdown(
                        "<div style='font-size:.72rem;color:#434651;padding:8px 0'>"
                        "⚡ Will auto-fetch: ATR-based targets & SL · RSI · ADX · EMA stack · "
                        "Volume ratio · AI Score · P/E · ROE · Analyst target · 52W range</div>",
                        unsafe_allow_html=True,
                    )

                if _do_add:
                    if not _add_sym:
                        st.error("Please enter a symbol.")
                    elif _add_price <= 0:
                        st.error("Entry price must be > 0.")
                    else:
                        # Check for duplicate (same symbol + same date)
                        _dup = any(
                            r["symbol"] == _add_sym and r["scan_date"] == str(_add_date)
                            for r in _tracked
                        )
                        if _dup:
                            st.warning(f"⚠️ {_add_sym} on {_add_date} already in tracker. Remove it first to re-add.")
                        else:
                            with st.spinner(f"🔍 Fetching live data for {_add_sym} and computing signal levels…"):
                                _new_rec = _tracker_enrich(_add_sym, _add_price, str(_add_date))

                            _tracked.append(_new_rec)
                            _tracker_save(_tracked)

                            _method = _new_rec.get("added_method","")
                            if "minimal" in _method:
                                st.warning(
                                    f"⚠️ Added {_add_sym} with basic levels only — "
                                    "live data fetch failed. Check symbol spelling or internet connection."
                                )
                            else:
                                st.success(
                                    f"✅ **{_add_sym}** added with full analysis:\n"
                                    f"- ST Target: ₹{_new_rec['st_target']:,.2f}  |  Stop: ₹{_new_rec['st_sl']:,.2f}  |  R:R {_new_rec['st_rr']}\n"
                                    f"- LT Target: ₹{_new_rec['lt_target']:,.2f}  |  AI Score: {_new_rec['ai_score']}%\n"
                                    f"- RSI: {_new_rec['rsi']}  |  ADX: {_new_rec['adx']}  |  Sector: {_new_rec['sector']}"
                                )
                            st.rerun()

            # ── REMOVE STOCK ──────────────────────────────────────────────
            with _rp_tab:
                if not _tracked:
                    st.info("No signals in tracker yet.")
                else:
                    st.markdown(
                        "<div style='font-size:.78rem;color:#787b86;padding:6px 0 8px'>"
                        "Select one or more signals to remove from the tracker.</div>",
                        unsafe_allow_html=True,
                    )
                    # Build display options: "RELIANCE (2026-03-20)"
                    _opts = [
                        f"{r['symbol']}  ·  {r['scan_date']}  ·  ₹{r.get('entry_price',0):,.2f}  ·  {r.get('status','OPEN')}"
                        for r in _tracked
                    ]
                    _to_remove = st.multiselect(
                        "Select signals to remove", _opts, key="remove_sel",
                        label_visibility="collapsed"
                    )
                    _rb1, _rb2 = st.columns([1, 2])
                    with _rb1:
                        if st.button("❌ Remove Selected", use_container_width=True,
                                     key="do_remove", disabled=not _to_remove):
                            # Build set of keys to remove
                            _remove_keys = set()
                            for opt in _to_remove:
                                parts = opt.split("  ·  ")
                                if len(parts) >= 2:
                                    _remove_keys.add((parts[0].strip(), parts[1].strip()))
                            _before = len(_tracked)
                            _tracked = [
                                r for r in _tracked
                                if (r["symbol"], r["scan_date"]) not in _remove_keys
                            ]
                            _tracker_save(_tracked)
                            st.success(f"Removed {_before - len(_tracked)} signal(s).")
                            st.rerun()
                    with _rb2:
                        if st.button("🗑️ Clear ALL Signals", use_container_width=True,
                                     key="clear_all", type="secondary"):
                            _tracked = []
                            _tracker_save(_tracked)
                            st.success("All signals cleared.")
                            st.rerun()

        # ── Update Status panel ───────────────────────────────────────────
        with st.expander("✏️ Update Signal Status (Mark as Closed / Target / SL)", expanded=False):
            _open_syms = [
                f"{r['symbol']}  ·  {r['scan_date']}  ·  ₹{r.get('entry_price',0):,.2f}"
                for r in _tracked if r.get("status", "OPEN") == "OPEN"
            ]
            if _open_syms:
                uc1, uc2, uc3 = st.columns([2.5, 1.5, 1])
                with uc1:
                    u_sel = st.selectbox("Select signal", _open_syms, key="u_sel",
                                         label_visibility="collapsed")
                with uc2:
                    u_status = st.selectbox("New status",
                        ["TARGET_HIT","SL_HIT","CLOSED","OPEN"],
                        key="u_status", label_visibility="collapsed")
                with uc3:
                    u_price = st.number_input("Close ₹", min_value=0.0, value=0.0,
                                               step=0.5, key="u_price",
                                               label_visibility="collapsed")
                if st.button("✅ Update Status", key="u_upd", use_container_width=True):
                    _sel_parts = u_sel.split("  ·  ")
                    _sel_sym   = _sel_parts[0].strip()
                    _sel_date  = _sel_parts[1].strip() if len(_sel_parts) > 1 else ""
                    for _rec in _tracked:
                        if _rec["symbol"] == _sel_sym and _rec.get("scan_date","") == _sel_date:
                            _rec["status"]       = u_status
                            _rec["close_date"]   = datetime.now().strftime("%Y-%m-%d")
                            _rec["close_price"]  = round(u_price, 2) if u_price > 0 else None
                            _rec["close_reason"] = u_status.lower().replace("_"," ")
                            break
                    _tracker_save(_tracked)
                    st.success(f"Updated {_sel_sym} → {u_status}")
                    st.rerun()
            else:
                st.info("No open signals to update.")

        if not _tracked:
            st.info(
                "📭 Tracker is empty.\n\n"
                "• **Auto-populated**: Run any scan → signals from March 20, 2026 onwards are saved automatically.\n"
                "• **Add manually**: Use the ➕ Add panel above — enter symbol, date, price and everything else is computed.\n",
                icon="📡",
            )
        else:
            # ── Live price fetch (5-min cache) ────────────────────────────
            @st.cache_data(ttl=300, show_spinner=False)
            def _live_prices_batch(sym_tuple):
                prices = {}
                import io as _iox, contextlib as _clx
                for sym in sym_tuple:
                    try:
                        buf = _iox.StringIO()
                        with _clx.redirect_stdout(buf), _clx.redirect_stderr(buf):
                            h = yf.Ticker(yf_ticker(sym)).history(period="2d", auto_adjust=True)
                        if h is not None and len(h) > 0:
                            _px = h["Close"].iloc[-1]
                            if _px == _px:   # NaN check
                                prices[sym] = round(float(_px), 2)
                    except Exception:
                        pass
                return prices

            _ref_btn = st.button("🔄 Refresh Live Prices", key="tr_ref",
                                  help="Auto-refreshes every 5 minutes")
            if _ref_btn:
                st.cache_data.clear()

            _all_syms_t = tuple(sorted({r["symbol"] for r in _tracked}))
            with st.spinner(f"Fetching live prices for {len(_all_syms_t)} symbols…"):
                _live_px = _live_prices_batch(_all_syms_t)

            # ── Build enriched rows ───────────────────────────────────────
            _rows = []
            for _rec in _tracked:
                sym    = _rec["symbol"]
                ep     = _safe_float(_rec.get("entry_price"), 0)
                stp    = _safe_float(_rec.get("st_target"),   0)
                sl     = _safe_float(_rec.get("st_sl"),       0)
                ltp    = _safe_float(_rec.get("lt_target"),   0)
                lt_sl_ = _safe_float(_rec.get("lt_sl"),       0)
                live   = _live_px.get(sym)
                status = _rec.get("status", "OPEN")

                # Auto-check targets
                if status == "OPEN" and live and ep > 0:
                    if stp > 0 and live >= stp:
                        status = "TARGET_HIT"
                        _rec["status"] = "TARGET_HIT"
                    elif sl > 0 and live <= sl:
                        status = "SL_HIT"
                        _rec["status"] = "SL_HIT"

                # P&L
                _ref_price = live if live else (_rec.get("close_price") or ep)
                pnl_pct  = round((_ref_price / ep - 1) * 100, 2) if ep > 0 and _ref_price else None
                pnl_rs   = round(_ref_price - ep, 2) if ep > 0 and _ref_price else None

                # Upside remaining
                upside_rem = round((stp / live - 1) * 100, 2) if (live and stp > 0 and status == "OPEN") else None

                # Progress 0→100 from entry to target
                if live and ep > 0 and stp > ep and sl < ep:
                    _rng  = stp - sl
                    prog  = round(min(max((live - sl) / _rng * 100, 0), 110), 1)
                    entry_pct = round((ep - sl) / _rng * 100, 1)
                else:
                    prog = None; entry_pct = 33.0

                # Holding days
                try:
                    import datetime as _dt
                    sig_dt    = _dt.datetime.strptime(_rec["scan_date"], "%Y-%m-%d")
                    hold_days = (datetime.now() - sig_dt).days
                except Exception:
                    hold_days = 0

                if status == "TARGET_HIT": scol="#26a69a"; sico="🚀 TARGET HIT"
                elif status == "SL_HIT":   scol="#ef5350"; sico="🔴 SL HIT"
                elif status == "CLOSED":   scol="#787b86"; sico="⬜ CLOSED"
                else:                      scol="#38bdf8"; sico="🔵 OPEN"

                _rows.append(dict(
                    _sym=sym, _ep=ep, _stp=stp, _sl=sl, _ltp=ltp, _lt_sl=lt_sl_,
                    _live=live, _pnl_pct=pnl_pct, _pnl_rs=pnl_rs,
                    _upside_rem=upside_rem, _progress=prog, _entry_pct=entry_pct,
                    _hold_days=hold_days, _status=status, _scol=scol, _sico=sico,
                    _scan_date=_rec.get("scan_date",""), _scan_time=_rec.get("scan_time",""),
                    _ai=_safe_float(_rec.get("ai_score"), 0),
                    _rsi=_safe_float(_rec.get("rsi"), 0),
                    _adx=_safe_float(_rec.get("adx"), 0),
                    _vol=_safe_float(_rec.get("vol_ratio"), 0),
                    _atr_pct=_safe_float(_rec.get("atr_pct"), 0),
                    _st_rr=_rec.get("st_rr",""),
                    _lt_rr=_rec.get("lt_rr",""),
                    _top_signal=_rec.get("top_signal",""),
                    _sector=_rec.get("sector","N/A"),
                    _is_fo=_rec.get("is_fo", False),
                    _stflip=_rec.get("st_flip", 0),
                    _pe=_rec.get("pe"), _roe=_rec.get("roe"),
                    _de=_rec.get("de"), _eps_g=_rec.get("eps_g"),
                    _rev_g=_rec.get("rev_g"), _mcap=_rec.get("mcap_cr"),
                    _analyst=_rec.get("analyst",""),
                    _analyst_tgt=_rec.get("analyst_target"),
                    _w52l=_rec.get("w52l"), _w52h=_rec.get("w52h"),
                    _ema_bull=_rec.get("ema_bull"),
                    _method=_rec.get("added_method","auto"),
                ))

            _tracker_save(_tracked)   # persist auto-updated statuses

            # ── Summary KPIs ──────────────────────────────────────────────
            _n_open   = sum(1 for r in _rows if r["_status"]=="OPEN")
            _n_hit    = sum(1 for r in _rows if r["_status"]=="TARGET_HIT")
            _n_sl     = sum(1 for r in _rows if r["_status"]=="SL_HIT")
            _n_closed = sum(1 for r in _rows if r["_status"]=="CLOSED")
            _w_pnl    = [r for r in _rows if r["_pnl_pct"] is not None]
            _win_rate = round(sum(1 for r in _w_pnl if _safe_float(r["_pnl_pct"],0)>0)/max(len(_w_pnl),1)*100,1)
            _avg_pnl  = round(sum(_safe_float(r["_pnl_pct"],0) for r in _w_pnl)/max(len(_w_pnl),1),2) if _w_pnl else 0
            _best     = max(_w_pnl, key=lambda r: _safe_float(r["_pnl_pct"],-999)) if _w_pnl else None
            _worst    = min(_w_pnl, key=lambda r: _safe_float(r["_pnl_pct"], 999)) if _w_pnl else None

            st.markdown(
                "<div style='display:grid;grid-template-columns:repeat(7,1fr);gap:9px;margin-bottom:16px'>"
                + "".join(
                    f"<div style='background:#131722;border-radius:4px;padding:10px;text-align:center;border-top:3px solid {c_}'>"
                    f"<div class='tv-label'>{l_}</div>"
                    f"<div style='font-family:Syne,sans-serif;font-weight:800;font-size:1.2rem;color:{c_}'>{v_}</div></div>"
                    for l_,v_,c_ in [
                        ("Total",        str(len(_rows)),          "#a855f7"),
                        ("🔵 Open",      str(_n_open),             "#38bdf8"),
                        ("🚀 Hit",        str(_n_hit),              "#26a69a"),
                        ("🔴 SL",         str(_n_sl),               "#ef5350"),
                        ("Win Rate",     f"{_win_rate:.1f}%",      "#26a69a" if _win_rate>=50 else "#f59e0b"),
                        ("Avg P&L",      f"{_avg_pnl:+.1f}%",     "#26a69a" if _avg_pnl>=0 else "#ef5350"),
                        ("Best",         _best["_sym"] if _best else "—", "#26a69a"),
                    ]
                ) + "</div>", unsafe_allow_html=True,
            )

            # ── Filter / Sort ─────────────────────────────────────────────
            fc1,fc2,fc3 = st.columns([2,1.5,1.5])
            with fc1:
                tr_sort = st.selectbox("Sort by", [
                    "P&L% Best First","P&L% Worst First","Signal Date Newest",
                    "Signal Date Oldest","AI Score","Hold Days",
                ], key="tr_sort", label_visibility="collapsed")
            with fc2:
                tr_filt = st.selectbox("Filter", [
                    "All","Open Only","Target Hit","SL Hit","Closed","Manual Only","Auto Only"
                ], key="tr_filt", label_visibility="collapsed")
            with fc3:
                tr_search = st.text_input("Search", placeholder="Symbol / Sector…",
                                           key="tr_srch", label_visibility="collapsed").strip().upper()

            # Apply
            _dr = _rows[:]
            if tr_filt == "Open Only":    _dr = [r for r in _dr if r["_status"]=="OPEN"]
            elif tr_filt == "Target Hit": _dr = [r for r in _dr if r["_status"]=="TARGET_HIT"]
            elif tr_filt == "SL Hit":     _dr = [r for r in _dr if r["_status"]=="SL_HIT"]
            elif tr_filt == "Closed":     _dr = [r for r in _dr if r["_status"]=="CLOSED"]
            elif tr_filt == "Manual Only":_dr = [r for r in _dr if "manual" in r.get("_method","")]
            elif tr_filt == "Auto Only":  _dr = [r for r in _dr if "manual" not in r.get("_method","")]
            if tr_search:
                _dr = [r for r in _dr if tr_search in r["_sym"] or tr_search in r["_sector"]]

            _sfn = {
                "P&L% Best First":     lambda r: -_safe_float(r["_pnl_pct"],-999),
                "P&L% Worst First":    lambda r:  _safe_float(r["_pnl_pct"],999),
                "Signal Date Newest":  lambda r: r["_scan_date"],
                "Signal Date Oldest":  lambda r: r["_scan_date"],
                "AI Score":            lambda r: -r["_ai"],
                "Hold Days":           lambda r: -r["_hold_days"],
            }
            _dr.sort(key=_sfn.get(tr_sort,_sfn["P&L% Best First"]),
                     reverse=(tr_sort not in ("Signal Date Oldest",)))

            st.markdown(
                f"<div class='tv-label' style='margin:4px 0 10px'>"
                f"Showing <b style='color:#a855f7'>{len(_dr)}</b> of {len(_rows)} signals</div>",
                unsafe_allow_html=True,
            )
            # ── Signal Tracker — Tabular Format ──────────────────────────
            # Assign serial numbers by original insertion order (immutable)
            _insert_order = {
                (rec["symbol"], rec.get("scan_date","")): idx + 1
                for idx, rec in enumerate(_tracked)
            }

            # Group repeated symbols, sort each group oldest-first
            _sym_groups: dict = {}
            for r in _dr:
                _sym_groups.setdefault(r["_sym"], []).append(r)
            for _k in _sym_groups:
                _sym_groups[_k].sort(
                    key=lambda r: _insert_order.get((r["_sym"], r["_scan_date"]), 9999)
                )

            # One group-level serial = smallest serial in the group (shared by all rows)
            _group_serial = {
                sym: min(_insert_order.get((r["_sym"], r["_scan_date"]), 9999) for r in grp)
                for sym, grp in _sym_groups.items()
            }

            # ORDER: multi-entry groups first, then singles — each sorted by group serial
            _multi_syms  = sorted([s for s,g in _sym_groups.items() if len(g)>1],  key=lambda s: _group_serial[s])
            _single_syms = sorted([s for s,g in _sym_groups.items() if len(g)==1], key=lambda s: _group_serial[s])
            _sym_order   = _multi_syms + _single_syms
            _n_unique    = len(_sym_groups)
            _n_dupes     = len(_multi_syms)

            # ── Build flat table rows (groups first, singles after) ───────
            _tbl = []
            for _gsym2 in _sym_order:
                _grp_serial = _group_serial[_gsym2]   # same # for every row in this group
                for r in _sym_groups[_gsym2]:
                    _serial2 = _grp_serial
                    _ep_t    = _safe_float(r["_ep"],  0)
                    _stp_t   = _safe_float(r["_stp"], 0)
                    _sl_t    = _safe_float(r["_sl"],  0)
                    _ltp_t   = _safe_float(r["_ltp"], 0)
                    _lv_t    = r["_live"] if r["_live"] else None
                    _pc_t    = r["_pnl_pct"]
                    _prog_t  = _safe_float(r["_progress"], 0.0)
                    _up_t    = r["_upside_rem"]
                    _prs_t   = r["_pnl_rs"]
                    _tbl.append({
                        "#":           _serial2,
                        "Symbol":      r["_sym"],
                        "Status":      r["_sico"],
                        "Signal Date": r["_scan_date"],
                        "Days":        r["_hold_days"],
                        "Entry ₹":    _ep_t,
                        "Live ₹":     _lv_t if _lv_t else "—",
                        "Chg ₹":      round(_safe_float(_lv_t,_ep_t) - _ep_t, 2) if _lv_t else "—",
                        "Chg %":      _pc_t,
                        "ST Target ₹":_stp_t,
                        "Stop Loss ₹":_sl_t,
                        "LT Target ₹":_ltp_t,
                        "Upside %":   _up_t,
                        "P&L ₹/sh":  _prs_t,
                        "Progress %": _prog_t,
                        "ST R:R":     r["_st_rr"],
                        "LT R:R":     r["_lt_rr"],
                        "AI %":       r["_ai"],
                        "RSI":        r["_rsi"] if r["_rsi"] else "—",
                        "ADX":        r["_adx"] if r["_adx"] else "—",
                        "Vol×":      r["_vol"] if r["_vol"] else "—",
                        "P/E":        r["_pe"] if r["_pe"] else "—",
                        "ROE %":      r["_roe"] if r["_roe"] else "—",
                        "D/E":        r["_de"] if r["_de"] else "—",
                        "EPS G%":     r["_eps_g"] if r["_eps_g"] else "—",
                        "Analyst":    r.get("_analyst","—").upper() if r.get("_analyst") else "—",
                        "Anlst Tgt":  f"₹{r['_analyst_tgt']:,.0f}" if r.get("_analyst_tgt") else "—",
                        "F&O":        "✅" if r["_is_fo"] else "—",
                        "Sector":     r["_sector"],
                        "Top Signal": r["_top_signal"][:40] if r["_top_signal"] else "—",
                        "Source":     r.get("_method","auto").replace("-"," ").title(),
                    })

            _tbl_df = pd.DataFrame(_tbl)

            # ── Stats bar ─────────────────────────────────────────────────
            st.markdown(
                f"<div style='display:flex;align-items:center;justify-content:space-between;"
                f"padding:9px 14px;background:#131722;border-radius:5px;"
                f"border:1px solid #2a3347;margin-bottom:10px'>"
                f"<div style='display:flex;gap:16px;font-size:.75rem'>"
                f"<span style='color:#787b86'><b style='color:#d1d4dc'>{len(_tbl)}</b> rows</span>"
                f"<span style='color:#787b86'><b style='color:#a855f7'>{_n_unique}</b> symbols</span>"
                + (f"<span style='color:#f59e0b'><b>{_n_dupes}</b> repeated</span>" if _n_dupes else "")
                + f"</div>"
                f"<span style='font-family:Syne,sans-serif;font-weight:700;font-size:.8rem;"
                f"color:#a855f7;letter-spacing:.06em'>📊 SIGNAL TRACKER</span>"
                f"</div>",
                unsafe_allow_html=True,
            )

            # ── Main table ────────────────────────────────────────────────
            st.dataframe(
                _tbl_df,
                use_container_width=True,
                height=520,
                hide_index=True,
                column_config={
                    "#":           st.column_config.NumberColumn("#",          format="%d",       width=42),
                    "Symbol":      st.column_config.TextColumn("Symbol",                          width=90),
                    "Status":      st.column_config.TextColumn("Status",                          width=120),
                    "Signal Date": st.column_config.TextColumn("Signal Date",                     width=100),
                    "Days":        st.column_config.NumberColumn("Days",       format="%d d",     width=55),
                    "Entry ₹":    st.column_config.NumberColumn("Entry ₹",    format="₹%.2f",    width=90),
                    "Live ₹":     st.column_config.NumberColumn("Live ₹",     format="₹%.2f",    width=90),
                    "Chg ₹":      st.column_config.NumberColumn("Chg ₹",      format="₹%+.2f",   width=80),
                    "Chg %":      st.column_config.NumberColumn("Chg %",       format="%+.2f%%",  width=75),
                    "ST Target ₹":st.column_config.NumberColumn("ST Target",   format="₹%.2f",    width=90),
                    "Stop Loss ₹":st.column_config.NumberColumn("Stop Loss",   format="₹%.2f",    width=90),
                    "LT Target ₹":st.column_config.NumberColumn("LT Target",   format="₹%.2f",    width=90),
                    "Upside %":   st.column_config.NumberColumn("Upside %",    format="+%.1f%%",  width=75),
                    "P&L ₹/sh":  st.column_config.NumberColumn("P&L ₹/sh",   format="₹%+.2f",   width=85),
                    "Progress %": st.column_config.ProgressColumn("Progress",
                                    min_value=0, max_value=110, format="%.0f%%", width=100),
                    "AI %":       st.column_config.NumberColumn("AI %",        format="%.0f%%",   width=60),
                    "RSI":        st.column_config.NumberColumn("RSI",         format="%.0f",     width=55),
                    "ADX":        st.column_config.NumberColumn("ADX",         format="%.0f",     width=55),
                },
            )

            st.caption(
                "Grouped by symbol · sorted by first signal date · "
                "Live prices refresh every 5 min · "
                "Chg% = (Live − Entry) / Entry"
            )

            # ═══════════════════════════════════════════════════════════════
            # STRATEGY PERFORMANCE ANALYTICS
            # ═══════════════════════════════════════════════════════════════
            st.markdown("<hr style='border-color:#2a3347;margin:18px 0'>", unsafe_allow_html=True)
            st.markdown("""
            <div style='padding:12px 18px;background:linear-gradient(90deg,#38bdf820,#0b0e11);
                        border-left:4px solid #38bdf8;border-radius:5px;margin-bottom:14px'>
              <div style='font-family:Syne,sans-serif;font-weight:800;font-size:1rem;color:#38bdf8;margin-bottom:5px'>
                📊 Strategy Performance Analytics — What Is Working
              </div>
              <div style='font-size:.75rem;color:#787b86;line-height:1.8'>
                Analyses every tracked signal to answer: <b style='color:#d1d4dc'>which patterns, sectors,
                RSI zones, AI bands and market conditions actually produce profits.</b>
                Each row shows: Verdict · Avg P&L · Signal Count · Win Rate · Best/Worst · Profit Factor
              </div>
            </div>""", unsafe_allow_html=True)

            # ── HOW TO READ THIS ─────────────────────────────────────────
            with st.expander("📖 How to read Strategy Analytics — Guide", expanded=False):
                st.markdown("""
**Verdict Labels:**
| Label | Meaning | What to do |
|-------|---------|------------|
| ✅ WORKING | Avg P&L ≥ +3% AND Win Rate ≥ 55% | **Keep using this signal** |
| 🟢 Positive | Avg P&L ≥ 0% AND Win Rate ≥ 45% | Good signal, can improve |
| 🟡 Mixed | Avg P&L between −2% and 0% | Use only with extra confirmation |
| 🔴 Avoid | Avg P&L below −2% | **Stop using this signal alone** |

**Column Meanings:**
- **Avg P&L** — average profit/loss % across all signals in that category. Most important number.
- **Win Rate** — % of signals that were profitable. Above 50% is good.
- **Best / Worst** — the single best gain and worst loss in the group.
- **Profit Factor (PF)** — total profits ÷ total losses. PF > 1.5 = excellent. PF < 1.0 = losing.
- **W: / L:** — average profit on winning trades vs average loss on losing trades.

**How to use the insights:**
1. **Signals table** → If *EMA Bull Stack* shows ✅ WORKING but *BB Squeeze* shows 🔴 Avoid → only trade EMA Bull Stack signals going forward
2. **Sector table** → If *Banking* shows +8% avg but *IT* shows −3% → focus on banking signals this market cycle
3. **AI Score band** → If *AI ≥ 80%* shows ✅ but *AI < 50%* shows 🔴 → raise your AI threshold in the sidebar to 0.20+
4. **RSI Zone** → If *RSI < 30* shows +12% avg → oversold bounces are your best entry condition
5. **ADX table** → If *ADX ≥ 28* works but *ADX < 20* doesn't → only trade in confirmed trending stocks
6. **Auto vs Manual** → tells you if the scan engine outperforms your manual picks or vice versa

**Profit Factor quick reference:**
- PF > 2.0 → Excellent strategy
- PF 1.5–2.0 → Good
- PF 1.0–1.5 → Marginal, needs improvement
- PF < 1.0 → Losing money overall
                """)

            # Build analytics from _rows (enriched, has live pnl)
            _an_rows = [r for r in _rows if r["_pnl_pct"] is not None]

            if len(_an_rows) < 3:
                st.info("Need at least 3 closed/tracked signals with live prices to show analytics.", icon="📊")
            else:
                # ── Helper: compute stats for a group of rows ─────────────
                def _grp_stats(rows):
                    pnls   = [_safe_float(r["_pnl_pct"], 0) for r in rows]
                    n      = len(pnls)
                    wins   = sum(1 for p in pnls if p > 0)
                    losses = sum(1 for p in pnls if p < 0)
                    avg    = round(sum(pnls) / n, 2) if n else 0
                    best   = round(max(pnls), 2) if pnls else 0
                    worst  = round(min(pnls), 2) if pnls else 0
                    wr     = round(wins / n * 100, 1) if n else 0
                    avg_win= round(sum(p for p in pnls if p>0)/max(wins,1), 2)
                    avg_los= round(sum(p for p in pnls if p<0)/max(losses,1), 2)
                    pf     = round(abs(avg_win*wins / (avg_los*losses)), 2) if losses>0 and avg_los!=0 else None
                    return dict(n=n, wins=wins, losses=losses, avg=avg,
                                best=best, worst=worst, wr=wr,
                                avg_win=avg_win, avg_los=avg_los, pf=pf)

                # ── Render one analytics table ────────────────────────────
                def _render_analytics(title, icon, grouped_data, sort_by="avg", min_n=1):
                    """grouped_data = {label: [rows]}"""
                    rows_out = []
                    for lbl, grp_rows in grouped_data.items():
                        if len(grp_rows) < min_n:
                            continue
                        _gs = _grp_stats(grp_rows)
                        rows_out.append({"_lbl": lbl, **_gs})
                    if not rows_out:
                        return
                    rows_out.sort(key=lambda x: -_safe_float(x.get(sort_by, 0), -999))

                    st.markdown(
                        f"<div style='font-size:.8rem;font-weight:700;color:#d1d4dc;"
                        f"letter-spacing:.06em;text-transform:uppercase;"
                        f"padding:8px 0 6px;border-bottom:1px solid #1c2030;margin:12px 0 10px'>"
                        f"{icon} {title}</div>",
                        unsafe_allow_html=True,
                    )
                    for r2 in rows_out:
                        _c  = "#26a69a" if r2["avg"] >= 0 else "#ef5350"
                        _wc = "#26a69a" if r2["wr"] >= 55 else "#f59e0b" if r2["wr"] >= 40 else "#ef5350"
                        _pfc= "#26a69a" if (r2["pf"] or 0) >= 1.5 else "#f59e0b" if (r2["pf"] or 0) >= 1 else "#ef5350"
                        _verdict = (
                            "✅ WORKING"   if r2["avg"] >= 3 and r2["wr"] >= 55 else
                            "🟢 Positive"  if r2["avg"] >= 0 and r2["wr"] >= 45 else
                            "🟡 Mixed"     if r2["avg"] >= -2 else
                            "🔴 Avoid"
                        )
                        _vc = "#26a69a" if "✅" in _verdict else "#4db6ac" if "🟢" in _verdict else "#f59e0b" if "🟡" in _verdict else "#ef5350"
                        _bar_w = _safe_int(min(abs(r2["avg"]) / 10 * 120, 120), 0)

                        st.markdown(
                            f"<div style='display:grid;grid-template-columns:180px 80px 1fr 60px 60px 70px 70px 80px 90px;"
                            f"gap:0;align-items:center;padding:7px 10px;margin-bottom:3px;"
                            f"background:#131722;border-left:3px solid {_c};border-radius:4px'>"
                            # Label
                            f"<span style='font-size:.8rem;color:#d1d4dc;font-weight:600;"
                            f"overflow:hidden;text-overflow:ellipsis;white-space:nowrap;padding-right:8px'>"
                            f"{str(r2['_lbl'])[:28]}</span>"
                            # Verdict
                            f"<span style='font-size:.72rem;font-weight:700;color:{_vc}'>{_verdict}</span>"
                            # Bar
                            f"<div style='display:flex;align-items:center;gap:6px'>"
                            f"<div style='height:6px;width:{_bar_w}px;background:{_c};border-radius:3px'></div>"
                            f"<span style='font-size:.82rem;font-weight:700;color:{_c}'>{r2['avg']:+.1f}%</span>"
                            f"</div>"
                            # N
                            f"<span style='font-size:.72rem;color:#434651;text-align:center'>{r2['n']} signals</span>"
                            # Win Rate
                            f"<span style='font-size:.78rem;font-weight:600;color:{_wc};text-align:center'>{r2['wr']:.0f}% WR</span>"
                            # Best
                            f"<span style='font-size:.72rem;color:#26a69a;text-align:center'>▲{r2['best']:+.1f}%</span>"
                            # Worst
                            f"<span style='font-size:.72rem;color:#ef5350;text-align:center'>▼{r2['worst']:+.1f}%</span>"
                            # Profit Factor
                            f"<span style='font-size:.72rem;color:{_pfc};text-align:center'>"
                            f"PF {r2['pf']:.2f}</span>" if r2["pf"] else f"<span style='color:#434651;font-size:.72rem'>PF —</span>"
                            # Avg win / avg loss
                            f"<span style='font-size:.68rem;color:#434651'>"
                            f"W:{r2['avg_win']:+.1f}% L:{r2['avg_los']:+.1f}%</span>"
                            f"</div>",
                            unsafe_allow_html=True,
                        )

                # Column header legend
                st.markdown(
                    "<div style='display:grid;grid-template-columns:180px 80px 1fr 60px 60px 70px 70px 80px 90px;"
                    "gap:0;padding:4px 10px;margin-bottom:2px'>"
                    + "".join(
                        f"<span style='font-size:.6rem;color:#434651;text-transform:uppercase;letter-spacing:.06em'>{h}</span>"
                        for h in ["Category","Verdict","Avg P&L","Count","Win Rate","Best","Worst","Prof.Factor","Avg W/L"]
                    )
                    + "</div>",
                    unsafe_allow_html=True,
                )

                # ── 1. By Top Signal Pattern ──────────────────────────────
                _by_signal = {}
                for r in _an_rows:
                    _sig = r.get("_top_signal","") or "Unknown"
                    # Simplify signal name
                    _sig = (_sig.split("·")[0].strip()
                                .replace("🌟","").replace("📈","").replace("⚡","")
                                .replace("🚀","").replace("🏆","").replace("🔊","")
                                .replace("🕯️","").replace("🔨","").replace("↩️","")
                                .strip()[:35])
                    _by_signal.setdefault(_sig, []).append(r)

                _render_analytics(
                    "Performance by Signal / Pattern Type",
                    "🎯", _by_signal, sort_by="avg", min_n=2
                )

                # ── 2. By Sector ──────────────────────────────────────────
                _by_sector = {}
                for r in _an_rows:
                    _by_sector.setdefault(r.get("_sector","N/A") or "N/A", []).append(r)
                _render_analytics(
                    "Performance by Sector",
                    "🏭", _by_sector, sort_by="avg", min_n=2
                )

                # ── 3. By AI Score range ──────────────────────────────────
                _by_ai = {}
                for r in _an_rows:
                    ai_v = _safe_float(r.get("_ai"), 0)
                    if ai_v >= 80:   bkt = "AI ≥ 80% (Very High)"
                    elif ai_v >= 65: bkt = "AI 65–79% (High)"
                    elif ai_v >= 50: bkt = "AI 50–64% (Medium)"
                    else:            bkt = "AI < 50% (Low)"
                    _by_ai.setdefault(bkt, []).append(r)
                _render_analytics(
                    "Performance by AI Score Band",
                    "🤖", _by_ai, sort_by="avg"
                )

                # ── 4. By RSI Zone at signal ──────────────────────────────
                _by_rsi = {}
                for r in _an_rows:
                    rv = _safe_float(r.get("_rsi"), 50)
                    if rv and rv < 30:   bkt = "RSI < 30 (Deeply Oversold)"
                    elif rv and rv < 40: bkt = "RSI 30–39 (Oversold)"
                    elif rv and rv < 50: bkt = "RSI 40–49 (Recovering)"
                    elif rv and rv < 65: bkt = "RSI 50–64 (Neutral)"
                    else:                bkt = "RSI ≥ 65 (Elevated)"
                    _by_rsi.setdefault(bkt, []).append(r)
                _render_analytics(
                    "Performance by RSI Zone at Signal",
                    "📉", _by_rsi, sort_by="avg"
                )

                # ── 5. By ADX strength ────────────────────────────────────
                _by_adx = {}
                for r in _an_rows:
                    av = _safe_float(r.get("_adx"), 0)
                    if av and av >= 40:  bkt = "ADX ≥ 40 (Very Strong Trend)"
                    elif av and av >= 28:bkt = "ADX 28–39 (Strong)"
                    elif av and av >= 20:bkt = "ADX 20–27 (Moderate)"
                    else:                bkt = "ADX < 20 (Weak / Sideways)"
                    _by_adx.setdefault(bkt, []).append(r)
                _render_analytics(
                    "Performance by ADX Trend Strength",
                    "💪", _by_adx, sort_by="avg"
                )

                # ── 6. By Hold Duration ──────────────────────────────────
                _by_hold = {}
                for r in _an_rows:
                    hd = _safe_float(r.get("_hold_days"), 0)
                    if hd <= 3:    bkt = "0–3 days (Very Short)"
                    elif hd <= 7:  bkt = "4–7 days (Short)"
                    elif hd <= 14: bkt = "8–14 days (Medium)"
                    elif hd <= 30: bkt = "15–30 days (Swing)"
                    else:          bkt = "30+ days (Long Hold)"
                    _by_hold.setdefault(bkt, []).append(r)
                _render_analytics(
                    "Performance by Hold Duration",
                    "⏱️", _by_hold, sort_by="avg"
                )

                # ── 7. By Source (Auto vs Manual) ─────────────────────────
                _by_src = {}
                for r in _an_rows:
                    _by_src.setdefault(r.get("_method","auto").replace("-"," ").title(), []).append(r)
                _render_analytics(
                    "Auto-Generated vs Manually Added Signals",
                    "⚙️", _by_src, sort_by="avg"
                )

                # ── 7. By F&O eligibility ─────────────────────────────────
                _by_fo = {}
                for r in _an_rows:
                    bkt = "F&O Eligible" if r.get("_is_fo") else "Non-F&O (Cash Only)"
                    _by_fo.setdefault(bkt, []).append(r)
                _render_analytics(
                    "F&O Eligible vs Non-F&O Stocks",
                    "🔰", _by_fo, sort_by="avg"
                )

                # ── 8. Top 10 best and worst individual signals ───────────
                st.markdown(
                    "<div style='font-size:.8rem;font-weight:700;color:#d1d4dc;"
                    "letter-spacing:.06em;text-transform:uppercase;"
                    "padding:8px 0 6px;border-bottom:1px solid #1c2030;margin:16px 0 10px'>"
                    "🏆 Top 10 Best Signals  &nbsp;|&nbsp;  💀 Top 10 Worst Signals</div>",
                    unsafe_allow_html=True,
                )
                _sorted_all = sorted(_an_rows, key=lambda r: -_safe_float(r["_pnl_pct"], -999))
                _top10 = _sorted_all[:10]
                _bot10 = _sorted_all[-10:][::-1]

                _t10_col, _b10_col = st.columns(2)
                for _col_w, _items, _title_c in [
                    (_t10_col, _top10, "#26a69a"),
                    (_b10_col, _bot10, "#ef5350"),
                ]:
                    with _col_w:
                        for idx_t, r in enumerate(_items, 1):
                            _pv   = _safe_float(r["_pnl_pct"], 0)
                            _pclr = "#26a69a" if _pv >= 0 else "#ef5350"
                            _bg   = "#0D2B1E" if _pv >= 0 else "#2B0D0D"
                            _col_w.markdown(
                                f"<div style='display:flex;justify-content:space-between;"
                                f"align-items:center;padding:6px 10px;margin-bottom:3px;"
                                f"background:{_bg};border-left:3px solid {_pclr};border-radius:3px'>"
                                f"<div>"
                                f"<span style='font-size:.68rem;color:#434651;margin-right:6px'>#{idx_t}</span>"
                                f"<span style='font-family:Syne,sans-serif;font-weight:700;color:#d1d4dc'>{r['_sym']}</span>"
                                f"<span style='font-size:.68rem;color:#434651;margin-left:6px'>{r['_scan_date']}</span>"
                                f"</div>"
                                f"<div style='text-align:right'>"
                                f"<span style='font-family:Syne,sans-serif;font-weight:800;"
                                f"color:{_pclr}'>{_pv:+.2f}%</span>"
                                f"<div style='font-size:.65rem;color:#434651'>"
                                f"{(r.get('_top_signal','') or '')[:30]}</div>"
                                f"</div>"
                                f"</div>",
                                unsafe_allow_html=True,
                            )

                # ── 9. Overall summary box ────────────────────────────────
                _all_stats = _grp_stats(_an_rows)
                _oc = "#26a69a" if _all_stats["avg"] >= 0 else "#ef5350"
                _overall_verdict = (
                    "✅ Strategy is PROFITABLE — keep running scans"
                        if _all_stats["avg"] >= 3 and _all_stats["wr"] >= 55 else
                    "🟢 Positive edge — refine entry criteria for better results"
                        if _all_stats["avg"] >= 0 and _all_stats["wr"] >= 45 else
                    "🟡 Mixed results — tighten filters (raise AI threshold)"
                        if _all_stats["avg"] >= -2 else
                    "🔴 Negative edge — review signal criteria urgently"
                )
                st.markdown(
                    f"<div style='background:#131722;border:1px solid {_oc}44;"
                    f"border-left:4px solid {_oc};border-radius:5px;padding:14px 18px;margin-top:16px'>"
                    f"<div style='font-family:Syne,sans-serif;font-weight:800;font-size:1rem;"
                    f"color:{_oc};margin-bottom:10px'>{_overall_verdict}</div>"
                    f"<div style='display:grid;grid-template-columns:repeat(7,1fr);gap:10px'>"
                    + "".join(
                        f"<div style='text-align:center'>"
                        f"<div style='font-size:.6rem;color:#434651;text-transform:uppercase;letter-spacing:.06em;margin-bottom:3px'>{_lbl}</div>"
                        f"<div style='font-family:Syne,sans-serif;font-weight:700;color:{_clr};font-size:.95rem'>{_val}</div>"
                        f"</div>"
                        for _lbl, _val, _clr in [
                            ("Total Signals",    str(_all_stats["n"]),                "#a855f7"),
                            ("Win Rate",         f"{_all_stats['wr']:.1f}%",          "#26a69a" if _all_stats["wr"]>=55 else "#f59e0b"),
                            ("Avg P&L",          f"{_all_stats['avg']:+.2f}%",        _oc),
                            ("Best Trade",       f"+{_all_stats['best']:.1f}%",       "#26a69a"),
                            ("Worst Trade",      f"{_all_stats['worst']:.1f}%",       "#ef5350"),
                            ("Profit Factor",    f"{_all_stats['pf']:.2f}" if _all_stats["pf"] else "—",
                                                 "#26a69a" if (_all_stats["pf"] or 0)>=1.5 else "#f59e0b"),
                            ("Avg Win / Loss",   f"{_all_stats['avg_win']:+.1f}% / {_all_stats['avg_los']:+.1f}%", "#787b86"),
                        ]
                    )
                    + "</div></div>",
                    unsafe_allow_html=True,
                )
                st.caption("Profit Factor > 1.5 = strategy is profitable. Win Rate > 55% + Avg P&L > 0 = reliable signal.")

            # Downloads
            _dl1, _dl2 = st.columns(2)
            with _dl1:
                st.download_button("⬇️ CSV", data=_tbl_df.to_csv(index=False).encode(),
                    file_name=f"tracker_{datetime.now().strftime('%Y%m%d_%H%M')}.csv",
                    mime="text/csv", key="tr_csv", use_container_width=True)
            with _dl2:
                try:
                    import io as _ioxl, openpyxl as _oxl
                    from openpyxl.styles import PatternFill, Font, Alignment
                    from openpyxl.utils import get_column_letter
                    _wb = _oxl.Workbook(); _ws = _wb.active; _ws.title="Tracker"
                    _hdr2 = list(_tbl_df.columns)
                    _ws.append(_hdr2)
                    for ci in range(1,len(_hdr2)+1):
                        _ws.cell(1,ci).fill = PatternFill("solid",fgColor="0D1117")
                        _ws.cell(1,ci).font = Font(bold=True,color="A855F7",size=9)
                        _ws.cell(1,ci).alignment = Alignment(horizontal="center")
                    _ws.row_dimensions[1].height=24
                    for rr_ in _tbl:
                        _ws.append(list(rr_.values()))
                        _ri=_ws.max_row
                        _fc_=("0D2B1E" if "TARGET" in str(rr_.get("Status",""))
                              else "2B0D0D" if "SL" in str(rr_.get("Status","")) else "0B0E11")
                        _fnc_=("26A69A" if "TARGET" in str(rr_.get("Status",""))
                               else "EF5350" if "SL" in str(rr_.get("Status","")) else "D1D4DC")
                        for ci in range(1,len(_hdr2)+1):
                            _ws.cell(_ri,ci).fill=PatternFill("solid",fgColor=_fc_)
                            _ws.cell(_ri,ci).font=Font(color=_fnc_,size=9)
                        _ws.row_dimensions[_ri].height=13
                    _cws2=[9,16,11,7,9,9,9,9,9,7,8,8,7,8,6,6,6,5,5,6,6,5,7,8,12,9,9,5,14,35,12]
                    for ci,w in enumerate(_cws2[:len(_hdr2)],1):
                        _ws.column_dimensions[get_column_letter(ci)].width=w
                    _ws.freeze_panes="A2"
                    _buf=_ioxl.BytesIO(); _wb.save(_buf); _buf.seek(0)
                    st.download_button("⬇️ Excel", data=_buf.getvalue(),
                        file_name=f"tracker_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        key="tr_xl", use_container_width=True)
                except Exception: pass

            st.markdown("<hr style='border-color:#2a3347;margin:14px 0'>", unsafe_allow_html=True)

            # ═══════════════════════════════════════════════════════════════
            # STRATEGY ANALYTICS — Which signals / strategies work best
            # ═══════════════════════════════════════════════════════════════
            st.markdown("""
            <div style='padding:12px 18px;background:linear-gradient(90deg,#38bdf820,#26a69a10,#0b0e11);
                        border-left:4px solid #38bdf8;border-radius:6px;margin-bottom:14px'>
              <div style='font-family:Syne,sans-serif;font-weight:800;font-size:1rem;color:#38bdf8'>
                📊 Strategy Performance Analytics — What is Working?
              </div>
              <div style='font-size:.74rem;color:#787b86;margin-top:3px'>
                Breakdown of signal performance by strategy · sector · AI score band · signal type · holding period
              </div>
            </div>""", unsafe_allow_html=True)

            # Build analytics only from signals that have P&L data
            _ana_rows = [r for r in _rows if r["_pnl_pct"] is not None]
            _closed   = [r for r in _rows if r["_status"] in ("TARGET_HIT","SL_HIT","CLOSED")]

            if len(_ana_rows) < 3:
                st.info("Not enough signals with live price data to analyse. Run a scan or refresh prices.", icon="📊")
            else:
                # ── Helper ────────────────────────────────────────────────
                def _stats(lst):
                    if not lst: return {}
                    pnls = [_safe_float(r["_pnl_pct"],0) for r in lst]
                    wins = [p for p in pnls if p > 0]
                    loss = [p for p in pnls if p < 0]
                    avg  = round(sum(pnls)/len(pnls),2)
                    wr   = round(len(wins)/len(pnls)*100,1)
                    best = round(max(pnls),2) if pnls else 0
                    worst= round(min(pnls),2) if pnls else 0
                    avg_w= round(sum(wins)/len(wins),2)  if wins else 0
                    avg_l= round(sum(loss)/len(loss),2)  if loss else 0
                    pf   = round(abs(avg_w*len(wins)/(avg_l*len(loss))),2) if loss and avg_l else None
                    return dict(n=len(lst),avg=avg,wr=wr,best=best,worst=worst,
                                avg_win=avg_w,avg_loss=avg_l,pf=pf,wins=len(wins),losses=len(loss))

                def _bar(avg, width=80):
                    col = "#26a69a" if avg>=0 else "#ef5350"
                    bw  = _safe_int(min(abs(avg)/15*width, width), 0)
                    return col, bw

                # ── Analytics tabs ────────────────────────────────────────
                _an_tab1, _an_tab2, _an_tab3, _an_tab4, _an_tab5 = st.tabs([
                    "🎯 By Signal Type",
                    "🏭 By Sector",
                    "🤖 By AI Score",
                    "⏱️ By Hold Period",
                    "📈 Overall Summary",
                ])

                # ══════════════════════════════════════════════════════════
                # TAB 1 — BY SIGNAL TYPE (top_signal text)
                # ══════════════════════════════════════════════════════════
                with _an_tab1:
                    st.markdown(
                        "<div style='font-size:.78rem;color:#787b86;margin-bottom:10px'>"
                        "Performance grouped by the <b style='color:#d1d4dc'>primary signal pattern</b> "
                        "that triggered the alert. Shows which candlestick / technical patterns deliver best returns.</div>",
                        unsafe_allow_html=True,
                    )
                    # Extract signal keyword
                    def _sig_key(sig):
                        sig = str(sig or "Unknown").strip()
                        for kw in ["SuperTrend","EMA Bull Stack","Full EMA","Golden Cross",
                                   "MACD Bull","RSI Oversold","Breakout","Volume Surge","Hammer",
                                   "Engulfing","Morning Star","Three White","Piercing",
                                   "BB Squeeze","Stochastic","ADX","Support Bounce","52W Low"]:
                            if kw.lower() in sig.lower():
                                return kw
                        return sig[:25] if sig else "Other"

                    _sig_groups: dict = {}
                    for r in _ana_rows:
                        k = _sig_key(r["_top_signal"])
                        _sig_groups.setdefault(k, []).append(r)

                    _sig_data = []
                    for k, grp in _sig_groups.items():
                        s = _stats(grp)
                        _sig_data.append((k, s))
                    _sig_data.sort(key=lambda x: (-x[1]["avg"], -x[1]["n"]))

                    # Header
                    st.markdown(
                        "<div style='display:grid;grid-template-columns:2fr 1fr 1fr 1fr 1fr 1fr 1.5fr;gap:4px;"
                        "padding:6px 10px;background:#1c2030;border-radius:4px;margin-bottom:4px;"
                        "font-size:.65rem;color:#434651;text-transform:uppercase;letter-spacing:.06em'>"
                        "<span>Signal Pattern</span><span style='text-align:center'>Count</span>"
                        "<span style='text-align:center'>Avg P&L</span><span style='text-align:center'>Win Rate</span>"
                        "<span style='text-align:center'>Best</span><span style='text-align:center'>Worst</span>"
                        "<span>Performance Bar</span></div>",
                        unsafe_allow_html=True,
                    )
                    for sig_name, s in _sig_data[:20]:
                        col_, bw_ = _bar(s["avg"])
                        verdict = ("🟢 Working" if s["avg"]>2 and s["wr"]>=55 else
                                   "🟡 Mixed"   if s["avg"]>0 else
                                   "🔴 Failing")
                        st.markdown(
                            f"<div style='display:grid;grid-template-columns:2fr 1fr 1fr 1fr 1fr 1fr 1.5fr;"
                            f"gap:4px;padding:7px 10px;background:#131722;border-radius:4px;"
                            f"border-left:3px solid {col_};margin-bottom:3px;align-items:center'>"
                            f"<span style='font-size:.78rem;color:#d1d4dc;font-weight:600'>{sig_name}</span>"
                            f"<span style='font-size:.75rem;color:#787b86;text-align:center'>{s['n']}</span>"
                            f"<span style='font-size:.8rem;font-weight:700;color:{col_};text-align:center'>{s['avg']:+.1f}%</span>"
                            f"<span style='font-size:.75rem;color:{'#26a69a' if s['wr']>=55 else '#ef5350'};text-align:center'>{s['wr']:.0f}%</span>"
                            f"<span style='font-size:.72rem;color:#26a69a;text-align:center'>+{s['best']:.1f}%</span>"
                            f"<span style='font-size:.72rem;color:#ef5350;text-align:center'>{s['worst']:.1f}%</span>"
                            f"<div style='display:flex;align-items:center;gap:5px'>"
                            f"<div style='height:5px;width:{bw_}px;background:{col_};border-radius:3px;min-width:3px'></div>"
                            f"<span style='font-size:.65rem;color:{col_}'>{verdict}</span></div>"
                            f"</div>",
                            unsafe_allow_html=True,
                        )

                # ══════════════════════════════════════════════════════════
                # TAB 2 — BY SECTOR
                # ══════════════════════════════════════════════════════════
                with _an_tab2:
                    st.markdown(
                        "<div style='font-size:.78rem;color:#787b86;margin-bottom:10px'>"
                        "Which <b style='color:#d1d4dc'>market sectors</b> are delivering positive returns "
                        "vs dragging the portfolio down.</div>",
                        unsafe_allow_html=True,
                    )
                    _sec_groups: dict = {}
                    for r in _ana_rows:
                        k = r["_sector"] or "N/A"
                        _sec_groups.setdefault(k, []).append(r)
                    _sec_data = [(k, _stats(g)) for k,g in _sec_groups.items()]
                    _sec_data.sort(key=lambda x: -x[1]["avg"])

                    _s1, _s2 = st.columns(2)
                    with _s1:
                        st.markdown("<div class='tv-label' style='margin-bottom:8px'>🟢 Best Performing Sectors</div>",
                                    unsafe_allow_html=True)
                        for sec, s in [x for x in _sec_data if x[1]["avg"]>=0][:8]:
                            col_, bw_ = _bar(s["avg"])
                            st.markdown(
                                f"<div style='display:flex;align-items:center;gap:8px;padding:6px 10px;"
                                f"background:#131722;border-left:3px solid {col_};border-radius:4px;margin-bottom:3px'>"
                                f"<span style='font-size:.78rem;color:#d1d4dc;min-width:120px'>{sec[:20]}</span>"
                                f"<span style='font-size:.68rem;color:#434651'>{s['n']} signals</span>"
                                f"<div style='height:5px;width:{bw_}px;background:{col_};border-radius:3px;flex-shrink:0'></div>"
                                f"<span style='font-family:Syne,sans-serif;font-weight:700;color:{col_};margin-left:auto'>{s['avg']:+.1f}%</span>"
                                f"<span style='font-size:.68rem;color:#26a69a'>{s['wr']:.0f}% WR</span>"
                                f"</div>",
                                unsafe_allow_html=True,
                            )
                    with _s2:
                        st.markdown("<div class='tv-label' style='margin-bottom:8px'>🔴 Underperforming Sectors</div>",
                                    unsafe_allow_html=True)
                        for sec, s in [x for x in _sec_data if x[1]["avg"]<0]:
                            col_, bw_ = _bar(s["avg"])
                            st.markdown(
                                f"<div style='display:flex;align-items:center;gap:8px;padding:6px 10px;"
                                f"background:#131722;border-left:3px solid {col_};border-radius:4px;margin-bottom:3px'>"
                                f"<span style='font-size:.78rem;color:#d1d4dc;min-width:120px'>{sec[:20]}</span>"
                                f"<span style='font-size:.68rem;color:#434651'>{s['n']} signals</span>"
                                f"<div style='height:5px;width:{bw_}px;background:{col_};border-radius:3px;flex-shrink:0'></div>"
                                f"<span style='font-family:Syne,sans-serif;font-weight:700;color:{col_};margin-left:auto'>{s['avg']:+.1f}%</span>"
                                f"<span style='font-size:.68rem;color:#ef5350'>{s['wr']:.0f}% WR</span>"
                                f"</div>",
                                unsafe_allow_html=True,
                            )

                # ══════════════════════════════════════════════════════════
                # TAB 3 — BY AI SCORE BAND
                # ══════════════════════════════════════════════════════════
                with _an_tab3:
                    st.markdown(
                        "<div style='font-size:.78rem;color:#787b86;margin-bottom:10px'>"
                        "Does a <b style='color:#d1d4dc'>higher AI confidence score</b> actually predict better outcomes? "
                        "This tells you whether to trust high-score signals more.</div>",
                        unsafe_allow_html=True,
                    )
                    _bands = [
                        ("90–100% (A+)",  lambda r: 90 <= r["_ai"] <= 100),
                        ("80–89%  (A)",   lambda r: 80 <= r["_ai"] < 90),
                        ("70–79%  (B+)",  lambda r: 70 <= r["_ai"] < 80),
                        ("60–69%  (B)",   lambda r: 60 <= r["_ai"] < 70),
                        ("50–59%  (C+)",  lambda r: 50 <= r["_ai"] < 60),
                        ("< 50%   (C)",   lambda r: r["_ai"] < 50),
                    ]
                    _band_data = []
                    for label, fn in _bands:
                        grp = [r for r in _ana_rows if fn(r)]
                        if grp:
                            _band_data.append((label, _stats(grp)))

                    # Header
                    st.markdown(
                        "<div style='display:grid;grid-template-columns:1.5fr 0.8fr 1fr 1fr 1fr 1fr 1fr 1.5fr;"
                        "gap:4px;padding:6px 10px;background:#1c2030;border-radius:4px;margin-bottom:4px;"
                        "font-size:.65rem;color:#434651;text-transform:uppercase;letter-spacing:.06em'>"
                        "<span>AI Band</span><span>Count</span><span>Avg P&L</span>"
                        "<span>Win Rate</span><span>Best</span><span>Worst</span>"
                        "<span>Profit Factor</span><span>Verdict</span></div>",
                        unsafe_allow_html=True,
                    )
                    for label, s in _band_data:
                        col_, _ = _bar(s["avg"])
                        pf_str = f"{s['pf']:.2f}" if s['pf'] else "—"
                        pf_col = "#26a69a" if s['pf'] and s['pf']>=1.5 else "#f59e0b" if s['pf'] and s['pf']>=1 else "#ef5350"
                        verdict = ("✅ High conviction" if s["avg"]>3 and s["wr"]>=60 else
                                   "✅ Working"         if s["avg"]>0 and s["wr"]>=50 else
                                   "⚠️ Mixed results"   if s["avg"]>-2 else
                                   "❌ Not predictive")
                        st.markdown(
                            f"<div style='display:grid;grid-template-columns:1.5fr 0.8fr 1fr 1fr 1fr 1fr 1fr 1.5fr;"
                            f"gap:4px;padding:7px 10px;background:#131722;border-left:3px solid {col_};"
                            f"border-radius:4px;margin-bottom:3px;align-items:center'>"
                            f"<span style='font-size:.78rem;color:#d1d4dc;font-weight:600'>{label}</span>"
                            f"<span style='font-size:.75rem;color:#787b86'>{s['n']}</span>"
                            f"<span style='font-size:.8rem;font-weight:700;color:{col_}'>{s['avg']:+.1f}%</span>"
                            f"<span style='font-size:.78rem;color:{'#26a69a' if s['wr']>=55 else '#ef5350'}'>{s['wr']:.0f}%</span>"
                            f"<span style='font-size:.72rem;color:#26a69a'>+{s['best']:.1f}%</span>"
                            f"<span style='font-size:.72rem;color:#ef5350'>{s['worst']:.1f}%</span>"
                            f"<span style='font-size:.75rem;color:{pf_col};font-weight:600'>{pf_str}</span>"
                            f"<span style='font-size:.72rem;color:{col_}'>{verdict}</span>"
                            f"</div>",
                            unsafe_allow_html=True,
                        )
                    st.caption("Profit Factor > 1.5 = strategy is profitable. Win Rate > 55% + Avg P&L > 0 = reliable signal.")

                # ══════════════════════════════════════════════════════════
                # TAB 4 — BY HOLDING PERIOD
                # ══════════════════════════════════════════════════════════
                with _an_tab4:
                    st.markdown(
                        "<div style='font-size:.78rem;color:#787b86;margin-bottom:10px'>"
                        "How long you hold trades affects returns. This shows the <b style='color:#d1d4dc'>optimal hold window</b> "
                        "based on actual results from your signals.</div>",
                        unsafe_allow_html=True,
                    )
                    _hold_bands = [
                        ("1–3 days   (Intraday/Quick)", lambda r: 1 <= r["_hold_days"] <= 3),
                        ("4–7 days   (Short swing)",    lambda r: 4 <= r["_hold_days"] <= 7),
                        ("8–14 days  (Medium swing)",   lambda r: 8 <= r["_hold_days"] <= 14),
                        ("15–30 days (Monthly hold)",   lambda r: 15 <= r["_hold_days"] <= 30),
                        ("> 30 days  (Position trade)", lambda r: r["_hold_days"] > 30),
                    ]
                    _hold_data = [(lbl, _stats([r for r in _ana_rows if fn(r)])) for lbl,fn in _hold_bands if [r for r in _ana_rows if fn(r)]]

                    for lbl, s in _hold_data:
                        col_, bw_ = _bar(s["avg"])
                        _hc1, _hc2, _hc3 = st.columns([2,3,1])
                        with _hc1:
                            st.markdown(
                                f"<div style='background:#131722;border-left:3px solid {col_};"
                                f"border-radius:4px;padding:8px 12px'>"
                                f"<div style='font-size:.75rem;color:#d1d4dc;font-weight:600;margin-bottom:4px'>{lbl}</div>"
                                f"<div style='font-size:.7rem;color:#434651'>{s['n']} signals</div>"
                                f"</div>",
                                unsafe_allow_html=True,
                            )
                        with _hc2:
                            st.markdown(
                                f"<div style='display:flex;gap:6px;flex-wrap:wrap;align-items:center;height:100%;padding-top:4px'>"
                                + "".join(
                                    f"<div style='background:#1c2030;border-radius:3px;padding:4px 8px;text-align:center'>"
                                    f"<div style='font-size:.58rem;color:#434651;text-transform:uppercase'>{ll}</div>"
                                    f"<div style='font-size:.8rem;font-weight:700;color:{lc}'>{lv}</div></div>"
                                    for ll,lv,lc in [
                                        ("Avg P&L",   f"{s['avg']:+.1f}%",    col_),
                                        ("Win Rate",  f"{s['wr']:.0f}%",      "#26a69a" if s["wr"]>=55 else "#ef5350"),
                                        ("Best",      f"+{s['best']:.1f}%",   "#26a69a"),
                                        ("Worst",     f"{s['worst']:.1f}%",   "#ef5350"),
                                        ("Avg Win",   f"+{s['avg_win']:.1f}%","#26a69a"),
                                        ("Avg Loss",  f"{s['avg_loss']:.1f}%","#ef5350"),
                                    ]
                                )
                                + "</div>",
                                unsafe_allow_html=True,
                            )
                        with _hc3:
                            verdict = ("✅ BEST" if s["avg"]>=3 and s["wr"]>=55 else
                                       "✅ Good" if s["avg"]>=0 else "❌ Avoid")
                            vc = "#26a69a" if "BEST" in verdict or "Good" in verdict else "#ef5350"
                            st.markdown(
                                f"<div style='background:{vc}15;border-radius:4px;padding:8px;text-align:center;height:100%;display:flex;align-items:center;justify-content:center'>"
                                f"<div style='font-size:.8rem;font-weight:700;color:{vc}'>{verdict}</div></div>",
                                unsafe_allow_html=True,
                            )

                # ══════════════════════════════════════════════════════════
                # TAB 5 — OVERALL SUMMARY
                # ══════════════════════════════════════════════════════════
                with _an_tab5:
                    _all_s = _stats(_ana_rows)
                    _open_s= _stats([r for r in _rows if r["_status"]=="OPEN" and r["_pnl_pct"] is not None])
                    _done_s= _stats([r for r in _rows if r["_status"] in ("TARGET_HIT","SL_HIT","CLOSED")])

                    # Top 5 and Bottom 5 performers
                    _sorted_by_pnl = sorted(_ana_rows, key=lambda r: _safe_float(r["_pnl_pct"],0), reverse=True)
                    _top5 = _sorted_by_pnl[:5]
                    _bot5 = list(reversed(_sorted_by_pnl[-5:]))

                    # Overall scorecards
                    _ov_cols = st.columns(3)
                    for _ovc, _ovtitle, _ovs, _ovcol in [
                        (_ov_cols[0], "📊 All Signals",    _all_s,  "#a855f7"),
                        (_ov_cols[1], "🔵 Open Positions", _open_s, "#38bdf8"),
                        (_ov_cols[2], "🏁 Closed Trades",  _done_s, "#26a69a"),
                    ]:
                        if not _ovs:
                            _ovc.info("No data")
                            continue
                        pf_s = f"{_ovs['pf']:.2f}" if _ovs.get('pf') else "—"
                        pf_c = "#26a69a" if _ovs.get('pf') and _ovs['pf']>=1.5 else "#f59e0b"
                        _ovc.markdown(
                            f"<div style='background:#131722;border-top:3px solid {_ovcol};"
                            f"border-radius:5px;padding:14px 16px'>"
                            f"<div style='font-family:Syne,sans-serif;font-weight:700;font-size:.85rem;"
                            f"color:{_ovcol};margin-bottom:10px'>{_ovtitle}</div>"
                            + "".join(
                                f"<div style='display:flex;justify-content:space-between;padding:4px 0;"
                                f"border-bottom:1px solid #1c2030'>"
                                f"<span style='font-size:.72rem;color:#434651'>{ll}</span>"
                                f"<span style='font-size:.78rem;font-weight:600;color:{lc}'>{lv}</span></div>"
                                for ll,lv,lc in [
                                    ("Total Signals",    str(_ovs["n"]),              "#d1d4dc"),
                                    ("Avg P&L",          f"{_ovs['avg']:+.2f}%",      "#26a69a" if _ovs["avg"]>=0 else "#ef5350"),
                                    ("Win Rate",         f"{_ovs['wr']:.1f}%",        "#26a69a" if _ovs["wr"]>=55 else "#ef5350"),
                                    ("Winners",          str(_ovs["wins"]),           "#26a69a"),
                                    ("Losers",           str(_ovs["losses"]),         "#ef5350"),
                                    ("Best Signal",      f"+{_ovs['best']:.1f}%",     "#26a69a"),
                                    ("Worst Signal",     f"{_ovs['worst']:.1f}%",     "#ef5350"),
                                    ("Avg Win",          f"+{_ovs['avg_win']:.1f}%",  "#26a69a"),
                                    ("Avg Loss",         f"{_ovs['avg_loss']:.1f}%",  "#ef5350"),
                                    ("Profit Factor",    pf_s,                         pf_c),
                                ]
                            )
                            + "</div>",
                            unsafe_allow_html=True,
                        )

                    st.markdown("<div style='margin:14px 0'></div>", unsafe_allow_html=True)

                    # Top/Bottom performers
                    _tb1, _tb2 = st.columns(2)
                    with _tb1:
                        st.markdown("<div class='tv-label' style='margin-bottom:8px'>🏆 Top 5 Best Signals</div>",
                                    unsafe_allow_html=True)
                        for i,r in enumerate(_top5):
                            pc = _safe_float(r["_pnl_pct"],0)
                            st.markdown(
                                f"<div style='display:flex;align-items:center;gap:8px;padding:6px 10px;"
                                f"background:#131722;border-left:3px solid #26a69a;border-radius:4px;margin-bottom:3px'>"
                                f"<span style='color:#434651;font-size:.72rem;min-width:18px'>#{i+1}</span>"
                                f"<span style='font-family:Syne,sans-serif;font-weight:700;color:#d1d4dc;min-width:80px'>{r['_sym']}</span>"
                                f"<span style='font-size:.7rem;color:#434651'>{r['_scan_date']}</span>"
                                f"<span style='font-size:.7rem;color:#434651;flex:1'>{r['_sector'][:14]}</span>"
                                f"<span style='font-family:Syne,sans-serif;font-weight:800;color:#26a69a'>{pc:+.1f}%</span>"
                                f"</div>",
                                unsafe_allow_html=True,
                            )
                    with _tb2:
                        st.markdown("<div class='tv-label' style='margin-bottom:8px'>📉 Bottom 5 Worst Signals</div>",
                                    unsafe_allow_html=True)
                        for i,r in enumerate(_bot5):
                            pc = _safe_float(r["_pnl_pct"],0)
                            st.markdown(
                                f"<div style='display:flex;align-items:center;gap:8px;padding:6px 10px;"
                                f"background:#131722;border-left:3px solid #ef5350;border-radius:4px;margin-bottom:3px'>"
                                f"<span style='color:#434651;font-size:.72rem;min-width:18px'>#{i+1}</span>"
                                f"<span style='font-family:Syne,sans-serif;font-weight:700;color:#d1d4dc;min-width:80px'>{r['_sym']}</span>"
                                f"<span style='font-size:.7rem;color:#434651'>{r['_scan_date']}</span>"
                                f"<span style='font-size:.7rem;color:#434651;flex:1'>{r['_sector'][:14]}</span>"
                                f"<span style='font-family:Syne,sans-serif;font-weight:800;color:#ef5350'>{pc:+.1f}%</span>"
                                f"</div>",
                                unsafe_allow_html=True,
                            )

                    # Key insight box
                    _best_sig  = max(_sig_groups.items(), key=lambda x: _stats(x[1])["avg"]) if _sig_groups else None
                    _worst_sig = min(_sig_groups.items(), key=lambda x: _stats(x[1])["avg"]) if _sig_groups else None
                    _best_sec  = max(_sec_groups.items(), key=lambda x: _stats(x[1])["avg"]) if "_sec_groups" in dir() and _sec_groups else None
                    if _best_sig or _best_sec:
                        st.markdown(
                            f"<div style='margin-top:12px;padding:12px 16px;background:#131722;"
                            f"border-left:4px solid #f59e0b;border-radius:5px'>"
                            f"<div style='font-family:Syne,sans-serif;font-weight:700;color:#f59e0b;"
                            f"margin-bottom:8px'>💡 Key Insights from Your Signals</div>"
                            f"<div style='font-size:.78rem;color:#787b86;line-height:1.9'>"
                            + (f"✅ <b style='color:#26a69a'>Best strategy:</b> <b style='color:#d1d4dc'>{_best_sig[0]}</b> — avg {_stats(_best_sig[1])['avg']:+.1f}% across {_stats(_best_sig[1])['n']} signals<br>" if _best_sig else "")
                            + (f"❌ <b style='color:#ef5350'>Worst strategy:</b> <b style='color:#d1d4dc'>{_worst_sig[0]}</b> — avg {_stats(_worst_sig[1])['avg']:+.1f}% — reduce position size or skip<br>" if _worst_sig else "")
                            + (f"📊 <b style='color:#38bdf8'>Overall win rate:</b> {_all_s['wr']:.1f}% — {'above 50%, strategy has edge ✅' if _all_s['wr']>=50 else 'below 50%, review signal filters ⚠️'}<br>" if _all_s else "")
                            + (f"💰 <b style='color:#38bdf8'>Avg profit factor:</b> {_all_s['pf']:.2f} — {'healthy edge, keep scaling ✅' if _all_s.get('pf') and _all_s['pf']>=1.5 else 'improve risk management'}" if _all_s.get('pf') else "")
                            + "</div></div>",
                            unsafe_allow_html=True,
                        )


    # ══════════════════════════════════════════════════════════════════════
    with tabs[10]:
        sc_c1, sc_c2, sc_c3 = st.columns([1,1,1])
        with sc_c1:
            sort_by=st.selectbox("Sort by",
                ["AI Score","Market Score","Pattern Score","Composite Score","RSI","ADX","Volume"],
                label_visibility="collapsed", key="scr_sort")
        with sc_c2:
            min_ai=st.slider("Min AI Score", 0, 100, 0, 5, key="scr_ai")
        with sc_c3:
            fo_only=st.toggle("F&O Only", value=False, key="scr_fo")

        smap={"AI Score":lambda r:-r["ai"]["ai_pct"],"Market Score":lambda r:-r["mkt"]["pct"],
              "Pattern Score":lambda r:-r["pat_conf"],"Composite Score":lambda r:-abs(r["score"]),
              "RSI":lambda r:r["rsi"],"ADX":lambda r:-r["adx"],"Volume":lambda r:-r["vol_ratio"]}
        filtered=[r for r in alerts if r["ai"]["ai_pct"]>=min_ai and (not fo_only or r["is_fo"])]
        sorted_a=sorted(filtered, key=smap.get(sort_by,smap["AI Score"]))

        rows=[]
        for r in sorted_a:
            ltr,_=_d_grade(r["ai"]["ai_pct"])
            stl2=r["levels"]["short_term"]; ltl2=r["levels"]["long_term"]
            rows.append({"Symbol":r["symbol"],"Grade":ltr,"AI%":round(r["ai"]["ai_pct"],1),
                "Mkt%":round(r["mkt"]["pct"],1),"Pat%":round(r["pat_conf"]*100,1),
                "Score":round(r["score"],4),"RSI":round(r["rsi"],1),"ADX":round(r["adx"],1),
                "Vol×":round(r["vol_ratio"],2),"ATR%":round(r["atr_pct"],2),
                "F&O":"✅" if r["is_fo"] else "—","Price ₹":r["last_close"],
                "ST Entry":stl2["entry"],"ST Target":stl2["tp"],"ST SL":stl2["sl"],
                "ST R:R":stl2["rr_str"],"LT Target":ltl2["tp"],"LT R:R":ltl2["rr_str"],
                "Top Signal":r["hits"][0][1] if r["hits"] else "—","Sector":r.get("sector","N/A"),
                "Indices":r["indices"]})

        df_sc=pd.DataFrame(rows)
        st.markdown(f"<div class='tv-label' style='margin-bottom:8px'>{len(df_sc)} signals shown</div>",
                    unsafe_allow_html=True)
        st.dataframe(df_sc, use_container_width=True, height=480, hide_index=True,
            column_config={
                "AI%":    st.column_config.ProgressColumn("AI%",   min_value=0,max_value=100,format="%.1f%%"),
                "Mkt%":   st.column_config.ProgressColumn("Mkt%",  min_value=0,max_value=100,format="%.1f%%"),
                "Pat%":   st.column_config.ProgressColumn("Pat%",  min_value=0,max_value=100,format="%.1f%%"),
                "Price ₹":st.column_config.NumberColumn("Price ₹", format="₹%.2f"),
                "ST Entry":st.column_config.NumberColumn("ST Entry",format="₹%.2f"),
                "ST Target":st.column_config.NumberColumn("ST Target",format="₹%.2f"),
                "ST SL":  st.column_config.NumberColumn("ST SL",   format="₹%.2f"),
                "LT Target":st.column_config.NumberColumn("LT Target",format="₹%.2f"),
                "Score":  st.column_config.NumberColumn("Score",    format="%.4f"),
            })

        st.markdown("<div class='tv-section' style='margin-top:8px'>🌐 TradingView NSE Screener</div>",
                    unsafe_allow_html=True)
        components.html(tv_screener(), height=620, scrolling=False)

        csv=df_sc.to_csv(index=False).encode()
        st.download_button("⬇️  Download CSV", data=csv,
            file_name=f"nse_signals_{datetime.now().strftime('%Y%m%d_%H%M')}.csv", mime="text/csv")

    # ══════════════════════════════════════════════════════════════════════
    # TAB 11 — NEWS & CALENDAR
    # ══════════════════════════════════════════════════════════════════════
    with tabs[11]:
        nc_sym=st.selectbox("Select Stock for News",
            [r["symbol"] for r in alerts], key="news_sym")
        nn1, nn2 = st.columns([1.2, 1])
        with nn1:
            st.markdown("<div class='tv-section'>📰 Latest News</div>", unsafe_allow_html=True)
            components.html(tv_news(nc_sym), height=520, scrolling=False)
        with nn2:
            st.markdown("<div class='tv-section'>📅 India Economic Calendar</div>", unsafe_allow_html=True)
            components.html(tv_economic_calendar(), height=520, scrolling=False)

    # ── Footer ─────────────────────────────────────────────────────────────
    st.markdown("""
    <hr class='tv-divider' style='margin-top:30px'>
    <div style='text-align:center;font-size:.65rem;color:#434651;padding:10px 0;letter-spacing:.05em'>
      ⚠️  Research &amp; Educational Use Only · Not Financial Advice ·
      Consult a SEBI-registered advisor before investing.
    </div>""", unsafe_allow_html=True)


# ══════════════════════════════════════════════════════════════════════════════
# §22  STREAMLIT SIDEBAR + RUNNER
# ══════════════════════════════════════════════════════════════════════════════

def _streamlit_sidebar() -> tuple:
    with st.sidebar:
        st.markdown("""
        <div style='text-align:center;padding:16px 0 8px'>
          <div style='font-family:Syne,sans-serif;font-weight:800;font-size:1.3rem;color:#38bdf8;letter-spacing:.1em'>📈 NSE PRO</div>
          <div style='font-size:.65rem;color:#434651;letter-spacing:.1em'>SWING TRADER v10</div>
        </div>""", unsafe_allow_html=True)
        st.divider()

        st.markdown("<p style='font-size:.65rem;letter-spacing:.1em;text-transform:uppercase;color:#787b86;margin-top:10px'>Universe</p>", unsafe_allow_html=True)
        group_sel=st.selectbox("Index Group",
            ["All Groups","NIFTY 50","NIFTY BANK","NIFTY IT","NIFTY ENERGY",
             "NIFTY AUTO","NIFTY INFRA","FO STOCKS","NIFTY NEXT 50","NIFTY MIDCAP 100"],
            label_visibility="collapsed")
        custom_syms=st.text_input("Or enter symbols",placeholder="TCS,INFY,HDFCBANK",
                                   label_visibility="collapsed")
        st.markdown("<p style='font-size:.65rem;letter-spacing:.1em;text-transform:uppercase;color:#787b86;margin-top:10px'>Signal</p>", unsafe_allow_html=True)
        threshold=st.slider("Score Threshold",0.08,0.35,0.16,0.01)
        min_rr=st.slider("Min Risk:Reward",1.0,3.0,1.2,0.1)
        st.markdown("<p style='font-size:.65rem;letter-spacing:.1em;text-transform:uppercase;color:#787b86;margin-top:10px'>Volume</p>", unsafe_allow_html=True)
        min_vol=st.select_slider("Min Daily Volume",
            options=[500_000,750_000,1_000_000,1_500_000,2_000_000,3_000_000],
            value=750_000,format_func=lambda x:f"{x/1e5:.1f}L shares",
            label_visibility="collapsed")
        min_tv2=st.slider("Min Traded Value (₹Cr/d)",0.5,20.0,2.0,0.5)
        period=st.select_slider("Lookback Period",["3mo","4mo","6mo","8mo","1y"],value="8mo",
                                 label_visibility="collapsed")
        capital_l=st.number_input("Portfolio ₹ (Lakhs)",min_value=1.0,max_value=1000.0,
                                   value=10.0,step=1.0,label_visibility="collapsed")
        capital=capital_l*1e5
        st.session_state["capital_val"]=capital
        top_n=st.slider("Top N Alerts",3,20,10)
        st.divider()
        run_btn=st.button("🚀  Run Scan", use_container_width=True)
        st.markdown("<div style='font-size:.62rem;text-align:center;color:#434651;margin-top:6px'>Results cached 1 hour</div>",
                    unsafe_allow_html=True)

    group_key="" if group_sel=="All Groups" else group_sel
    syms_key=custom_syms.strip()

    cfg=Cfg()
    cfg.use_sample=False; cfg.live_period=period; cfg.top_n=top_n
    cfg.min_avg_vol=min_vol; cfg.min_traded_val_cr=min_tv2; cfg.min_rr=min_rr
    cfg.base_threshold=threshold; cfg.bear_threshold=threshold+0.08
    cfg.capital=capital; cfg.output_dir=Path("nse_v10_output")

    if syms_key:
        cfg.symbols=[s.strip().upper() for s in syms_key.split(",") if s.strip()]
    elif group_key:
        gk=group_key.strip().upper()
        matched=[sl for grp,sl in _UNIVERSE.items() if gk in grp.upper()]
        if matched:
            cfg.symbols=sorted({s for sub in matched for s in sub}-_SKIP_SYMBOLS)

    return cfg, run_btn


def _run_scan_cached(cfg: Cfg):
    import glob
    buf=io.StringIO()
    with contextlib.redirect_stdout(buf), contextlib.redirect_stderr(buf):
        alerts_out, bt_out = run(cfg)
    nifty_out=nifty50_state()
    feat_df_out=pd.DataFrame()
    try:
        cfg.output_dir.mkdir(parents=True,exist_ok=True)
        csvs=sorted(glob.glob(str(cfg.output_dir/"alerts_*.csv")))
        if csvs: feat_df_out=pd.read_csv(csvs[-1])
    except Exception:
        pass
    # Auto-persist signals to tracker (from March 20 2026 onwards)
    try:
        _tracker_ingest(alerts_out)
    except Exception:
        pass
    return alerts_out, bt_out, nifty_out, feat_df_out


def _streamlit_main():
    cfg, run_btn = _streamlit_sidebar()
    cache_key=(str(sorted(cfg.symbols)),cfg.base_threshold,cfg.min_avg_vol,
               cfg.min_traded_val_cr,cfg.min_rr,cfg.live_period,cfg.top_n)

    if run_btn or "scan_data" not in st.session_state or st.session_state.get("scan_key")!=cache_key:
        with st.spinner("🔴  Running live scan..."):
            alerts,bt,nifty,feat_df=_run_scan_cached(cfg)
        st.session_state["scan_data"]=(alerts,bt,nifty,feat_df)
        st.session_state["scan_key"]=cache_key
    else:
        alerts,bt,nifty,feat_df=st.session_state["scan_data"]

    if alerts is None:
        st.error("Scan failed — check yfinance is installed."); return

    run_dashboard(alerts,bt,nifty,feat_df)


# ══════════════════════════════════════════════════════════════════════════════
# §23  TERMINAL CLI
# ══════════════════════════════════════════════════════════════════════════════

def main():
    p=argparse.ArgumentParser(
        description="NSE Swing Trader v10.0 — Unified (terminal + Streamlit)",
        formatter_class=argparse.RawTextHelpFormatter)
    p.add_argument("--symbols",   type=str,   default="")
    p.add_argument("--group",     type=str,   default="")
    p.add_argument("--top-n",     type=int,   default=10)
    p.add_argument("--sample",    action="store_true")
    p.add_argument("--prices-csv",type=Path,  default=Path("data/prices.csv"))
    p.add_argument("--output-dir",type=Path,  default=Path("nse_v10_output"))
    p.add_argument("--period",    type=str,   default="8mo")
    p.add_argument("--min-vol",   type=int,   default=1_500_000)
    p.add_argument("--min-tv",    type=float, default=5.0)
    p.add_argument("--min-rr",    type=float, default=1.5)
    p.add_argument("--threshold", type=float, default=0.22)
    p.add_argument("--no-fund",   action="store_true")
    p.add_argument("--capital",   type=float, default=1_000_000)
    a=p.parse_args()
    cfg=Cfg(); cfg.use_sample=a.sample; cfg.output_dir=a.output_dir
    cfg.prices_csv=a.prices_csv; cfg.live_period=a.period; cfg.top_n=a.top_n
    cfg.min_avg_vol=a.min_vol; cfg.min_traded_val_cr=a.min_tv; cfg.min_rr=a.min_rr
    cfg.base_threshold=a.threshold; cfg.fetch_fundamentals=not a.no_fund; cfg.capital=a.capital
    if a.symbols:
        cfg.symbols=[s.strip().upper() for s in a.symbols.split(",") if s.strip()]
    elif a.group:
        gk=a.group.strip().upper()
        matched=[sl for grp,sl in _UNIVERSE.items() if gk in grp.upper()]
        if matched:
            cfg.symbols=sorted({s for sub in matched for s in sub}-_SKIP_SYMBOLS)
        else:
            print(f"Group '{a.group}' not found."); sys.exit(1)
    run(cfg)


# ══════════════════════════════════════════════════════════════════════════════
# §24  ENTRY POINT DISPATCHER
# ══════════════════════════════════════════════════════════════════════════════

if _STREAMLIT:
    _streamlit_main()
elif __name__ == "__main__":
    main()
